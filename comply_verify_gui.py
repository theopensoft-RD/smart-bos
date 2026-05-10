#!/usr/bin/env python3
"""
Comply Verify GUI — visual verification tool for Smart Plant 1 Comply spec.

Reads:
  - output/Comply spec Smart Plant 1.xlsx
  - output/**/*.pdf (catalog PDFs with annotations)
  - TOR/*.pdf (optional)
  - BOQ/*.xlsx (optional)

Usage:
  python3 comply_verify_gui.py
  → open http://127.0.0.1:5173 in browser

Persistence:
  output/verification_status.json
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import urllib.parse
import webbrowser
from datetime import datetime
from pathlib import Path

import openpyxl
from flask import Flask, Response, abort, jsonify, render_template_string, request, send_file

try:
    import fitz  # PyMuPDF
except ImportError:
    sys.stderr.write("PyMuPDF (fitz) is required. Install: pip3 install --user pymupdf\n")
    sys.exit(1)

try:
    from PIL import Image, ImageDraw
except ImportError:
    sys.stderr.write("Pillow is required. Install: pip3 install --user pillow\n")
    sys.exit(1)

from app import catalog, core, learning
from app import database as db

ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "output"


def _load_dotenv(path: Path) -> None:
    """Minimal .env loader so we don't pull in python-dotenv as a dep.
    Lines like KEY=value (no quotes needed). Skips comments + blanks.
    Existing env vars take precedence (so shell exports win)."""
    if not path.exists():
        return
    try:
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            os.environ.setdefault(k, v)
    except Exception as e:
        sys.stderr.write(f"[_load_dotenv] {e}\n")

# Layout detection — supports both:
#   1. comply-module/ layout (current): TOR, BOQ, scripts, _versions sit
#      alongside comply_verify_gui.py (PROJECT == ROOT)
#   2. Legacy Tools/ layout: TOR, BOQ live in the parent (PROJECT ==
#      ROOT.parent)
PROJECT = ROOT if (ROOT / "TOR").exists() else ROOT.parent
TOR_DIR = PROJECT / "TOR"
BOQ_DIR = PROJECT / "BOQ"

# Project-level snapshot system (scripts/version.py + _versions/snapshots/)
SCRIPTS_DIR    = ROOT / "scripts"
VERSIONS_DIR   = ROOT / "_versions"
SNAPS_DIR      = VERSIONS_DIR / "snapshots"
VERSION_SCRIPT = SCRIPTS_DIR / "version.py"

XLSX_PATH = OUTPUT / "Comply spec Smart Plant 1.xlsx"
STATUS_PATH = OUTPUT / "verification_status.json"

# SQLite database — local cache + audit log + FTS search index. Lives in
# the project; survives across boots; safe to delete (rebuilt from xlsx +
# filesystem on next boot).
DB_DIR  = ROOT / "_db"
DB_PATH = DB_DIR / "comply.db"

# Project singleton — OOP entry point that owns everything else. The legacy
# globals (ROWS, PDF_INDEX, etc.) coexist for now; new code should prefer
# this object. ``project.rows`` mirrors ROWS automatically via load_rows().
PROJECT_OBJ = core.Project(ROOT)

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False


# ---------------------------------------------------------------------------
# Indexing
# ---------------------------------------------------------------------------

# ref-key (e.g. "5.1.2-2") → list[Path]
PDF_INDEX: dict[str, list[Path]] = {}
# section-key (e.g. "5.1.2") → list[Path]   (section-level matches, no -N)
SECTION_INDEX: dict[str, list[Path]] = {}
# model-substring → list[Path]   (for Col D that is just a model name)
MODEL_INDEX: dict[str, list[Path]] = {}
# rows from xlsx
ROWS: list[dict] = []
# header info (TOR PDF, BOQ xlsx)
EXTRA_REFS: dict = {}


def build_pdf_index() -> None:
    """Index PDFs by ref-key (5.1.2-1), section (5.1.6.2), and model substrings.

    Folder convention:
      - Parent rack folder: '{section}.-{N}'   (e.g. '5.1.2.-1', '5.2.1.-9')
      - Sub-item PDFs sit beside parent in same folder: '{section}.M-{n} ...pdf'
      - Sensor/LED/เสา folder: '{section}. {thai-name}' (no -N suffix)
    """
    PDF_INDEX.clear()
    SECTION_INDEX.clear()
    MODEL_INDEX.clear()
    if not OUTPUT.exists():
        return
    for pdf in OUTPUT.rglob("*.pdf"):
        if pdf.name.startswith("~"):
            continue
        if pdf.name == "Comply spec Smart Plant 1.pdf":
            continue
        stem = pdf.stem
        parent = pdf.parent.name

        # Folder pattern A: "{section}.-{N}" — parent rack folder
        # (allow optional whitespace before the dash, e.g. "5.1.3.1. -1")
        fm = re.match(r"^(\d+(?:\.\d+){1,3})\.?\s*-(\d+)\b", parent)
        if fm:
            sec, n = fm.group(1), fm.group(2)
            key_parent = f"{sec}-{n}"
            # Is this the parent file? filename ends with "-N"
            tail = re.search(r"-(\d+)$", stem)
            if tail and tail.group(1) == n:
                PDF_INDEX.setdefault(key_parent, []).insert(0, pdf)
            else:
                # Sub-item PDF — filename starts with sub-section then -n
                im = re.match(r"^(\d+(?:\.\d+){2,3})-(\d+)\b", stem)
                if im:
                    PDF_INDEX.setdefault(f"{im.group(1)}-{im.group(2)}", []).append(pdf)
                else:
                    PDF_INDEX.setdefault(key_parent, []).append(pdf)
        else:
            # Folder pattern B: "{section}. {thai-name}" — section-only ref (sensor/LED/เสา)
            sm = re.match(r"^(\d+(?:\.\d+){1,3})", parent)
            if sm:
                SECTION_INDEX.setdefault(sm.group(1), []).append(pdf)
            else:
                # Fallback: index by stem section
                sm2 = re.match(r"^(\d+(?:\.\d+){1,3})", stem)
                if sm2:
                    SECTION_INDEX.setdefault(sm2.group(1), []).append(pdf)

        # Model index — for Col D that names just a model (e.g. "UFC9312A")
        text = f"{parent} {stem}"
        for tok in re.findall(r"[A-Za-z][\w\-]{2,}", text):
            if len(tok) >= 3 and not tok.isdigit():
                MODEL_INDEX.setdefault(tok.lower(), []).append(pdf)


def parse_col_d(d: str | None) -> dict:
    """Parse Col D — return dict with type + extracted fields."""
    out: dict = {"raw": d or "", "type": "empty"}
    if not d:
        return out
    s = d.strip()
    out["raw"] = s
    # Type classification
    if s.startswith("ยี่ห้อ"):
        out["type"] = "brand_model"
        m = re.match(r"ยี่ห้อ\s+(.+?)\s+รุ่น\s+(.+)$", s)
        if m:
            out["brand"] = m.group(1).strip()
            out["model"] = m.group(2).strip()
    elif s.startswith("เทียบเท่าข้อกำหนด"):
        out["type"] = "equivalent"
    elif s.startswith("สูงกว่าข้อกำหนด"):
        out["type"] = "higher"
    elif s.startswith("ยินดีปฏิบัติ"):
        out["type"] = "commitment"
    elif re.match(r"^\d+(?:\.\d+){2,3}", s):
        out["type"] = "filename_format"
    else:
        # fall back — could be model-only string (e.g., "UFC9312A")
        out["type"] = "model_only"

    # Extract reference fields if present
    m = re.search(r"เอกสาร\s+(\d+(?:\.\d+){1,3}(?:-\d+)?)", s)
    if m:
        out["ref"] = m.group(1)
    m = re.search(r"หน้า\s+(\d+)", s)
    if m:
        out["page"] = int(m.group(1))
    m = re.search(r"ข้อ\s+(\d+)\)\s*ข้อย่อย\s+(\d+)", s)
    if m:
        out["item"] = int(m.group(1))
        out["subitem"] = int(m.group(2))
    else:
        m = re.search(r"ข้อ\s+(\d+)\)", s)
        if m:
            out["item"] = int(m.group(1))
        m = re.search(r"ข้อย่อย\s+(\d+)", s)
        if m:
            out["subitem"] = int(m.group(1))

    return out


def _section_to_folder_key(section: str) -> str | None:
    """Map a section number to the PDF_INDEX key for its folder.

    "5.2.1.9" → "5.2.1-9"  (single-row item under 5.2.1.-9/ folder)
    "5.1.1.2" → "5.1.1-2"  (parent rack folder 5.1.1.-2/)
    """
    parts = section.split(".")
    if len(parts) < 3:
        return None
    return f"{'.'.join(parts[:-1])}-{parts[-1]}"


def _pick_best_in_folder(paths: list[Path], section: str, raw: str) -> Path | None:
    """When PDF_INDEX[key] has multiple PDFs, prefer the one whose filename
    matches Col D's tokens most distinctively (avoids picking sub-item PDFs
    when the row points to the parent, or vice versa)."""
    if not paths:
        return None
    if len(paths) == 1:
        return paths[0]
    # Distinctive tokens from Col D (English/numeric, ≥4 chars or has digit)
    raw_tokens = []
    for tok in re.findall(r"[A-Za-z][\w\-]{2,}", raw or ""):
        if len(tok) >= 4 or any(c.isdigit() for c in tok):
            raw_tokens.append(tok.lower())
    # Score each PDF by token overlap with its filename
    scored = []
    for p in paths:
        stem_l = p.stem.lower()
        score = sum(1 for t in raw_tokens if t in stem_l)
        # Bonus: filename starts with the exact section prefix
        if p.stem.startswith(section + " ") or p.stem.startswith(section + "."):
            score += 5
        # Bonus: filename ends with `-N` matching last section digit (parent file)
        if section.split(".")[-1] in p.stem.rsplit("-", 1)[-1]:
            score += 1
        scored.append((score, p))
    scored.sort(key=lambda x: (-x[0], str(x[1])))
    return scored[0][1]


def _ref_to_folder_keys(ref: str) -> list[str]:
    """Yield candidate PDF_INDEX/SECTION_INDEX keys for a Col D ref.

    Users mix two conventions:
      • dash form  "5.1.1-2"   → PDF_INDEX["5.1.1-2"]   (parent rack folder)
      • dot form   "5.1.1.2"   → folder "5.1.1.-2/"     → "5.1.1-2"
      • dot form   "5.1.6.2"   → SECTION_INDEX["5.1.6.2"]  (sensor/LED)
    The clone/auto-annotate scripts emit dot form; the legacy convention is
    dash form. Try both.
    """
    if not ref:
        return []
    out = [ref]
    # If ref is dot form depth ≥3, also try last-segment-as-N
    if "-" not in ref:
        parts = ref.split(".")
        if len(parts) >= 3:
            try:
                int(parts[-1])
                folder = ".".join(parts[:-1]) + "-" + parts[-1]
                out.append(folder)
            except ValueError:
                pass
    # If ref is dash form, also try expanded dot form
    if "-" in ref:
        m = re.match(r"^(\d+(?:\.\d+)*)-(\d+)$", ref)
        if m:
            out.append(m.group(1) + "." + m.group(2))
    return out


def resolve_ref_to_pdf(parsed: dict, raw: str) -> Path | None:
    """Find PDF path from parsed Col D."""
    ref = parsed.get("ref")
    if ref:
        for k in _ref_to_folder_keys(ref):
            if k in PDF_INDEX:
                return _pick_best_in_folder(PDF_INDEX[k], k.replace("-", "."), raw)
            if k in SECTION_INDEX:
                return SECTION_INDEX[k][0]

    if parsed.get("type") == "filename_format":
        # (a) Direct {section}-{N} sub-item pattern at start of D
        m = re.match(r"^(\d+(?:\.\d+){2,3}-\d+)", raw)
        if m and m.group(1) in PDF_INDEX:
            return _pick_best_in_folder(PDF_INDEX[m.group(1)],
                                        m.group(1).replace("-", "."), raw)
        # (b) Single-row format "5.2.1.9 ..." → check folder 5.2.1.-9
        msec = re.match(r"^(\d+(?:\.\d+){2,3})\b", raw)
        if msec:
            sec = msec.group(1)
            # Section-only folder (sensor, LED, เสา) takes priority
            if sec in SECTION_INDEX:
                return SECTION_INDEX[sec][0]
            # Otherwise look in `parent.-N/` folder
            key = _section_to_folder_key(sec)
            if key and key in PDF_INDEX:
                return _pick_best_in_folder(PDF_INDEX[key], sec, raw)
        # (c) Older shorter section "5.1.3.2." → SECTION_INDEX
        m2 = re.match(r"^(\d+(?:\.\d+){1,3})", raw)
        if m2 and m2.group(1) in SECTION_INDEX:
            return SECTION_INDEX[m2.group(1)][0]

    # model-only: scan tokens, but ONLY accept a hit that's also unique
    # (multiple PDFs share many tokens — picking the first match silently
    # routes rows to the wrong catalog).
    if parsed.get("type") in ("model_only", "filename_format"):
        candidate_pdfs: list[Path] = []
        for tok in re.findall(r"[A-Za-z][\w\-]{2,}", raw):
            paths = MODEL_INDEX.get(tok.lower())
            if paths:
                # Only trust unique matches OR matches that contain the
                # exact token as a "model" suffix (e.g. UF-2010A)
                if len(set(paths)) == 1:
                    return paths[0]
                candidate_pdfs.extend(paths)
        # Score remaining candidates (same heuristic as folder match)
        if candidate_pdfs:
            return _pick_best_in_folder(list(set(candidate_pdfs)), "", raw)
    return None


def load_rows() -> None:
    ROWS.clear()
    if not XLSX_PATH.exists():
        sys.stderr.write(f"WARN: {XLSX_PATH} not found\n")
        return
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    # rows 1-4 are header — start at row 5
    for r in range(5, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        e = ws.cell(r, 5).value
        f = ws.cell(r, 6).value
        if not any([a, b, c, d, e, f]):
            continue
        parsed = parse_col_d(d)
        pdf = resolve_ref_to_pdf(parsed, str(d) if d else "")
        # Detect section from B (e.g., "5.1.2." or "          1) ..." etc.)
        section = None
        if a:
            section = str(a).strip()
        elif b:
            mb = re.match(r"\s*(\d+(?:\.\d+)+)\.?", str(b))
            if mb:
                section = mb.group(1)
        ROWS.append({
            "row": r,
            "A": a,
            "B": b,
            "C": c,
            "D": d,
            "E": e,
            "F": f,                                       # existing status
            "parsed": parsed,
            "pdf_rel": str(pdf.relative_to(OUTPUT)) if pdf else None,
            "pdf_inherited": False,
            "section_inferred": section,
        })

    # Second pass: brand_model parent rows inherit PDF from next sub-item row
    # (so clicking "ยี่ห้อ Schneider รุ่น QO..." shows the Schneider/QO catalog page)
    for i, r in enumerate(ROWS):
        if r["parsed"].get("type") == "brand_model" and not r["pdf_rel"]:
            for j in range(i + 1, min(i + 6, len(ROWS))):
                nxt = ROWS[j]
                if nxt["pdf_rel"]:
                    r["pdf_rel"] = nxt["pdf_rel"]
                    r["pdf_inherited"] = True
                    r["parsed"].setdefault("page", 1)
                    break
            # Fallback: use the row's own section to find the parent rack
            # folder (handles brand_model headers when sub-items haven't been
            # auto-annotated yet)
            if not r["pdf_rel"]:
                sec = r.get("section_inferred")
                if sec:
                    for k in _ref_to_folder_keys(sec):
                        if k in PDF_INDEX:
                            r["pdf_rel"] = str(PDF_INDEX[k][0].relative_to(OUTPUT))
                            r["pdf_inherited"] = True
                            r["parsed"].setdefault("page", 1)
                            break
                        if k in SECTION_INDEX:
                            r["pdf_rel"] = str(SECTION_INDEX[k][0].relative_to(OUTPUT))
                            r["pdf_inherited"] = True
                            r["parsed"].setdefault("page", 1)
                            break

    # Third pass: propagate inferred section to sub-item rows without one.
    # When inheriting, prefer the MORE SPECIFIC of:
    #   • last_section (from preceding section-header row)
    #   • D-ref-derived section (e.g. ref "5.1.2-2" → "5.1.2")
    # Otherwise a row in section 5.1.1.2 with D ref "5.1.1-2" would be
    # mis-inferred as "5.1.1" (less specific) instead of "5.1.1.2".
    def _section_depth(s: str) -> int:
        return len(s.split("."))

    last_section: str | None = None
    for r in ROWS:
        if r["section_inferred"]:
            last_section = r["section_inferred"]
            continue

        candidates: list[str] = []
        if last_section:
            candidates.append(last_section)
        d_ref = r["parsed"].get("ref")
        if d_ref:
            m = re.match(r"^(\d+(?:\.\d+){1,3})", d_ref)
            if m:
                candidates.append(m.group(1))

        if candidates:
            # Pick deepest section number; ties broken by D-ref preference
            chosen = max(candidates, key=lambda s: (_section_depth(s), s != last_section))
            r["section_inferred"] = chosen
            r["section_inherited"] = True

    # Fourth pass: rows that don't yet have a pdf_rel but DO have a section
    # → try to find a candidate catalog by folder convention.
    #
    # Also runs for "ยินดีปฏิบัติตามข้อกำหนด" (commitment) rows so the user
    # can SEE the catalog when invoking 📍 Mark in the manual-annotate
    # workflow. Col D itself isn't changed — only the surface pdf_rel mapping.
    #
    # Uses _candidate_pdfs_for_section which tries multiple lookup strategies:
    #   • exact section ("5.1.6.2" → SECTION_INDEX)
    #   • folder convention ("5.1.1.2" → PDF_INDEX["5.1.1-2"])
    #   • wildcard for sections with multiple sub-catalogs
    #     ("5.1.2"  → first of 5.1.2-1, 5.1.2-2, …)
    #     ("5.1.6.1" → first of 5.1.6.1-1, …)
    #   • parent fallback ("5.1.6.1" → 5.1.6 sensors/เสา)
    for r in ROWS:
        if r["pdf_rel"]:
            continue
        d_value = (r.get("D") or "").strip()
        is_commit = d_value.startswith("ยินดีปฏิบัติ")
        if d_value and not is_commit:
            continue
        sec = r.get("section_inferred")
        if not sec:
            continue
        candidates = _candidate_pdfs_for_section(sec)
        if not candidates:
            continue
        r["pdf_rel"] = str(candidates[0].relative_to(OUTPUT))
        r["pdf_inherited"] = True
        if not d_value:
            r["needs_col_d"] = True

    # Mirror into the OOP Project singleton so new code paths can use it.
    PROJECT_OBJ.set_rows([core.from_legacy_dict(r) for r in ROWS])


def collect_extra_refs() -> None:
    EXTRA_REFS.clear()
    EXTRA_REFS["tor_pdfs"] = sorted(str(p.relative_to(PROJECT)) for p in TOR_DIR.glob("*.pdf")) if TOR_DIR.exists() else []
    EXTRA_REFS["boq_xlsx"] = sorted(str(p.relative_to(PROJECT)) for p in BOQ_DIR.glob("*.xlsx")) if BOQ_DIR.exists() else []
    EXTRA_REFS["comply_pdf"] = "Comply spec Smart Plant 1.pdf" if (OUTPUT / "Comply spec Smart Plant 1.pdf").exists() else None


# ---------------------------------------------------------------------------
# Tree builder — section hierarchy + sub-items grouped under parent
# ---------------------------------------------------------------------------

def _section_parts(sec: str) -> list[str]:
    return [p for p in sec.split(".") if p]


def build_tree() -> dict:
    """Return tree of {key, label, children, rows[], type}.

    Hierarchy:
      Smart Plant 1
      └─ 5
         └─ 5.1
            └─ 5.1.2
               ├─ R62 (parent header)
               ├─ R63 (item 1))
               ├─ R64 (item 2))
               │  ├─ R65 (sub-item 1.)
               │  ├─ R66 (sub-item 2.)
               │  ...
    """
    root = {"key": "ROOT", "label": "Smart Plant 1", "children": {}, "rows": [], "type": "root"}

    def get_or_create(node, key, label, ntype="section"):
        if key not in node["children"]:
            node["children"][key] = {"key": key, "label": label,
                                     "children": {}, "rows": [], "type": ntype}
        return node["children"][key]

    # First pass: section nodes
    last_section_node = root
    last_item_node = None  # within section, current ข้อ N) parent
    for r in ROWS:
        sec = r.get("section_inferred")
        b = (r.get("B") or "").strip()
        d = r.get("D") or ""
        if sec:
            parts = _section_parts(sec)
            node = root
            for i, _ in enumerate(parts):
                key = ".".join(parts[: i + 1])
                node = get_or_create(node, key, key)
            # this row IS the section header (or first row of section)
            node["rows"].append(r["row"])
            last_section_node = node
            last_item_node = None
        else:
            # sub-row — could be "1) ..." (parent item) or "1. ..." (sub-item)
            m_parent = re.match(r"^\s*(\d+)\s*\)", b)
            m_sub = re.match(r"^\s*(\d+)\s*\.", b)
            if m_parent:
                # ข้อ N) — child of last_section_node
                key = f"{last_section_node['key']}#{m_parent.group(1)})"
                item = get_or_create(last_section_node, key,
                                     f"ข้อ {m_parent.group(1)})", ntype="item")
                item["rows"].append(r["row"])
                last_item_node = item
            elif m_sub and last_item_node is not None:
                # ข้อย่อย N. — child of last_item_node
                key = f"{last_item_node['key']}#{m_sub.group(1)}."
                sub = get_or_create(last_item_node, key,
                                    f"ข้อย่อย {m_sub.group(1)}.", ntype="sub")
                sub["rows"].append(r["row"])
            else:
                # fall-through — attach to nearest known parent
                target = last_item_node or last_section_node
                target["rows"].append(r["row"])

    return root


# ---------------------------------------------------------------------------
# TOR text search (find page by Col B content)
# ---------------------------------------------------------------------------

TOR_PATH: Path | None = None
TOR_CACHE: dict[int, dict] = {}              # row_num → {page, rects}
TOR_SECTION_INDEX: dict[str, int] = {}       # section → first page (1-indexed)
TOR_PAGE_TEXTS: dict[int, str] = {}          # page → normalized full text (for substring search)


def normalize_thai_text(s: str) -> str:
    """Aggressive normalization for cross-document Thai matching.

    Handles:
      - SARA AM (precomposed ำ U+0E33) ↔ NIKHAHIT + SARA AA (ํา) — TOR
        often uses the decomposed form while xlsx uses precomposed.
      - Combining-mark order: TOR sometimes writes NIKHAHIT before tone
        marks (`นํ้า`), xlsx writes tone before SARA AM (`น้ำ`). Unicode
        canonical ordering doesn't fix this because NIKHAHIT has
        combining class 0. We swap explicitly.
      - Whitespace collapse + case fold.
    """
    if not s:
        return ""
    # SARA AM decomposition equivalence
    s = s.replace("ำ", "ํา")  # always force decomposed form
    # NIKHAHIT (U+0E4D) before tone marks (0E48..0E4B) → tone before NIKHAHIT
    # (this normalizes to the order typically produced by xlsx)
    s = re.sub(r"ํ([่-๋])", r"\1ํ", s)
    # collapse all whitespace (incl. zero-width and NBSP) to single space
    s = re.sub(r"[\s ​]+", " ", s)
    return s.strip().lower()


def index_tor_text() -> None:
    """Pre-extract + normalize every TOR page so we can do substring search
    independent of Unicode form."""
    global TOR_PATH, TOR_PAGE_TEXTS
    TOR_PAGE_TEXTS = {}
    if TOR_PATH is None:
        TOR_PATH = find_tor_pdf()
    if TOR_PATH is None:
        return
    try:
        doc = fitz.open(TOR_PATH)
    except Exception:
        return
    for pno in range(len(doc)):
        try:
            txt = doc[pno].get_text("text")
        except Exception:
            txt = ""
        TOR_PAGE_TEXTS[pno + 1] = normalize_thai_text(txt)


def find_tor_pdf() -> Path | None:
    if TOR_DIR.exists():
        pdfs = sorted(TOR_DIR.glob("*.pdf"))
        if pdfs:
            return pdfs[0]
    return None


# Match a section heading at line start. Anchor with a dot or whitespace
# AFTER the section to avoid catching arbitrary numbers like "85" or "5mm".
# Acceptable: "5.", "5.1", "5.1.2.", "5.1.2 ตู้..." (digit + dot/space + non-digit).
_SECTION_HEAD = re.compile(
    r"(?m)^\s*(\d+(?:\.\d+){0,4})\s*[\.\)]?\s+(?=[^\d])"
)


def _is_valid_section_for_index(sec: str) -> bool:
    """Reject "85" "100.00" etc. — keep only realistic section paths."""
    parts = sec.split(".")
    if len(parts) > 5:
        return False
    for p in parts:
        if not p.isdigit():
            return False
        n = int(p)
        if n < 1 or n > 50:
            return False
    return True


def build_tor_section_index() -> None:
    """Scan TOR once, map each section heading to its first occurrence page."""
    global TOR_PATH, TOR_SECTION_INDEX
    TOR_SECTION_INDEX = {}
    if TOR_PATH is None:
        TOR_PATH = find_tor_pdf()
    if TOR_PATH is None:
        return
    try:
        doc = fitz.open(TOR_PATH)
    except Exception:
        return
    for pno in range(len(doc)):
        try:
            text = doc[pno].get_text("text")
        except Exception:
            continue
        for m in _SECTION_HEAD.finditer(text):
            sec = m.group(1)
            if not _is_valid_section_for_index(sec):
                continue
            # Only record FIRST page of each section
            if sec not in TOR_SECTION_INDEX:
                TOR_SECTION_INDEX[sec] = pno + 1


def get_tor_section_range(section: str | None) -> tuple[int, int] | None:
    """Return (start, end) inclusive page range of TOR for the given section.

    Falls back to parent section if the exact section isn't indexed.
    The end page is determined by the next section at the same depth or
    shallower; if nothing follows, default to start + 6.
    """
    if not section or not TOR_SECTION_INDEX or TOR_PATH is None:
        return None
    parts = section.split(".")
    start = TOR_SECTION_INDEX.get(section)
    used_section = section
    if start is None:
        for i in range(len(parts) - 1, 0, -1):
            parent = ".".join(parts[:i])
            if parent in TOR_SECTION_INDEX:
                start = TOR_SECTION_INDEX[parent]
                used_section = parent
                break
    if start is None:
        return None

    # End: smallest page > start belonging to a section at depth <= used_section
    target_depth = len(used_section.split("."))
    end = None
    for sec, pg in TOR_SECTION_INDEX.items():
        if sec == used_section:
            continue
        if pg <= start:
            continue
        sec_depth = len(sec.split("."))
        # Sibling or shallower → defines the boundary
        if sec_depth <= target_depth:
            if end is None or pg < end:
                end = pg - 1
    if end is None:
        try:
            doc = fitz.open(TOR_PATH)
            end = min(start + 6, len(doc))
        except Exception:
            end = start + 6
    return (start, end)


def _normalize_for_search(s: str) -> str:
    """Strip leading whitespace + numbering prefix; pick a stable substring."""
    s = (s or "").strip()
    # remove leading "N)" or "N." or "5.1.2."
    s = re.sub(r"^\s*\d+(?:\.\d+)*\.?\s*", "", s)
    s = re.sub(r"^\s*\d+\s*[\)\.]\s*", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _thai_variants(s: str) -> list[str]:
    """TOR PDFs sometimes use decomposed Thai (NIKHAHIT + SARA AA), while xlsx
    uses precomposed SARA AM (U+0E33). NFD doesn't break U+0E33 (it's only a
    compatibility decomposition) — must use NFKD."""
    import unicodedata
    out = [s]
    for form in ("NFC", "NFD", "NFKC", "NFKD"):
        v = unicodedata.normalize(form, s)
        if v not in out:
            out.append(v)
    # also: explicit U+0E33 → U+0E4D + U+0E32 (TOR's preferred form)
    explicit = s.replace("ำ", "ํา")
    if explicit not in out:
        out.append(explicit)
    return out


def _try_get_rects(page, needle: str) -> list:
    """Best-effort rect lookup for `needle` on a page. Try the original needle
    plus several Thai variants. Return the first non-empty result."""
    for variant in _thai_variants(needle):
        try:
            rects = page.search_for(variant, quads=False)
        except Exception:
            rects = []
        if rects:
            return rects
    return []


def _extract_search_tokens(text: str) -> list[str]:
    """Distinctive tokens used as a fallback when the full phrase doesn't
    match. Prefer English words, model codes, numeric values with units."""
    s = re.sub(r"^\s*\d+\s*[\)\.]\s*", "", text or "")
    tokens: list[str] = []
    seen = set()

    def add(t):
        t = t.strip()
        if len(t) < 3:
            return
        k = t.lower()
        if k in seen:
            return
        seen.add(k)
        tokens.append(t)

    # English words / model codes (≥3 chars, must contain a letter)
    for m in re.finditer(r"[A-Za-z][A-Za-z0-9]{2,}(?:[-+/.][A-Za-z0-9]+)*", s):
        tok = m.group(0)
        if len(tok) >= 4 or any(c.isdigit() for c in tok):
            add(tok)
    # Standards
    for m in re.finditer(r"(?:IEC|ISO|ASTM|ANSI|UL|TIS|มอก\.?)\s*\d+(?:[-/.\s]\d+)*", s):
        add(m.group(0))
    # IP rating / DC voltage / typical units
    for m in re.finditer(r"\b(?:IP\d+|DC\s*\d+(?:-\d+)?\s*V|AC\s*\d+\s*V)\b", s):
        add(m.group(0))
    # Numeric + unit
    for m in re.finditer(r"\d+(?:[\.,]\d+)?\s*(?:mA|MHz|GHz|kHz|Hz|°C|°F|mm|cm|m|kg|hr|min|kV|V|A|W|kW)", s, re.IGNORECASE):
        add(m.group(0))
    return tokens


def _next_sibling_section(section: str) -> str | None:
    """Compute the next sibling section number at the same depth.
    "5.1.1.2" → "5.1.1.3". Used as a y-bound when filtering matches."""
    if not section:
        return None
    parts = section.split(".")
    try:
        last = int(parts[-1])
    except ValueError:
        return None
    return ".".join(parts[:-1] + [str(last + 1)])


def _section_y_bounds_on_page(page, section: str | None,
                              next_section: str | None) -> tuple[float, float] | None:
    """Find the y-range of `section`'s content on this page.

    Returns (y_top, y_bottom). y_top = y0 of section header marker (if found
    on this page; else 0). y_bottom = y0 of next-section marker (if any on
    this page; else page height).

    A rect entirely below y_bottom (or above y_top) belongs to a different
    section's content — the matcher should reject it.
    """
    if not section:
        return None
    page_h = page.rect.height
    y_top = 0.0
    y_bottom = page_h

    # Section header on this page? (e.g. "5.1.1.2." or "5.1.1.2 ")
    sec_hits: list = []
    for needle in (f"{section}.", f"{section} "):
        try:
            sec_hits.extend(page.search_for(needle, quads=False))
        except Exception:
            pass
    if sec_hits:
        # Section may have multiple hits (header + citations) — take topmost
        y_top = min(h.y0 for h in sec_hits)

    # Next sibling header on this page = upper bound for section's content
    if next_section:
        nx_hits: list = []
        for needle in (f"{next_section}.", f"{next_section} "):
            try:
                nx_hits.extend(page.search_for(needle, quads=False))
            except Exception:
                pass
        # Only consider next-section hits that come AFTER section header
        nx_hits = [h for h in nx_hits if h.y0 > y_top + 1]
        if nx_hits:
            y_bottom = min(h.y0 for h in nx_hits)

    return (y_top, y_bottom)


def _filter_rects_by_bounds(rects: list, bounds: tuple[float, float] | None,
                            margin: float = 4.0) -> list:
    if not bounds or not rects:
        return rects
    y_top, y_bottom = bounds
    return [r for r in rects
            if (r.y0 >= y_top - margin and r.y1 <= y_bottom + margin)]


def _densest_y_cluster(rects: list, max_span: float = 35.0) -> list:
    """Among scattered rects, return the tightest y-range cluster.

    When the full phrase can't be located by ``search_for`` (e.g. Thai
    combining-mark mismatches), the token fallback collects rects from many
    individual matches across the page. Most of those tokens repeat in
    headers/footers/citations; only the cluster on the actual spec line is
    relevant. This function picks the tightest band containing the most
    rects so we highlight a single line, not every occurrence on the page.
    """
    if len(rects) <= 1:
        return rects
    def y0(r):
        return r.y0 if hasattr(r, "y0") else r[1]
    sorted_rects = sorted(rects, key=y0)
    best_start, best_count = 0, 0
    for i, r in enumerate(sorted_rects):
        top = y0(r)
        count = sum(1 for r2 in sorted_rects if top <= y0(r2) <= top + max_span)
        if count > best_count:
            best_count = count
            best_start = i
    top = y0(sorted_rects[best_start])
    return [r for r in sorted_rects if top <= y0(r) <= top + max_span]


def _anchor_cluster_by_rarest(token_hits: dict, line_span: float = 14.0) -> list:
    """Pick the line containing the rarest token, then keep only token rects
    on that same line (within ``line_span`` of the anchor's y).

    ``token_hits`` is ``{token: [rect, ...]}``. Generic words like "Firewall"
    repeat across the section header + multiple sub-items, so densest-cluster
    can still span 3 lines. Anchoring on the RAREST token (the one that
    appears once on the whole page) pinpoints the actual content line.
    """
    if not token_hits:
        return []
    # Sort tokens by hit count ascending (rarest first); break ties by token
    # length (longer = more distinctive).
    ranked = sorted(token_hits.items(),
                    key=lambda kv: (len(kv[1]), -len(kv[0])))
    # If even the rarest appears many times, fall back to densest cluster
    rarest_tok, rarest_rects = ranked[0]
    if len(rarest_rects) > 4:
        flat = [r for rs in token_hits.values() for r in rs]
        return _densest_y_cluster(flat, max_span=line_span * 2)

    def y0(r):
        return r.y0 if hasattr(r, "y0") else r[1]

    # Try each occurrence of the rarest token as a potential anchor and
    # pick the anchor that gathers the most other-token rects on its line.
    best: tuple[list, int] = ([], -1)
    for anchor in rarest_rects:
        anchor_y = y0(anchor)
        line_rects = []
        for tok, rs in token_hits.items():
            for r in rs:
                if abs(y0(r) - anchor_y) <= line_span:
                    line_rects.append(r)
        if len(line_rects) > best[1]:
            best = (line_rects, len(line_rects))
    return best[0]


def _search_tor_in_pages(text: str, page_nums: list[int],
                          section: str | None = None) -> dict | None:
    """Two-stage search across the given page list:

    1. **Normalized substring** match against the pre-indexed TOR text.
       Robust to Thai Unicode form differences (handles SARA AM ↔ NIKHAHIT
       and combining-mark order).
    2. Once we know which page contains the match, use ``page.search_for``
       with multiple variants to recover rect coordinates for highlight.
    3. **Filter rects to within the section's y-bounds** on that page so
       phrases that repeat across sections don't bleed into the highlight.
    """
    if TOR_PATH is None or not text or not page_nums or not TOR_PAGE_TEXTS:
        return None
    needle_raw = _normalize_for_search(text)
    if len(needle_raw) < 6:
        return None
    try:
        doc = fitz.open(TOR_PATH)
    except Exception:
        return None

    next_section = _next_sibling_section(section) if section else None

    # Try full needle, then progressively shorter prefixes
    for length in (90, 60, 40, 25, 15):
        head = needle_raw[:length].strip()
        if not head:
            continue
        head_n = normalize_thai_text(head)
        if not head_n:
            continue
        for pno in page_nums:
            if pno < 1 or pno > len(doc):
                continue
            page_text = TOR_PAGE_TEXTS.get(pno, "")
            if not page_text:
                continue
            if head_n in page_text:
                page = doc[pno - 1]
                bounds = _section_y_bounds_on_page(page, section, next_section)
                rects = _try_get_rects(page, head)
                rects = _filter_rects_by_bounds(rects, bounds)
                if not rects:
                    for sub_len in (30, 20, 12):
                        if sub_len >= len(head):
                            continue
                        rects = _try_get_rects(page, head[:sub_len])
                        rects = _filter_rects_by_bounds(rects, bounds)
                        if rects:
                            break
                if not rects:
                    # Token-rect fallback — anchor on the rarest token to
                    # pin a single line. Generic words ("Firewall", "Power")
                    # repeat across the section header + every sub-item; an
                    # uncommon token like "Throughput" or a numeric+unit
                    # appears only on the actual spec line.
                    token_hits: dict = {}
                    for tok in _extract_search_tokens(needle_raw)[:8]:
                        rs = _filter_rects_by_bounds(
                            _try_get_rects(page, tok), bounds)
                        if rs:
                            token_hits[tok] = rs
                    rects = _anchor_cluster_by_rarest(token_hits, line_span=14)
                if not rects and bounds:
                    # If page contains the phrase but every match is outside
                    # the section's y-range, skip this page (it's a different
                    # section's repetition of the phrase) — keep looking on
                    # other pages.
                    continue
                return {"page": pno,
                        "rects": [[r.x0, r.y0, r.x1, r.y1] for r in rects],
                        "needle": head}

    # Phrase fallback failed — try distinctive-token scoring
    tokens = _extract_search_tokens(needle_raw)
    if tokens:
        best = None  # (page, score, rects)
        for pno in page_nums:
            if pno < 1 or pno > len(doc):
                continue
            page_text = TOR_PAGE_TEXTS.get(pno, "")
            if not page_text:
                continue
            score = 0
            matched_tokens = []
            for tok in tokens:
                if normalize_thai_text(tok) in page_text:
                    score += 1
                    matched_tokens.append(tok)
            if score and (best is None or score > best[1]):
                best = (pno, score, matched_tokens)
        if best and best[1] >= max(2, min(3, len(tokens))):
            page = doc[best[0] - 1]
            bounds = _section_y_bounds_on_page(page, section, next_section)
            token_hits: dict = {}
            for tok in best[2][:6]:
                rs = _filter_rects_by_bounds(_try_get_rects(page, tok), bounds)
                if rs:
                    token_hits[tok] = rs
            rects = _anchor_cluster_by_rarest(token_hits, line_span=14)
            if rects:
                return {"page": best[0],
                        "rects": [[r.x0, r.y0, r.x1, r.y1] for r in rects],
                        "needle": " + ".join(best[2])}
    return None


def _chapter_root(section: str | None) -> str | None:
    """Top-level chapter, e.g. "5.1.2" → "5"."""
    if not section:
        return None
    return section.split(".")[0]


def find_in_tor(text: str, section: str | None = None) -> dict | None:
    """Search TOR PDF for `text`, biased to the row's section.

    Strategy (in priority order):
      1. Within the row's exact section range (e.g. 5.1.2 → pages 12–13).
      2. Section range widened by ±3 pages.
      3. From section start forward to end of the chapter root (e.g. all of 5).
      4. **No** full-document fallback — wrong-chapter matches are worse than
         no match. If nothing is found, default to the section's first page.
    """
    global TOR_PATH
    if TOR_PATH is None:
        TOR_PATH = find_tor_pdf()
    if TOR_PATH is None or not text:
        return None
    try:
        doc = fitz.open(TOR_PATH)
    except Exception:
        return None
    n_pages = len(doc)

    rng = get_tor_section_range(section)
    chapter_rng = get_tor_section_range(_chapter_root(section))

    page_groups: list[list[int]] = []
    if rng:
        s, e = rng
        # 1) exact section pages, in order
        page_groups.append(list(range(s, e + 1)))
        # 2) widen FORWARD only (content often spills past the section heading
        #    of the next sub-section, but rarely appears before the heading)
        page_groups.append(list(range(s, min(n_pages, e + 5) + 1)))
    if chapter_rng:
        cs, ce = chapter_rng
        # 3) forward from section start (or chapter start) to chapter end
        forward_start = rng[0] if rng else cs
        page_groups.append(list(range(forward_start, min(n_pages, ce) + 1)))
        # 4) backward within chapter, only if forward yielded nothing
        if rng:
            page_groups.append(list(range(cs, rng[0])))
        else:
            page_groups.append(list(range(cs, min(n_pages, ce) + 1)))
    if not page_groups:
        # Section unknown — search full doc as low-confidence fallback
        page_groups.append(list(range(1, n_pages + 1)))

    seen: set[int] = set()
    for group in page_groups:
        fresh = [p for p in group if p not in seen]
        seen.update(group)
        if not fresh:
            continue
        result = _search_tor_in_pages(text, fresh, section=section)
        if result:
            result["section_range"] = rng
            return result

    # Nothing matched within the chapter — return the section's first page
    # so the user lands in the right neighbourhood (no highlight).
    if rng:
        return {"page": rng[0], "rects": [], "section_range": rng,
                "fallback": "section_start"}
    return None


def render_tor_page(page_num: int, highlight_rects: list | None = None,
                    dpi: int = 110) -> bytes:
    if TOR_PATH is None:
        return b""
    doc = fitz.open(TOR_PATH)
    if page_num < 1 or page_num > len(doc):
        page_num = max(1, min(len(doc), page_num))
    page = doc[page_num - 1]
    pix = page.get_pixmap(dpi=dpi, annots=False)
    img_bytes = pix.tobytes("png")
    if not highlight_rects:
        return img_bytes
    im = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
    overlay = Image.new("RGBA", im.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    scale = dpi / 72.0
    for r in highlight_rects:
        x0, y0, x1, y1 = r[0] * scale, r[1] * scale, r[2] * scale, r[3] * scale
        # yellow translucent fill + outline
        draw.rectangle([x0, y0, x1, y1], fill=(255, 230, 0, 90),
                       outline=(255, 180, 0, 220), width=3)
    im2 = Image.alpha_composite(im, overlay)
    buf = io.BytesIO()
    im2.convert("RGB").save(buf, "PNG")
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Auto checks (light — flag obvious problems without judging content)
# ---------------------------------------------------------------------------

SOFTWARE_KEYWORDS = ["ซอฟต์แวร์", "Software", "Package Application"]
INSTALL_KEYWORDS = ["งานติดตั้ง", "ติดตั้งและทดสอบ", "งานเดินสาย"]
COMPARE_PHRASES = ["หรือดีกว่า", "ไม่น้อยกว่า", "ไม่น้อยไปกว่า", "ไม่มากกว่า", "ต้องสามารถ", "จะต้อง"]


def auto_check_row(row: dict) -> list[dict]:
    """Return list of {level, message} flags."""
    flags = []
    b, c, d = (row.get("B") or ""), (row.get("C") or ""), (row.get("D") or "")
    parsed = row.get("parsed", {})
    # Col D format
    if d and parsed.get("type") == "empty":
        flags.append({"level": "warn", "msg": "Col D ไม่อยู่ใน 6 รูปแบบที่กำหนด"})
    # Software/install rows should be commitment + no PDF rect
    is_software = any(k in b for k in SOFTWARE_KEYWORDS) or any(k in c for k in SOFTWARE_KEYWORDS) if (b or c) else False
    is_install = any(k in b for k in INSTALL_KEYWORDS) or any(k in c for k in INSTALL_KEYWORDS) if (b or c) else False
    if (is_software or is_install) and parsed.get("type") not in ("commitment", "empty"):
        flags.append({"level": "info", "msg": f"Software/install row — Col D ควรเป็น 'ยินดีปฏิบัติตามข้อกำหนด' (พบ: {parsed.get('type')})"})
    # Col C still has compare phrases?
    if c:
        for ph in COMPARE_PHRASES:
            if ph in c:
                flags.append({"level": "warn", "msg": f"Col C ยังมีคำเปรียบเทียบ: '{ph}'"})
    # Col B copy from TOR — Col B should have leading whitespace (TOR indentation), Col C should not
    # gentle hint only
    if b and c and isinstance(b, str) and isinstance(c, str):
        if b.lstrip() == c.lstrip() and parsed.get("type") in ("equivalent", "higher", "brand_model"):
            # Col C identical to Col B (after trim) — not necessarily wrong but worth noting
            pass
    # PDF resolution
    if parsed.get("type") in ("equivalent", "higher", "filename_format", "model_only") and not row.get("pdf_rel"):
        flags.append({"level": "warn", "msg": "ไม่พบ PDF ที่ Col D อ้างอิง"})
    # Catalog exists but Col D is still empty (user added catalogs but hasn't
    # updated the spreadsheet yet — we resolve the PDF via folder convention)
    if row.get("needs_col_d"):
        flags.append({"level": "warn",
                      "msg": "พบ catalog ใน folder แต่ Col D ว่าง — ต้องเติม Col D ให้ตรงตาม convention"})
    return flags


# ---------------------------------------------------------------------------
# Status persistence
# ---------------------------------------------------------------------------

def load_status() -> dict:
    """Verification status now lives in SQLite (canonical source). The legacy
    verification_status.json is migrated on boot and kept as backup."""
    if db.db_path() is not None:
        try:
            return db.get_all_status()
        except Exception as e:
            sys.stderr.write(f"WARN: db get_all_status: {e}\n")
    # Fallback: read legacy JSON if DB isn't ready yet
    if STATUS_PATH.exists():
        try:
            return json.loads(STATUS_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_status(data: dict) -> None:
    """Backup write to JSON for compat with manual diffs. Not the source of
    truth — DB is."""
    try:
        STATUS_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                               encoding="utf-8")
    except Exception as e:
        sys.stderr.write(f"WARN: cannot save status JSON: {e}\n")


# ---------------------------------------------------------------------------
# PDF rendering
# ---------------------------------------------------------------------------

def safe_iter_annots(page) -> list:
    """Iterate annotations skipping ones with broken appearance streams.

    PyMuPDF's ``page.annots()`` generator can fail entirely (returning zero
    annots) when the page mixes drawable annots with broken Image/Form
    annotations — common in our catalog PDFs. To stay robust we bypass the
    generator and call ``load_annot(xref)`` per entry from ``annot_xrefs()``
    with individual try/except. A single bad annot only loses itself, never
    the rest of the page.
    """
    out = []
    try:
        xrefs = page.annot_xrefs()
    except Exception:
        # fall back to the old generator path if annot_xrefs is unavailable
        try:
            gen = page.annots()
            if gen is None:
                return out
            while True:
                try:
                    ann = next(gen)
                except StopIteration:
                    break
                except Exception:
                    continue
                if ann is None:
                    continue
                out.append(ann)
        except Exception:
            pass
        return out

    for entry in xrefs:
        # entry is typically (xref, anntype, name); fall back gracefully
        if isinstance(entry, tuple):
            xref = entry[0]
        else:
            try:
                xref = int(entry)
            except Exception:
                continue
        if not xref:
            # xref=0 means an inline annot — handled by parse_inline_annots
            continue
        try:
            ann = page.load_annot(int(xref))
        except Exception:
            continue
        if ann is None:
            continue
        out.append(ann)
    return out


# ---------------------------------------------------------------------------
# Inline annotation parser
#
# Some PDFs (notably the เสา catalogs in 5.1.3.2 / 5.1.4.3 / 5.1.6.3) embed
# their annotations DIRECTLY in the page's /Annots array as literal
# dictionaries instead of indirect references. PyMuPDF's standard
# Page.annots() iterator skips these (xref reported as 0), so they were
# invisible to the highlight matcher and to the front-end annotation list.
# This parser extracts them by reading the raw /Annots value and decoding
# Subtype, Rect, and Contents fields out of each <<...>> block.
# ---------------------------------------------------------------------------

def _split_annots_array_items(arr_text: str) -> list[tuple[str, str]]:
    """Split a page /Annots array into ordered items, preserving the
    distinction between inline blocks and indirect references.

    Returns list of (kind, text) where kind ∈ {'ref', 'inline'} so callers
    can splice and rebuild the array correctly when removing an inline
    block (the indirect refs must keep their positions).
    """
    s = arr_text.strip()
    if s.startswith("["):
        s = s[1:]
    if s.endswith("]"):
        s = s[:-1]
    items: list[tuple[str, str]] = []
    i = 0
    L = len(s)
    while i < L:
        c = s[i]
        if c.isspace():
            i += 1
            continue
        if c == "<" and i + 1 < L and s[i + 1] == "<":
            depth = 1
            j = i + 2
            while j + 1 < L and depth > 0:
                if s[j] == "<" and s[j + 1] == "<":
                    depth += 1; j += 2
                elif s[j] == ">" and s[j + 1] == ">":
                    depth -= 1; j += 2
                else:
                    j += 1
            items.append(("inline", s[i:j]))
            i = j
        else:
            m = re.match(r"\d+\s+\d+\s+R", s[i:])
            if m:
                items.append(("ref", m.group(0)))
                i += m.end()
            else:
                i += 1
    return items


def _delete_inline_annot(doc, page, inline_index: int) -> bool:
    """Remove the N-th inline annotation from this page's /Annots array.

    PyMuPDF can't address inline annots via load_annot (they have xref=0),
    so we rewrite the array string in place with xref_set_key.
    """
    try:
        page_xref = page.xref
        kind, value = doc.xref_get_key(page_xref, "Annots")
    except Exception:
        return False
    if kind != "array" or not value:
        return False
    items = _split_annots_array_items(value)
    seen_inline = -1
    target_pos = -1
    for k, (knd, _val) in enumerate(items):
        if knd == "inline":
            seen_inline += 1
            if seen_inline == inline_index:
                target_pos = k
                break
    if target_pos < 0:
        return False
    items.pop(target_pos)
    new_value = "[" + " ".join(v for _, v in items) + "]"
    try:
        doc.xref_set_key(page_xref, "Annots", new_value)
    except Exception:
        return False
    return True


def _split_dict_blocks(arr_text: str) -> list[str]:
    """Split '[<<...>><<...>>]' into ['<<...>>', '<<...>>'] at depth 1."""
    s = arr_text.strip()
    if s.startswith("["):
        s = s[1:]
    if s.endswith("]"):
        s = s[:-1]
    blocks: list[str] = []
    depth = 0
    start = -1
    i = 0
    L = len(s)
    while i < L:
        if i + 1 < L and s[i] == "<" and s[i + 1] == "<":
            if depth == 0:
                start = i
            depth += 1
            i += 2
        elif i + 1 < L and s[i] == ">" and s[i + 1] == ">":
            depth -= 1
            i += 2
            if depth == 0 and start >= 0:
                blocks.append(s[start:i])
                start = -1
        else:
            i += 1
    return blocks


def _decode_pdf_string_literal(s: str) -> str:
    """Decode a (literal-string) PDF value with backslash escapes."""
    out = []
    i = 0
    L = len(s)
    while i < L:
        c = s[i]
        if c == "\\" and i + 1 < L:
            nxt = s[i + 1]
            if nxt in "()\\":
                out.append(nxt); i += 2; continue
            if nxt == "n": out.append("\n"); i += 2; continue
            if nxt == "r": out.append("\r"); i += 2; continue
            if nxt == "t": out.append("\t"); i += 2; continue
            if nxt.isdigit():
                # octal escape, up to 3 digits
                j = i + 1
                while j < L and j - i <= 3 and s[j].isdigit():
                    j += 1
                try:
                    out.append(chr(int(s[i + 1:j], 8)))
                except Exception:
                    pass
                i = j
                continue
            out.append(nxt); i += 2
        else:
            out.append(c); i += 1
    return "".join(out)


def _decode_pdf_hex_string(h: str) -> str:
    h = re.sub(r"\s+", "", h)
    if not h:
        return ""
    if len(h) % 2:
        h += "0"
    try:
        b = bytes.fromhex(h)
    except ValueError:
        return ""
    if b[:2] == b"\xfe\xff":
        try:
            return b[2:].decode("utf-16-be", errors="replace")
        except Exception:
            return ""
    if b[:2] == b"\xff\xfe":
        try:
            return b[2:].decode("utf-16-le", errors="replace")
        except Exception:
            return ""
    try:
        return b.decode("latin-1", errors="replace")
    except Exception:
        return ""


_RECT_RE = re.compile(
    r"/Rect\s*\[\s*([\-\d.]+)\s+([\-\d.]+)\s+([\-\d.]+)\s+([\-\d.]+)\s*\]"
)
_SUBTYPE_RE = re.compile(r"/Subtype\s*/(\w+)")
_CONTENTS_LITERAL_RE = re.compile(r"/Contents\s*\((.*?)(?<!\\)\)", re.DOTALL)
_CONTENTS_HEX_RE = re.compile(r"/Contents\s*<([0-9A-Fa-f\s]+)>")


def _parse_one_inline_annot(block: str) -> dict | None:
    """Parse a single <<...>> annotation dict block into our metadata shape."""
    if "/Type/Annot" not in block and "/Type /Annot" not in block:
        return None
    m_subtype = _SUBTYPE_RE.search(block)
    m_rect = _RECT_RE.search(block)
    if not m_subtype or not m_rect:
        return None
    out = {
        "type": m_subtype.group(1),
        "rect": [float(m_rect.group(i)) for i in (1, 2, 3, 4)],
        "contents": "",
    }
    # Hex string takes precedence (Thai content is usually hex-encoded)
    m_hex = _CONTENTS_HEX_RE.search(block)
    if m_hex:
        out["contents"] = _decode_pdf_hex_string(m_hex.group(1))
    else:
        m_lit = _CONTENTS_LITERAL_RE.search(block)
        if m_lit:
            out["contents"] = _decode_pdf_string_literal(m_lit.group(1))
    return out


def parse_inline_annots(doc, page) -> list[dict]:
    """Return inline-annotation dicts for a page (those without xrefs).

    PDF annotation /Rect values are in user-space (Y axis goes BOTTOM-UP in
    the standard PDF coordinate system). PyMuPDF's regular ``Annot.rect``
    auto-transforms to page-space (top-down) using
    ``page.transformation_matrix``; we replicate that transformation here so
    the rects we return are directly comparable to image-pixel coords.
    """
    try:
        page_xref = page.xref
        kind, value = doc.xref_get_key(page_xref, "Annots")
    except Exception:
        return []
    if kind != "array" or not value:
        return []
    try:
        mat = page.transformation_matrix
    except Exception:
        mat = None
    out = []
    for block in _split_dict_blocks(value):
        ann = _parse_one_inline_annot(block)
        if ann is None:
            continue
        if mat is not None:
            x0, y0, x1, y1 = ann["rect"]
            try:
                p0 = fitz.Point(x0, y0) * mat
                p1 = fitz.Point(x1, y1) * mat
                ann["rect"] = [
                    min(p0.x, p1.x), min(p0.y, p1.y),
                    max(p0.x, p1.x), max(p0.y, p1.y),
                ]
            except Exception:
                pass
        out.append(ann)
    return out


def all_annots_on_page(doc, page, page_num: int) -> list[dict]:
    """Combine real (xref'd) annotations and inline ones into a single list of
    dicts shaped like the rest of the codebase expects. Real annots come first
    so callers that key by xref don't lose them."""
    out: list[dict] = []
    seen_signatures: set[tuple] = set()  # to deduplicate
    for ann in safe_iter_annots(page):
        try:
            xref = ann.xref
            t = ann.type[1]
            r = [round(c, 2) for c in ann.rect]
            c = ann.info.get("content", "") or ""
        except Exception:
            continue
        sig = (t, tuple(r), c)
        seen_signatures.add(sig)
        out.append({
            "xref": int(xref),
            "page": page_num,
            "type": t,
            "rect": r,
            "contents": c,
        })
    inline_idx = 0
    for ann in parse_inline_annots(doc, page):
        r = [round(c, 2) for c in ann["rect"]]
        sig = (ann["type"], tuple(r), ann["contents"])
        # Even if the signature matches a real annot we keep the inline_index
        # counter advancing — its value must mirror the position of the inline
        # block in the page's /Annots array for delete-path lookups.
        if sig in seen_signatures:
            inline_idx += 1
            continue
        out.append({
            "xref": 0,            # 0 = inline (not editable via load_annot)
            "page": page_num,
            "type": ann["type"],
            "rect": r,
            "contents": ann["contents"],
            "_inline": True,
            "inline_index": inline_idx,
        })
        inline_idx += 1
    return out


# Patterns used to parse annotation labels into structured fields:
#   "5.1.2 ข้อ 3) ข้อย่อย 1."   → {item: 3, subitem: 1}
#   "5.1.2 ข้อ 3)"              → {item: 3, subitem: None}
#   "5.1.2 ข้อย่อย 4."          → {item: None, subitem: 4}
#   "ยี่ห้อ" / "รุ่น"            → {kind: 'brand_or_model', literal: ...}
_LABEL_ITEM_RE    = re.compile(r"ข้อ\s*(\d+)\s*\)")
_LABEL_SUBITEM_RE = re.compile(r"ข้อย่อย\s*(\d+)")
_LABEL_LITERALS   = ("ยี่ห้อ", "รุ่น")


def _parse_label(s: str) -> dict:
    """Parse query OR annotation content into structured fields.

    Returns ``{}`` for unrecognized strings (caller can fall back to substring).
    """
    out: dict = {}
    if not s:
        return out
    s = s.strip()
    # Brand/model literal (label content is typically just "ยี่ห้อ" or "รุ่น")
    if s in _LABEL_LITERALS:
        return {"kind": "literal", "literal": s}
    item = _LABEL_ITEM_RE.search(s)
    sub  = _LABEL_SUBITEM_RE.search(s)
    if item:
        out["item"] = int(item.group(1))
    if sub:
        out["subitem"] = int(sub.group(1))
    if out:
        out["kind"] = "section_label"
    return out


def _match_annot_label(query: str, content: str) -> bool:
    """Decide whether an annotation's content matches a highlight query.

    Goals:
      • exact digit match (so "ข้อย่อย 1" doesn't accidentally hit "ข้อย่อย 10")
      • when the query is a sub-item, also light up the parent rect
        ("ข้อ X)") so the user gets context — but NOT siblings
      • when the query is brand/model, only literal labels match (the paired
        Square gets added separately by spatial pairing in the caller)
    """
    qf = _parse_label(query)
    cf = _parse_label(content)

    # Brand/model literal — exact only
    if qf.get("kind") == "literal":
        return cf.get("kind") == "literal" and cf["literal"] == qf["literal"]

    # If we couldn't parse the query as a section label, fall back to
    # boundary-aware substring matching (e.g. tag content "5.1.2(2-model)").
    if qf.get("kind") != "section_label":
        if query == content:
            return True
        # Use word-boundary check so "ข้อย่อย 1" doesn't match "ข้อย่อย 10"
        # when callers pass raw substrings.
        try:
            return re.search(re.escape(query) + r"(?!\d)", content) is not None
        except re.error:
            return query in content

    if cf.get("kind") != "section_label":
        return False

    # Now both sides parsed.
    q_item, q_sub = qf.get("item"), qf.get("subitem")
    c_item, c_sub = cf.get("item"), cf.get("subitem")

    # SUB-ITEM QUERY (e.g., "ข้อ 3) ข้อย่อย 1.")
    if q_sub is not None:
        if c_sub is None:
            # Light up the parent rect of the SAME item (gives visual context)
            if q_item is not None and c_item == q_item:
                return True
            return False
        # Both have subitem — must match exactly
        if c_sub != q_sub:
            return False
        # If query specifies item, content's item must match too
        if q_item is not None and c_item is not None and c_item != q_item:
            return False
        return True

    # PARENT-ONLY QUERY ("ข้อ X)")
    if q_item is not None:
        if c_item != q_item:
            return False
        # Match the parent rect itself (no subitem in content). Skipping
        # subitem rects keeps the highlight focused.
        return c_sub is None

    return False


def render_pdf_page_png(pdf_path: Path, page_num: int, dpi: int = 130,
                        highlight: str | None = None,
                        no_annots: bool = False) -> tuple[bytes, dict]:
    """Render PDF page → (png_bytes, info).

    info = {"y0": float|None, "y1": float|None}  — image-pixel Y range of
    matched annotations (None if no highlight match).

    When ``no_annots`` is True the page is rendered WITHOUT baked annotations
    so the frontend can draw an interactive SVG overlay on top.
    """
    info: dict = {"y0": None, "y1": None}
    doc = fitz.open(pdf_path)
    if page_num < 1 or page_num > len(doc):
        page_num = max(1, min(len(doc), page_num))
    page = doc[page_num - 1]

    # Render page (annotations included unless explicitly suppressed)
    pix = page.get_pixmap(dpi=dpi, annots=not no_annots)
    img_bytes = pix.tobytes("png")

    if no_annots or not highlight:
        return img_bytes, info

    # Highlight may be a single label or "|"-separated alternatives
    queries = [q.strip() for q in highlight.split("|") if q.strip()]

    # Collect ALL annotations on this page — both real (xref'd) and inline
    # (literal-dict) so the matcher can find e.g. เสา catalog rects.
    doc_for_annots = page.parent
    annots = all_annots_on_page(doc_for_annots, page, page_num)

    # Find matching annotations using a structured, boundary-aware matcher.
    matched: list[dict] = []
    seen_ids: set[int] = set()
    for ann in annots:
        c = (ann.get("contents") or "").strip()
        if not c:
            continue
        for q in queries:
            if _match_annot_label(q, c):
                aid = id(ann)
                if aid not in seen_ids:
                    matched.append(ann)
                    seen_ids.add(aid)
                break

    # For each matched FreeText, pair with the nearest Square (by minimum
    # rect-edge distance). Handles all label layouts: right of square, below
    # square (เสา P2), above, or overlapping.
    def _rect_edge_distance(r1: list, r2: list) -> float:
        x_gap = max(0.0, r1[0] - r2[2], r2[0] - r1[2])
        y_gap = max(0.0, r1[1] - r2[3], r2[1] - r1[3])
        return (x_gap * x_gap + y_gap * y_gap) ** 0.5

    paired: list[dict] = []
    for m in matched:
        if m.get("type") != "FreeText":
            continue
        best = None
        best_d = 1e9
        for o in annots:
            if o.get("type") != "Square":
                continue
            d_ = _rect_edge_distance(m["rect"], o["rect"])
            if d_ < best_d:
                best_d, best = d_, o
        # Threshold: 60pt — close enough to be the labelled square but far
        # enough to not pair with random squares elsewhere on the page.
        if best is not None and best_d <= 60.0 and best not in matched and best not in paired:
            paired.append(best)

    if not matched and not paired:
        return img_bytes, info

    # Overlay highlights + collect Y range
    im = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
    overlay = Image.new("RGBA", im.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    scale = dpi / 72.0
    ys: list[float] = []
    for ann in matched + paired:
        r = ann.get("rect")
        if not r or len(r) < 4:
            continue
        x0, y0, x1, y1 = r[0] * scale, r[1] * scale, r[2] * scale, r[3] * scale
        ys.append(y0); ys.append(y1)
        for w, alpha in [(8, 60), (5, 100), (2, 220)]:
            draw.rectangle([x0 - w, y0 - w, x1 + w, y1 + w],
                           outline=(255, 220, 0, alpha), width=w)
    if ys:
        info["y0"] = min(ys)
        info["y1"] = max(ys)
    im2 = Image.alpha_composite(im, overlay)
    buf = io.BytesIO()
    im2.convert("RGB").save(buf, "PNG")
    buf.seek(0)
    return buf.read(), info


# ---------------------------------------------------------------------------
# PDF editing — apply edits + per-PDF versioning
# ---------------------------------------------------------------------------

PDF_HISTORY = OUTPUT / "_pdf_history"


def _flatten_rel(rel: str) -> str:
    """Turn a relative PDF path into a single safe directory name."""
    return re.sub(r"[^\w\-]", "_", rel)[:160]


def snapshot_pdf(pdf_path: Path, tag: str = "") -> Path | None:
    """Copy the PDF into _pdf_history/<flat>/<timestamp>.pdf."""
    try:
        rel = str(pdf_path.relative_to(OUTPUT))
    except ValueError:
        return None
    flat = _flatten_rel(rel)
    hist_dir = PDF_HISTORY / flat
    hist_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    name = f"{ts}{('_' + tag) if tag else ''}.pdf"
    target = hist_dir / name
    try:
        shutil.copy2(pdf_path, target)
    except Exception as e:
        sys.stderr.write(f"[snapshot] {e}\n")
        return None
    return target


def list_pdf_snapshots(pdf_path: Path) -> list[dict]:
    try:
        rel = str(pdf_path.relative_to(OUTPUT))
    except ValueError:
        return []
    flat = _flatten_rel(rel)
    hist_dir = PDF_HISTORY / flat
    if not hist_dir.exists():
        return []
    out = []
    for p in sorted(hist_dir.glob("*.pdf"), reverse=True):
        st = p.stat()
        out.append({
            "name": p.name,
            "id": p.stem,
            "size": st.st_size,
            "mtime": st.st_mtime,
        })
    return out


def restore_pdf_snapshot(pdf_path: Path, snapshot_id: str) -> bool:
    rel = str(pdf_path.relative_to(OUTPUT))
    flat = _flatten_rel(rel)
    src = PDF_HISTORY / flat / f"{snapshot_id}.pdf"
    if not src.exists():
        return False
    # Snapshot current state before overwriting
    snapshot_pdf(pdf_path, tag="pre_restore")
    shutil.copy2(src, pdf_path)
    return True


# Standard appearance string for newly-created Square / FreeText annots,
# matching the convention defined in SKILL.md (red, Helvetica-Bold 9pt).
_DA_BRAND   = "1 0 0 rg /HeBo 9 Tf"
_DS_BRAND   = "font: bold Helvetica,sans-serif 9.0pt; text-align:left; color:#FF0000"


def apply_pdf_edits(pdf_path: Path, edits: list[dict]) -> dict:
    """Apply a list of edit ops to a PDF. Snapshots before write.

    Each edit is one of:
      {action: "update",  xref: int, rect?: [x0,y0,x1,y1], contents?: str}
      {action: "delete",  xref: int}
      {action: "create",  page: int, type: "Square"|"FreeText",
                          rect: [...], contents?: str, fontsize?: int}
    """
    if not edits:
        return {"applied": 0, "errors": 0, "snapshot": None}

    snap = snapshot_pdf(pdf_path, tag="pre_edit")

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        return {"applied": 0, "errors": len(edits), "error": str(e)}

    # Map xref → page_index without keeping annot references alive (those go
    # stale after intervening operations). When we need an annot we re-load
    # it fresh via page.load_annot(int_xref).
    xref_to_page: dict[int, int] = {}
    for pno in range(len(doc)):
        try:
            for entry in doc[pno].annot_xrefs():
                xref = entry[0] if isinstance(entry, tuple) else int(entry)
                xref_to_page[int(xref)] = pno
        except Exception:
            continue

    # Cache page objects so the annots returned by load_annot stay bound
    # (PyMuPDF detaches Annot objects when their page goes out of scope).
    _pages: dict[int, fitz.Page] = {}

    def _get_page(pno: int):
        if pno not in _pages:
            _pages[pno] = doc[pno]
        return _pages[pno]

    def _find_annot(xref: int):
        pno = xref_to_page.get(xref)
        if pno is None:
            return None, None, None
        page = _get_page(pno)
        try:
            ann = page.load_annot(int(xref))
        except Exception:
            return None, None, None
        return pno, page, ann

    applied = 0
    errors = 0
    error_msgs: list[str] = []
    new_xrefs: dict[str, int] = {}  # client_id → real xref (for "create" results)

    for edit in edits:
        try:
            action = edit.get("action")
            if action == "update":
                xref = int(edit["xref"])
                _pno, _page, ann = _find_annot(xref)
                if ann is None:
                    errors += 1
                    error_msgs.append(f"update: xref {xref} not found")
                    continue
                if "rect" in edit:
                    ann.set_rect(fitz.Rect(*edit["rect"]))
                if "contents" in edit:
                    info = ann.info
                    info["content"] = str(edit["contents"])
                    ann.set_info(info)
                # Re-assert standard appearance so legacy annotations
                # (saved with default colors / no explicit stroke) pick up
                # the red theme next time the user edits them.
                try:
                    kind = ann.type[1] if isinstance(ann.type, tuple) else str(ann.type)
                except Exception:
                    kind = ""
                if kind == "Square":
                    try:
                        ann.set_colors(stroke=(1, 0, 0))
                        ann.set_border(width=1)
                    except Exception: pass
                elif kind == "FreeText":
                    try:
                        # In PyMuPDF FreeText, "stroke" is the TEXT color
                        ann.set_colors(stroke=(1, 0, 0))
                    except Exception: pass
                ann.update()
                applied += 1

            elif action == "delete":
                xref = int(edit["xref"])
                pno, page, ann = _find_annot(xref)
                if ann is None:
                    errors += 1
                    error_msgs.append(f"delete: xref {xref} not found")
                    continue
                page.delete_annot(ann)
                # remove from map so future ops don't try to use it
                xref_to_page.pop(xref, None)
                applied += 1

            elif action == "delete_inline":
                # Remove an inline annotation (xref=0) by rewriting the page's
                # /Annots array string. inline_index identifies which inline
                # block to drop, preserving order of indirect refs.
                pno = int(edit["page"]) - 1
                if pno < 0 or pno >= len(doc):
                    errors += 1
                    error_msgs.append("delete_inline: bad page")
                    continue
                page = _get_page(pno)
                idx = int(edit.get("inline_index", -1))
                if idx < 0:
                    errors += 1
                    error_msgs.append("delete_inline: missing inline_index")
                    continue
                if _delete_inline_annot(doc, page, idx):
                    applied += 1
                else:
                    errors += 1
                    error_msgs.append(
                        f"delete_inline: page {pno + 1} idx {idx} not found")

            elif action == "create":
                pno = int(edit["page"]) - 1
                if pno < 0 or pno >= len(doc):
                    errors += 1
                    continue
                page = _get_page(pno)
                rect = fitz.Rect(*edit["rect"])
                ann_type = edit.get("type", "Square")
                if ann_type == "Square":
                    a = page.add_rect_annot(rect)
                    a.set_colors(stroke=(1, 0, 0))
                    a.set_border(width=1)
                    info = a.info
                    info["content"] = str(edit.get("contents", ""))
                    a.set_info(info)
                    a.update()
                elif ann_type == "FreeText":
                    # Compute fontsize from rect height if not provided —
                    # matches the SVG overlay's clamp formula so what the
                    # user sees in edit mode is what they get on disk.
                    rect_h = max(0.0, rect.y1 - rect.y0)
                    default_fs = max(6.0, min(14.0, rect_h * 0.65))
                    fs_raw = edit.get("fontsize")
                    fs = float(fs_raw) if fs_raw not in (None, "") else default_fs
                    fs = max(6.0, min(14.0, fs))
                    # NOTE: PyMuPDF rejects border_color/fill_color unless
                    # rich_text=True. We deliberately render FreeText labels
                    # WITHOUT a border — the SVG overlay matches by drawing
                    # a transparent label rect (just the red text shows).
                    # The paired Square already provides the visible frame.
                    a = page.add_freetext_annot(
                        rect,
                        str(edit.get("contents", "")),
                        fontsize=fs,
                        fontname="helv",
                        text_color=(1, 0, 0),
                        align=0,
                    )
                    a.set_info({"content": str(edit.get("contents", ""))})
                    a.update()
                else:
                    errors += 1
                    continue
                client_id = edit.get("client_id")
                if client_id is not None:
                    new_xrefs[str(client_id)] = a.xref
                applied += 1

            else:
                errors += 1
                error_msgs.append(f"unknown action: {action}")
        except Exception as e:
            errors += 1
            error_msgs.append(f"{edit.get('action')}: {e}")
            sys.stderr.write(f"[apply_pdf_edits] {e}\n")

    # Save with garbage cleanup so removed annots actually leave the file
    tmp = pdf_path.with_suffix(".pdf.tmp")
    try:
        doc.save(str(tmp), garbage=4, clean=True, deflate=True)
    finally:
        doc.close()

    # Atomic-replace via os.open + write (avoids permission quirks on cloud
    # mounts like Google Drive)
    try:
        with open(tmp, "rb") as src:
            data = src.read()
        fd = os.open(str(pdf_path), os.O_WRONLY | os.O_TRUNC)
        try:
            os.write(fd, data)
        finally:
            os.close(fd)
    except Exception:
        # fallback: replace
        shutil.move(str(tmp), str(pdf_path))
    else:
        try:
            tmp.unlink()
        except Exception:
            pass

    return {
        "applied": applied,
        "errors": errors,
        "error_msgs": error_msgs,
        "snapshot": snap.name if snap else None,
        "new_xrefs": new_xrefs,
    }


# ---------------------------------------------------------------------------

def list_pdf_annots(pdf_path: Path) -> list[dict]:
    """Return annotations metadata for all pages.

    Includes both real (xref'd) and inline (literal-dict) annotations so the
    highlight matcher and frontend can see annotations from PDFs that embed
    their /Annots inline (e.g. the เสา catalogs).
    """
    doc = fitz.open(pdf_path)
    out: list[dict] = []
    for pno in range(len(doc)):
        page = doc[pno]
        out.extend(all_annots_on_page(doc, page, pno + 1))
    return out


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template_string(INDEX_HTML)


@app.route("/api/index")
def api_index():
    rows_payload = []
    for r in ROWS:
        # Derive structural role from Col B (mirrors detect_row_role) so the
        # frontend can branch on role without needing a separate API call.
        b_full = r.get("B") or ""
        b = b_full.strip()
        role = "unknown"
        if re.match(r"^\d+(?:\.\d+){2,3}\.\s+", b):
            role = "section_header"
        else:
            m_item = re.match(r"^(\d+)\s*\)\s*", b)
            m_sub = re.match(r"^(\d+)\s*\.\s*", b)
            if m_item and (not m_sub or m_item.end() <= m_sub.end()):
                role = "item"
            elif m_sub:
                role = "sub_item"
        rows_payload.append({
            "row": r["row"],
            "A": r["A"],
            "B": r["B"],
            "C": r["C"],
            "D": r["D"],
            "E": r["E"],
            "F": r["F"],
            "parsed": r["parsed"],
            "pdf_rel": r["pdf_rel"],
            "pdf_inherited": r.get("pdf_inherited", False),
            "section": r["section_inferred"],
            "role": role,
            "auto_flags": auto_check_row(r),
        })
    sections = sorted({r["section_inferred"] for r in ROWS if r.get("section_inferred")},
                      key=lambda s: [int(x) for x in re.findall(r"\d+", s)])
    # serialize tree (children dict → list, preserve insertion order)
    def _serialize(node):
        return {
            "key": node["key"],
            "label": node["label"],
            "type": node["type"],
            "rows": node["rows"],
            "children": [_serialize(c) for c in node["children"].values()],
        }
    return jsonify({
        "rows": rows_payload,
        "sections": sections,
        "tree": _serialize(build_tree()),
        "extra": EXTRA_REFS,
        "tor_available": TOR_PATH is not None or find_tor_pdf() is not None,
        "status": load_status(),
        "version_sync": get_version_sync_status(),
        "stats": {
            "total": len(ROWS),
            "with_pdf_ref": sum(1 for r in ROWS if r["pdf_rel"]),
            "by_type": _counter([r["parsed"].get("type") for r in ROWS]),
        },
    })


def _counter(seq):
    out = {}
    for s in seq:
        out[s] = out.get(s, 0) + 1
    return out


@app.route("/api/status", methods=["GET", "POST"])
def api_status():
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        try:
            row_num = int(data.get("row"))
        except (TypeError, ValueError):
            return jsonify({"error": "row required"}), 400
        # Write through DB (canonical) — this also writes an audit entry
        # whenever the status actually changes.
        entry = db.set_status(
            row_num,
            status=data.get("status") if "status" in data else None,
            notes=data.get("notes") if "notes" in data else None,
            actor="user",
        )
        # Mirror into legacy JSON for compatibility
        full = db.get_all_status()
        save_status(full)
        return jsonify({"ok": True, "entry": entry})
    return jsonify(load_status())


@app.route("/api/pdf_page")
def api_pdf_page():
    rel = request.args.get("rel")
    page = int(request.args.get("page", 1))
    dpi = int(request.args.get("dpi", 130))
    highlight = request.args.get("highlight")
    edit_mode = request.args.get("edit") == "1"
    # WYSIWYG: even in edit mode we bake annots into the rendered image so
    # the user sees the SAME thing they'd see in preview (yellow header
    # banners, custom-appearance labels, etc.). The SVG overlay only adds
    # invisible hit areas + handles for selected annots, plus visible
    # rect/text for newly-drawn (unsaved) ones. Suppressing annots
    # altogether (the old behaviour) made edit-mode look totally different
    # from preview because PyMuPDF's appearance streams (/AP) — including
    # backgrounds and custom fonts — were stripped.
    # An explicit ?bake=0 still forces stripping (used by the wizard when
    # actively editing existing annots so the user doesn't see ghost
    # copies during drag).
    bake_param = request.args.get("bake")
    if bake_param == "0":
        no_annots = True
    else:
        no_annots = False
    if not rel:
        abort(400, "rel required")
    p = (OUTPUT / rel).resolve()
    if not str(p).startswith(str(OUTPUT.resolve())):
        abort(403)
    if not p.exists():
        abort(404)
    png, info = render_pdf_page_png(p, page, dpi=dpi,
                                    highlight=highlight,
                                    no_annots=no_annots)
    resp = send_file(io.BytesIO(png), mimetype="image/png")
    if info.get("y0") is not None:
        resp.headers["X-Highlight-Y0"] = f"{info['y0']:.0f}"
        resp.headers["X-Highlight-Y1"] = f"{info['y1']:.0f}"
    resp.headers["Access-Control-Expose-Headers"] = "X-Highlight-Y0, X-Highlight-Y1"
    # Cache-bust on edit-mode changes by setting no-store
    if edit_mode:
        resp.headers["Cache-Control"] = "no-store"
    return resp


def _resolve_pdf_rel(rel: str) -> Path:
    if not rel:
        abort(400, "rel required")
    p = (OUTPUT / rel).resolve()
    if not str(p).startswith(str(OUTPUT.resolve())):
        abort(403)
    if not p.exists():
        abort(404)
    return p


@app.route("/api/pdf_save", methods=["POST"])
def api_pdf_save():
    """Apply a list of annotation edits to a PDF and snapshot the prior state.

    Body: {"rel": "...", "edits": [...]}
    """
    data = request.get_json(silent=True) or {}
    rel = data.get("rel")
    edits = data.get("edits") or []
    if not rel or not isinstance(edits, list):
        return jsonify({"error": "rel and edits[] required"}), 400
    p = _resolve_pdf_rel(rel)
    result = apply_pdf_edits(p, edits)
    try:
        db.log_audit(action="pdf_edit", target_type="pdf", target_id=rel,
                     details={"edits": len(edits),
                              "applied": result.get("applied", 0),
                              "errors": result.get("errors", 0),
                              "snapshot": result.get("snapshot")},
                     actor="user")
    except Exception: pass
    return jsonify({"ok": True, **result})


@app.route("/api/pdf_history")
def api_pdf_history():
    rel = request.args.get("rel")
    if not rel:
        abort(400)
    p = _resolve_pdf_rel(rel)
    return jsonify({
        "rel": rel,
        "snapshots": list_pdf_snapshots(p),
    })


@app.route("/api/pdf_restore", methods=["POST"])
def api_pdf_restore():
    data = request.get_json(silent=True) or {}
    rel = data.get("rel")
    snap_id = data.get("snapshot")
    if not rel or not snap_id:
        return jsonify({"error": "rel + snapshot required"}), 400
    p = _resolve_pdf_rel(rel)
    ok = restore_pdf_snapshot(p, snap_id)
    if not ok:
        return jsonify({"error": "snapshot not found"}), 404
    return jsonify({"ok": True})


@app.route("/api/pdf_snapshot", methods=["POST"])
def api_pdf_snapshot():
    """Manual snapshot trigger ('save version' button)."""
    data = request.get_json(silent=True) or {}
    rel = data.get("rel")
    tag = (data.get("tag") or "manual").strip()
    tag = re.sub(r"[^\w\-]", "_", tag)[:40]
    if not rel:
        return jsonify({"error": "rel required"}), 400
    p = _resolve_pdf_rel(rel)
    snap = snapshot_pdf(p, tag=tag)
    return jsonify({"ok": True, "name": snap.name if snap else None})


@app.route("/api/pdf_meta")
def api_pdf_meta():
    rel = request.args.get("rel")
    if not rel:
        abort(400)
    p = (OUTPUT / rel).resolve()
    if not str(p).startswith(str(OUTPUT.resolve())):
        abort(403)
    if not p.exists():
        abort(404)
    doc = fitz.open(p)
    meta = {
        "rel": rel,
        "pages": len(doc),
        "annots": list_pdf_annots(p),
        "page_sizes": [(round(doc[i].rect.width, 1), round(doc[i].rect.height, 1))
                       for i in range(len(doc))],
    }
    return jsonify(meta)


@app.route("/api/raw_pdf")
def api_raw_pdf():
    rel = request.args.get("rel")
    if not rel:
        abort(400)
    # Allow PDFs from output, TOR, BOQ
    candidates = [OUTPUT / rel, PROJECT / rel]
    for p in candidates:
        if p.exists() and p.is_file() and p.suffix.lower() == ".pdf":
            return send_file(p, mimetype="application/pdf")
    abort(404)


@app.route("/api/tor_page")
def api_tor_page():
    """Render TOR PDF page with highlight on Col B text from a given row."""
    row = int(request.args.get("row", 0))
    page = request.args.get("page")
    dpi = int(request.args.get("dpi", 110))

    target = next((r for r in ROWS if r["row"] == row), None)
    if not target:
        abort(404, "row not found")
    text = target.get("B") or target.get("C") or ""
    section = target.get("section_inferred")

    # Cache lookup
    if row in TOR_CACHE:
        info = TOR_CACHE[row]
    else:
        info = find_in_tor(text, section=section) or {"page": 1, "rects": []}
        TOR_CACHE[row] = info

    use_page = int(page) if page else info["page"]
    rects = info["rects"] if (use_page == info["page"]) else []
    png = render_tor_page(use_page, rects, dpi=dpi)
    if not png:
        abort(404, "TOR not found")

    resp = send_file(io.BytesIO(png), mimetype="image/png")
    resp.headers["X-TOR-Page"] = str(info["page"])
    resp.headers["X-TOR-Hits"] = str(len(info["rects"]))
    rng = info.get("section_range") or get_tor_section_range(section)
    if rng:
        resp.headers["X-Section-Range"] = f"{rng[0]}-{rng[1]}"
    if section:
        resp.headers["X-Section"] = section
    # image-pixel Y of first matched rect → for auto-scroll
    if rects:
        scale = dpi / 72.0
        y0 = min(r[1] for r in rects) * scale
        y1 = max(r[3] for r in rects) * scale
        resp.headers["X-Highlight-Y0"] = f"{y0:.0f}"
        resp.headers["X-Highlight-Y1"] = f"{y1:.0f}"
    # Allow client JS to read these custom headers (CORS isn't an issue
    # since same-origin, but some browsers still gate non-standard headers)
    resp.headers["Access-Control-Expose-Headers"] = (
        "X-TOR-Page, X-TOR-Hits, X-Highlight-Y0, X-Highlight-Y1, "
        "X-Section, X-Section-Range"
    )
    return resp


@app.route("/api/tor_meta")
def api_tor_meta():
    global TOR_PATH
    if TOR_PATH is None:
        TOR_PATH = find_tor_pdf()
    if TOR_PATH is None:
        return jsonify({"available": False})
    doc = fitz.open(TOR_PATH)
    return jsonify({
        "available": True,
        "filename": TOR_PATH.name,
        "pages": len(doc),
    })


@app.route("/api/row/col_d", methods=["POST"])
def api_row_col_d():
    """Inline Col D edit (from xlsx preview double-click). Records the
    change as HITL feedback so the learner can mine repeated patterns."""
    data = request.get_json(silent=True) or {}
    try:
        row_num = int(data["row"])
    except (TypeError, ValueError, KeyError):
        return jsonify({"ok": False, "error": "row required"}), 400
    new_d = (data.get("col_d") or "").strip()
    original = (data.get("original") or "").strip()

    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404

    # Pre-snapshot for safety
    _run_version_cmd(["snap", f"pre-inline-row-{row_num}"], timeout=60)

    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        ws.cell(row_num, 4).value = new_d
        tmp = XLSX_PATH.with_suffix(".xlsx.tmp")
        wb.save(str(tmp))
        with open(tmp, "rb") as f:
            buf = f.read()
        fd = os.open(str(XLSX_PATH), os.O_WRONLY | os.O_TRUNC)
        try:
            os.write(fd, buf)
        finally:
            os.close(fd)
        tmp.unlink(missing_ok=True)
    except Exception as e:
        return jsonify({"ok": False, "error": f"xlsx write failed: {e}"}), 500

    # Refresh + log
    load_rows()
    sync_db_from_memory()
    try:
        db.log_audit(action="col_d_inline_edit",
                     target_type="row", target_id=str(row_num),
                     before=original, after=new_d, actor="user")
    except Exception: pass
    try:
        learning.record_feedback(
            row_num=row_num,
            section=row.get("section_inferred"),
            input_b=row.get("B", "") or "",
            input_pdf_rel=row.get("pdf_rel"),
            input_role=detect_row_role(row_num).get("role", ""),
            input_filename=(Path(row["pdf_rel"]).name if row.get("pdf_rel") else None),
            suggested_c=row.get("C", "") or "",
            suggested_d=original,
            suggested_annots=[],
            confidence=0.0,
            generator="prior",
            provenance={},
            user_action="edited",
            final_c=row.get("C", "") or "",
            final_d=new_d,
            final_annots=[],
        )
    except Exception as e:
        sys.stderr.write(f"[col_d_inline] feedback: {e}\n")

    return jsonify({"ok": True, "row": row_num, "new_d": new_d})


@app.route("/api/row/col_d/suggest")
def api_row_col_d_suggest():
    """Phase B3: live autocomplete for the Col D inline editor.

    Returns up to 6 ranked candidates: the AI's auto-annotate proposal
    (top-priority), Col D shapes from same-section verified rows, and
    a couple of generic shape templates so users can tab-complete.

    Cheap path — no LLM call here. The expensive proposal comes from
    auto_annotate_plan() which is already cache-friendly via section
    + filename derivations.
    """
    try:
        row_num = int(request.args.get("row", "0"))
    except ValueError:
        return jsonify({"ok": False, "error": "row required"}), 400
    q = (request.args.get("q") or "").strip()
    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404

    suggestions: list[dict] = []
    section = row.get("section_inferred") or ""

    # 1. AI proposal from auto_annotate_plan (rule+pattern, no LLM)
    try:
        plan = auto_annotate_plan(row_num)
        if plan.get("ok") and plan.get("proposed_d"):
            suggestions.append({
                "text": plan["proposed_d"],
                "kind": "ai",
                "label": "AI proposal",
                "confidence": round(float(plan.get("confidence") or 0.0), 2),
                "generator": plan.get("generator", "rules"),
            })
    except Exception:
        pass

    # 2. Col D values from verified rows in the same section root
    #    (e.g. for section 5.1.2, suggest Col D values from any verified
    #    5.1.* row). This gives users a "match the neighbor's style"
    #    completion. De-dupe on text.
    seen = {s["text"].strip() for s in suggestions}
    if section:
        section_root = ".".join(section.split(".")[:2])  # "5.1.2" → "5.1"
        try:
            statuses = db.get_all_status()
        except Exception:
            statuses = {}
        for r in ROWS:
            if r["row"] == row_num: continue
            sec = r.get("section_inferred") or ""
            if not sec.startswith(section_root): continue
            d = (r.get("D") or "").strip()
            if not d or d in seen: continue
            # Prefer rows with a non-unverified verdict
            st = (statuses.get(str(r["row"]), {}) or {}).get("status", "unverified")
            if st in ("pass", "need_fix"):  # known good/edited shapes
                suggestions.append({
                    "text": d, "kind": "neighbor",
                    "label": f"R{r['row']} · {sec}",
                    "confidence": 0.7 if st == "pass" else 0.5,
                    "generator": f"section:{sec}",
                })
                seen.add(d)
            if len(suggestions) >= 6: break

    # 3. Shape templates if we still have room — handy when a user
    #    types a section ref and wants to expand to canonical form.
    if len(suggestions) < 6:
        templates = [
            {"text": f"เอกสาร {section} ... หน้า ?", "kind": "shape",
             "label": "doc-ref (dot form)", "confidence": 0.3},
            {"text": "ยินดีปฏิบัติตามข้อกำหนด", "kind": "shape",
             "label": "commitment", "confidence": 0.3},
        ]
        for t in templates:
            if t["text"] in seen: continue
            suggestions.append(t)
            seen.add(t["text"])
            if len(suggestions) >= 6: break

    # If user has typed a query, prefer suggestions whose text contains q
    # (case-insensitive), but always keep AI proposal at top.
    if q:
        ql = q.casefold()
        def _score(s):
            t = s["text"].casefold()
            kind_w = {"ai": 100, "neighbor": 50, "shape": 10}.get(s["kind"], 0)
            match_w = 30 if ql in t else 0
            start_w = 20 if t.startswith(ql) else 0
            return -(kind_w + match_w + start_w + s.get("confidence", 0) * 10)
        suggestions.sort(key=_score)

    return jsonify({"ok": True, "row": row_num, "section": section,
                    "suggestions": suggestions[:6]})


@app.route("/api/row_context")
def api_row_context():
    """Return rows surrounding the target row for the xlsx preview pane."""
    row = int(request.args.get("row", 0))
    radius = int(request.args.get("radius", 6))
    idx = next((i for i, r in enumerate(ROWS) if r["row"] == row), -1)
    if idx < 0:
        abort(404)
    lo = max(0, idx - radius)
    hi = min(len(ROWS), idx + radius + 1)
    out = []
    for i in range(lo, hi):
        r = ROWS[i]
        out.append({
            "row": r["row"],
            "A": r["A"], "B": r["B"], "C": r["C"],
            "D": r["D"], "E": r["E"], "F": r["F"],
            "type": r["parsed"].get("type"),
        })
    return jsonify({"target": row, "rows": out})


@app.route("/api/refresh")
def api_refresh():
    """Re-scan filesystem + xlsx (in case user edited files)."""
    build_pdf_index()
    load_rows()
    collect_extra_refs()
    build_tor_section_index()
    index_tor_text()
    TOR_CACHE.clear()
    sync_db_from_memory()
    try:
        db.log_audit(action="refresh", target_type="project",
                     details={"rows": len(ROWS), "pdfs": len(PDF_INDEX)},
                     actor="user")
    except Exception: pass
    return jsonify({"ok": True, "rows": len(ROWS), "pdfs": len(PDF_INDEX),
                    "tor_sections": len(TOR_SECTION_INDEX),
                    "tor_pages_indexed": len(TOR_PAGE_TEXTS)})


# ---------------------------------------------------------------------------
# Auto-annotate pipeline
#
# For rows that have a resolved catalog PDF but empty Col D, we can:
#   1. Extract brand/model from the catalog filename
#   2. Detect the row's role (section header / item / sub-item)
#   3. For sub-items: text-search the PDF for distinctive Col B tokens to
#      pinpoint which page the spec lives on; compute a content rect.
#   4. Generate proposed Col D + (optional) PDF Square + FreeText annotations
#   5. Snapshot before writing, then commit changes to xlsx + PDF
#
# Operates in dry-run by default (returns a "plan") so the user can review
# each row before applying. Batch mode processes all rows with the
# `needs_col_d` flag.
# ---------------------------------------------------------------------------

# Tokens that look like brand names in filenames but actually describe the
# product category. Skipped when picking the brand.
_FILENAME_UNITS = {
    "kVA", "KVA", "kW", "kWh", "kHz", "MHz", "GHz", "Hz", "kV", "mV", "mA",
    "mm", "cm", "kg", "ms", "us", "ns", "Mbps", "Gbps", "DC", "AC", "VAC",
    "VDC", "ms.", "rpm", "RPM", "lm", "lx", "ID", "OD",
    # Generic product-category acronyms (NAS / PoE / AP / etc.)
    "NAS", "PoE", "POE", "AP", "IP", "USB", "HDMI", "RJ45", "SFP", "SFP+",
    "L2", "L3", "PC", "TB", "GB", "MB", "RAM", "ROM", "CPU", "GPU", "UPS",
    "NVR", "DVR", "CCTV", "PDU", "FO", "VCT", "EMT", "HDPE",
}

# Compare phrases per SKILL.md (used when generating Col C from Col B)
_COMPARE_PHRASES_STRIP = [
    ("หรือดีกว่า", ""), ("ไม่น้อยไปกว่า", ""), ("ไม่น้อยกว่า", ""),
    ("ไม่มากกว่า", ""), ("ต้องสามารถ", "สามารถ"), ("จะต้อง", ""),
]


def parse_brand_model_from_filename(stem: str) -> tuple[str, str]:
    """Heuristic: strip section + Thai description, take first Latin token as
    brand and the rest as model.

    Examples:
      '5.1.1.2. เครื่องคอมพิวเตอร์แม่ข่าย แบบที่ 2 Lenovo ThinkSystem SR630 V4'
        → ('Lenovo', 'ThinkSystem SR630 V4')
      '5.1.1.7. เครื่องสำรองไฟฟ้า ขนาด 10 kVA (...) cleanline t10k33lv2'
        → ('Cleanline', 't10k33lv2')
    """
    s = stem.strip()
    # Strip leading section number
    s = re.sub(r"^\d+(?:\.\d+){1,3}\.?\s+", "", s)
    # Strip "แบบที่ N" markers (variant labels)
    s = re.sub(r"แบบที่\s*\d+\s*", "", s)
    # Strip parenthesised phrases (often Thai notes)
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"\s+", " ", s).strip()

    tokens = s.split()
    brand_idx = -1
    for i, tok in enumerate(tokens):
        if tok in _FILENAME_UNITS:
            continue
        # Token must start with ≥2 ASCII letters (skips raw digits + units)
        m = re.match(r"^[A-Za-z]{2,}", tok)
        if m:
            brand_idx = i
            break
    if brand_idx < 0:
        return ("", "")
    brand = tokens[brand_idx]
    # Title-case the brand if it's all-lowercase (e.g. "cleanline" → "Cleanline")
    if brand and brand.islower():
        brand = brand[0].upper() + brand[1:]
    model = " ".join(tokens[brand_idx + 1:]).strip()
    return (brand, model)


def extract_section_name(stem: str, section: str) -> str:
    """Strip section prefix + brand/model tail from filename, leaving the
    Thai description (used in Col D between the catalog ref and the page).

    '5.1.1.2. เครื่องคอมพิวเตอร์แม่ข่าย แบบที่ 2 Lenovo ThinkSystem SR630 V4'
      → 'เครื่องคอมพิวเตอร์แม่ข่าย แบบที่ 2'
    """
    brand, model = parse_brand_model_from_filename(stem)
    s = stem.strip()
    s = re.sub(r"^\d+(?:\.\d+){1,3}\.?\s+", "", s)
    if brand:
        # remove brand and everything after
        idx = s.find(brand)
        if idx > 0:
            s = s[:idx].strip()
    # Strip trailing "-N" if filename was a parent rack
    s = re.sub(r"\s*-\s*\d+\s*$", "", s).strip()
    return s


def detect_row_role(row_num: int) -> dict:
    """Determine whether this row is a section header, an item (N)), or a
    sub-item (N.). Returns the structural location used by Col D format
    generation."""
    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return {"role": "unknown"}
    b_full = row.get("B") or ""
    b = b_full.strip()
    section = row.get("section_inferred", "")

    # Section header: B starts with full section number + period (3+ levels)
    if re.match(r"^\d+(?:\.\d+){2,3}\.\s+", b):
        return {"role": "section_header", "section": section}

    # Item: B starts with "N)" (after some indent)
    m_item = re.match(r"^(\d+)\s*\)\s*", b)
    # Sub-item: B starts with "N." (after deeper indent)
    m_sub = re.match(r"^(\d+)\s*\.\s*", b)

    # Distinguish item vs sub-item by indent depth in original Col B
    indent_a = len(b_full) - len(b_full.lstrip())
    if m_item and (not m_sub or m_item.end() <= m_sub.end()):
        return {"role": "item", "item_num": int(m_item.group(1)),
                "section": section, "indent": indent_a}
    if m_sub:
        # Find the most recent preceding item-row in the same section
        parent_item = None
        for prev in ROWS:
            if prev["row"] >= row_num:
                continue
            if prev.get("section_inferred") != section:
                continue
            pb = (prev.get("B") or "").strip()
            pm = re.match(r"^(\d+)\s*\)\s*", pb)
            if pm:
                parent_item = int(pm.group(1))
        return {"role": "sub_item", "sub_num": int(m_sub.group(1)),
                "parent_item": parent_item, "section": section, "indent": indent_a}

    return {"role": "unknown", "section": section}


def col_b_search_tokens(b: str) -> list[str]:
    """Distinctive tokens from Col B that should also appear in catalog text.
    Prefers English/code/numeric+unit because Thai compound matching is
    fragile across catalog encodings."""
    s = re.sub(r"^\s*\d+\s*[\)\.]\s*", "", b or "")
    s = re.sub(r"^\s*\d+(?:\.\d+)*\.\s+", "", s)
    seen: set[str] = set()
    out: list[str] = []

    def add(t: str):
        t = t.strip()
        if len(t) < 3 or t.lower() in seen:
            return
        seen.add(t.lower())
        out.append(t)

    for m in re.finditer(r"[A-Za-z][A-Za-z0-9]{2,}(?:[-+/.][A-Za-z0-9]+)*", s):
        add(m.group(0))
    for m in re.finditer(r"(?:IEC|ISO|ASTM|ANSI|UL|TIS|มอก\.?)\s*\d+(?:[-/.\s]\d+)*", s):
        add(m.group(0))
    for m in re.finditer(r"\b(?:IP\d+|DC\s*\d+(?:-\d+)?\s*V|AC\s*\d+\s*V)\b", s):
        add(m.group(0))
    for m in re.finditer(
        r"\d+(?:[\.,]\d+)?\s*(?:mA|MHz|GHz|kHz|Hz|°C|°F|mm|cm|m|kg|hr|min|kV|V|A|W|kW|Mbps|Gbps|MB|GB|TB|byte)",
        s, re.IGNORECASE,
    ):
        add(m.group(0))
    return out


def find_text_match_in_pdf(pdf_path: Path, b_text: str,
                            min_score: int = 2) -> dict | None:
    """Find best-matching page for `b_text` in `pdf_path`. Returns
    {page, rects, score, matched_tokens} or None."""
    tokens = col_b_search_tokens(b_text)
    if not tokens or not pdf_path.exists():
        return None
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return None
    best = None
    for pno in range(len(doc)):
        page = doc[pno]
        score = 0
        rects: list = []
        matched: list[str] = []
        for tok in tokens:
            for variant in _thai_variants(tok):
                try:
                    rs = page.search_for(variant, quads=False)
                except Exception:
                    rs = []
                if rs:
                    score += 1
                    matched.append(tok)
                    rects.extend(rs[:2])
                    break
        if score and (best is None or score > best["score"]):
            best = {"page": pno + 1, "score": score,
                    "matched_tokens": matched,
                    "rects": [[r.x0, r.y0, r.x1, r.y1] for r in rects]}
    if best and best["score"] >= max(min_score, len(tokens) // 4):
        return best
    return None


def merge_rects_to_content(rects: list, padding: float = 4.0,
                           max_height: float = 80.0) -> list:
    """Combine matched-token rects into a single content rect. If the rects
    span too tall a region (>max_height), use only the topmost cluster to
    avoid over-broad highlights."""
    if not rects:
        return [0, 0, 0, 0]
    sorted_by_y = sorted(rects, key=lambda r: r[1])
    cluster = [sorted_by_y[0]]
    for r in sorted_by_y[1:]:
        if r[1] - cluster[-1][3] < 25:  # within 25pt of previous → same cluster
            cluster.append(r)
        else:
            break
    x0 = min(r[0] for r in cluster) - padding
    y0 = min(r[1] for r in cluster) - padding
    x1 = max(r[2] for r in cluster) + padding
    y1 = max(r[3] for r in cluster) + padding
    return [max(0, x0), max(0, y0), x1, min(y1, y0 + max_height)]


def compute_label_rect(content_rect: list, page_w: float, page_h: float,
                       label_w: float = 130, label_h: float = 14) -> list:
    """Place a FreeText label in whitespace next to the content rect.
    Prefers right side; falls back to below if right doesn't fit."""
    x0, y0, x1, y1 = content_rect
    # Right of rect
    rx0 = x1 + 5
    if rx0 + label_w <= page_w - 5:
        ry0 = max(0, (y0 + y1) / 2 - label_h / 2)
        return [rx0, ry0, rx0 + label_w, ry0 + label_h]
    # Below rect
    by0 = y1 + 4
    if by0 + label_h <= page_h - 5:
        bx0 = x0
        return [bx0, by0, bx0 + label_w, by0 + label_h]
    # Above rect
    ay1 = y0 - 4
    return [x0, ay1 - label_h, x0 + label_w, ay1]


def make_col_d_for_row(row: dict, role_info: dict, pdf_path: Path,
                       page: int | None) -> str:
    """Compose Col D string per SKILL.md convention."""
    role = role_info.get("role")
    section = role_info.get("section") or ""
    stem = pdf_path.stem
    cat_full = stem  # full filename (no .pdf)

    # Determine the "ref" portion: prefer "{parent}-{N}" form when filename
    # ends in -N, else use the full filename (newer convention from
    # clone_*.py scripts).
    ref = cat_full
    sec_name = ""
    parent_folder = pdf_path.parent.name
    fm = re.match(r"^(\d+(?:\.\d+){1,3})\.?\s*-(\d+)\b", parent_folder)
    if fm:
        ref = f"{fm.group(1)}-{fm.group(2)}"
        sec_name = extract_section_name(stem, section)

    if role == "section_header":
        brand, model = parse_brand_model_from_filename(stem)
        if brand and model:
            return f"ยี่ห้อ {brand} รุ่น {model}"
        if brand:
            return f"ยี่ห้อ {brand}"
        return ""

    if role == "item":
        n = role_info.get("item_num", 1)
        if page is None:
            return ""
        if sec_name:
            return f"เทียบเท่าข้อกำหนด เอกสาร {ref} {sec_name} หน้า {page} ข้อ {section} ข้อ {n})"
        return f"เทียบเท่าข้อกำหนด เอกสาร {ref} หน้า {page} ข้อ {section} ข้อ {n})"

    if role == "sub_item":
        sub = role_info.get("sub_num", 1)
        parent = role_info.get("parent_item")
        if page is None:
            return ""
        if parent is not None:
            tail = f"ข้อ {section} ข้อ {parent}) ข้อย่อย {sub}."
        else:
            tail = f"ข้อ {section} ข้อย่อย {sub}."
        if sec_name:
            return f"เทียบเท่าข้อกำหนด เอกสาร {ref} {sec_name} หน้า {page} {tail}"
        return f"เทียบเท่าข้อกำหนด เอกสาร {ref} หน้า {page} {tail}"

    return ""


def make_col_c_from_b(b: str) -> str:
    """Strip indentation, leading numbering, and compare phrases."""
    if not b:
        return ""
    s = b.strip()
    for needle, repl in _COMPARE_PHRASES_STRIP:
        s = s.replace(needle, repl)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def auto_annotate_plan(row_num: int) -> dict:
    """Generate a dry-run plan for one row. Caller can inspect + apply.

    HITL pipeline:
      1. Apply learned patterns first (highest priority — user-validated).
      2. Fall through to rule-based generation (filename parsing, text
         search, etc.) for the long tail.
      3. (Optional) ask the LLM provider when both above produce low
         confidence.
      4. Compute a confidence score + provenance trail for the UI.
    """
    row = next((r for r in ROWS if r["row"] == row_num), None)
    if row is None:
        return {"ok": False, "error": f"row {row_num} not found"}

    plan: dict = {
        "ok": True,
        "row": row_num,
        "section": row.get("section_inferred"),
        "current_d": row.get("D") or "",
        "proposed_d": "",
        "current_c": row.get("C") or "",
        "proposed_c": "",
        "annotations": [],
        "warnings": [],
        "generator": "rules",
        "provenance": {},
        "confidence": 0.0,
    }

    pdf_rel = row.get("pdf_rel")
    if not pdf_rel:
        plan["ok"] = False
        plan["warnings"].append("ไม่มี PDF ที่อ้างอิง — ไม่สามารถ auto-annotate ได้")
        return plan
    pdf_path = OUTPUT / pdf_rel
    plan["pdf_rel"] = pdf_rel
    plan["pdf_filename"] = pdf_path.name

    role_info = detect_row_role(row_num)
    plan["role"] = role_info

    b = row.get("B") or ""
    proposed_c = make_col_c_from_b(b)
    plan["proposed_c"] = proposed_c

    role = role_info.get("role")

    if role == "section_header":
        # 1. learned filename_brand pattern wins if it fires
        learned_brand, prov = learning.apply_learned_brand(pdf_path.stem)
        if learned_brand:
            # use the rule-based model parser to get the model part, but
            # override the brand with what the user historically picked
            _, model = parse_brand_model_from_filename(pdf_path.stem)
            proposed_d = (f"ยี่ห้อ {learned_brand} รุ่น {model}".strip()
                          if model else f"ยี่ห้อ {learned_brand}")
            plan["proposed_d"] = proposed_d
            plan["generator"] = "rules+pattern"
            plan["provenance"] = {"brand": prov}
        else:
            proposed_d = make_col_d_for_row(row, role_info, pdf_path, None)
            plan["proposed_d"] = proposed_d
        plan["meta"] = {"brand_model_only": True}
        if not plan["proposed_d"]:
            plan["warnings"].append("ไม่สามารถ parse brand/model จาก filename")
        plan["confidence"] = learning.confidence_score(
            generator=plan["generator"], provenance=plan["provenance"],
            role="section_header", has_match=bool(plan["proposed_d"]),
            warnings=len(plan["warnings"]),
        )
        # Claude refinement for section headers (brand/model decomposition)
        plan = _maybe_refine_with_claude(plan, row, role_info)
        return plan

    if role in ("item", "sub_item"):
        match = find_text_match_in_pdf(pdf_path, b)
        if not match:
            plan["warnings"].append("ไม่เจอข้อความ Col B ใน catalog (อาจเป็น commitment)")
            plan["proposed_d"] = "ยินดีปฏิบัติตามข้อกำหนด"
            plan["fallback"] = "commitment"
            plan["confidence"] = learning.confidence_score(
                generator=plan["generator"], provenance=plan["provenance"],
                role=role, has_match=False, warnings=len(plan["warnings"]),
            )
            # Claude refinement — text search missed but Claude may know better
            plan = _maybe_refine_with_claude(plan, row, role_info)
            return plan
        page = match["page"]
        plan["match"] = match
        plan["proposed_d"] = make_col_d_for_row(row, role_info, pdf_path, page)

        content_rect = merge_rects_to_content(match["rects"])
        try:
            doc = fitz.open(pdf_path)
            pg = doc[page - 1]
            page_w, page_h = pg.rect.width, pg.rect.height
        except Exception:
            page_w, page_h = 595, 842
        label_rect = compute_label_rect(content_rect, page_w, page_h)
        if role == "sub_item" and role_info.get("parent_item") is not None:
            label = f"{role_info['section']} ข้อ {role_info['parent_item']}) ข้อย่อย {role_info['sub_num']}."
        elif role == "sub_item":
            label = f"{role_info['section']} ข้อย่อย {role_info['sub_num']}."
        else:
            label = f"{role_info['section']} ข้อ {role_info['item_num']})"

        plan["annotations"] = [
            {"page": page, "type": "Square",
             "rect": [round(v, 2) for v in content_rect], "contents": ""},
            {"page": page, "type": "FreeText",
             "rect": [round(v, 2) for v in label_rect], "contents": label},
        ]
        plan["confidence"] = learning.confidence_score(
            generator=plan["generator"], provenance=plan["provenance"],
            role=role, has_match=True, warnings=len(plan["warnings"]),
        )
        # Claude refinement (if installed)
        plan = _maybe_refine_with_claude(plan, row, role_info)
        return plan

    plan["warnings"].append(f"row role '{role}' — ไม่รองรับการ auto-annotate")
    plan["ok"] = False
    plan["confidence"] = 0.0
    return plan


def _maybe_refine_with_claude(plan: dict, row: dict, role_info: dict) -> dict:
    """If a Claude provider is installed AND the rule-based plan has low
    confidence, ask Claude for a structured proposal and merge into the plan.

    Strategy:
      • plan.confidence ≥ 0.85 → keep rules (no spend)
      • plan.confidence < 0.85 → ask Claude, override col_d/col_c if Claude
        returns a propose_col_d tool call with higher confidence
      • Always record the Claude call's audit for analytics

    Failure modes are non-fatal — Claude errors fall through to rule output.
    """
    try:
        from app import anthropic_provider as ap
    except Exception:
        return plan

    provider = ap.get_provider()
    if provider is None:
        return plan

    if plan.get("confidence", 0) >= 0.85:
        # rules already confident — skip the spend
        plan["llm"] = {"skipped": "high_confidence_rules"}
        return plan

    # Find a few similar past corrections for in-context learning
    few_shot: list[dict] = []
    try:
        few_shot = _claude_few_shot_for(row, role_info, limit=4)
    except Exception as e:
        sys.stderr.write(f"[claude-refine] few_shot failed: {e}\n")

    row_context = {
        "row": row.get("row"),
        "section": row.get("section_inferred"),
        "role": role_info.get("role"),
        "col_a": row.get("A"),
        "col_b": row.get("B") or "",
        "col_c_current": row.get("C") or "",
        "col_d_current": row.get("D") or "",
        "col_e": row.get("E") or "",
        "pdf_rel": row.get("pdf_rel"),
        "pdf_filename": Path(row["pdf_rel"]).name if row.get("pdf_rel") else None,
        "tor_excerpt": _claude_tor_excerpt(row),
        "rule_proposal": {
            "proposed_d": plan.get("proposed_d", ""),
            "generator": plan.get("generator", "rules"),
            "confidence": plan.get("confidence", 0),
        },
        "_few_shot": few_shot,
    }

    try:
        result = provider.propose(row_context=row_context, few_shot=few_shot)
    except ap.BudgetExceededError as e:
        plan["llm"] = {"error": "budget_exceeded", "msg": str(e)}
        plan["warnings"].append(f"Claude skipped: {e}")
        return plan
    except Exception as e:
        plan["llm"] = {"error": str(e)}
        sys.stderr.write(f"[claude-refine] {e}\n")
        return plan

    if not result.get("ok"):
        plan["llm"] = {"error": result.get("error", "unknown")}
        return plan

    # Merge Claude's proposal into plan
    plan["llm"] = {
        "model": result.get("model"),
        "tokens": result.get("usage"),
        "cost_usd": round(result.get("cost_usd", 0), 6),
        "elapsed_ms": result.get("elapsed_ms"),
        "tool_calls": result.get("tool_calls", []),
        "rationale": "",
        "claude_confidence": 0.0,
        "escalation": None,
    }

    for tc in result.get("tool_calls", []):
        name = tc.get("name")
        inp = tc.get("input", {}) or {}
        if name == "propose_col_d":
            claude_d = (inp.get("col_d_text") or "").strip()
            claude_c = (inp.get("col_c_proposed") or "").strip()
            claude_conf = float(inp.get("confidence", 0))
            plan["llm"]["rationale"] = inp.get("rationale", "")
            plan["llm"]["claude_confidence"] = claude_conf
            plan["llm"]["pattern"] = inp.get("pattern", "")
            # Override only if Claude's confidence beats the rule
            if claude_d and claude_conf > plan.get("confidence", 0):
                plan["proposed_d"] = claude_d
                plan["generator"] = f"claude+{plan.get('generator', 'rules')}"
                plan["confidence"] = max(plan.get("confidence", 0), claude_conf)
                plan["provenance"] = {
                    **plan.get("provenance", {}),
                    "claude_pattern": inp.get("pattern", ""),
                    "claude_rationale": inp.get("rationale", ""),
                }
                if claude_c:
                    plan["proposed_c"] = claude_c
                if inp.get("page_in_catalog"):
                    plan["claude_page"] = int(inp["page_in_catalog"])
        elif name == "propose_brand_model":
            brand = (inp.get("brand") or "").strip()
            model = (inp.get("model") or "").strip()
            if brand and brand != "-" and model:
                # For section_header rows, override Col D with Claude's brand+model
                claude_d = f"ยี่ห้อ {brand} รุ่น {model}"
                if float(inp.get("confidence", 0)) > plan.get("confidence", 0):
                    plan["proposed_d"] = claude_d
                    plan["generator"] = "claude+brand_model"
                    plan["confidence"] = float(inp.get("confidence", 0))
                    plan["provenance"]["claude_brand"] = brand
                    plan["provenance"]["claude_model"] = model
            elif brand == "-" and model:
                plan["proposed_d"] = f"ยี่ห้อ - รุ่น {model}"
                plan["generator"] = "claude+brand_model_fabricate"
                plan["confidence"] = float(inp.get("confidence", 0))
            plan["llm"]["rationale"] = inp.get("rationale", "")
            plan["llm"]["claude_confidence"] = float(inp.get("confidence", 0))
        elif name == "escalate_to_user":
            plan["llm"]["escalation"] = {
                "question": inp.get("question", ""),
                "options": inp.get("options", []),
                "context": inp.get("context", ""),
            }
            plan["warnings"].append(
                f"Claude escalated: {inp.get('question', '')[:100]}"
            )

    return plan


def _claude_few_shot_for(row: dict, role_info: dict, limit: int = 4) -> list[dict]:
    """Pull the most-similar past corrections from learning_feedback to use
    as in-context examples. Heuristic: same section root + same role +
    user_action='edited' (i.e. the user actually fixed something useful)."""
    section = row.get("section_inferred") or ""
    sec_root = ".".join(section.split(".")[:2]) if section else ""
    role = role_info.get("role") or ""
    out: list[dict] = []
    try:
        with db.conn() as c:
            rows = c.execute(
                """SELECT input_b, final_d, generator, correction_kind
                   FROM learning_feedback
                   WHERE user_action = 'edited'
                     AND section LIKE ?
                     AND input_role = ?
                   ORDER BY ts DESC LIMIT ?""",
                (sec_root + ".%", role, limit),
            ).fetchall()
            for r in rows:
                out.append({
                    "input_b": r["input_b"],
                    "final_d": r["final_d"],
                    "generator": r["generator"],
                    "correction_kind": r["correction_kind"],
                })
    except Exception as e:
        sys.stderr.write(f"[claude-fewshot] {e}\n")
    return out


def _claude_tor_excerpt(row: dict, max_chars: int = 800) -> str:
    """Pull the TOR text snippet for the row's section, if available."""
    try:
        section = row.get("section_inferred") or ""
        if not section:
            return ""
        with db.conn() as c:
            r = c.execute(
                """SELECT page_text FROM tor_pages
                   JOIN tor_sections ON tor_sections.page_num = tor_pages.page_num
                   WHERE tor_sections.section = ? LIMIT 1""",
                (section,),
            ).fetchone()
            if r and r["page_text"]:
                return r["page_text"][:max_chars]
    except Exception:
        pass
    return ""


def apply_auto_annotate_plan(plan: dict, write_pdf: bool = True,
                              write_xlsx: bool = True) -> dict:
    """Commit a plan: snapshot first, then write Col C/D in xlsx and add
    annotations to the catalog PDF."""
    if not plan.get("ok"):
        return {"ok": False, "error": "plan is not OK"}

    result = {"ok": True, "row": plan.get("row"),
              "xlsx_written": False, "pdf_written": False, "snapshot": None}

    # Take a quick snap first (xlsx only — fast)
    snap = _run_version_cmd(["snap", f"pre-auto-row-{plan.get('row')}"], timeout=60)
    if snap.get("ok"):
        result["snapshot"] = "pre-snap created"

    # Write xlsx
    if write_xlsx and plan.get("proposed_d"):
        try:
            wb = openpyxl.load_workbook(XLSX_PATH)
            ws = wb.active
            r = plan["row"]
            if plan.get("proposed_d"):
                ws.cell(r, 4).value = plan["proposed_d"]
            if plan.get("proposed_c") and not (ws.cell(r, 3).value or "").strip():
                ws.cell(r, 3).value = plan["proposed_c"]
            # Save with O_TRUNC workaround for Google Drive
            tmp = XLSX_PATH.with_suffix(".xlsx.tmp")
            wb.save(str(tmp))
            with open(tmp, "rb") as f:
                data = f.read()
            fd = os.open(str(XLSX_PATH), os.O_WRONLY | os.O_TRUNC)
            try:
                os.write(fd, data)
            finally:
                os.close(fd)
            tmp.unlink(missing_ok=True)
            result["xlsx_written"] = True
        except Exception as e:
            result["ok"] = False
            result["xlsx_error"] = str(e)

    # Write PDF annotations
    if write_pdf and plan.get("annotations") and plan.get("pdf_rel"):
        try:
            pdf_path = OUTPUT / plan["pdf_rel"]
            edits = []
            for a in plan["annotations"]:
                edits.append({
                    "action": "create",
                    "page": a["page"],
                    "type": a["type"],
                    "rect": a["rect"],
                    "contents": a["contents"],
                    "fontsize": 9,
                })
            r = apply_pdf_edits(pdf_path, edits)
            result["pdf_written"] = (r.get("applied", 0) > 0)
            result["pdf_result"] = r
        except Exception as e:
            result["ok"] = False
            result["pdf_error"] = str(e)

    return result


@app.route("/api/auto_annotate/preview")
def api_auto_annotate_preview():
    row_num = int(request.args.get("row", 0))
    if not row_num:
        return jsonify({"ok": False, "error": "row required"}), 400
    return jsonify(auto_annotate_plan(row_num))


@app.route("/api/auto_annotate/apply", methods=["POST"])
def api_auto_annotate_apply():
    data = request.get_json(silent=True) or {}
    row_num = int(data.get("row", 0))
    if not row_num:
        return jsonify({"ok": False, "error": "row required"}), 400
    plan = auto_annotate_plan(row_num)
    if not plan.get("ok"):
        return jsonify({"ok": False, "error": "no plan", "plan": plan}), 400
    write_pdf = bool(data.get("write_pdf", True))
    write_xlsx = bool(data.get("write_xlsx", True))
    # Record the plan first so even failed applies leave a trace
    try:
        plan_id = db.record_plan(row_num, plan)
    except Exception:
        plan_id = None
    result = apply_auto_annotate_plan(plan, write_pdf=write_pdf, write_xlsx=write_xlsx)
    if result.get("ok"):
        load_rows()
        TOR_CACHE.clear()
        sync_db_from_memory()
        if plan_id:
            try: db.mark_plan_applied(plan_id, result)
            except Exception: pass
        # HITL: record feedback. The user can override the suggested Col D
        # by passing `final_d` in the body; otherwise we treat the apply
        # as "accepted as proposed".
        try:
            final_d = data.get("final_d") if "final_d" in data else plan.get("proposed_d")
            final_c = data.get("final_c") if "final_c" in data else plan.get("proposed_c")
            user_action = data.get("user_action") or (
                "edited" if final_d != plan.get("proposed_d") else "accepted"
            )
            learning.record_feedback(
                row_num=row_num,
                section=plan.get("section"),
                input_b=(next((r["B"] for r in ROWS if r["row"] == row_num), "") or ""),
                input_pdf_rel=plan.get("pdf_rel"),
                input_role=(plan.get("role") or {}).get("role") or "",
                input_filename=plan.get("pdf_filename"),
                suggested_c=plan.get("proposed_c") or "",
                suggested_d=plan.get("proposed_d") or "",
                suggested_annots=plan.get("annotations") or [],
                confidence=plan.get("confidence", 0),
                generator=plan.get("generator", "rules"),
                provenance=plan.get("provenance"),
                user_action=user_action,
                final_c=final_c or "",
                final_d=final_d or "",
                final_annots=plan.get("annotations") or [],
            )
        except Exception as e:
            sys.stderr.write(f"[learning] record_feedback: {e}\n")
        try:
            db.log_audit(action="auto_annotate_apply",
                         target_type="row", target_id=str(row_num),
                         before=plan.get("current_d") or "",
                         after=plan.get("proposed_d") or "",
                         details={"role": (plan.get("role") or {}).get("role"),
                                  "annotations": len(plan.get("annotations") or []),
                                  "write_pdf": write_pdf,
                                  "write_xlsx": write_xlsx,
                                  "plan_id": plan_id,
                                  "confidence": plan.get("confidence")},
                         actor="user")
        except Exception: pass
    return jsonify({"ok": result.get("ok"), "result": result, "plan": plan})


@app.route("/api/auto_annotate/batch_preview")
def api_auto_annotate_batch_preview():
    """Return plans for all rows that look auto-annotateable (have a
    catalog PDF + needs Col D fill)."""
    plans = []
    for r in ROWS:
        if not r.get("needs_col_d") and (r.get("D") or "").strip():
            continue
        if not r.get("pdf_rel"):
            continue
        p = auto_annotate_plan(r["row"])
        plans.append(p)
    return jsonify({
        "ok": True,
        "count": len(plans),
        "by_role": {role: sum(1 for p in plans if (p.get("role") or {}).get("role") == role)
                    for role in ("section_header", "item", "sub_item", "unknown")},
        "plans": plans,
    })


@app.route("/api/auto_annotate/batch_apply", methods=["POST"])
def api_auto_annotate_batch_apply():
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    write_pdf = bool(data.get("write_pdf", True))
    write_xlsx = bool(data.get("write_xlsx", True))
    if not isinstance(rows, list) or not rows:
        return jsonify({"ok": False, "error": "rows[] required"}), 400

    # One snapshot at start of the batch
    _run_version_cmd(["snap", f"pre-batch-auto-{len(rows)}-rows"], timeout=60)

    results = []
    for rn in rows:
        try:
            plan = auto_annotate_plan(int(rn))
        except Exception as e:
            results.append({"row": rn, "ok": False, "error": str(e)})
            continue
        if not plan.get("ok"):
            results.append({"row": rn, "ok": False, "warnings": plan.get("warnings", [])})
            continue
        r = apply_auto_annotate_plan(plan, write_pdf=write_pdf, write_xlsx=write_xlsx)
        results.append({"row": rn, "ok": r.get("ok"),
                        "proposed_d": plan.get("proposed_d"),
                        "annotations_added": len(plan.get("annotations") or [])
                                              if r.get("pdf_written") else 0})
    # refresh once at end
    load_rows()
    TOR_CACHE.clear()
    return jsonify({"ok": True, "count": len(results),
                    "applied_ok": sum(1 for x in results if x.get("ok")),
                    "results": results})


# ---------------------------------------------------------------------------
# Manual-annotate workflow loop
#
# When the AI couldn't find Col B content in the catalog and fell back to
# "ยินดีปฏิบัติตามข้อกำหนด", but the user notices the spec IS in the catalog,
# they enter manual-annotate mode:
#   1. /api/manual_annotate/context returns the row's role + a suggested
#      FreeText label content (using SKILL.md format).
#   2. User draws a rect on the catalog PDF (via existing edit-mode SVG
#      overlay). The frontend auto-pairs a draggable FreeText label.
#   3. /api/manual_annotate/save accepts the rect + label positions, writes
#      both annotations into the PDF, computes the proper Col D string from
#      the page where the rect sits, updates the xlsx, and records the
#      manual override as strong learning feedback.
# ---------------------------------------------------------------------------

def _label_for_row(role_info: dict) -> str:
    """SKILL.md label for a Square's FreeText partner."""
    role = role_info.get("role")
    section = role_info.get("section") or ""
    if role == "sub_item":
        if role_info.get("parent_item") is not None:
            return (f"{section} ข้อ {role_info['parent_item']}) "
                    f"ข้อย่อย {role_info['sub_num']}.")
        return f"{section} ข้อย่อย {role_info['sub_num']}."
    if role == "item":
        return f"{section} ข้อ {role_info['item_num']})"
    return section


def _candidate_pdfs_for_section(section: str | None) -> list[Path]:
    """Find catalog PDFs that could correspond to a section. Used by the
    manual-annotate flow to populate a picker for commitment rows whose
    pdf_rel wasn't resolved.

    Wildcard matches are sorted by trailing numeric suffix so the parent
    rack catalog (-1) is picked first, not whichever sub-catalog Python
    happened to insert first in the dict.
    """
    if not section:
        return []
    out: list[Path] = []
    seen: set[Path] = set()

    def _add(paths):
        for p in paths:
            if p not in seen:
                seen.add(p); out.append(p)

    def _sort_key(k: str) -> tuple:
        # Sort "5.1.1-1", "5.1.1-2", ... → 1, 2, ... so rack parent (-1) wins
        m = re.search(r"-(\d+)$", k)
        return (int(m.group(1)) if m else 0, k)

    # 1. Direct keys + dot/dash translation (highest priority)
    for k in _ref_to_folder_keys(section):
        if k in PDF_INDEX: _add(PDF_INDEX[k])
        if k in SECTION_INDEX: _add(SECTION_INDEX[k])

    # 2. Wildcard: any sub-catalog under this section, sorted ascending so the
    #    parent rack (-1) comes first (e.g. for 5.1.1 → 5.1.1-1 before -2..-7)
    matches = sorted(
        [k for k in PDF_INDEX if k.startswith(section + "-") or k == section],
        key=_sort_key,
    )
    for k in matches:
        _add(PDF_INDEX[k])

    # 3. Parent section fallback (e.g. 5.1.6.1 → 5.1.6) so a row referring to
    #    a whole sub-section can find its parent rack catalog.
    parts = section.split(".")
    if len(parts) >= 3:
        parent = ".".join(parts[:-1])
        for k in _ref_to_folder_keys(parent):
            if k in PDF_INDEX: _add(PDF_INDEX[k])
            if k in SECTION_INDEX: _add(SECTION_INDEX[k])
        parent_matches = sorted(
            [k for k in PDF_INDEX if k.startswith(parent + "-")],
            key=_sort_key,
        )
        for k in parent_matches:
            _add(PDF_INDEX[k])
    return out


@app.route("/api/manual_annotate/context")
def api_manual_context():
    row_num = int(request.args.get("row", 0))
    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404

    role_info = detect_row_role(row_num)
    pdf_rel = row.get("pdf_rel")
    section = row.get("section_inferred")

    # If row didn't auto-resolve a PDF (e.g. commitment row), surface a
    # picker of candidate catalogs from the row's section so the user can
    # choose where to mark.
    candidates: list[dict] = []
    if not pdf_rel and section:
        for p in _candidate_pdfs_for_section(section):
            candidates.append({
                "rel": str(p.relative_to(OUTPUT)),
                "name": p.name,
                "folder": p.parent.name,
            })
        if candidates:
            pdf_rel = candidates[0]["rel"]

    def _meta(rel: str) -> dict | None:
        try:
            doc = fitz.open(OUTPUT / rel)
            return {
                "rel": rel,
                "pages": len(doc),
                "page_sizes": [(round(doc[i].rect.width, 1),
                                round(doc[i].rect.height, 1))
                               for i in range(len(doc))],
            }
        except Exception:
            return None

    return jsonify({
        "ok": True,
        "row": row_num,
        "section": section,
        "role": role_info,
        "col_b": row.get("B") or "",
        "col_d_current": row.get("D") or "",
        "is_commitment": (row.get("D") or "").strip().startswith("ยินดีปฏิบัติ"),
        "pdf_rel": pdf_rel,
        "pdf_meta": _meta(pdf_rel) if pdf_rel else None,
        "candidates": candidates,
        "suggested_label": _label_for_row(role_info),
    })


@app.route("/api/manual_annotate/save", methods=["POST"])
def api_manual_save():
    """Body: {row, page, content_rect:[x0,y0,x1,y1], label_rect:[...],
             label_text, pdf_rel?}"""
    data = request.get_json(silent=True) or {}
    try:
        row_num = int(data["row"])
        page = int(data["page"])
        content_rect = list(map(float, data["content_rect"]))
        label_rect = list(map(float, data["label_rect"]))
    except Exception as e:
        return jsonify({"ok": False, "error": f"bad payload: {e}"}), 400

    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404

    pdf_rel = data.get("pdf_rel") or row.get("pdf_rel")
    if not pdf_rel:
        return jsonify({"ok": False, "error": "no PDF for row"}), 404
    pdf_path = (OUTPUT / pdf_rel).resolve()
    if not str(pdf_path).startswith(str(OUTPUT.resolve())) or not pdf_path.exists():
        return jsonify({"ok": False, "error": "PDF not found"}), 404

    role_info = detect_row_role(row_num)
    label_text = (data.get("label_text") or _label_for_row(role_info)).strip()

    # Pre-snapshot
    _run_version_cmd(["snap", f"pre-manual-row-{row_num}"], timeout=60)

    # Apply PDF edits — Square + FreeText pair. fontsize is left to
    # apply_pdf_edits which clamps from rect height (matches SVG overlay).
    edits = [
        {"action": "create", "page": page, "type": "Square",
         "rect": content_rect, "contents": ""},
        {"action": "create", "page": page, "type": "FreeText",
         "rect": label_rect, "contents": label_text},
    ]
    pdf_result = apply_pdf_edits(pdf_path, edits)

    # Compute new Col D using the same generator as auto-annotate so format
    # stays consistent with the rest of the spreadsheet.
    new_d = make_col_d_for_row(row, role_info, pdf_path, page)
    if not new_d:
        # Fallback (should rarely hit since make_col_d_for_row covers most)
        new_d = (f"เทียบเท่าข้อกำหนด เอกสาร {pdf_path.stem} หน้า {page} "
                 f"ข้อ {role_info.get('section', '')}")

    old_d = row.get("D") or ""

    # Update xlsx Col D (Google Drive O_TRUNC workaround)
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        ws.cell(row_num, 4).value = new_d
        # If Col C is still the raw Col B, tighten it via make_col_c_from_b
        cur_c = (ws.cell(row_num, 3).value or "").strip()
        cur_b = (ws.cell(row_num, 2).value or "").strip()
        if not cur_c or cur_c == cur_b:
            ws.cell(row_num, 3).value = make_col_c_from_b(cur_b)
        tmp = XLSX_PATH.with_suffix(".xlsx.tmp")
        wb.save(str(tmp))
        with open(tmp, "rb") as f:
            data_bytes = f.read()
        fd = os.open(str(XLSX_PATH), os.O_WRONLY | os.O_TRUNC)
        try:
            os.write(fd, data_bytes)
        finally:
            os.close(fd)
        tmp.unlink(missing_ok=True)
    except Exception as e:
        return jsonify({"ok": False, "error": f"xlsx write failed: {e}",
                        "pdf_result": pdf_result}), 500

    # Refresh internal caches
    try:
        load_rows()
        TOR_CACHE.clear()
        sync_db_from_memory()
    except Exception as e:
        sys.stderr.write(f"[manual_save] refresh: {e}\n")

    # Audit + learning feedback (strong "edited" signal — manual override)
    try:
        db.log_audit(action="manual_annotate",
                     target_type="row", target_id=str(row_num),
                     before=old_d, after=new_d,
                     details={"page": page, "pdf_rel": pdf_rel,
                              "annots_applied": pdf_result.get("applied", 0),
                              "annots_errors": pdf_result.get("errors", 0)},
                     actor="user")
    except Exception: pass
    try:
        learning.record_feedback(
            row_num=row_num,
            section=row.get("section_inferred"),
            input_b=row.get("B", "") or "",
            input_pdf_rel=pdf_rel,
            input_role=role_info.get("role", ""),
            input_filename=pdf_path.name,
            suggested_c=row.get("C", "") or "",
            suggested_d=old_d,
            suggested_annots=[],
            confidence=0.0,
            generator="commitment_fallback",
            provenance={},
            user_action="edited",
            final_c=row.get("C", "") or "",
            final_d=new_d,
            final_annots=edits,
        )
    except Exception as e:
        sys.stderr.write(f"[manual_save] feedback: {e}\n")

    return jsonify({"ok": True, "row": row_num, "page": page,
                    "old_d": old_d, "new_d": new_d,
                    "label_text": label_text,
                    "pdf_result": pdf_result})


# ---------------------------------------------------------------------------
# Re-annotate Wizard
#
# Extends the manual-annotate flow into a stepwise wizard that:
#   • for brand_model section_header rows, expects 2 rect+label pairs
#     (ยี่ห้อ, รุ่น) per SKILL.md §"Label DA / DS standard (ยี่ห้อ/รุ่น)"
#   • for item / sub_item rows, expects 1 pair (existing behaviour)
#   • lets the user switch the catalog PDF before annotating
#   • optionally deletes existing annotations whose label starts with the
#     same prefix(es) before writing the new ones — a clean re-annotate
# ---------------------------------------------------------------------------

def _expected_steps_for_row(row: dict, role_info: dict) -> list[dict]:
    """Plan the step labels the user must produce. Pairs the SKILL.md
    label convention with the row's structural role."""
    role = role_info.get("role")
    section = role_info.get("section") or ""
    parsed = row.get("parsed") or {}

    # Section header that's brand_model → 2 steps (ยี่ห้อ + รุ่น)
    if role == "section_header":
        # parsed.type is the cheapest signal; fallback to filename inspection
        is_bm = (parsed.get("type") == "brand_model")
        if not is_bm:
            pdf_rel = row.get("pdf_rel") or ""
            if pdf_rel:
                stem = Path(pdf_rel).stem
                b, m = parse_brand_model_from_filename(stem)
                is_bm = bool(b and m)
        if is_bm:
            return [
                {"label": "ยี่ห้อ",
                 "hint": "ลาก rect รอบ <strong>โลโก้ยี่ห้อ</strong> ใน catalog",
                 "kind": "brand"},
                {"label": "รุ่น",
                 "hint": "ลาก rect รอบ <strong>ชื่อรุ่น</strong> (model number) ใน catalog",
                 "kind": "model"},
            ]
        return [
            {"label": section,
             "hint": "ลาก rect รอบเนื้อหาที่เกี่ยวกับ section นี้",
             "kind": "section"},
        ]

    # Item / sub_item — same single-step flow as manual-annotate
    label = _label_for_row(role_info)
    if role == "sub_item":
        hint = (f"ลาก rect รอบเนื้อหาของ <strong>ข้อย่อย "
                f"{role_info.get('sub_num', '?')}</strong>")
    elif role == "item":
        hint = (f"ลาก rect รอบเนื้อหาของ <strong>ข้อ "
                f"{role_info.get('item_num', '?')})</strong>")
    else:
        hint = "ลาก rect รอบเนื้อหาที่เกี่ยวกับ row นี้"
    return [{"label": label, "hint": hint, "kind": role or "row"}]


def _delete_label_prefixes_for_row(row: dict, role_info: dict) -> list[str]:
    """Prefixes whose existing FreeText annotations should be wiped before
    we re-annotate. Conservative — only deletes labels that are *clearly*
    this row's prior work."""
    role = role_info.get("role")
    section = role_info.get("section") or ""
    parsed = row.get("parsed") or {}

    if role == "section_header":
        if parsed.get("type") == "brand_model":
            return ["ยี่ห้อ", "รุ่น"]
        if section:
            return [section]   # exact section header label
        return []

    if role == "item":
        n = role_info.get("item_num")
        if section and n is not None:
            # Match "5.1.1 ข้อ 1)" but NOT "5.1.1 ข้อ 1) ข้อย่อย 1." — sub-items
            # share the parent prefix, so we anchor on a closing ')' boundary.
            return [f"{section} ข้อ {n})"]
        return []

    if role == "sub_item":
        sub = role_info.get("sub_num")
        parent = role_info.get("parent_item")
        if section and sub is not None:
            if parent is not None:
                return [f"{section} ข้อ {parent}) ข้อย่อย {sub}."]
            return [f"{section} ข้อย่อย {sub}."]
        return []

    return []


def _annots_to_delete_by_prefix(pdf_path: Path, page_num: int,
                                 prefixes: list[str]) -> list[int]:
    """Return xrefs of FreeText annotations whose contents begin with any
    of the given prefixes, plus the xref of any nearby Square that looks
    paired with them (within 12pt edge-to-edge on any side).

    Conservative: a Square is only paired if it shares a side with the
    label rect — that's how SKILL.md and our generator place them.
    """
    if not prefixes or not pdf_path.exists():
        return []
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return []
    if page_num < 1 or page_num > len(doc):
        doc.close()
        return []

    page = doc[page_num - 1]
    annots = safe_iter_annots(page)
    label_to_xref: list[tuple[float, float, float, float, int]] = []  # x0,y0,x1,y1,xref
    square_xrefs: list[tuple[float, float, float, float, int]] = []
    for ann in annots:
        try:
            kind = ann.type[1] if isinstance(ann.type, tuple) else str(ann.type)
        except Exception:
            kind = ""
        try:
            r = ann.rect
            xref = ann.xref
            contents = (ann.info.get("content") or "").strip()
        except Exception:
            continue
        if kind == "FreeText":
            for pre in prefixes:
                if contents.startswith(pre):
                    label_to_xref.append((r.x0, r.y0, r.x1, r.y1, int(xref)))
                    break
        elif kind == "Square":
            square_xrefs.append((r.x0, r.y0, r.x1, r.y1, int(xref)))

    out: set[int] = set()
    for (lx0, ly0, lx1, ly1, lxref) in label_to_xref:
        out.add(lxref)
        # Find the closest Square that shares a side within 12pt
        best_xref = None
        best_dist = 1e9
        for (sx0, sy0, sx1, sy1, sxref) in square_xrefs:
            # Right side of Square ↔ Left side of label
            dx = abs(sx1 - lx0)
            dy = abs((sy0 + sy1) / 2 - (ly0 + ly1) / 2)
            d = dx + dy
            if dx <= 12 and dy <= 30 and d < best_dist:
                best_dist = d; best_xref = sxref
            # Below Square ↔ above label
            dy2 = abs(sy1 - ly0)
            dx2 = abs((sx0 + sx1) / 2 - (lx0 + lx1) / 2)
            d2 = dy2 + dx2
            if dy2 <= 12 and dx2 <= 80 and d2 < best_dist:
                best_dist = d2; best_xref = sxref
            # Left side of Square ↔ Right side of label (label on left)
            dx3 = abs(sx0 - lx1)
            d3 = dx3 + dy
            if dx3 <= 12 and dy <= 30 and d3 < best_dist:
                best_dist = d3; best_xref = sxref
            # Above Square ↔ below label (label above)
            dy4 = abs(sy0 - ly1)
            d4 = dy4 + dx2
            if dy4 <= 12 and dx2 <= 80 and d4 < best_dist:
                best_dist = d4; best_xref = sxref
        if best_xref is not None:
            out.add(best_xref)

    doc.close()
    return sorted(out)


@app.route("/api/reannotate/context")
def api_reannotate_context():
    """Returns the wizard plan for a row.

    Response: {ok, row, section, role, col_b, col_d_current, parsed_type,
               pdf_rel, pdf_meta, candidates, steps, delete_label_prefixes,
               brand_hint, model_hint}
    """
    try:
        row_num = int(request.args.get("row", 0))
    except Exception:
        return jsonify({"ok": False, "error": "bad row"}), 400
    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404

    role_info = detect_row_role(row_num)
    section = row.get("section_inferred")
    pdf_rel = row.get("pdf_rel")

    # Always surface candidates so the user can swap PDFs even when one is
    # already assigned (key request from the user).
    candidates: list[dict] = []
    cand_paths = _candidate_pdfs_for_section(section)
    for p in cand_paths:
        try:
            rel = str(p.relative_to(OUTPUT))
        except Exception:
            continue
        candidates.append({
            "rel": rel,
            "name": p.name,
            "folder": p.parent.name,
            "is_current": (rel == pdf_rel),
        })

    # If the row had no PDF, default to the first candidate
    if not pdf_rel and candidates:
        pdf_rel = candidates[0]["rel"]

    def _meta(rel: str) -> dict | None:
        try:
            doc = fitz.open(OUTPUT / rel)
            return {
                "rel": rel,
                "pages": len(doc),
                "page_sizes": [(round(doc[i].rect.width, 1),
                                round(doc[i].rect.height, 1))
                               for i in range(len(doc))],
            }
        except Exception:
            return None

    # Filename-derived brand/model hints (defaults the wizard pre-fills)
    brand_hint, model_hint = ("", "")
    if pdf_rel:
        stem = Path(pdf_rel).stem
        brand_hint, model_hint = parse_brand_model_from_filename(stem)

    return jsonify({
        "ok": True,
        "row": row_num,
        "section": section,
        "role": role_info,
        "col_b": row.get("B") or "",
        "col_d_current": row.get("D") or "",
        "parsed_type": (row.get("parsed") or {}).get("type"),
        "pdf_rel": pdf_rel,
        "pdf_meta": _meta(pdf_rel) if pdf_rel else None,
        "candidates": candidates,
        "steps": _expected_steps_for_row(row, role_info),
        "delete_label_prefixes": _delete_label_prefixes_for_row(row, role_info),
        "brand_hint": brand_hint,
        "model_hint": model_hint,
    })


@app.route("/api/reannotate/save", methods=["POST"])
def api_reannotate_save():
    """Body:
      {row, pdf_rel, page, steps:[{content_rect, label_rect, label_text}, ...],
       delete_existing: bool, delete_prefixes: [str, ...],
       col_d_override: str|null}
    """
    data = request.get_json(silent=True) or {}
    try:
        row_num = int(data["row"])
        pdf_rel = str(data["pdf_rel"])
        page = int(data["page"])
        steps = list(data.get("steps") or [])
    except Exception as e:
        return jsonify({"ok": False, "error": f"bad payload: {e}"}), 400

    if not steps:
        return jsonify({"ok": False, "error": "no steps to save"}), 400

    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404

    pdf_path = (OUTPUT / pdf_rel).resolve()
    if (not str(pdf_path).startswith(str(OUTPUT.resolve())) or
            not pdf_path.exists()):
        return jsonify({"ok": False, "error": "PDF not found"}), 404

    role_info = detect_row_role(row_num)

    # Pre-snapshot — `pre-reannotate-row-N` shows up in the snapshots list
    _run_version_cmd(["snap", f"pre-reannotate-row-{row_num}"], timeout=60)

    # 1. Compute deletions (only if requested AND prefixes provided)
    edits: list[dict] = []
    deleted_xrefs: list[int] = []
    if data.get("delete_existing"):
        prefixes = list(data.get("delete_prefixes") or [])
        if prefixes:
            deleted_xrefs = _annots_to_delete_by_prefix(pdf_path, page, prefixes)
            for xref in deleted_xrefs:
                edits.append({"action": "delete", "xref": xref})

    # 2. Append create edits (Square + FreeText pairs)
    for st in steps:
        try:
            crect = list(map(float, st["content_rect"]))
            lrect = list(map(float, st["label_rect"]))
            ltext = (st.get("label_text") or "").strip()
        except Exception as e:
            return jsonify({"ok": False, "error": f"bad step: {e}"}), 400
        edits.append({"action": "create", "page": page, "type": "Square",
                      "rect": crect, "contents": ""})
        edits.append({"action": "create", "page": page, "type": "FreeText",
                      "rect": lrect, "contents": ltext})

    # 3. Apply
    pdf_result = apply_pdf_edits(pdf_path, edits)

    # 4. Compute new Col D
    col_d_override = (data.get("col_d_override") or "").strip()
    if col_d_override:
        new_d = col_d_override
    else:
        # If we're reannotating a brand_model section header AND the user
        # passed brand/model strings via override, use those; otherwise fall
        # through to the file-name-based generator.
        brand = (data.get("brand") or "").strip()
        model = (data.get("model") or "").strip()
        if brand and model:
            new_d = f"ยี่ห้อ {brand} รุ่น {model}"
        elif brand:
            new_d = f"ยี่ห้อ {brand}"
        else:
            new_d = make_col_d_for_row(row, role_info, pdf_path, page)
            if not new_d:
                new_d = (f"เทียบเท่าข้อกำหนด เอกสาร {pdf_path.stem} หน้า {page} "
                         f"ข้อ {role_info.get('section', '')}")

    old_d = row.get("D") or ""
    old_pdf_rel = row.get("pdf_rel") or ""

    # 5. Persist Col D + (when PDF changed) update xlsx
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        ws.cell(row_num, 4).value = new_d
        cur_c = (ws.cell(row_num, 3).value or "").strip()
        cur_b = (ws.cell(row_num, 2).value or "").strip()
        if not cur_c or cur_c == cur_b:
            ws.cell(row_num, 3).value = make_col_c_from_b(cur_b)
        tmp = XLSX_PATH.with_suffix(".xlsx.tmp")
        wb.save(str(tmp))
        with open(tmp, "rb") as f:
            data_bytes = f.read()
        fd = os.open(str(XLSX_PATH), os.O_WRONLY | os.O_TRUNC)
        try:
            os.write(fd, data_bytes)
        finally:
            os.close(fd)
        tmp.unlink(missing_ok=True)
    except Exception as e:
        return jsonify({"ok": False, "error": f"xlsx write failed: {e}",
                        "pdf_result": pdf_result}), 500

    # 6. Refresh + audit + learning feedback
    try:
        load_rows()
        TOR_CACHE.clear()
        sync_db_from_memory()
    except Exception as e:
        sys.stderr.write(f"[reannotate] refresh: {e}\n")

    try:
        db.log_audit(action="reannotate",
                     target_type="row", target_id=str(row_num),
                     before=old_d, after=new_d,
                     details={"page": page, "pdf_rel": pdf_rel,
                              "old_pdf_rel": old_pdf_rel,
                              "n_steps": len(steps),
                              "deleted_xrefs": deleted_xrefs,
                              "annots_applied": pdf_result.get("applied", 0),
                              "annots_errors": pdf_result.get("errors", 0)},
                     actor="user")
    except Exception: pass

    try:
        learning.record_feedback(
            row_num=row_num,
            section=row.get("section_inferred"),
            input_b=row.get("B", "") or "",
            input_pdf_rel=pdf_rel,
            input_role=role_info.get("role", ""),
            input_filename=pdf_path.name,
            suggested_c=row.get("C", "") or "",
            suggested_d=old_d,
            suggested_annots=[],
            confidence=0.0,
            generator="reannotate_wizard",
            provenance={"old_pdf_rel": old_pdf_rel,
                        "n_steps": len(steps),
                        "n_deleted": len(deleted_xrefs)},
            user_action="edited",
            final_c=row.get("C", "") or "",
            final_d=new_d,
            final_annots=edits,
        )
    except Exception as e:
        sys.stderr.write(f"[reannotate] feedback: {e}\n")

    return jsonify({
        "ok": True, "row": row_num, "page": page,
        "old_d": old_d, "new_d": new_d,
        "pdf_rel": pdf_rel,
        "pdf_changed": pdf_rel != old_pdf_rel,
        "deleted_xrefs": deleted_xrefs,
        "pdf_result": pdf_result,
    })


# ---------------------------------------------------------------------------
# Project-level versioning — wraps scripts/version.py
#
# version.py is the canonical snapshot tool (snap, snap-full, list, restore,
# diff, prune, auto-snap). Rather than reimplementing it, we shell out via
# subprocess so behavior stays in lock-step with the CLI tool.
# ---------------------------------------------------------------------------

def _version_script_available() -> bool:
    return VERSION_SCRIPT.exists()


def _read_snapshot(snap_dir: Path) -> dict | None:
    m_path = snap_dir / "manifest.json"
    if not m_path.exists():
        return None
    try:
        m = json.loads(m_path.read_text(encoding="utf-8"))
    except Exception:
        return None
    try:
        size = sum(p.stat().st_size for p in snap_dir.rglob("*") if p.is_file())
    except Exception:
        size = 0
    return {
        "id": m.get("id", snap_dir.name),
        "tag": m.get("tag", ""),
        "kind": m.get("kind", "?"),
        "timestamp": m.get("timestamp", ""),
        "size": size,
        "n_tracked": len(m.get("files", {})),
        "n_output": len(m.get("output_tree", [])),
        "has_tarball": (snap_dir / "output.tar.gz").exists(),
        "files": [
            {"path": k, "size": v.get("size", 0), "kind": v.get("kind", "")}
            for k, v in (m.get("files", {}) or {}).items()
        ],
    }


def _run_version_cmd(args: list[str], timeout: int = 600) -> dict:
    """Invoke version.py with the given args and return a JSON-serialisable
    summary."""
    if not _version_script_available():
        return {"ok": False, "error": "scripts/version.py not found",
                "stdout": "", "stderr": ""}
    cmd = [sys.executable, str(VERSION_SCRIPT)] + args
    try:
        r = subprocess.run(cmd, cwd=str(ROOT),
                           capture_output=True, text=True, timeout=timeout)
    except subprocess.TimeoutExpired:
        return {"ok": False, "error": "timeout", "stdout": "", "stderr": ""}
    except Exception as e:
        return {"ok": False, "error": str(e), "stdout": "", "stderr": ""}
    return {
        "ok": r.returncode == 0,
        "returncode": r.returncode,
        "stdout": r.stdout,
        "stderr": r.stderr,
    }


@app.route("/api/versions")
def api_versions():
    """List project snapshots from _versions/snapshots/."""
    snaps = []
    if SNAPS_DIR.exists():
        for d in sorted(SNAPS_DIR.iterdir(), reverse=True):
            if not d.is_dir():
                continue
            info = _read_snapshot(d)
            if info:
                snaps.append(info)
    return jsonify({
        "available": _version_script_available(),
        "root": str(ROOT),
        "snapshots": snaps,
    })


@app.route("/api/versions/snap", methods=["POST"])
def api_versions_snap():
    """Create a new snapshot. Body: {tag?: str, full?: bool}."""
    data = request.get_json(silent=True) or {}
    tag = (data.get("tag") or "").strip()
    full = bool(data.get("full"))
    args = ["snap-full" if full else "snap"]
    if tag:
        args.append(tag)
    result = _run_version_cmd(args, timeout=600 if full else 60)
    if result.get("ok"):
        try:
            db.log_audit(action="snapshot",
                         target_type="project", target_id=tag or "(no tag)",
                         details={"kind": "full" if full else "quick", "tag": tag},
                         actor="user")
        except Exception: pass
    return jsonify(result)


@app.route("/api/versions/restore", methods=["POST"])
def api_versions_restore():
    """Restore from snapshot. Body: {id: str, full?: bool}.

    Always takes an auto-snap of the current state first (handled by version.py
    only when --pre-snap is used in some setups; here we explicitly create one
    via auto-snap to be safe).
    """
    data = request.get_json(silent=True) or {}
    snap_id = (data.get("id") or "").strip()
    full = bool(data.get("full"))
    if not snap_id:
        return jsonify({"ok": False, "error": "id required"}), 400
    # Take a safety snapshot of the current state first
    _run_version_cmd(["snap", "pre-restore"], timeout=60)
    result = _run_version_cmd(
        ["restore-full" if full else "restore", snap_id, "-y"],
        timeout=600 if full else 60,
    )
    if result.get("ok"):
        # Refresh internal caches after restore
        try:
            build_pdf_index()
            load_rows()
            collect_extra_refs()
            build_tor_section_index()
            index_tor_text()
            TOR_CACHE.clear()
            sync_db_from_memory()
        except Exception as e:
            sys.stderr.write(f"[restore refresh] {e}\n")
        try:
            db.log_audit(action="restore",
                         target_type="project", target_id=snap_id,
                         details={"full": full}, actor="user")
        except Exception: pass
    return jsonify(result)


@app.route("/api/versions/diff")
def api_versions_diff():
    id1 = (request.args.get("id1") or "").strip()
    id2 = (request.args.get("id2") or "").strip()
    if not id1 or not id2:
        return jsonify({"ok": False, "error": "id1 + id2 required"}), 400
    result = _run_version_cmd(["diff", id1, id2], timeout=60)
    return jsonify(result)


@app.route("/api/versions/auto-snap", methods=["POST"])
def api_versions_auto_snap():
    """Snap only if the xlsx changed since the most recent snapshot."""
    result = _run_version_cmd(["auto-snap"], timeout=60)
    return jsonify(result)


@app.route("/api/versions/prune", methods=["POST"])
def api_versions_prune():
    data = request.get_json(silent=True) or {}
    keep = max(1, int(data.get("keep", 10)))
    result = _run_version_cmd(["prune", "--keep", str(keep), "-y"], timeout=60)
    return jsonify(result)


@app.route("/api/versions/show")
def api_versions_show():
    snap_id = (request.args.get("id") or "").strip()
    if not snap_id:
        return jsonify({"ok": False, "error": "id required"}), 400
    # Find snap by exact ID first, then by prefix
    candidates = []
    if SNAPS_DIR.exists():
        for d in SNAPS_DIR.iterdir():
            if d.is_dir() and (d.name == snap_id or d.name.startswith(snap_id)):
                candidates.append(d)
    if len(candidates) != 1:
        return jsonify({"ok": False, "error": f"{len(candidates)} matches"}), 404
    info = _read_snapshot(candidates[0])
    return jsonify({"ok": True, **(info or {})})


# ---------------------------------------------------------------------------
# "Always work on latest" — sync the live working files with the most recent
# snapshot
# ---------------------------------------------------------------------------

def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
    except Exception:
        return ""
    return h.hexdigest()


def _find_latest_snapshot() -> tuple[Path, dict] | None:
    if not SNAPS_DIR.exists():
        return None
    snaps = sorted(
        [d for d in SNAPS_DIR.iterdir() if d.is_dir()],
        reverse=True,
    )
    for snap_dir in snaps:
        m_path = snap_dir / "manifest.json"
        if not m_path.exists():
            continue
        try:
            return snap_dir, json.loads(m_path.read_text(encoding="utf-8"))
        except Exception:
            continue
    return None


def get_version_sync_status() -> dict:
    """Compare the live working files against the most recent valid snapshot.

    Returns a state describing the relationship:

    - ``no_snapshots``      — versioning system is empty
    - ``in_sync``           — every tracked file matches the latest snapshot
    - ``working_ahead``     — files differ AND working mtimes ≥ snapshot ts
                              (the user has been editing since last snapshot;
                              the boot hook auto-snaps to bring "latest"
                              forward to match)
    - ``working_behind``    — files differ AND working mtimes < snapshot ts
                              (something rolled the working dir back; UI
                              prompts the user to restore)
    - ``divergent``         — mixed (some newer, some older) — UI prompts
    - ``incomplete_local``  — a tracked file is missing from the working dir
    """
    out: dict = {
        "has_snapshots": False,
        "state": "no_snapshots",
        "latest": None,
        "files": [],
    }
    found = _find_latest_snapshot()
    if found is None:
        return out
    snap_dir, m = found

    out["has_snapshots"] = True
    out["latest"] = {
        "id": m.get("id", snap_dir.name),
        "tag": m.get("tag", ""),
        "kind": m.get("kind", "?"),
        "timestamp": m.get("timestamp", ""),
    }

    snap_ts = 0.0
    if m.get("timestamp"):
        try:
            snap_ts = datetime.fromisoformat(m["timestamp"]).timestamp()
        except Exception:
            pass

    all_match = True
    any_newer = False
    any_older = False
    any_missing = False

    for rel, info in (m.get("files") or {}).items():
        if info.get("kind") == "tarball":
            continue
        full = ROOT / rel
        if not full.exists():
            all_match = False
            any_missing = True
            out["files"].append({"path": rel, "status": "missing_local"})
            continue
        cur_hash = _file_sha256(full)
        snap_hash = info.get("sha256", "")
        match = (cur_hash == snap_hash)
        cur_mtime = full.stat().st_mtime
        if not match:
            all_match = False
            if cur_mtime >= snap_ts:
                any_newer = True
            else:
                any_older = True
        out["files"].append({
            "path": rel,
            "status": "match" if match else "differ",
            "current_size": full.stat().st_size,
            "snapshot_size": info.get("size", 0),
            "current_mtime": cur_mtime,
            "snapshot_mtime": snap_ts,
        })

    if all_match:
        out["state"] = "in_sync"
    elif any_missing:
        out["state"] = "incomplete_local"
    elif any_newer and not any_older:
        out["state"] = "working_ahead"
    elif any_older and not any_newer:
        out["state"] = "working_behind"
    else:
        out["state"] = "divergent"
    return out


@app.route("/api/versions/sync")
def api_versions_sync():
    return jsonify(get_version_sync_status())


def boot_sync_check() -> dict:
    """Run at server start. Tries to keep the invariant ``latest = current``:

    - If no snapshot exists yet → take a baseline snap so versioning is
      bootstrapped.
    - If working state is **ahead** of latest (the common case after a work
      session) → take an auto-snap so latest reflects the new state.
    - If working state is **behind** or **divergent** → DO NOTHING
      automatically (would risk overwriting in-progress work). Surface the
      condition through the UI so the user can decide.

    Set env ``COMPLY_NO_BOOT_SNAP=1`` to disable boot auto-snapping.
    """
    result = {"performed": None, "status": None}
    if not _version_script_available():
        result["status"] = "version_script_missing"
        return result
    status = get_version_sync_status()
    result["status"] = status["state"]

    if os.environ.get("COMPLY_NO_BOOT_SNAP") == "1":
        return result

    state = status["state"]
    if state == "no_snapshots":
        r = _run_version_cmd(["snap", "boot-baseline"], timeout=60)
        result["performed"] = "baseline_snap" if r.get("ok") else f"failed: {r.get('error')}"
    elif state == "working_ahead":
        # auto-snap creates a tagged snapshot only if xlsx actually changed
        # (it's idempotent)
        r = _run_version_cmd(["snap", "boot-auto"], timeout=60)
        result["performed"] = "boot_auto_snap" if r.get("ok") else f"failed: {r.get('error')}"
    elif state in ("working_behind", "divergent", "incomplete_local"):
        result["performed"] = "skipped (manual review required)"
    else:
        result["performed"] = "no-op (already in sync)"

    return result


# ---------------------------------------------------------------------------
# DB sync + DB-backed API endpoints
# ---------------------------------------------------------------------------

def _build_pdf_records_for_db() -> list[dict]:
    """Snapshot the PDF index into dict records for the DB.

    Annotations are NOT pre-listed — they used to be (via list_pdf_annots
    per PDF), which made boot ~60-90 sec on Google Drive filesystems
    because PyMuPDF opens each of 101 PDFs and walks every annot. The
    pdf_annotations table is only ever WRITTEN (and counted), never
    queried — annotations are fetched live via /api/pdf_meta when a row
    is selected. Keeping it empty saves ~minute on every boot.
    """
    seen: dict[Path, dict] = {}
    # Combine both indexes
    for ref_key, paths in PDF_INDEX.items():
        for p in paths:
            if p in seen:
                continue
            seen[p] = {"folder_key": ref_key}
    for sec_key, paths in SECTION_INDEX.items():
        for p in paths:
            seen.setdefault(p, {"folder_key": None,
                                "section_prefix": sec_key})
            seen[p].setdefault("section_prefix", sec_key)

    records = []
    for p, meta in seen.items():
        try:
            stat = p.stat()
            rel = str(p.relative_to(OUTPUT))
        except Exception:
            continue
        # Try detect brand/model from filename (fast, no PDF open)
        try:
            brand, model = parse_brand_model_from_filename(p.stem)
        except Exception:
            brand, model = ("", "")
        records.append({
            "rel_path": rel,
            "folder_key": meta.get("folder_key"),
            "section_prefix": meta.get("section_prefix"),
            "size": stat.st_size,
            "mtime": stat.st_mtime,
            "num_pages": None,        # filled lazily via /api/pdf_meta
            "brand": brand or None,
            "model": model or None,
            "annotations": [],         # empty — see docstring
        })
    return records


def sync_db_from_memory() -> None:
    """Mirror the in-memory rows + PDFs + TOR + snapshots into the DB."""
    if db.db_path() is None:
        return
    try:
        db.sync_rows(ROWS)
    except Exception as e:
        sys.stderr.write(f"[db] sync_rows: {e}\n")
    try:
        db.sync_pdfs(_build_pdf_records_for_db())
    except Exception as e:
        sys.stderr.write(f"[db] sync_pdfs: {e}\n")
    try:
        db.sync_tor(TOR_SECTION_INDEX, TOR_PAGE_TEXTS)
    except Exception as e:
        sys.stderr.write(f"[db] sync_tor: {e}\n")
    # Snapshots — read from filesystem
    try:
        snaps = []
        if SNAPS_DIR.exists():
            for d in SNAPS_DIR.iterdir():
                if d.is_dir():
                    info = _read_snapshot(d)
                    if info:
                        snaps.append(info)
        db.sync_snapshots(snaps)
    except Exception as e:
        sys.stderr.write(f"[db] sync_snapshots: {e}\n")
    # PDF history (output/_pdf_history/<flat>/*.pdf)
    try:
        records = []
        if PDF_HISTORY.exists():
            for flat_dir in PDF_HISTORY.iterdir():
                if not flat_dir.is_dir():
                    continue
                for snap in flat_dir.glob("*.pdf"):
                    st = snap.stat()
                    records.append({
                        "pdf_rel": flat_dir.name,
                        "snapshot_filename": snap.name,
                        "ts": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
                        "tag": "",
                        "size": st.st_size,
                    })
        db.sync_pdf_history(records)
    except Exception as e:
        sys.stderr.write(f"[db] sync_pdf_history: {e}\n")


@app.route("/api/db/stats")
def api_db_stats():
    return jsonify(db.stats_summary())


@app.route("/api/db/audit")
def api_db_audit():
    limit = int(request.args.get("limit", 100))
    action = request.args.get("action")
    return jsonify({
        "entries": db.recent_audit(limit=min(500, max(1, limit)),
                                   action_filter=action),
    })


@app.route("/api/db/search")
def api_db_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"results": []})
    return jsonify({"results": db.fts_search(q, limit=80)})


@app.route("/api/db/section_progress")
def api_db_section_progress():
    return jsonify({"sections": db.section_progress()})


# ---------------------------------------------------------------------------
# HITL learning endpoints
# ---------------------------------------------------------------------------

@app.route("/api/learn/stats")
def api_learn_stats():
    days = int(request.args.get("days", 30))
    return jsonify(learning.feedback_stats(window_days=max(1, min(365, days))))


@app.route("/api/learn/patterns")
def api_learn_patterns():
    ptype = request.args.get("type")
    return jsonify({"patterns": learning.list_patterns(pattern_type=ptype)})


@app.route("/api/learn/retrain", methods=["POST"])
def api_learn_retrain():
    result = learning.retrain_patterns()
    try:
        db.log_audit(action="retrain", target_type="learning",
                     details=result, actor="user")
    except Exception: pass
    return jsonify({"ok": True, **result})


@app.route("/api/learn/pin_pattern", methods=["POST"])
def api_learn_pin_pattern():
    """Force-learn a pattern from a single row (Sprint S2.1).

    User clicks 'Pin as template' on a row they've just verified. We
    register the row's (filename, section, role, final Col D) tuple as
    learned_patterns with confidence=1.0 and samples_total=1, regardless
    of the usual PROMOTION_THRESHOLD=2. Future rows matching the trigger
    will use this template directly.
    """
    data = request.get_json(silent=True) or {}
    try:
        row_num = int(data.get("row", 0))
    except Exception:
        return jsonify({"ok": False, "error": "bad row"}), 400
    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404

    final_d = (row.get("D") or "").strip()
    if not final_d:
        return jsonify({"ok": False, "error": "row has no Col D to pin"}), 400

    section = row.get("section_inferred")
    pdf_rel = row.get("pdf_rel")
    role_info = detect_row_role(row_num)
    role = role_info.get("role", "unknown")

    pinned = []
    with db.conn() as c:
        # 1. filename_brand pattern (if Col D is brand_model)
        if final_d.startswith("ยี่ห้อ ") and pdf_rel:
            stem = Path(pdf_rel).stem
            tokens = re.findall(r"[A-Za-z][A-Za-z0-9]{2,}", stem)
            m = re.match(r"ยี่ห้อ\s+(\S+)\s+รุ่น\s+", final_d)
            if m and tokens:
                trigger = next((t for t in tokens if len(t) >= 3), tokens[0]).lower()
                brand = m.group(1)
                _upsert_pattern(c, "filename_brand", trigger, None, brand,
                                samples=1, correct=1, confidence=1.0,
                                note=f"pinned from R{row_num}")
                pinned.append(("filename_brand", trigger, brand))

        # 2. row_format_d pattern (always — captures shape preference)
        sec_root = ".".join(section.split(".")[:2]) if section else "unknown"
        shape = learning._shape_of_col_d(final_d)
        _upsert_pattern(c, "row_format_d", role, json.dumps([sec_root]),
                        shape, samples=1, correct=1, confidence=1.0,
                        note=f"pinned from R{row_num}")
        pinned.append(("row_format_d", f"{role}/{sec_root}", shape))

    try:
        db.log_audit(action="pin_pattern", target_type="row",
                     target_id=str(row_num),
                     details={"pinned": pinned, "final_d": final_d},
                     actor="user")
    except Exception: pass

    return jsonify({"ok": True, "row": row_num, "pinned": pinned,
                    "n_patterns": len(pinned)})


def _upsert_pattern(c, ptype, trigger, trigger_extra, value,
                     samples=1, correct=1, confidence=1.0, note=None):
    """Helper for pin_pattern — insert or update a learned_patterns row."""
    existing = c.execute(
        """SELECT pattern_id, samples_total, samples_correct
           FROM learned_patterns
           WHERE pattern_type=? AND trigger_key=?
             AND ifnull(trigger_extra,'')=ifnull(?, '')""",
        (ptype, trigger, trigger_extra),
    ).fetchone()
    if existing:
        c.execute(
            """UPDATE learned_patterns
               SET output_value=?, samples_total=samples_total+?,
                   samples_correct=samples_correct+?,
                   confidence=?, enabled=1,
                   note=COALESCE(?, note),
                   last_used_at=CURRENT_TIMESTAMP
               WHERE pattern_id=?""",
            (value, samples, correct, confidence, note, existing["pattern_id"]),
        )
    else:
        c.execute(
            """INSERT INTO learned_patterns
               (pattern_type, trigger_key, trigger_extra, output_value,
                samples_total, samples_correct, confidence, enabled, note)
               VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)""",
            (ptype, trigger, trigger_extra, value, samples, correct, confidence, note),
        )


@app.route("/api/settings/api_key", methods=["POST"])
def api_settings_save_api_key():
    """Save the Anthropic API key from the frontend.

    The key is written to ROOT/.env (gitignored), the Anthropic provider
    singleton is reset, and we reinstall it into the learning hook so
    subsequent auto-annotate calls go through Claude immediately — no
    restart needed.
    """
    data = request.get_json(silent=True) or {}
    key = (data.get("api_key") or "").strip()
    model = (data.get("model") or "").strip()
    budget = data.get("budget_usd_per_day")

    if key and not key.startswith("sk-ant-"):
        return jsonify({"ok": False, "error": "API key must start with 'sk-ant-'"}), 400

    env_path = ROOT / ".env"
    # Read existing .env (if present) and replace/append the keys
    lines: list[str] = []
    have = {"ANTHROPIC_API_KEY": False, "COMPLY_LLM": False,
            "COMPLY_LLM_MODEL": False, "COMPLY_LLM_BUDGET_USD_PER_DAY": False}
    try:
        if env_path.exists():
            for raw in env_path.read_text(encoding="utf-8").splitlines():
                stripped = raw.strip()
                if stripped.startswith("#") or "=" not in stripped:
                    lines.append(raw); continue
                k = stripped.split("=", 1)[0].strip()
                if k == "ANTHROPIC_API_KEY":
                    if key:
                        lines.append(f"ANTHROPIC_API_KEY={key}")
                        have[k] = True
                    # else: drop the line (user cleared the key)
                elif k == "COMPLY_LLM":
                    lines.append("COMPLY_LLM=anthropic" if key else "COMPLY_LLM=")
                    have[k] = True
                elif k == "COMPLY_LLM_MODEL":
                    if model:
                        lines.append(f"COMPLY_LLM_MODEL={model}"); have[k] = True
                    else:
                        lines.append(raw); have[k] = True
                elif k == "COMPLY_LLM_BUDGET_USD_PER_DAY":
                    if budget is not None:
                        lines.append(f"COMPLY_LLM_BUDGET_USD_PER_DAY={float(budget):.2f}")
                        have[k] = True
                    else:
                        lines.append(raw); have[k] = True
                else:
                    lines.append(raw)
    except Exception as e:
        return jsonify({"ok": False, "error": f"failed to read .env: {e}"}), 500

    # Append any missing keys
    if key and not have["ANTHROPIC_API_KEY"]:
        lines.append(f"ANTHROPIC_API_KEY={key}")
    if not have["COMPLY_LLM"]:
        lines.append("COMPLY_LLM=anthropic" if key else "COMPLY_LLM=")
    if not have["COMPLY_LLM_MODEL"] and model:
        lines.append(f"COMPLY_LLM_MODEL={model}")
    if not have["COMPLY_LLM_BUDGET_USD_PER_DAY"] and budget is not None:
        lines.append(f"COMPLY_LLM_BUDGET_USD_PER_DAY={float(budget):.2f}")

    # Write back atomically
    try:
        tmp = env_path.with_suffix(".env.tmp")
        tmp.write_text("\n".join(lines) + "\n", encoding="utf-8")
        tmp.replace(env_path)
        # Also restrict perms (key is sensitive)
        try: env_path.chmod(0o600)
        except Exception: pass
    except Exception as e:
        return jsonify({"ok": False, "error": f"failed to write .env: {e}"}), 500

    # Live-reload env vars + provider
    if key:
        os.environ["ANTHROPIC_API_KEY"] = key
        os.environ["COMPLY_LLM"] = "anthropic"
        if model: os.environ["COMPLY_LLM_MODEL"] = model
        if budget is not None: os.environ["COMPLY_LLM_BUDGET_USD_PER_DAY"] = str(float(budget))
    else:
        os.environ.pop("ANTHROPIC_API_KEY", None)
        os.environ["COMPLY_LLM"] = ""

    # Reset and reinstall the provider singleton
    try:
        from app import anthropic_provider as ap
        ap._provider = None
        installed = ap.install_into_learning() if key else False
        status = ap.get_provider().budget_status() if installed else {"available": False}
    except Exception as e:
        return jsonify({"ok": False, "error": f"provider init: {e}"}), 500

    try:
        db.log_audit(action="set_api_key", target_type="settings",
                     details={"model": model, "budget": budget,
                              "cleared": not bool(key)},
                     actor="user")
    except Exception: pass

    return jsonify({"ok": True, "installed": bool(key) and installed,
                    "status": status})


@app.route("/api/learn/llm_status")
def api_learn_llm_status():
    """LLM provider info + budget snapshot (Sprint S3 — surfaced in Settings).

    Phase 1 (Claude Code as core): try Claude Code provider FIRST, fall back
    to Anthropic API provider. Returns a ``provider_kind`` field so the UI
    can render the right copy ('Claude Max OAuth' vs 'API key + budget').
    """
    base = learning.llm_status()
    base["provider_kind"] = "off"
    # Try Claude Code first
    try:
        from app import claude_code_provider as cp
        p = cp.get_provider()
        if p:
            base.update(p.budget_status())
            base["provider_kind"] = "claude_code"
    except Exception as e:
        base["error_claude_code"] = str(e)
    # Fall back to Anthropic API
    if base.get("provider_kind") == "off":
        try:
            from app import anthropic_provider as ap
            p = ap.get_provider()
            if p:
                base.update(p.budget_status())
                base["provider_kind"] = "anthropic_api"
        except Exception as e:
            base["error_anthropic"] = str(e)
    # Today's call count + last call info (works for both providers)
    if base.get("provider_kind") != "off":
        try:
            with db.conn() as c:
                row = c.execute(
                    """SELECT COUNT(*) AS n,
                              COALESCE(SUM(input_tokens),0) AS in_tok,
                              COALESCE(SUM(output_tokens),0) AS out_tok,
                              COALESCE(SUM(cache_read_tokens),0) AS cache_read,
                              COALESCE(MAX(ts),'') AS last_ts
                       FROM llm_calls
                       WHERE substr(ts, 1, 10) = strftime('%Y-%m-%d', 'now')""",
                ).fetchone()
                if row:
                    base.update({
                        "calls_today": row["n"],
                        "tokens_in_today": row["in_tok"],
                        "tokens_out_today": row["out_tok"],
                        "tokens_cache_read_today": row["cache_read"],
                        "last_call_ts": row["last_ts"],
                    })
        except Exception as e:
            base["error_calls"] = str(e)
    return jsonify(base)


@app.route("/api/claude/stream")
def api_claude_stream():
    """Phase 1 (Claude Code as core): stream Claude's reasoning live via SSE.

    Frontend opens an EventSource on this endpoint per row. Each line is a
    JSON event from ``ClaudeCodeProvider.propose_streaming``:
      • {type:"thinking", text:...}
      • {type:"tool_use",  name:..., input:{...}}     ← Read/Grep/propose_*
      • {type:"tool_result", name:..., text:...}
      • {type:"text",      content:...}
      • {type:"result",    proposal:{...}, usage:{...}, cost_usd, elapsed_ms}
      • {type:"error",     error:...}

    The endpoint runs the async generator on a private event loop in this
    request's thread (Flask is sync). For one user that's plenty; if we
    ever scale, swap to Hypercorn/Quart.
    """
    try:
        row_num = int(request.args.get("row", "0"))
    except ValueError:
        return jsonify({"ok": False, "error": "row required"}), 400
    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404

    # Lazy-import to keep boot fast when SDK isn't installed
    try:
        from app import claude_code_provider as cp
    except ImportError:
        return jsonify({"ok": False, "error": "claude-agent-sdk not installed"}), 503
    p = cp.get_provider()
    if p is None:
        return jsonify({
            "ok": False,
            "error": "Claude Code provider unavailable",
            "hint": "Set COMPLY_LLM=claude_code and ensure 'claude' CLI is logged in",
        }), 503

    # Build row_context the same shape AnthropicProvider expects
    role_info = detect_row_role(row_num)
    pdf_filename = None
    if row.get("pdf_rel"):
        pdf_filename = Path(row["pdf_rel"]).name
    # Lightweight rule-based proposal as a hint for Claude
    rule_plan = None
    try:
        rp = auto_annotate_plan(row_num)
        if rp.get("ok"):
            rule_plan = {
                "proposed_d": rp.get("proposed_d", ""),
                "generator": rp.get("generator", ""),
                "confidence": rp.get("confidence", 0),
            }
    except Exception:
        pass
    row_context = {
        "row": row_num,
        "section": row.get("section_inferred"),
        "role": role_info.get("role", "unknown"),
        "col_a": row.get("A"),
        "col_b": row.get("B", ""),
        "col_c_current": row.get("C", ""),
        "col_d_current": row.get("D", ""),
        "col_e": row.get("E", ""),
        "pdf_rel": row.get("pdf_rel"),
        "pdf_filename": pdf_filename,
        "rule_proposal": rule_plan,
    }

    import asyncio as _aio

    def event_stream():
        loop = _aio.new_event_loop()
        try:
            agen = p.propose_streaming(row_context=row_context)
            agen_iter = agen.__aiter__()
            while True:
                try:
                    event = loop.run_until_complete(agen_iter.__anext__())
                except StopAsyncIteration:
                    break
                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
            yield "event: done\ndata: {}\n\n"
        except Exception as e:
            err = {"type": "error", "error": str(e)}
            yield f"data: {json.dumps(err, ensure_ascii=False)}\n\n"
        finally:
            try:
                loop.close()
            except Exception:
                pass

    headers = {
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",       # disable nginx buffering if proxied
        "Connection": "keep-alive",
    }
    return Response(event_stream(), mimetype="text/event-stream", headers=headers)


# ─────────────────────────────────────────────────────────────────────
# Phase 2 — Catalog library API (multi-company / multi-project)
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/catalogs")
def api_catalogs_list():
    """List catalogs with optional filters: ?brand=&category=&section=&q=&archived=0."""
    archived = request.args.get("archived") == "1"
    items = catalog.list_catalogs(
        brand=request.args.get("brand") or None,
        category=request.args.get("category") or None,
        section=request.args.get("section") or None,
        q=request.args.get("q") or None,
        archived=archived,
        limit=int(request.args.get("limit", 200)),
    )
    return jsonify({"ok": True, "items": items, "count": len(items)})


@app.route("/api/catalogs/stats")
def api_catalogs_stats():
    """Quick counts for the catalog rail badge + dashboard."""
    return jsonify(catalog.stats())


@app.route("/api/catalogs/<int:catalog_id>")
def api_catalog_get(catalog_id):
    """Full detail for a single catalog (incl. annotations + page text)."""
    item = catalog.get_catalog(catalog_id)
    if not item:
        return jsonify({"ok": False, "error": "not found"}), 404
    return jsonify({"ok": True, "catalog": item})


@app.route("/api/catalogs/<int:catalog_id>", methods=["PATCH"])
def api_catalog_update(catalog_id):
    """Update editable metadata fields. Body = JSON with any of:
    {brand, model, category, section_hint, description, metadata_json, archived}.
    """
    if not catalog.get_catalog(catalog_id):
        return jsonify({"ok": False, "error": "not found"}), 404
    data = request.get_json(silent=True) or {}
    try:
        catalog.update_catalog(catalog_id, **data)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    db.log_audit(action="catalog_update", target_type="catalog",
                 target_id=str(catalog_id), details=data, actor="user")
    return jsonify({"ok": True, "catalog": catalog.get_catalog(catalog_id)})


@app.route("/api/catalogs/<int:catalog_id>/links")
def api_catalog_links(catalog_id):
    """Which rows currently use this catalog?"""
    return jsonify({"ok": True, "links": catalog.list_links_for_catalog(catalog_id)})


@app.route("/api/catalogs/reingest", methods=["POST"])
def api_catalogs_reingest():
    """Re-scan output/ to pick up new PDFs (or refresh all metadata when
    ?force=1)."""
    force = request.args.get("force") == "1"
    result = catalog.ingest_output_dir(OUTPUT, force=force)
    return jsonify(result)


# ─── Catalog annotations (DB-stored, editable independent of PDF) ────

@app.route("/api/catalogs/<int:catalog_id>/annotations")
def api_catalog_annots(catalog_id):
    page = request.args.get("page")
    page_n = int(page) if page else None
    return jsonify({"ok": True,
                    "annotations": catalog.list_annotations(catalog_id, page=page_n)})


@app.route("/api/catalogs/<int:catalog_id>/annotations", methods=["POST"])
def api_catalog_annots_add(catalog_id):
    data = request.get_json(silent=True) or {}
    try:
        annot_id = catalog.add_annotation(
            catalog_id=catalog_id,
            page=int(data["page"]),
            type=data["type"],
            rect=data["rect"],
            contents=data.get("contents", ""),
            color=data.get("color"),
            border_width=float(data.get("border_width", 1.0)),
            anchor_text=data.get("anchor_text"),
        )
    except (ValueError, KeyError) as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    db.log_audit(action="catalog_annot_add", target_type="catalog",
                 target_id=str(catalog_id),
                 details={"annot_id": annot_id, "page": data.get("page")},
                 actor="user")
    return jsonify({"ok": True, "annot_id": annot_id})


@app.route("/api/catalogs/<int:catalog_id>/annotations/<int:annot_id>",
           methods=["PATCH"])
def api_catalog_annots_update(catalog_id, annot_id):
    data = request.get_json(silent=True) or {}
    try:
        catalog.update_annotation(annot_id, **data)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    return jsonify({"ok": True})


@app.route("/api/catalogs/<int:catalog_id>/annotations/<int:annot_id>",
           methods=["DELETE"])
def api_catalog_annots_delete(catalog_id, annot_id):
    catalog.delete_annotation(annot_id)
    return jsonify({"ok": True})


# ─── Companies / projects ────────────────────────────────────────────

@app.route("/api/companies")
def api_companies():
    return jsonify({"ok": True, "items": catalog.list_companies()})


@app.route("/api/companies", methods=["POST"])
def api_companies_add():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "name required"}), 400
    cid = catalog.upsert_company(name=name, code=data.get("code"))
    return jsonify({"ok": True, "company_id": cid})


@app.route("/api/projects")
def api_projects():
    company_id = request.args.get("company_id")
    cid = int(company_id) if company_id else None
    return jsonify({"ok": True,
                    "items": catalog.list_projects(company_id=cid),
                    "active": catalog.get_active_project()})


@app.route("/api/projects", methods=["POST"])
def api_projects_add():
    data = request.get_json(silent=True) or {}
    try:
        pid = catalog.upsert_project(
            company_id=int(data["company_id"]),
            name=data["name"],
            code=data.get("code"),
            xlsx_rel=data.get("xlsx_rel"),
            output_rel=data.get("output_rel", "output"),
        )
    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    return jsonify({"ok": True, "project_id": pid})


@app.route("/api/projects/<int:project_id>/activate", methods=["POST"])
def api_projects_activate(project_id):
    catalog.set_active_project(project_id)
    return jsonify({"ok": True, "active": catalog.get_active_project()})


# ─── Apply catalog → row (writes Col D + records link) ───────────────

@app.route("/api/row/apply_catalog", methods=["POST"])
def api_row_apply_catalog():
    """Bind a catalog (and optional page) to a project row, generating Col D.

    Body: {row, catalog_id, page?, col_d_text?}

    If ``col_d_text`` is omitted we synthesize one using the same convention
    as the existing rule-based generator (``make_col_d_for_row``).
    """
    data = request.get_json(silent=True) or {}
    try:
        row_num = int(data["row"])
        cat_id = int(data["catalog_id"])
    except (KeyError, ValueError):
        return jsonify({"ok": False, "error": "row, catalog_id required"}), 400

    row = next((r for r in ROWS if r["row"] == row_num), None)
    if not row:
        return jsonify({"ok": False, "error": "row not found"}), 404
    cat = catalog.get_catalog(cat_id)
    if not cat:
        return jsonify({"ok": False, "error": "catalog not found"}), 404

    page = int(data["page"]) if data.get("page") else None
    col_d = (data.get("col_d_text") or "").strip()
    if not col_d:
        # Synthesize: use existing generator if possible
        try:
            role_info = detect_row_role(row_num)
            pdf_path = OUTPUT / cat["pdf_rel"]
            col_d = make_col_d_for_row(row, role_info, pdf_path, page) or ""
        except Exception as e:
            sys.stderr.write(f"[apply_catalog] synth col_d failed: {e}\n")
            col_d = f"เอกสาร {cat.get('section_hint', '?')} {cat.get('brand') or ''} {cat.get('model') or ''}".strip()
            if page:
                col_d += f" หน้า {page}"

    proj = catalog.get_active_project()
    if not proj:
        return jsonify({"ok": False, "error": "no active project"}), 500

    # Pre-snap, write xlsx Col D, then refresh + record link
    _run_version_cmd(["snap", f"pre-apply-catalog-row-{row_num}"], timeout=60)

    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        original = (ws.cell(row_num, 4).value or "").strip()
        ws.cell(row_num, 4).value = col_d
        tmp = XLSX_PATH.with_suffix(".xlsx.tmp")
        wb.save(str(tmp))
        with open(tmp, "rb") as f:
            buf = f.read()
        fd = os.open(str(XLSX_PATH), os.O_WRONLY | os.O_TRUNC)
        try:
            os.write(fd, buf)
        finally:
            os.close(fd)
        tmp.unlink(missing_ok=True)
    except Exception as e:
        return jsonify({"ok": False, "error": f"xlsx write failed: {e}"}), 500

    # Update memory + DB mirror
    load_rows()
    sync_db_from_memory()

    # Record link
    catalog.bind_row_to_catalog(
        project_id=int(proj["project_id"]),
        row_num=row_num, catalog_id=cat_id,
        page=page, col_d_text=col_d,
    )
    db.log_audit(action="apply_catalog", target_type="row",
                 target_id=str(row_num),
                 before=original, after=col_d,
                 details={"catalog_id": cat_id, "page": page}, actor="user")

    return jsonify({"ok": True, "row": row_num, "col_d": col_d,
                    "catalog_id": cat_id, "page": page})


@app.route("/api/learn/feedback", methods=["POST"])
def api_learn_feedback():
    """Direct feedback recorder (e.g., when user edits Col D outside the
    auto-annotate flow but still wants the system to learn)."""
    data = request.get_json(silent=True) or {}
    required = ("row_num", "user_action", "final_d")
    if any(k not in data for k in required):
        return jsonify({"ok": False, "error": f"missing {required}"}), 400
    fb_id = learning.record_feedback(
        row_num=int(data["row_num"]),
        section=data.get("section"),
        input_b=data.get("input_b", ""),
        input_pdf_rel=data.get("input_pdf_rel"),
        input_role=data.get("input_role", ""),
        input_filename=data.get("input_filename"),
        suggested_c=data.get("suggested_c", ""),
        suggested_d=data.get("suggested_d", ""),
        suggested_annots=data.get("suggested_annots") or [],
        confidence=float(data.get("confidence", 0)),
        generator=data.get("generator", "manual"),
        provenance=data.get("provenance"),
        user_action=data["user_action"],
        final_c=data.get("final_c", ""),
        final_d=data["final_d"],
        final_annots=data.get("final_annots") or [],
    )
    return jsonify({"ok": True, "fb_id": fb_id})


@app.route("/api/learn/pattern/<int:pattern_id>", methods=["POST"])
def api_learn_toggle_pattern(pattern_id: int):
    """Enable/disable a learned pattern (or override its output)."""
    data = request.get_json(silent=True) or {}
    fields = []
    params: list = []
    if "enabled" in data:
        fields.append("enabled = ?"); params.append(1 if data["enabled"] else 0)
    if "output_value" in data:
        fields.append("output_value = ?"); params.append(str(data["output_value"]))
    if "note" in data:
        fields.append("note = ?"); params.append(str(data["note"]))
    if not fields:
        return jsonify({"ok": False, "error": "nothing to update"}), 400
    params.append(pattern_id)
    with db.conn() as c:
        c.execute(f"UPDATE learned_patterns SET {', '.join(fields)} WHERE pattern_id = ?",
                  params)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# HTML template
# ---------------------------------------------------------------------------

INDEX_HTML = r"""<!doctype html>
<html lang="th">
<head>
<meta charset="utf-8">
<title>Comply Verify GUI — Smart Plant 1</title>
<style>
/* =====================================================================
 * Comply Verify — Production UI v2
 * Design tokens · components · layout · dark mode · motion
 * Class names referenced by JS are preserved verbatim.
 * ================================================================== */

/* ── Tokens ─────────────────────────────────────────────────────── */
:root {
  /* Brand & semantic palette (light) */
  --c-bg:           #f6f7f9;
  --c-bg-subtle:    #eceef2;
  --c-surface:      #ffffff;
  --c-surface-2:    #f9fafb;
  --c-surface-3:    #f3f4f6;
  --c-overlay:      rgba(15, 23, 42, 0.55);

  --c-border:       #e5e7eb;
  --c-border-strong:#d1d5db;
  --c-divider:      #f1f2f4;

  --c-text:         #0f172a;
  --c-text-muted:   #475569;
  --c-text-soft:    #64748b;
  --c-text-faint:   #94a3b8;
  --c-text-invert:  #ffffff;

  --c-primary:      #2563eb;
  --c-primary-hover:#1d4ed8;
  --c-primary-soft: #dbeafe;
  --c-primary-text: #1e3a8a;

  --c-success:      #10b981;
  --c-success-hover:#059669;
  --c-success-soft: #d1fae5;
  --c-success-text: #065f46;

  --c-warn:         #f59e0b;
  --c-warn-hover:   #d97706;
  --c-warn-soft:    #fef3c7;
  --c-warn-text:    #92400e;

  --c-danger:       #ef4444;
  --c-danger-hover: #dc2626;
  --c-danger-soft:  #fee2e2;
  --c-danger-text:  #991b1b;

  --c-info:         #3b82f6;
  --c-info-soft:    #dbeafe;
  --c-info-text:    #1e40af;

  --c-purple:       #a855f7;
  --c-purple-soft:  #f3e8ff;
  --c-purple-text:  #6b21a8;

  --c-teal:         #14b8a6;
  --c-teal-soft:    #ccfbf1;
  --c-teal-text:    #115e59;

  --c-pink-soft:    #fce7f3;
  --c-pink-text:    #9d174d;

  /* PDF viewer chrome */
  --c-canvas-bg:    #2c2e33;
  --c-canvas-edit:  #3a3c42;

  /* Focus ring */
  --c-focus:        #3b82f6;
  --c-focus-ring:   0 0 0 3px rgba(59, 130, 246, 0.30);

  /* Spacing scale (4px grid) */
  --s-0: 0;
  --s-1: 2px;
  --s-2: 4px;
  --s-3: 6px;
  --s-4: 8px;
  --s-5: 10px;
  --s-6: 12px;
  --s-7: 16px;
  --s-8: 20px;
  --s-9: 24px;
  --s-10: 32px;
  --s-12: 48px;

  /* Radius */
  --r-sm: 4px;
  --r-md: 6px;
  --r-lg: 8px;
  --r-xl: 12px;
  --r-pill: 999px;

  /* Elevation */
  --e-0: none;
  --e-1: 0 1px 2px rgba(15, 23, 42, 0.06), 0 1px 1px rgba(15, 23, 42, 0.04);
  --e-2: 0 2px 6px rgba(15, 23, 42, 0.08), 0 1px 2px rgba(15, 23, 42, 0.04);
  --e-3: 0 6px 16px rgba(15, 23, 42, 0.10), 0 1px 3px rgba(15, 23, 42, 0.06);
  --e-4: 0 12px 28px rgba(15, 23, 42, 0.14), 0 2px 6px rgba(15, 23, 42, 0.06);
  --e-5: 0 24px 48px rgba(15, 23, 42, 0.20), 0 4px 12px rgba(15, 23, 42, 0.08);

  /* Type */
  --f-sans: ui-sans-serif, -apple-system, BlinkMacSystemFont, "Inter", "Segoe UI",
            "Sukhumvit Set", "Noto Sans Thai", "Thonburi", system-ui, sans-serif;
  --f-mono: ui-monospace, "SF Mono", Menlo, Consolas, "Roboto Mono", monospace;

  /* Type scale (bumped for production density 2026-05) */
  --t-xs:   11px;
  --t-sm:   12px;
  --t-base: 13px;
  --t-md:   14px;   /* body default */
  --t-lg:   15px;
  --t-xl:   17px;
  --t-2xl:  21px;
  --t-3xl:  26px;

  --lh-tight:  1.25;
  --lh-normal: 1.45;
  --lh-loose:  1.6;

  /* Motion */
  --ease-out:  cubic-bezier(0.16, 1, 0.3, 1);
  --ease-in:   cubic-bezier(0.4, 0, 1, 1);
  --ease-std:  cubic-bezier(0.4, 0, 0.2, 1);

  --d-fast:    120ms;
  --d-base:    180ms;
  --d-slow:    260ms;

  /* Layout */
  --topbar-h:    52px;
  --ribbon-h:    52px;             /* mode-tabs + sub-toolbar */
  --statusbar-h: 38px;             /* bottom status strip — tall enough for verdict pills (Phase A6) */
  --rail-w:      48px;             /* left activity rail (icon-only) */
  --rail-panel-w: 280px;           /* expanded panel that slides next to rail */
  --ai-w:        340px;            /* right AI pane */
  --pane-head-h: 40px;
  --toolbar-h:   40px;            /* unified canvas/edit toolbar height */
  --action-bar-h: 76px;            /* fixed; tall enough for verdict + notes */
  --tree-w:      320px;
  --tree-w-md:   280px;
  --tree-row-h:  32px;
  --btn-h:       32px;
  --btn-h-sm:    28px;
  --btn-h-toolbar: 28px;

  color-scheme: light;
}

/* Dark mode — auto + manual */
@media (prefers-color-scheme: dark) {
  :root:not([data-theme="light"]) {
    --c-bg:           #0b0d12;
    --c-bg-subtle:    #11141b;
    --c-surface:      #161a22;
    --c-surface-2:    #1c2029;
    --c-surface-3:    #232834;
    --c-overlay:      rgba(0, 0, 0, 0.7);

    --c-border:       #2a2f3a;
    --c-border-strong:#3a4150;
    --c-divider:      #1f2430;

    --c-text:         #e8ecf3;
    --c-text-muted:   #b6bdca;
    --c-text-soft:    #8a93a4;
    --c-text-faint:   #5a6373;

    --c-primary:      #60a5fa;
    --c-primary-hover:#3b82f6;
    --c-primary-soft: rgba(59, 130, 246, 0.18);
    --c-primary-text: #bfdbfe;

    --c-success-soft: rgba(16, 185, 129, 0.18);
    --c-success-text: #6ee7b7;

    --c-warn-soft:    rgba(245, 158, 11, 0.18);
    --c-warn-text:    #fcd34d;

    --c-danger-soft:  rgba(239, 68, 68, 0.18);
    --c-danger-text:  #fca5a5;

    --c-info-soft:    rgba(59, 130, 246, 0.18);
    --c-info-text:    #bfdbfe;

    --c-purple-soft:  rgba(168, 85, 247, 0.18);
    --c-purple-text:  #d8b4fe;

    --c-teal-soft:    rgba(20, 184, 166, 0.18);
    --c-teal-text:    #5eead4;

    --c-pink-soft:    rgba(236, 72, 153, 0.18);
    --c-pink-text:    #fbcfe8;

    --c-canvas-bg:    #06080c;
    --c-canvas-edit:  #0e1118;

    --e-1: 0 1px 2px rgba(0, 0, 0, 0.40);
    --e-2: 0 2px 6px rgba(0, 0, 0, 0.45);
    --e-3: 0 6px 16px rgba(0, 0, 0, 0.55);
    --e-4: 0 12px 28px rgba(0, 0, 0, 0.65);
    --e-5: 0 24px 48px rgba(0, 0, 0, 0.75);

    color-scheme: dark;
  }
}
:root[data-theme="dark"] {
  --c-bg:           #0b0d12;
  --c-bg-subtle:    #11141b;
  --c-surface:      #161a22;
  --c-surface-2:    #1c2029;
  --c-surface-3:    #232834;
  --c-overlay:      rgba(0, 0, 0, 0.7);
  --c-border:       #2a2f3a;
  --c-border-strong:#3a4150;
  --c-divider:      #1f2430;
  --c-text:         #e8ecf3;
  --c-text-muted:   #b6bdca;
  --c-text-soft:    #8a93a4;
  --c-text-faint:   #5a6373;
  --c-primary:      #60a5fa;
  --c-primary-hover:#3b82f6;
  --c-primary-soft: rgba(59, 130, 246, 0.18);
  --c-primary-text: #bfdbfe;
  --c-success-soft: rgba(16, 185, 129, 0.18);
  --c-success-text: #6ee7b7;
  --c-warn-soft:    rgba(245, 158, 11, 0.18);
  --c-warn-text:    #fcd34d;
  --c-danger-soft:  rgba(239, 68, 68, 0.18);
  --c-danger-text:  #fca5a5;
  --c-info-soft:    rgba(59, 130, 246, 0.18);
  --c-info-text:    #bfdbfe;
  --c-purple-soft:  rgba(168, 85, 247, 0.18);
  --c-purple-text:  #d8b4fe;
  --c-teal-soft:    rgba(20, 184, 166, 0.18);
  --c-teal-text:    #5eead4;
  --c-pink-soft:    rgba(236, 72, 153, 0.18);
  --c-pink-text:    #fbcfe8;
  --c-canvas-bg:    #06080c;
  --c-canvas-edit:  #0e1118;
  --e-1: 0 1px 2px rgba(0, 0, 0, 0.40);
  --e-2: 0 2px 6px rgba(0, 0, 0, 0.45);
  --e-3: 0 6px 16px rgba(0, 0, 0, 0.55);
  --e-4: 0 12px 28px rgba(0, 0, 0, 0.65);
  --e-5: 0 24px 48px rgba(0, 0, 0, 0.75);
  color-scheme: dark;
}

/* ── Reset ──────────────────────────────────────────────────────── */
*, *::before, *::after { box-sizing: border-box; }
html, body {
  margin: 0; padding: 0; height: 100%;
  overflow: hidden;             /* lock page scroll on BOTH — prevents
                                   scrollIntoView() from cascading up to
                                   the document, which would push the
                                   topbar off-screen on row navigation */
  overscroll-behavior: none;    /* no rubber-band on macOS trackpad */
}
html { -webkit-font-smoothing: antialiased; -moz-osx-font-smoothing: grayscale; text-size-adjust: 100%; }
body {
  font-family: var(--f-sans);
  font-size: var(--t-md);
  line-height: var(--lh-normal);
  color: var(--c-text);
  background: var(--c-bg);
}
button, input, select, textarea { font-family: inherit; color: inherit; }
button { cursor: pointer; }
button:disabled { cursor: not-allowed; }
img { display: block; max-width: 100%; }
a { color: var(--c-primary); text-decoration: none; }
a:hover { text-decoration: underline; }
::selection { background: var(--c-primary-soft); color: var(--c-primary-text); }

/* Custom scrollbars (subtle, persistent) */
@supports (scrollbar-width: thin) {
  * { scrollbar-width: thin; scrollbar-color: var(--c-border-strong) transparent; }
}
::-webkit-scrollbar { width: 10px; height: 10px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb {
  background: var(--c-border-strong);
  border: 2px solid transparent; background-clip: padding-box;
  border-radius: 999px;
}
::-webkit-scrollbar-thumb:hover { background: var(--c-text-faint); background-clip: padding-box; border: 2px solid transparent; }

/* Focus-visible ring (universal) */
:focus { outline: none; }
:focus-visible {
  outline: 2px solid var(--c-focus);
  outline-offset: 2px;
  border-radius: var(--r-sm);
}
button:focus-visible, .btn:focus-visible, .ab-btn:focus-visible,
.versions-btn:focus-visible, .canvas-toolbar button:focus-visible,
.edit-toolbar button:focus-visible, .edit-toggle-btn:focus-visible,
.col-d-menu button:focus-visible, .mobile-tabs button:focus-visible {
  outline: 2px solid var(--c-focus);
  outline-offset: 2px;
}
input:focus-visible, select:focus-visible, textarea:focus-visible {
  outline: none; border-color: var(--c-primary);
  box-shadow: var(--c-focus-ring);
}

/* Visually-hidden helper for SR-only labels */
.sr-only {
  position: absolute; width: 1px; height: 1px; padding: 0; margin: -1px;
  overflow: hidden; clip: rect(0, 0, 0, 0); white-space: nowrap; border: 0;
}

/* SVG icon system — single source of glyphs (replaces inline emojis) */
.ico {
  width: 16px; height: 16px;
  fill: none; stroke: currentColor;
  stroke-width: 2;
  stroke-linecap: round; stroke-linejoin: round;
  display: inline-block; vertical-align: -3px;
  flex-shrink: 0;
}
.ico-sm { width: 14px; height: 14px; vertical-align: -2px; }
.ico-xs { width: 12px; height: 12px; vertical-align: -2px; stroke-width: 2.4; }
.ico-lg { width: 20px; height: 20px; vertical-align: -4px; }
.ico-xl { width: 24px; height: 24px; }

/* Skeleton loader (used during async fetches) */
.skel {
  background: linear-gradient(90deg,
    var(--c-surface-2) 0%,
    var(--c-surface-3) 50%,
    var(--c-surface-2) 100%);
  background-size: 200% 100%;
  animation: skel-shimmer 1.6s ease-in-out infinite;
  border-radius: var(--r-md);
}
@keyframes skel-shimmer {
  0%   { background-position: 200% 0; }
  100% { background-position: -200% 0; }
}
.skel-line { height: 12px; margin: var(--s-3) 0; }
.skel-line.tall { height: 16px; }
.skel-block { height: 80px; }
.skel-img { aspect-ratio: 595 / 842; max-width: 70%; margin: 0 auto; box-shadow: var(--e-2); }
.skel-stack { display: flex; flex-direction: column; gap: var(--s-3); padding: var(--s-7); }
@media (prefers-reduced-motion: reduce) {
  .skel { animation: none; background: var(--c-surface-2); }
}

/* ── App shell ─────────────────────────────────────────────────────
   Acrobat-style layout — 5 rows × 5 cols:
     safetop : host-app overlay clearance (Claude Preview, etc.)
     topbar  : brand + breadcrumb + cmdK
     ribbon  : mode tabs + sub-toolbar (Verify / Edit / Re-annotate / Apply)
     content : LEFT-RAIL · TREE · CENTER · PDF · AI-PANE
     status  : current row · verdict · progress · Claude online · save state
*/
:root { --safe-top: 0px; }
body[data-embedded="1"] { --safe-top: 56px; }

#app {
  display: grid;
  grid-template-rows:
    var(--safe-top)
    var(--topbar-h)
    var(--ribbon-h)
    1fr
    auto                    /* action bar (legacy, will fold into status bar) */
    var(--statusbar-h);
  grid-template-columns:
    var(--rail-w)
    minmax(var(--tree-w-md), var(--tree-w))
    1fr
    1fr
    var(--ai-w);
  grid-template-areas:
    "safetop safetop safetop safetop safetop"
    "topbar  topbar  topbar  topbar  topbar"
    "ribbon  ribbon  ribbon  ribbon  ribbon"
    "rail    tree    center  pdf     ai"
    "rail    action  action  action  ai"
    "status  status  status  status  status";
  height: 100vh;
  background: var(--c-bg);
}
#app::before {
  content: '';
  grid-area: safetop;
  background: var(--c-bg);
}
#app > .action-bar  { grid-area: action; }
#app > .activity-rail { grid-area: rail; }
#app > .context-ribbon { grid-area: ribbon; }
#app > .ai-pane     { grid-area: ai; }
#app > .status-bar  { grid-area: status; }

/* When the AI pane is collapsed, shrink that column to 0 */
body[data-ai-pane="0"] #app { grid-template-columns:
  var(--rail-w) minmax(var(--tree-w-md), var(--tree-w)) 1fr 1fr 0; }
body[data-ai-pane="0"] .ai-pane { display: none; }

/* When the rail-panel is showing instead of the tree, swap tree out */
body[data-rail-panel] .tree-pane { display: none; }
body[data-rail-panel] #app { grid-template-columns:
  var(--rail-w) var(--rail-panel-w) 1fr 1fr var(--ai-w); }
body[data-rail-panel="0"] .tree-pane { display: flex; }

#app.has-sync-banner { padding-top: 32px; }

/* ── Activity Rail (left, icon-only) ─────────────────────────────
   Persistent vertical rail with icons. Click an icon to expand a
   panel beside the rail (overlays/replaces the tree pane on small
   screens, sits beside it on large ones). Click again to collapse. */
.activity-rail {
  display: flex; flex-direction: column;
  background: var(--c-surface);
  border-right: 1px solid var(--c-border);
  padding: var(--s-3) 0;
  gap: 2px;
  overflow-y: auto;
}
.activity-rail .rail-btn {
  width: var(--rail-w); height: var(--rail-w);
  display: inline-flex; align-items: center; justify-content: center;
  background: transparent; border: 0;
  color: var(--c-text-soft);
  cursor: pointer;
  position: relative;
  border-radius: 0;
  transition: color var(--d-fast) var(--ease-std), background var(--d-fast) var(--ease-std);
}
.activity-rail .rail-btn:hover { color: var(--c-text); background: var(--c-surface-2); }
.activity-rail .rail-btn .ico { width: 20px; height: 20px; }
.activity-rail .rail-btn.active {
  color: var(--c-primary);
  background: var(--c-primary-soft);
}
.activity-rail .rail-btn.active::before {
  content: ''; position: absolute;
  left: 0; top: 8px; bottom: 8px; width: 3px;
  background: var(--c-primary);
  border-radius: 0 2px 2px 0;
}
.activity-rail .rail-spacer { flex: 1; }
.activity-rail .rail-sep {
  margin: var(--s-3) auto;
  width: 24px; height: 1px;
  background: var(--c-border);
}

/* Rail panel (slides in beside rail, replaces the tree pane area) */
.rail-panel {
  background: var(--c-surface);
  border-right: 1px solid var(--c-border);
  display: none;
  flex-direction: column;
  min-width: 0;
}
body[data-rail-panel] .rail-panel { display: flex; grid-area: tree; }
.rail-panel-head {
  display: flex; align-items: center; gap: var(--s-3);
  padding: 0 var(--s-5);
  height: var(--pane-head-h);
  border-bottom: 1px solid var(--c-border);
  font-size: var(--t-base); font-weight: 600;
  color: var(--c-text);
}
.rail-panel-head .panel-title { flex: 1; }
.rail-panel-head .panel-close {
  background: transparent; border: 0;
  color: var(--c-text-soft);
  width: 24px; height: 24px;
  display: inline-flex; align-items: center; justify-content: center;
  border-radius: var(--r-sm);
  cursor: pointer;
}
.rail-panel-head .panel-close:hover { background: var(--c-surface-2); color: var(--c-text); }
.rail-panel-body {
  flex: 1; overflow: auto;
  padding: var(--s-5);
  font-size: var(--t-sm);
}

/* ── Context Ribbon (mode tabs + sub-toolbar) ───────────────────── */
.context-ribbon {
  background: var(--c-surface);
  border-bottom: 1px solid var(--c-border);
  display: flex; flex-direction: column;
  z-index: 8;
}
.ribbon-tabs {
  display: flex; gap: 2px;
  padding: 6px var(--s-7) 0 var(--s-7);
  background: var(--c-surface);
  height: 32px;
  align-items: center;
}
.ribbon-tab {
  display: inline-flex; align-items: center; gap: var(--s-2);
  padding: 4px var(--s-5);
  height: 26px;
  border: 1px solid transparent;
  background: transparent;
  color: var(--c-text-muted);
  border-radius: var(--r-md) var(--r-md) 0 0;
  font-size: var(--t-sm); font-weight: 600;
  cursor: pointer;
  position: relative; top: 1px;
  transition: all var(--d-fast) var(--ease-std);
}
.ribbon-tab:hover { color: var(--c-text); background: var(--c-surface-2); }
.ribbon-tab.active {
  color: var(--c-primary-text);
  background: var(--c-surface);
  border-color: var(--c-border);
  border-bottom-color: var(--c-surface);
  z-index: 2;
}
.ribbon-tab .ico { width: 14px; height: 14px; }

.ribbon-subtoolbar {
  flex: 1;
  display: flex; align-items: center; gap: var(--s-3);
  padding: 0 var(--s-7);
  border-top: 1px solid var(--c-border);
  font-size: var(--t-sm);
  background: var(--c-surface);
  min-height: 0;
  overflow-x: auto;
  scrollbar-width: none;
}
.ribbon-subtoolbar::-webkit-scrollbar { display: none; }
.ribbon-subtoolbar > * { flex-shrink: 0; }
.ribbon-subtoolbar .rb-group { display: inline-flex; gap: 2px; align-items: center; }
.ribbon-subtoolbar .rb-sep {
  width: 1px; height: 18px;
  background: var(--c-border);
  margin: 0 var(--s-3);
}
.ribbon-subtoolbar .rb-spacer { flex: 1; }
.ribbon-subtoolbar button {
  display: inline-flex; align-items: center; gap: var(--s-2);
  height: 28px; padding: 0 var(--s-4);
  border: 1px solid transparent;
  background: transparent;
  color: var(--c-text-muted);
  border-radius: var(--r-md);
  font-size: var(--t-sm); font-weight: 500;
  cursor: pointer;
  white-space: nowrap;
  transition: all var(--d-fast) var(--ease-std);
}
.ribbon-subtoolbar button:hover { background: var(--c-surface-2); color: var(--c-text); }
.ribbon-subtoolbar button.active { background: var(--c-warn); color: white; }
.ribbon-subtoolbar button:disabled { opacity: 0.4; cursor: not-allowed; }
.ribbon-subtoolbar label {
  display: inline-flex; align-items: center; gap: var(--s-2);
  font-size: var(--t-sm); color: var(--c-text-muted);
  cursor: pointer;
}
.ribbon-subtoolbar .rb-info {
  font-size: var(--t-sm); color: var(--c-text-soft);
  font-variant-numeric: tabular-nums;
}

/* Show only the active mode's sub-toolbar */
.ribbon-mode-bar { display: none; flex: 1; align-items: center; gap: var(--s-3); padding: 0 var(--s-7); }
body[data-mode="verify"] .ribbon-mode-bar.mode-verify { display: flex; }
body[data-mode="edit"] .ribbon-mode-bar.mode-edit { display: flex; }
body[data-mode="reannotate"] .ribbon-mode-bar.mode-reannotate { display: flex; }
body[data-mode="apply"] .ribbon-mode-bar.mode-apply { display: flex; }

/* ── AI Pane (right, collapsible) ──────────────────────────────── */
.ai-pane {
  background: var(--c-surface);
  border-left: 1px solid var(--c-border);
  display: flex; flex-direction: column;
  overflow: hidden;
  min-width: 0;
}
.ai-pane-head {
  display: flex; align-items: center; gap: var(--s-3);
  padding: 0 var(--s-5);
  height: var(--pane-head-h);
  border-bottom: 1px solid var(--c-border);
  background: var(--c-surface);
}
.ai-pane-head .ai-title {
  flex: 1; min-width: 0;
  display: flex; flex-direction: column;
  font-size: var(--t-sm);
}
.ai-pane-head .ai-title .name { font-weight: 700; color: var(--c-text); }
.ai-pane-head .ai-title .sub  { font-size: var(--t-xs); color: var(--c-text-soft);
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.ai-pane-head .ai-status-dot {
  width: 8px; height: 8px; border-radius: 50%;
  background: var(--c-text-faint);
  flex-shrink: 0;
}
.ai-pane-head[data-status="online"] .ai-status-dot { background: var(--c-success); animation: ai-pulse 2s ease-in-out infinite; }
.ai-pane-head[data-status="offline"] .ai-status-dot { background: var(--c-danger); }
@keyframes ai-pulse {
  0%, 100% { box-shadow: 0 0 0 0 rgba(16,185,129,0.4); }
  50%      { box-shadow: 0 0 0 4px rgba(16,185,129,0); }
}
.ai-pane-head .ai-collapse {
  background: transparent; border: 0;
  color: var(--c-text-soft);
  width: 24px; height: 24px;
  display: inline-flex; align-items: center; justify-content: center;
  border-radius: var(--r-sm);
  cursor: pointer;
}
.ai-pane-head .ai-collapse:hover { background: var(--c-surface-2); color: var(--c-text); }

.ai-pane-body {
  flex: 1; overflow-y: auto;
  padding: var(--s-5);
  display: flex; flex-direction: column; gap: var(--s-5);
}

.ai-section {
  background: var(--c-surface-2);
  border: 1px solid var(--c-border);
  border-radius: var(--r-md);
  padding: var(--s-4) var(--s-5);
}
.ai-section h4 {
  margin: 0 0 var(--s-3);
  font-size: var(--t-xs); font-weight: 700;
  color: var(--c-text-faint);
  text-transform: uppercase; letter-spacing: 0.06em;
  display: flex; align-items: center; gap: var(--s-2);
}
.ai-section h4 .ico { color: var(--c-text-soft); width: 14px; height: 14px; }

/* AI Proposal block */
.ai-proposal-text {
  font-size: var(--t-base); color: var(--c-text);
  background: var(--c-surface); padding: var(--s-3) var(--s-4);
  border: 1px solid var(--c-border); border-radius: var(--r-sm);
  white-space: pre-wrap; word-break: break-word;
  font-family: var(--f-mono); font-size: var(--t-sm);
  margin-bottom: var(--s-3);
}
.ai-conf-bar {
  height: 6px; border-radius: 3px; background: var(--c-surface-3);
  overflow: hidden; margin-bottom: var(--s-2);
}
.ai-conf-bar > span { display: block; height: 100%; background: var(--c-success); transition: width 0.3s; }
.ai-conf-bar.med > span  { background: var(--c-warn); }
.ai-conf-bar.low > span  { background: var(--c-danger); }
.ai-conf-text {
  display: flex; justify-content: space-between;
  font-size: var(--t-xs); color: var(--c-text-soft);
  margin-bottom: var(--s-3);
}
.ai-rationale {
  font-size: var(--t-xs); color: var(--c-text-muted);
  font-style: italic;
  border-left: 2px solid var(--c-info);
  padding: 2px var(--s-3);
  margin-bottom: var(--s-3);
}

/* ── Phase B5: Patterns triggered (subsection inside Proposal) ──── */
.ai-patterns {
  margin: var(--s-3) 0;
  padding: var(--s-3);
  border: 1px solid var(--c-border);
  border-radius: var(--r-md);
  background: var(--c-bg-soft);
}
.ai-patterns-head {
  display: flex; align-items: center; gap: 6px;
  font-size: var(--t-xs); font-weight: 700;
  color: var(--c-text-soft);
  text-transform: uppercase; letter-spacing: 0.04em;
  margin-bottom: var(--s-2);
}
.ai-patterns-head .ico { color: var(--c-text-soft); }
.ai-patterns-count {
  margin-left: auto;
  padding: 1px 6px;
  font-size: 10px;
  background: var(--c-surface-3);
  border-radius: 999px;
  color: var(--c-text-soft);
}
.ai-pattern-row {
  display: grid;
  grid-template-columns: auto 1fr auto;
  gap: 6px;
  padding: 4px 0;
  font-size: var(--t-xs);
  align-items: baseline;
}
.ai-pattern-row + .ai-pattern-row {
  border-top: 1px dashed var(--c-divider);
}
.ai-pattern-type {
  font-family: var(--f-mono);
  color: var(--c-primary, var(--c-text));
  font-weight: 600;
  font-size: 11px;
}
.ai-pattern-trigger {
  color: var(--c-text);
  font-family: var(--f-mono);
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}
.ai-pattern-meta {
  display: inline-flex; gap: 6px;
  color: var(--c-text-faint);
  font-variant-numeric: tabular-nums;
}
.ai-pattern-conf {
  padding: 0 5px;
  border-radius: 999px;
  background: var(--c-success-soft);
  color: var(--c-success-text);
  font-weight: 600;
}
.ai-actions {
  display: flex; gap: var(--s-2);
}
.ai-actions button {
  flex: 1;
  height: 32px;
  border: 1px solid var(--c-border-strong);
  background: var(--c-surface);
  color: var(--c-text);
  border-radius: var(--r-md);
  font-size: var(--t-sm); font-weight: 600;
  cursor: pointer;
  display: inline-flex; align-items: center; justify-content: center; gap: var(--s-2);
  transition: all var(--d-fast) var(--ease-std);
}
.ai-actions button.ai-accept { background: var(--c-success); color: white; border-color: var(--c-success); }
.ai-actions button.ai-accept:hover { background: var(--c-success-hover); }
.ai-actions button.ai-edit:hover { background: var(--c-surface-2); border-color: var(--c-primary); color: var(--c-primary-text); }
.ai-actions button.ai-reject:hover { background: var(--c-danger-soft); border-color: var(--c-danger); color: var(--c-danger-text); }

/* ── Phase 1: Claude Code live runner ───────────────────────── */
.ai-claude-live { position: relative; }
.ai-claude-actions {
  display: flex; gap: var(--s-2);
  margin-bottom: var(--s-3);
}
.ai-claude-actions button {
  display: inline-flex; align-items: center; gap: 6px;
  padding: 6px 10px;
  border: 1px solid var(--c-border-strong);
  border-radius: var(--r-md);
  background: var(--c-surface);
  color: var(--c-text);
  font-size: var(--t-sm); font-weight: 600;
  cursor: pointer;
  transition: all var(--d-fast) var(--ease-std);
}
.ai-claude-actions button.ai-claude-run {
  background: var(--c-primary, #4f46e5);
  color: white; border-color: var(--c-primary, #4f46e5);
}
.ai-claude-actions button.ai-claude-run:hover:not(:disabled) {
  filter: brightness(1.08);
}
.ai-claude-actions button:disabled {
  opacity: 0.5; cursor: not-allowed;
}
.ai-claude-log {
  max-height: 360px; overflow-y: auto;
  border: 1px solid var(--c-border);
  border-radius: var(--r-md);
  background: var(--c-bg-soft);
  padding: var(--s-2);
  font-family: var(--f-mono, "SF Mono", monospace);
  font-size: 11.5px;
  display: flex; flex-direction: column; gap: 3px;
}
.ai-claude-log:empty::before {
  content: '— กด Run เพื่อให้ Claude วิเคราะห์ row นี้ —';
  display: block;
  padding: var(--s-3);
  color: var(--c-text-faint);
  font-style: italic;
  font-family: var(--f-base);
  font-size: var(--t-xs);
  text-align: center;
}
.ai-claude-log.streaming {
  border-color: var(--c-primary, #4f46e5);
  box-shadow: 0 0 0 2px var(--c-primary-soft, rgba(79,70,229,0.12));
}
.ai-claude-log .aclog {
  padding: 4px 8px;
  border-radius: 4px;
  border-left: 2px solid var(--c-border);
  background: var(--c-surface);
  white-space: pre-wrap; word-break: break-word;
}
.ai-claude-log .aclog.sys      { color: var(--c-text-faint); border-left-color: var(--c-text-faint); font-style: italic; }
.ai-claude-log .aclog.thinking { color: var(--c-text-soft); border-left-color: var(--c-info, #06b6d4); background: rgba(6,182,212,0.04); }
.ai-claude-log .aclog.tool     { color: var(--c-text); border-left-color: var(--c-warn); background: var(--c-warn-soft); font-weight: 500; }
.ai-claude-log .aclog.tool::before { content: '→ '; color: var(--c-warn-text); }
.ai-claude-log .aclog.tool-res { color: var(--c-text-soft); border-left-color: var(--c-success); padding-left: 16px; }
.ai-claude-log .aclog.tool-res::before { content: '← '; color: var(--c-success); }
.ai-claude-log .aclog.tool-err { color: var(--c-danger-text); border-left-color: var(--c-danger); background: var(--c-danger-soft); }
.ai-claude-log .aclog.text     { color: var(--c-text); border-left-color: var(--c-primary, #4f46e5); }
.ai-claude-log .aclog.err      { color: var(--c-danger-text); border-left-color: var(--c-danger); background: var(--c-danger-soft); font-weight: 600; }
.ai-claude-log .aclog.result {
  margin-top: var(--s-2);
  padding: var(--s-3);
  border: 1px solid var(--c-success);
  border-left-width: 4px;
  background: var(--c-success-soft);
  border-radius: var(--r-md);
  font-family: var(--f-base);
  font-size: var(--t-sm);
}
.ai-claude-log .aclog-result-head {
  display: flex; align-items: center; gap: 6px;
  margin-bottom: var(--s-2);
  color: var(--c-success-text);
}
.ai-claude-log .aclog-result-head .ac-meta {
  margin-left: auto;
  font-size: 11px; color: var(--c-text-faint);
  font-family: var(--f-mono);
}
.ai-claude-log .aclog-result-text {
  font-family: var(--f-mono);
  padding: var(--s-2) var(--s-3);
  background: var(--c-surface);
  border-radius: var(--r-sm);
  margin-bottom: var(--s-2);
  word-break: break-all;
}
.ai-claude-log .aclog-result-conf {
  font-size: 11px; color: var(--c-text-soft);
  margin-bottom: var(--s-2);
}
.ai-claude-log .aclog-result-rat {
  font-size: var(--t-xs);
  font-style: italic;
  color: var(--c-text-muted);
  margin-bottom: var(--s-2);
}
.ai-claude-log .aclog-result-actions {
  display: flex; gap: var(--s-2);
}
.ai-claude-log .aclog-result-actions button {
  flex: 1;
  display: inline-flex; align-items: center; justify-content: center; gap: 4px;
  padding: 5px 10px;
  border: 1px solid var(--c-border-strong);
  border-radius: var(--r-sm);
  background: var(--c-surface);
  color: var(--c-text);
  font-size: var(--t-xs); font-weight: 600;
  cursor: pointer;
}
.ai-claude-log .aclog-result-actions button.ai-accept {
  background: var(--c-success); color: white; border-color: var(--c-success);
}
.ai-claude-log .aclog-result-actions button.ai-reject:hover {
  background: var(--c-danger-soft); border-color: var(--c-danger); color: var(--c-danger-text);
}

/* Teach-back textarea */
.ai-teach textarea {
  width: 100%; min-height: 64px; resize: vertical;
  font: inherit; font-size: var(--t-sm);
  padding: var(--s-3) var(--s-4);
  border: 1px solid var(--c-border-strong);
  border-radius: var(--r-md);
  background: var(--c-surface);
  margin-bottom: var(--s-3);
}
.ai-teach .ai-tags {
  display: flex; flex-wrap: wrap; gap: 4px;
  margin-bottom: var(--s-3);
}
.ai-teach .ai-tag {
  font-size: var(--t-xs);
  padding: 2px var(--s-3);
  border-radius: var(--r-pill);
  background: var(--c-surface);
  border: 1px solid var(--c-border);
  color: var(--c-text-muted);
  cursor: pointer;
  font-family: var(--f-mono);
  user-select: none;
  transition: all var(--d-fast) var(--ease-std);
}
.ai-teach .ai-tag:hover { background: var(--c-info-soft); border-color: var(--c-info); color: var(--c-info-text); }
.ai-teach .ai-tag.active { background: var(--c-info); color: white; border-color: var(--c-info); }
.ai-teach button.ai-send {
  height: 28px; padding: 0 var(--s-5);
  background: var(--c-primary); color: white; border: 0;
  border-radius: var(--r-md);
  font-size: var(--t-sm); font-weight: 600;
  cursor: pointer;
}
.ai-teach button.ai-send:hover { background: var(--c-primary-hover); }

.ai-empty {
  text-align: center;
  padding: var(--s-7) var(--s-4);
  color: var(--c-text-faint);
  font-size: var(--t-sm);
}

/* ── Status Bar (bottom) ──────────────────────────────────────── */
.status-bar {
  display: flex; align-items: center; gap: var(--s-5);
  padding: 0 var(--s-5);
  height: var(--statusbar-h);
  background: var(--c-surface);
  border-top: 1px solid var(--c-border);
  font-size: var(--t-xs);
  color: var(--c-text-muted);
}
.status-bar .sb-section {
  display: inline-flex; align-items: center; gap: var(--s-2);
  white-space: nowrap;
}
.status-bar .sb-spacer { flex: 1; }
.status-bar .sb-sep { width: 1px; height: 16px; background: var(--c-border); }
.status-bar strong { color: var(--c-text); font-weight: 700; }
.status-bar .sb-pill {
  display: inline-flex; align-items: center; gap: 4px;
  padding: 2px var(--s-3);
  border-radius: var(--r-pill);
  font-size: var(--t-xs); font-weight: 600;
}
.status-bar .sb-pill.pass     { background: var(--c-success-soft); color: var(--c-success-text); }
.status-bar .sb-pill.fail     { background: var(--c-danger-soft); color: var(--c-danger-text); }
.status-bar .sb-pill.need_fix { background: var(--c-warn-soft); color: var(--c-warn-text); }
.status-bar .sb-pill.skip     { background: var(--c-surface-3); color: var(--c-text-soft); }
.status-bar .sb-progress {
  display: inline-flex; align-items: center; gap: var(--s-3);
  font-variant-numeric: tabular-nums;
}
.status-bar .sb-progress .bar {
  width: 80px; height: 4px; border-radius: 2px;
  background: var(--c-surface-3); overflow: hidden;
}
.status-bar .sb-progress .bar > span {
  display: block; height: 100%;
  background: var(--c-primary);
  transition: width 0.3s var(--ease-out);
}
/* Phase A6: verdict pills moved from action bar to status bar */
.status-bar .sb-row-info {
  max-width: 38ch; overflow: hidden; text-overflow: ellipsis;
}
.status-bar .sb-verdict {
  display: inline-flex; align-items: center; gap: 2px;
  padding: 3px;
  background: var(--c-bg-soft);
  border: 1px solid var(--c-border);
  border-radius: var(--r-pill);
}
.status-bar .sb-vbtn {
  display: inline-flex; align-items: center; gap: 4px;
  padding: 3px 8px;
  height: 24px;
  border: 0; background: transparent;
  border-radius: 999px;
  color: var(--c-text-soft);
  font-size: var(--t-xs); font-weight: 600;
  cursor: pointer;
  white-space: nowrap;
  transition: background 120ms ease, color 120ms ease;
}
.status-bar .sb-vbtn:hover { background: var(--c-surface-3); color: var(--c-text); }
.status-bar .sb-vbtn .kbd {
  margin-left: 2px;
  padding: 0 4px;
  font-size: 10px;
  background: var(--c-surface-3);
  color: var(--c-text-soft);
  border-radius: 3px;
  font-weight: 500;
}
.status-bar .sb-vbtn[aria-checked="true"].pass     { background: var(--c-success-soft); color: var(--c-success-text); }
.status-bar .sb-vbtn[aria-checked="true"].fail     { background: var(--c-danger-soft);  color: var(--c-danger-text); }
.status-bar .sb-vbtn[aria-checked="true"].fix      { background: var(--c-warn-soft);    color: var(--c-warn-text); }
.status-bar .sb-vbtn[aria-checked="true"].skip     { background: var(--c-surface-3);    color: var(--c-text); }
.status-bar .sb-vbtn[aria-checked="true"] .kbd     { background: rgba(0,0,0,0.10); color: inherit; }
.status-bar .sb-vbtn.reset { padding: 3px 6px; color: var(--c-text-faint); }
.status-bar .sb-vbtn.reset:hover { color: var(--c-text-soft); }
/* Disable verdict pills when no row selected */
body[data-row-selected="0"] .status-bar .sb-vbtn { opacity: 0.4; pointer-events: none; }

.status-bar .sb-claude {
  display: inline-flex; align-items: center; gap: var(--s-2);
}
.status-bar .sb-claude .dot {
  width: 6px; height: 6px; border-radius: 50%;
  background: var(--c-text-faint);
}
.status-bar .sb-claude.online .dot { background: var(--c-success); }
.status-bar .sb-claude.offline .dot { background: var(--c-text-faint); }

.topbar {
  grid-area: topbar;
  display: flex; align-items: center; gap: var(--s-6);
  padding: 0 var(--s-7);
  background: var(--c-surface);
  border-bottom: 1px solid var(--c-border);
  z-index: 10;
}
.topbar-brand {
  display: flex; align-items: center; gap: var(--s-4);
  font-weight: 700; font-size: var(--t-lg); letter-spacing: -0.01em;
  color: var(--c-text);
}
.topbar-brand .logo {
  width: 24px; height: 24px; border-radius: var(--r-md);
  background: linear-gradient(135deg, var(--c-primary), var(--c-purple));
  display: inline-flex; align-items: center; justify-content: center;
  color: white; font-weight: 800; font-size: 12px;
  box-shadow: var(--e-1);
}
.topbar-brand .sub { font-weight: 400; color: var(--c-text-soft); font-size: var(--t-base); }
.topbar-spacer { flex: 1; }
.topbar-actions { display: flex; gap: var(--s-2); align-items: center; }
.topbar-actions .stats-pill {
  display: inline-flex; align-items: center; gap: var(--s-3);
  padding: 5px var(--s-5);
  background: var(--c-surface-2); border: 1px solid var(--c-border);
  border-radius: var(--r-pill);
  font-size: var(--t-sm); color: var(--c-text-muted);
  font-variant-numeric: tabular-nums;
}
.topbar-actions .stats-pill strong { color: var(--c-text); font-weight: 700; }
.topbar-actions .stats-pill .sep { color: var(--c-text-faint); margin: 0 2px; }

/* Topbar — current row context chip (breadcrumb-style) */
.topbar-context {
  display: inline-flex; align-items: center; gap: var(--s-3);
  height: var(--btn-h);
  padding: 0 var(--s-5);
  margin-left: var(--s-5);
  background: var(--c-surface-2);
  border: 1px solid var(--c-border);
  border-radius: var(--r-md);
  color: var(--c-text);
  font-size: var(--t-sm);
  cursor: pointer;
  transition: all var(--d-fast) var(--ease-std);
  max-width: 380px;
  overflow: hidden;
  white-space: nowrap;
}
.topbar-context:hover {
  background: var(--c-surface);
  border-color: var(--c-border-strong);
  transform: translateY(-1px);
  box-shadow: var(--e-1);
}
.topbar-context-label {
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
  font-variant-numeric: tabular-nums;
}
.topbar-context-label .row-num {
  font-family: var(--f-mono); font-weight: 700;
  background: var(--c-primary-soft); color: var(--c-primary-text);
  padding: 1px 6px; border-radius: var(--r-sm);
  margin-right: var(--s-3);
}
.topbar-context-label .sep {
  color: var(--c-text-faint); margin: 0 var(--s-2);
}
.topbar-context-label .ctx-section {
  color: var(--c-text-muted); font-weight: 500;
}
.topbar-context-label .ctx-meta {
  color: var(--c-text-soft);
  font-style: italic;
}

/* Topbar buttons (icon + label, ghost-default) */
.topbar-btn {
  display: inline-flex; align-items: center; gap: var(--s-3);
  padding: 0 var(--s-5);
  height: var(--btn-h);
  background: transparent;
  border: 1px solid transparent;
  color: var(--c-text-muted);
  border-radius: var(--r-md);
  font-size: var(--t-sm); font-weight: 500;
  letter-spacing: 0;
  transition: all var(--d-fast) var(--ease-std);
  cursor: pointer;
}
.topbar-btn:hover {
  background: var(--c-surface-2); color: var(--c-text);
  border-color: var(--c-border);
}
.topbar-btn:active { transform: translateY(1px); }
.topbar-btn.icon-only { padding: 0; width: var(--btn-h); justify-content: center; }
.topbar-btn .ico { width: 16px; height: 16px; }

.pane {
  overflow: hidden;
  display: flex; flex-direction: column;
  background: var(--c-surface);
  border-right: 1px solid var(--c-border);
  min-width: 0; min-height: 0;
}
.pane:last-child { border-right: none; }
.tree-pane   { grid-area: tree; }
.center-pane { grid-area: center; }
.pdf-pane    { grid-area: pdf; }

/* Pane headers */
.pane > h2 {
  margin: 0;
  padding: 0 var(--s-7);
  height: var(--pane-head-h);
  flex-shrink: 0;
  font-size: var(--t-sm); font-weight: 600;
  background: var(--c-surface);
  border-bottom: 1px solid var(--c-border);
  color: var(--c-text);
  display: flex; justify-content: space-between; align-items: center;
  letter-spacing: 0.02em;
  gap: var(--s-4);
}
.pane > h2 .pane-title { display: inline-flex; align-items: center; gap: var(--s-3); }
.pane > h2 .pane-title .icon { font-size: 14px; }

/* ── Buttons (canonical) ────────────────────────────────────────── */
.btn {
  display: inline-flex; align-items: center; justify-content: center;
  gap: var(--s-3);
  padding: 6px 12px;
  border: 1px solid var(--c-border-strong);
  background: var(--c-surface);
  color: var(--c-text);
  border-radius: var(--r-md);
  font-size: var(--t-base);
  font-weight: 500;
  line-height: 1.4;
  transition: background var(--d-fast) var(--ease-std),
              border-color var(--d-fast) var(--ease-std),
              color var(--d-fast) var(--ease-std),
              transform var(--d-fast) var(--ease-out),
              box-shadow var(--d-fast) var(--ease-std);
  white-space: nowrap;
}
.btn:hover { background: var(--c-surface-2); border-color: var(--c-text-faint); }
.btn:active { transform: translateY(1px); }
.btn:disabled, .btn.is-disabled {
  opacity: 0.5; pointer-events: none;
}
.btn.btn-primary {
  background: var(--c-primary); color: var(--c-text-invert); border-color: var(--c-primary);
}
.btn.btn-primary:hover { background: var(--c-primary-hover); border-color: var(--c-primary-hover); }
.btn.btn-success, .btn.save-btn {
  background: var(--c-success); color: white; border-color: var(--c-success);
  font-weight: 600;
}
.btn.btn-success:hover, .btn.save-btn:hover:not(:disabled) {
  background: var(--c-success-hover); border-color: var(--c-success-hover);
}
.btn.save-btn:disabled { background: var(--c-border-strong); color: var(--c-text-soft); border-color: var(--c-border-strong); }
.btn.btn-ghost { background: transparent; border-color: transparent; color: var(--c-text-muted); }
.btn.btn-ghost:hover { background: var(--c-surface-2); color: var(--c-text); }
.btn.btn-danger { color: var(--c-danger-text); border-color: var(--c-danger); background: var(--c-surface); }
.btn.btn-danger:hover { background: var(--c-danger-soft); }

/* ── Unified toolbar system ──────────────────────────────────────
   Used by canvas-toolbar (TOR/PDF/xlsx) and edit-toolbar.
   Visual rules:
     • height: var(--toolbar-h)        (40px — consistent across panes)
     • button: 28×28 icon-only OR 28×auto with label, padded
     • groups: separated by .tb-sep    (1px vertical divider)
     • info chips:  .info               (text in muted color)
     • flex distribution: groups → spacer → groups
*/
.canvas-toolbar {
  padding: 0 var(--s-6);
  height: var(--toolbar-h);
  background: var(--c-surface);
  border-bottom: 1px solid var(--c-border);
  display: flex; gap: var(--s-2); align-items: center;
  font-size: var(--t-sm); flex-wrap: nowrap;
  flex-shrink: 0;
  overflow-x: auto;
  scrollbar-width: none;
}
.canvas-toolbar::-webkit-scrollbar { display: none; }
.canvas-toolbar button,
.canvas-toolbar .tb-btn {
  display: inline-flex; align-items: center; justify-content: center; gap: var(--s-2);
  min-width: var(--btn-h-toolbar); height: var(--btn-h-toolbar);
  padding: 0 var(--s-4);
  border: 1px solid transparent;
  background: transparent;
  color: var(--c-text-muted);
  border-radius: var(--r-md);
  font-size: var(--t-sm); font-weight: 500;
  cursor: pointer;
  transition: background var(--d-fast) var(--ease-std),
              color var(--d-fast) var(--ease-std),
              border-color var(--d-fast) var(--ease-std);
  white-space: nowrap;
  flex-shrink: 0;
}
.canvas-toolbar button:hover,
.canvas-toolbar .tb-btn:hover {
  background: var(--c-surface-2);
  color: var(--c-text);
}
.canvas-toolbar button:active,
.canvas-toolbar .tb-btn:active { transform: translateY(1px); }
.canvas-toolbar button:disabled { opacity: 0.4; cursor: not-allowed; }
.canvas-toolbar .info {
  color: var(--c-text-soft); font-size: var(--t-sm);
  font-variant-numeric: tabular-nums;
  display: inline-flex; align-items: center; gap: var(--s-2);
  flex-shrink: 0;
}
.canvas-toolbar label {
  display: inline-flex; align-items: center; gap: 4px;
  font-size: var(--t-sm); color: var(--c-text-muted);
  cursor: pointer; user-select: none;
  flex-shrink: 0;
}
/* Group + separator — visual chunking of related actions */
.canvas-toolbar .tb-group {
  display: inline-flex; gap: 2px; align-items: center;
  flex-shrink: 0;
}
.canvas-toolbar .tb-sep {
  width: 1px; height: 18px;
  background: var(--c-border);
  margin: 0 var(--s-3);
  flex-shrink: 0;
}
.canvas-toolbar .tb-spacer { flex: 1; }
/* Page indicator chip (e.g. "3 / 12") */
.canvas-toolbar .page-ind {
  display: inline-flex; align-items: center;
  height: var(--btn-h-toolbar);
  padding: 0 var(--s-4);
  background: var(--c-surface-2);
  border: 1px solid var(--c-border);
  border-radius: var(--r-md);
  font-family: var(--f-mono); font-size: var(--t-sm);
  color: var(--c-text);
  min-width: 64px; justify-content: center;
}

/* ── Inputs ─────────────────────────────────────────────────────── */
input[type="text"], input[type="search"], select, textarea {
  font-family: inherit; font-size: var(--t-base);
  padding: 6px var(--s-5);
  border: 1px solid var(--c-border-strong);
  background: var(--c-surface);
  color: var(--c-text);
  border-radius: var(--r-md);
  transition: border-color var(--d-fast) var(--ease-std), box-shadow var(--d-fast) var(--ease-std);
  width: 100%;
}
input[type="checkbox"] {
  accent-color: var(--c-primary);
  cursor: pointer;
}
input::placeholder, textarea::placeholder { color: var(--c-text-faint); }
select { cursor: pointer; }

/* ── kbd ────────────────────────────────────────────────────────── */
.kbd {
  display: inline-block;
  padding: 0 5px; min-width: 14px;
  font-family: var(--f-mono); font-size: var(--t-xs);
  background: var(--c-surface-3); color: var(--c-text-muted);
  border: 1px solid var(--c-border);
  border-bottom-width: 2px;
  border-radius: var(--r-sm);
  line-height: 16px; text-align: center;
}

/* ── Mobile-tab nav ─────────────────────────────────────────────── */
.mobile-tabs {
  display: none;
  background: var(--c-surface); border-bottom: 1px solid var(--c-border);
  padding: 4px;
  grid-area: topbar;
}
.mobile-tabs button {
  flex: 1; padding: 8px 4px; font-size: var(--t-base); font-weight: 600;
  border: 0; background: transparent; cursor: pointer;
  color: var(--c-text-soft);
  border-radius: var(--r-md);
  transition: background var(--d-fast) var(--ease-std), color var(--d-fast) var(--ease-std);
}
.mobile-tabs button:hover { background: var(--c-surface-2); color: var(--c-text); }
.mobile-tabs button.active {
  background: var(--c-primary-soft); color: var(--c-primary-text);
}

/* ── Stats bar (top of tree) ────────────────────────────────────── */
.stats-bar {
  padding: var(--s-3) var(--s-7);
  background: var(--c-surface-2);
  font-size: var(--t-sm);
  color: var(--c-text-muted);
  display: flex; gap: var(--s-6);
  border-bottom: 1px solid var(--c-border);
  flex-wrap: wrap;
}
.stats-bar strong { color: var(--c-text); font-weight: 700; }

/* ── Tree pane ──────────────────────────────────────────────────── */
.tree-pane .toolbar {
  padding: var(--s-3) var(--s-5);
  border-bottom: 1px solid var(--c-border);
  background: var(--c-surface);
  display: flex; flex-direction: column; gap: var(--s-3);
}
.tree-pane .toolbar input,
.tree-pane .toolbar select {
  font-size: var(--t-base);
  padding: 6px var(--s-4);
  height: var(--btn-h-sm);
}
.tree-pane .toolbar input[type="text"]::-webkit-search-cancel-button { cursor: pointer; }

/* Search input with prefix icon */
.search-wrap {
  position: relative;
}
.search-wrap .search-ico {
  position: absolute; left: var(--s-4); top: 50%;
  transform: translateY(-50%);
  color: var(--c-text-faint);
  pointer-events: none;
}
.search-wrap input {
  padding-left: 28px !important;
}
.search-wrap input:focus + .search-ico,
.search-wrap input:focus ~ .search-ico { color: var(--c-primary); }

.filter-row { display: flex; gap: var(--s-3); align-items: center; }
.filter-row select { font-size: var(--t-sm); padding: 4px var(--s-3); flex: 1; min-width: 0; }
.filter-row label {
  display: inline-flex; align-items: center; gap: 4px;
  font-size: var(--t-sm); color: var(--c-text-muted); flex: 1; cursor: pointer;
  user-select: none;
}
.filter-row .mini-btn {
  font-size: var(--t-sm); padding: 3px var(--s-4);
  background: var(--c-surface); border: 1px solid var(--c-border);
  border-radius: var(--r-sm); color: var(--c-text-muted);
  transition: background var(--d-fast) var(--ease-std), color var(--d-fast) var(--ease-std);
}
.filter-row .mini-btn:hover { background: var(--c-surface-2); color: var(--c-text); }

.tree-scroll { flex: 1; overflow: auto; padding: var(--s-3) 0; font-size: var(--t-base); }
.tree-node { user-select: none; }
.tree-row {
  display: flex; align-items: center;
  padding: 0 var(--s-3) 0 0;
  cursor: pointer; gap: var(--s-2);
  line-height: 1.4; min-height: var(--tree-row-h);
  position: relative;
  transition: background var(--d-fast) var(--ease-std);
}
/* Status indicator strip (left edge, 4px wide) — replaces emoji icon */
.tree-row::before {
  content: ''; position: absolute;
  left: 0; top: 4px; bottom: 4px;
  width: 3px; border-radius: 0 2px 2px 0;
  background: transparent;
  transition: background var(--d-fast) var(--ease-std);
}
.tree-row[data-status="pass"]::before     { background: var(--c-success); }
.tree-row[data-status="fail"]::before     { background: var(--c-danger); }
.tree-row[data-status="need_fix"]::before { background: var(--c-warn); }
.tree-row[data-status="skip"]::before     { background: var(--c-text-faint); }
.tree-row:hover { background: var(--c-surface-2); }
.tree-row.selected {
  background: var(--c-primary-soft);
  box-shadow: inset 3px 0 0 var(--c-primary);
}
.tree-row.selected::before { background: var(--c-primary) !important; }
.tree-row.selected .tree-label { color: var(--c-primary-text); font-weight: 600; }
.tree-chev {
  width: 16px; flex-shrink: 0; text-align: center; color: var(--c-text-faint);
  font-size: 9px;
  transition: transform var(--d-fast) var(--ease-out);
}
.tree-chev.empty { visibility: hidden; }
.tree-node.expanded > .tree-row .tree-chev:not(.empty) { transform: rotate(0deg); }
.tree-icon { width: 16px; flex-shrink: 0; font-size: 12px; text-align: center; color: var(--c-text-faint); }
.tree-icon.section { color: var(--c-primary); }
.tree-icon.item    { color: var(--c-warn); }
.tree-icon.sub     { color: var(--c-success); }
.tree-label {
  flex: 1; min-width: 0;
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
  font-size: var(--t-base); color: var(--c-text);
}
.tree-status { display: none; }   /* Replaced by left-edge strip via [data-status] */
.tree-flags  { font-size: var(--t-xs); color: var(--c-warn-text); }
.tree-children { display: none; }
.tree-node.expanded > .tree-children { display: block; }
.tree-row.match { background: var(--c-warn-soft); }
.tree-row.path-match .tree-label { font-weight: 600; }

/* Confidence dots */
.tree-row .conf-dot {
  width: 8px; height: 8px; border-radius: 50%;
  display: inline-block; flex-shrink: 0; margin-left: 2px;
  box-shadow: 0 0 0 2px var(--c-surface);
}
.conf-dot.high { background: var(--c-success); }
.conf-dot.med  { background: var(--c-warn); }
.conf-dot.low  { background: var(--c-danger); }
.conf-dot.none { background: var(--c-border-strong); }

/* ── Center pane (TOR / xlsx) ───────────────────────────────────── */
.center-pane { display: flex; flex-direction: column; }
.split-top, .split-bot { display: flex; flex-direction: column; min-height: 0; overflow: hidden; }
.split-top { flex: 1 1 50%; }
.split-bot { flex: 1 1 50%; border-top: 1px solid var(--c-border); position: relative; }
.split-bot::before {
  content: ''; position: absolute; top: -2px; left: 0; right: 0; height: 4px;
  cursor: ns-resize; z-index: 2;
  background: transparent;
  transition: background var(--d-base) var(--ease-std);
}
.split-bot:hover::before { background: var(--c-primary-soft); }

.tor-canvas, .pdf-canvas {
  flex: 1; overflow: auto;
  background: var(--c-canvas-bg);
  padding: var(--s-7);
  text-align: center;
  scroll-behavior: smooth;
}
.tor-canvas img, .pdf-canvas img {
  max-width: 100%;
  box-shadow: 0 4px 16px rgba(0,0,0,0.45);
  background: white;
  border-radius: 2px;
}
.empty-canvas {
  color: var(--c-text-faint);
  padding: var(--s-12) var(--s-7);
  text-align: center;
  font-size: var(--t-md);
  display: flex; flex-direction: column; align-items: center; gap: var(--s-4);
}
.empty-canvas::before {
  content: '◌';
  font-size: 32px; opacity: 0.5;
}
.empty-canvas.loading::before {
  content: '';
  width: 28px; height: 28px;
  border: 3px solid var(--c-border);
  border-top-color: var(--c-primary);
  border-radius: 50%;
  animation: spin 0.8s linear infinite;
}

/* ── XLSX preview ───────────────────────────────────────────────── */
.xlsx-wrap { flex: 1; overflow: auto; background: var(--c-bg); }
.xlsx-table { width: 100%; border-collapse: separate; border-spacing: 0; font-size: var(--t-sm); }
.xlsx-table thead th {
  position: sticky; top: 0; z-index: 1;
  background: var(--c-surface-3);
  color: var(--c-text);
  padding: var(--s-3) var(--s-5);
  border-bottom: 1px solid var(--c-border-strong);
  border-right: 1px solid var(--c-border);
  font-size: var(--t-xs);
  font-weight: 700;
  text-align: left; white-space: nowrap;
  letter-spacing: 0.04em; text-transform: uppercase;
  color: var(--c-text-muted);
}
.xlsx-table thead th:last-child { border-right: none; }
.xlsx-table td {
  border-bottom: 1px solid var(--c-divider);
  border-right: 1px solid var(--c-divider);
  padding: var(--s-4) var(--s-5);
  vertical-align: top; white-space: pre-wrap; word-break: break-word;
  background: var(--c-surface);
  transition: background var(--d-fast) var(--ease-std);
}
.xlsx-table td:last-child { border-right: none; }
.xlsx-table .col-A { width: 56px; font-family: var(--f-mono); color: var(--c-text-soft); }
.xlsx-table .col-B { width: 28%; }
.xlsx-table .col-C { width: 28%; }
.xlsx-table .col-D { width: 24%; }
.xlsx-table .col-E { width: 80px; font-size: var(--t-xs); }
.xlsx-table .col-F { width: 110px; font-size: var(--t-xs); color: var(--c-text-soft); }
.xlsx-table .row-num {
  width: 42px; font-family: var(--f-mono); font-size: var(--t-xs);
  color: var(--c-text-faint); background: var(--c-surface-2);
}
.xlsx-table tr:hover td { background: var(--c-surface-2); }
.xlsx-table tr.target td {
  background: var(--c-warn-soft) !important;
  font-weight: 500;
}
.xlsx-table tr.target .col-A,
.xlsx-table tr.target .row-num { color: var(--c-text); font-weight: 700; }
.xlsx-table tr.target:hover td { background: #fde68a !important; }

/* Col D — right-click opens menu, double-click edits inline.
   Subtle right-edge hint stripe so users know it's interactive. */
.xlsx-table td.col-D.editable {
  cursor: context-menu; position: relative;
}
.xlsx-table td.col-D.editable::after {
  content: ''; position: absolute;
  right: 0; top: 4px; bottom: 4px; width: 2px;
  background: var(--c-purple);
  opacity: 0;
  border-radius: 2px;
  transition: opacity var(--d-fast) var(--ease-std);
}
.xlsx-table td.col-D.editable:hover { background: var(--c-surface-2); }
.xlsx-table td.col-D.editable:hover::after { opacity: 0.4; }
.xlsx-table td.col-D.commitment .d-text { color: var(--c-text-soft); font-style: italic; }
.xlsx-table td.col-D.commitment::before {
  content: '⚠ '; color: var(--c-warn); font-style: normal;
}
.xlsx-table td.col-D.editing {
  background: var(--c-purple-soft) !important;
  outline: 2px solid var(--c-purple);
  outline-offset: -2px;
}
.xlsx-table td.col-D.editing::after { display: none; }

/* Vendor tags */
.vendor-tag {
  display: inline-flex; align-items: center;
  padding: 1px var(--s-4);
  border-radius: var(--r-sm);
  font-size: var(--t-xs); font-weight: 700;
  letter-spacing: 0.04em;
}
.vendor-tag.SMART { background: var(--c-primary-soft); color: var(--c-primary-text); }
.vendor-tag.TRIO  { background: var(--c-success-soft); color: var(--c-success-text); }
.vendor-tag.SR    { background: var(--c-pink-soft); color: var(--c-pink-text); }

/* ── Col D dropdown menu ────────────────────────────────────────── */
.col-d-menu {
  position: fixed; z-index: 250;
  background: var(--c-surface);
  border-radius: var(--r-lg);
  box-shadow: var(--e-4);
  border: 1px solid var(--c-border);
  padding: 5px; min-width: 260px;
  font-size: var(--t-base);
  animation: menu-pop var(--d-fast) var(--ease-out);
}
@keyframes menu-pop {
  from { opacity: 0; transform: scale(0.96) translateY(-4px); }
  to   { opacity: 1; transform: scale(1) translateY(0); }
}
.col-d-menu .menu-header {
  padding: var(--s-4) var(--s-5) var(--s-3);
  font-size: var(--t-xs);
  color: var(--c-text-faint);
  text-transform: uppercase; letter-spacing: 0.06em; font-weight: 700;
  border-bottom: 1px solid var(--c-divider);
  margin-bottom: var(--s-2);
}
.col-d-menu button {
  display: flex; align-items: center; gap: var(--s-4);
  width: 100%; padding: 7px var(--s-5);
  border: 0; background: transparent;
  border-radius: var(--r-sm);
  text-align: left;
  font-size: var(--t-base);
  color: var(--c-text);
  transition: background var(--d-fast) var(--ease-std);
}
.col-d-menu button:hover { background: var(--c-surface-2); }
.col-d-menu button .icon { font-size: 14px; width: 18px; text-align: center; }
.col-d-menu button .label { flex: 1; }

/* ── Phase B3: Col D autocomplete panel ────────────────────────── */
.col-d-ac-panel {
  position: fixed; z-index: 260;
  background: var(--c-surface);
  border: 1px solid var(--c-border);
  border-radius: var(--r-md);
  box-shadow: var(--e-3);
  padding: 4px;
  min-width: 320px; max-width: 720px;
  max-height: 280px; overflow-y: auto;
  font-size: var(--t-sm);
}
.col-d-ac-panel .ac-item {
  display: flex; align-items: center; gap: 8px;
  padding: 6px 8px;
  border-radius: var(--r-sm);
  cursor: pointer;
  white-space: nowrap;
}
.col-d-ac-panel .ac-item.active { background: var(--c-primary-soft); }
.col-d-ac-panel .ac-item:hover  { background: var(--c-surface-2); }
.col-d-ac-panel .ac-item.active:hover { background: var(--c-primary-soft); }
.col-d-ac-panel .ac-kind {
  display: inline-block;
  padding: 1px 6px;
  font-size: 10px; font-weight: 700;
  border-radius: var(--r-sm);
  background: var(--c-bg-soft); color: var(--c-text-soft);
  flex-shrink: 0;
}
.col-d-ac-panel .ac-kind.ai       { background: var(--c-primary-soft); color: var(--c-primary-text, var(--c-primary)); }
.col-d-ac-panel .ac-kind.neighbor { background: var(--c-success-soft); color: var(--c-success-text); }
.col-d-ac-panel .ac-kind.shape    { background: var(--c-bg-soft); color: var(--c-text-faint); }
.col-d-ac-panel .ac-text {
  flex: 1; min-width: 0;
  overflow: hidden; text-overflow: ellipsis;
  color: var(--c-text);
}
.col-d-ac-panel .ac-meta {
  flex-shrink: 0;
  font-size: var(--t-xs); color: var(--c-text-faint);
  display: inline-flex; align-items: center; gap: 6px;
}
.col-d-ac-panel .ac-conf {
  padding: 1px 5px;
  border-radius: 999px;
  background: var(--c-surface-3); color: var(--c-text-soft);
  font-variant-numeric: tabular-nums;
}
.col-d-menu button .hint { font-size: var(--t-xs); color: var(--c-text-faint); }
.col-d-menu button.danger { color: var(--c-danger-text); }
.col-d-menu button.danger:hover { background: var(--c-danger-soft); }
.col-d-menu button.primary { color: var(--c-success-text); font-weight: 600; }
.col-d-menu button.primary:hover { background: var(--c-success-soft); }
.col-d-menu .sep { height: 1px; background: var(--c-divider); margin: var(--s-2) 0; }

/* ── Phase 2: Catalog browser ────────────────────────────────────── */
#catalog-list .cat-item {
  padding: 8px 10px;
  border-radius: var(--r-sm);
  cursor: pointer;
  margin-bottom: 2px;
  border: 1px solid transparent;
  font-size: var(--t-sm);
  transition: background 100ms, border-color 100ms;
}
#catalog-list .cat-item:hover { background: var(--c-surface-2); border-color: var(--c-border); }
#catalog-list .cat-item.active {
  background: var(--c-primary-soft);
  border-color: var(--c-primary);
}
#catalog-list .cat-item .cat-head {
  display: flex; align-items: center; gap: 6px; margin-bottom: 2px;
}
#catalog-list .cat-item .cat-section {
  font-family: var(--f-mono); font-size: 11px;
  background: var(--c-bg-soft); color: var(--c-text-soft);
  padding: 1px 5px; border-radius: 3px; flex-shrink: 0;
}
#catalog-list .cat-item .cat-brand { font-weight: 600; color: var(--c-text); }
#catalog-list .cat-item .cat-model { color: var(--c-text-soft); flex: 1; min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
#catalog-list .cat-item .cat-pages { font-size: 11px; color: var(--c-text-faint); }
#catalog-list .cat-item .cat-rel {
  font-size: 11px; color: var(--c-text-faint);
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}

#catalog-detail h3 { margin: 0 0 var(--s-3); display: flex; align-items: center; gap: 8px; }
#catalog-detail h3 .cat-section-pill {
  font-family: var(--f-mono); font-size: 12px;
  background: var(--c-primary-soft); color: var(--c-primary);
  padding: 2px 8px; border-radius: 999px; font-weight: 600;
}
#catalog-detail .meta-grid {
  display: grid; grid-template-columns: 110px 1fr; gap: 6px 12px;
  margin-bottom: var(--s-4);
  font-size: var(--t-sm);
}
#catalog-detail .meta-grid label {
  color: var(--c-text-soft); font-weight: 600;
  align-self: center;
}
#catalog-detail .meta-grid input,
#catalog-detail .meta-grid textarea,
#catalog-detail .meta-grid select {
  font-size: var(--t-sm);
  padding: 4px 8px;
  border: 1px solid var(--c-border);
  border-radius: var(--r-sm);
  background: var(--c-surface);
  font-family: inherit;
}
#catalog-detail .meta-grid textarea { min-height: 50px; resize: vertical; }
#catalog-detail .detail-actions {
  display: flex; gap: var(--s-2); margin-top: var(--s-3); flex-wrap: wrap;
}
#catalog-detail .detail-actions button {
  padding: 6px 12px; font-size: var(--t-sm); font-weight: 600;
  border-radius: var(--r-sm); cursor: pointer;
  border: 1px solid var(--c-border-strong); background: var(--c-surface);
  display: inline-flex; align-items: center; gap: 4px;
}
#catalog-detail .detail-actions button.primary {
  background: var(--c-primary, #4f46e5); color: white; border-color: var(--c-primary, #4f46e5);
}
#catalog-detail .detail-actions button.primary:hover { filter: brightness(1.08); }
#catalog-detail .detail-actions button.success {
  background: var(--c-success); color: white; border-color: var(--c-success);
}
#catalog-detail .detail-section {
  margin-top: var(--s-4);
  padding-top: var(--s-4);
  border-top: 1px solid var(--c-divider);
}
#catalog-detail .detail-section h4 {
  font-size: var(--t-xs); text-transform: uppercase; letter-spacing: 0.05em;
  color: var(--c-text-soft); margin: 0 0 var(--s-2);
}
#catalog-detail .annot-row {
  display: grid; grid-template-columns: 50px 70px 1fr auto;
  gap: 6px; padding: 4px 6px;
  font-size: var(--t-xs); font-family: var(--f-mono);
  border-radius: var(--r-sm);
  align-items: center;
}
#catalog-detail .annot-row:hover { background: var(--c-surface-2); }
#catalog-detail .annot-row .pg { color: var(--c-text-soft); }
#catalog-detail .annot-row .type {
  background: var(--c-bg-soft); padding: 1px 5px; border-radius: 3px;
  text-align: center; color: var(--c-text-soft);
}
#catalog-detail .annot-row .contents { overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
#catalog-detail .annot-row button {
  padding: 2px 6px; font-size: 10px; cursor: pointer;
  border: 1px solid var(--c-border); border-radius: 3px;
  background: var(--c-surface); color: var(--c-danger-text);
}
#catalog-detail .links-row {
  display: flex; gap: 8px; align-items: center;
  padding: 4px 0;
  font-size: var(--t-xs);
}
#catalog-detail .links-row .row-num {
  font-family: var(--f-mono); font-weight: 600;
  background: var(--c-primary-soft); color: var(--c-primary);
  padding: 1px 6px; border-radius: 3px; cursor: pointer;
}
#catalog-detail .links-row .row-num:hover { filter: brightness(1.08); }

/* ── Catalog pane ───────────────────────────────────────────────── */
.pdf-pane > h2 { gap: var(--s-4); }
.pdf-pane > h2 .filename {
  font-weight: 400;
  text-transform: none;
  letter-spacing: 0;
  color: var(--c-text-soft);
  font-size: var(--t-sm);
  font-family: var(--f-mono);
  max-width: 320px;
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
  flex: 1;
}
.edit-toggle-btn {
  font-size: var(--t-sm); padding: 4px 10px;
  background: var(--c-surface); border: 1px solid var(--c-border-strong);
  color: var(--c-text-muted);
  border-radius: var(--r-md);
  text-transform: none; letter-spacing: 0; font-weight: 600;
  display: inline-flex; align-items: center; gap: 4px;
  transition: all var(--d-fast) var(--ease-std);
}
.edit-toggle-btn:hover { background: var(--c-surface-2); color: var(--c-text); }
.edit-toggle-btn.active {
  background: var(--c-warn); color: white; border-color: var(--c-warn);
  box-shadow: var(--e-2);
}

/* Edit toolbar — extends canvas-toolbar, adds amber accent */
.edit-toolbar.canvas-toolbar {
  background: linear-gradient(180deg, var(--c-warn-soft) 0%, var(--c-surface) 200%);
  border-bottom: 1px solid var(--c-warn);
}
.edit-toolbar button.tool.active {
  background: var(--c-warn); color: white; border-color: var(--c-warn);
  box-shadow: var(--e-1);
}
.edit-toolbar button.tool.active:hover {
  background: var(--c-warn-hover); color: white;
}
.edit-toolbar .save-btn {
  background: var(--c-success); color: white !important; border-color: var(--c-success) !important;
  font-weight: 600;
  padding: 0 var(--s-5) !important;
}
.edit-toolbar .save-btn:hover:not(:disabled) {
  background: var(--c-success-hover); border-color: var(--c-success-hover);
}
.edit-toolbar .save-btn:disabled {
  background: var(--c-border-strong) !important; color: var(--c-text-soft) !important;
  border-color: var(--c-border-strong) !important;
}
.edit-toolbar .dirty-indicator {
  color: var(--c-warn-hover); font-size: var(--t-sm); font-weight: 600;
  padding: 0 var(--s-3);
  display: inline-flex; align-items: center;
  flex-shrink: 0;
}

/* PDF canvas wrapper for SVG overlay */
.pdf-page-host { position: relative; display: inline-block; }
.pdf-page-host img.pdf-page-img {
  max-width: 100%;
  box-shadow: 0 4px 16px rgba(0,0,0,0.5);
  display: block; background: white;
  border-radius: 2px;
}
.pdf-overlay { position: absolute; top: 0; left: 0; width: 100%; height: 100%; pointer-events: none; }
.pdf-overlay.editable { pointer-events: all; }
.pdf-overlay g.annot { pointer-events: all; cursor: pointer; }
.pdf-overlay g.annot.selected { cursor: move; }
/* Annotation rendering — fully WYSIWYG with preview.
   The PDF page is now rendered WITH annots baked in (same as preview),
   so SVG ann-rect is TRANSPARENT for existing annots. We only paint
   visible borders for:
     • NEWLY drawn (unsaved) annots — class .is-new
     • Currently SELECTED annot — class .selected (warn-orange handles)
   Both cases show on top of the baked image as overlays. */
.pdf-overlay rect.ann-rect {
  fill: transparent;
  stroke: none;                          /* transparent by default */
}
/* Newly drawn annot — show red border so user knows it's there
   (the saved PDF won't have it baked yet) */
.pdf-overlay g.annot.is-new rect.ann-rect {
  stroke: rgb(255, 0, 0);
  stroke-width: 1;
}
.pdf-overlay g.annot.is-new rect.ann-rect.freetext {
  stroke: none;     /* FreeText label rect has no border in saved PDF either */
}
/* Selected — warn-orange highlight on top of whatever's there */
.pdf-overlay g.annot.selected rect.ann-rect {
  stroke: var(--c-warn);
  stroke-width: 2;
}
.pdf-overlay g.annot.selected rect.ann-rect.freetext {
  stroke: var(--c-warn);
  stroke-width: 1.5;
}
.pdf-overlay text.ann-text {
  fill: rgb(255, 0, 0);
  /* Helvetica family — closest to PyMuPDF's "helv" rendered glyphs */
  font-family: Helvetica, "Helvetica Neue", Arial, sans-serif;
  font-weight: normal;
  user-select: none; pointer-events: none;
}
.pdf-overlay rect.ann-hit { fill: rgba(0,0,0,0); stroke: none; }
.pdf-overlay rect.ann-hit:hover { fill: rgba(245,158,11,0.18); }
.pdf-overlay g.annot.selected rect.ann-hit { fill: rgba(245,158,11,0.10); }
.pdf-overlay .handle { fill: white; stroke: var(--c-warn); stroke-width: 1.5; cursor: nwse-resize; }
.pdf-overlay .handle.h-n, .pdf-overlay .handle.h-s { cursor: ns-resize; }
.pdf-overlay .handle.h-e, .pdf-overlay .handle.h-w { cursor: ew-resize; }
.pdf-overlay .handle.h-ne, .pdf-overlay .handle.h-sw { cursor: nesw-resize; }
.pdf-overlay rect.draw-preview {
  fill: rgba(245,158,11,0.20); stroke: var(--c-warn);
  stroke-width: 2; stroke-dasharray: 4 2; pointer-events: none;
}
.pdf-canvas.edit-mode { background: var(--c-canvas-edit); }

/* ── Phase A5: floating annotation toolbar (Acrobat-style) ───────────
   Appears above the currently selected annotation in edit mode. Lives
   on body to escape the catalog pane's stacking context, positioned
   in viewport coordinates by JS. */
.float-annot-toolbar {
  position: fixed;
  z-index: 320;             /* above topbar (10), below modals (350+) */
  display: none;
  align-items: center;
  gap: 6px;
  padding: 6px 8px;
  background: var(--c-surface);
  border: 1px solid var(--c-border);
  border-radius: 8px;
  box-shadow: 0 8px 24px rgba(0,0,0,0.18), 0 2px 4px rgba(0,0,0,0.08);
  font-size: var(--t-sm);
  user-select: none;
  white-space: nowrap;
  pointer-events: auto;
  opacity: 0;
  transform: translateY(2px);
  transition: opacity 120ms ease, transform 120ms ease;
}
.float-annot-toolbar.visible {
  display: inline-flex;
  opacity: 1;
  transform: translateY(0);
}
.float-annot-toolbar .fat-type {
  display: inline-flex; align-items: center; gap: 4px;
  padding: 2px 6px; border-radius: 4px;
  background: var(--c-bg-soft); color: var(--c-text-soft);
  font-weight: 600; font-size: 11px;
}
.float-annot-toolbar .fat-sep {
  width: 1px; align-self: stretch; background: var(--c-border); margin: 0 2px;
}
.float-annot-toolbar .fat-swatch {
  display: inline-block;
  width: 14px; height: 14px;
  border-radius: 3px;
  border: 1px solid rgba(0,0,0,0.15);
  vertical-align: middle;
}
.float-annot-toolbar .fat-meta {
  color: var(--c-text-soft); font-size: 11px;
}
.float-annot-toolbar button.fat-btn {
  display: inline-flex; align-items: center; gap: 4px;
  padding: 4px 8px; border-radius: 5px;
  border: 1px solid transparent; background: transparent;
  color: var(--c-text); cursor: pointer;
  font-size: var(--t-sm);
}
.float-annot-toolbar button.fat-btn:hover {
  background: var(--c-bg-soft); border-color: var(--c-border);
}
.float-annot-toolbar button.fat-btn.danger { color: var(--c-danger-text); }
.float-annot-toolbar button.fat-btn.danger:hover {
  background: var(--c-danger-soft); border-color: var(--c-danger);
}
.float-annot-toolbar .fat-arrow {
  position: absolute;
  bottom: -6px; left: 16px;
  width: 10px; height: 10px;
  background: var(--c-surface);
  border-right: 1px solid var(--c-border);
  border-bottom: 1px solid var(--c-border);
  transform: rotate(45deg);
}
.float-annot-toolbar.below .fat-arrow {
  bottom: auto; top: -6px;
  border-right: none; border-bottom: none;
  border-left: 1px solid var(--c-border);
  border-top: 1px solid var(--c-border);
}
.pdf-canvas.edit-mode.tool-drawRect { cursor: crosshair; }
.pdf-canvas.edit-mode.tool-addText { cursor: text; }

/* Inline text editor (for FreeText) */
.text-editor {
  position: absolute; box-sizing: border-box;
  background: white; color: red; font-weight: bold;
  border: 2px solid var(--c-warn); border-radius: 2px;
  padding: 2px 4px; outline: none; resize: none;
  z-index: 50; font-family: inherit;
  box-shadow: var(--e-3);
}

.pdf-annots {
  max-height: 140px; overflow: auto;
  background: var(--c-surface-2);
  border-top: 1px solid var(--c-border);
  padding: var(--s-3) var(--s-7);
  font-size: var(--t-xs);
  font-family: var(--f-mono);
  color: var(--c-text-muted);
}
.pdf-annots .ann-row { padding: 1px 0; }
.pdf-annots .ann-row.matched {
  background: var(--c-warn-soft);
  font-weight: 600;
  padding-left: var(--s-3);
  border-left: 2px solid var(--c-warn);
  color: var(--c-warn-text);
  border-radius: 0 var(--r-sm) var(--r-sm) 0;
}

/* Manual-annotate / Re-annotate wizard banner */
.manual-mode-banner {
  display: none;
  padding: var(--s-3) var(--s-7);
  background: linear-gradient(90deg, var(--c-success-soft), rgba(167,243,208,0.5));
  border-bottom: 2px solid var(--c-success);
  font-size: var(--t-base);
  align-items: center; gap: var(--s-4);
  flex-wrap: wrap;
}
.manual-mode-banner.show { display: flex; animation: slide-down var(--d-base) var(--ease-out); }

/* Wizard stepper progress (chips) */
.wiz-progress {
  display: inline-flex; gap: 4px; align-items: center;
  flex-shrink: 0;
}
.wiz-progress .step {
  display: inline-flex; align-items: center; gap: 4px;
  padding: 2px var(--s-4);
  background: rgba(255,255,255,0.6);
  border: 1px solid var(--c-success);
  border-radius: var(--r-pill);
  font-size: var(--t-xs); font-weight: 600;
  color: var(--c-success-text);
  transition: all var(--d-fast) var(--ease-std);
  white-space: nowrap;
}
.wiz-progress .step.active {
  background: var(--c-success); color: white;
  box-shadow: var(--e-2);
}
.wiz-progress .step.done {
  background: var(--c-success-soft);
  color: var(--c-success-text);
  opacity: 0.85;
}
.wiz-progress .step .num {
  display: inline-block; min-width: 14px; text-align: center;
  font-family: var(--f-mono);
}
.wiz-progress .step .check { font-size: 9px; }
.wiz-progress .arrow {
  color: var(--c-success-text); opacity: 0.5;
  font-size: var(--t-xs);
}

/* Wizard action bar within banner */
.wiz-actions {
  display: inline-flex; gap: var(--s-3); align-items: center;
  margin-left: auto; flex-wrap: wrap;
}

/* Wizard PDF picker button */
.wiz-pdf-btn {
  display: inline-flex; align-items: center; gap: 4px;
  padding: 4px var(--s-4);
  background: rgba(255,255,255,0.85);
  border: 1px solid var(--c-success);
  color: var(--c-success-text);
  border-radius: var(--r-md);
  font-size: var(--t-xs); font-weight: 600;
  font-family: var(--f-mono);
  max-width: 280px;
  transition: all var(--d-fast) var(--ease-std);
  position: relative;
}
.wiz-pdf-btn:hover { background: white; border-color: var(--c-success-hover); }
.wiz-pdf-btn .wiz-pdf-icon { flex-shrink: 0; font-size: var(--t-sm); }
.wiz-pdf-btn .wiz-pdf-label {
  flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
  font-size: var(--t-xs);
}
.wiz-pdf-btn .wiz-pdf-caret { flex-shrink: 0; font-size: 9px; opacity: 0.7; }

.manual-mode-banner .back,
.manual-mode-banner .cancel,
.manual-mode-banner .save { white-space: nowrap; }
.manual-mode-banner .back:disabled {
  opacity: 0.4; cursor: not-allowed;
  background: rgba(255,255,255,0.4);
}

/* Floating PDF picker dropdown */
.wiz-pdf-menu {
  position: fixed; z-index: 250;
  background: var(--c-surface);
  border-radius: var(--r-lg);
  box-shadow: var(--e-4);
  border: 1px solid var(--c-border);
  padding: 4px;
  min-width: 320px; max-width: 480px;
  max-height: 60vh; overflow: auto;
  font-size: var(--t-base);
  animation: menu-pop var(--d-fast) var(--ease-out);
}
.wiz-pdf-menu .menu-header {
  padding: 6px var(--s-5) 4px;
  font-size: var(--t-xs); color: var(--c-text-faint);
  text-transform: uppercase; letter-spacing: 0.06em; font-weight: 700;
  border-bottom: 1px solid var(--c-divider);
  margin-bottom: 2px;
}
.wiz-pdf-menu button.pdf-opt {
  display: flex; flex-direction: column;
  width: 100%; padding: 6px var(--s-5);
  border: 0; background: transparent;
  border-radius: var(--r-sm);
  text-align: left;
  font-size: var(--t-sm);
  color: var(--c-text);
  transition: background var(--d-fast) var(--ease-std);
  gap: 1px;
}
.wiz-pdf-menu button.pdf-opt:hover { background: var(--c-surface-2); }
.wiz-pdf-menu button.pdf-opt.is-current {
  background: var(--c-success-soft); color: var(--c-success-text);
  font-weight: 600;
}
.wiz-pdf-menu button.pdf-opt .pdf-name {
  font-family: var(--f-mono); font-size: var(--t-sm);
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}
.wiz-pdf-menu button.pdf-opt .pdf-folder {
  font-size: var(--t-xs); color: var(--c-text-soft);
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}
.wiz-pdf-menu button.pdf-opt.is-current .pdf-folder { color: var(--c-success); }
@keyframes slide-down {
  from { opacity: 0; transform: translateY(-6px); }
  to   { opacity: 1; transform: translateY(0); }
}
.manual-mode-banner .target-info { flex: 1; color: var(--c-success-text); }
.manual-mode-banner .target-info strong { font-weight: 700; }
.manual-mode-banner .target-info code {
  background: rgba(255,255,255,0.7);
  padding: 1px 4px; border-radius: var(--r-sm);
  font-size: var(--t-sm); font-family: var(--f-mono);
}
.manual-mode-banner button {
  padding: 5px var(--s-5);
  border: 1px solid var(--c-success); background: var(--c-surface);
  color: var(--c-success-text);
  border-radius: var(--r-md); cursor: pointer;
  font-size: var(--t-sm); font-weight: 600;
  transition: all var(--d-fast) var(--ease-std);
}
.manual-mode-banner button:hover { background: var(--c-success-soft); }
.manual-mode-banner button.save {
  background: var(--c-success); color: white; border-color: var(--c-success);
}
.manual-mode-banner button.save:hover:not(:disabled) {
  background: var(--c-success-hover); border-color: var(--c-success-hover);
}
.manual-mode-banner button.save:disabled {
  background: var(--c-border-strong); border-color: var(--c-border-strong);
  color: var(--c-text-soft); cursor: not-allowed;
}
.manual-mode-banner button.cancel {
  color: var(--c-danger-text); border-color: var(--c-danger);
}
.manual-mode-banner button.cancel:hover { background: var(--c-danger-soft); }

/* ── Docked bottom action bar (locked height — no flicker) ─────── */
/* Always occupies the same grid 'action' row, ALWAYS at fixed height.
   Internal sections (top + bottom) sized exactly so notes textarea
   never expands. Empty state is an absolute overlay, doesn't affect
   the box. */
.action-bar {
  position: relative;
  background: var(--c-surface);
  border-top: 1px solid var(--c-border);
  padding: var(--s-3) var(--s-7);
  z-index: 60;
  box-shadow: 0 -2px 12px rgba(15, 23, 42, 0.06);
  width: 100%;
  height: var(--action-bar-h);    /* FIXED — not min-height */
  max-height: var(--action-bar-h);
  display: flex; flex-direction: column;
  justify-content: center;
  gap: var(--s-2);
  overflow: hidden;
  contain: layout size;            /* tells the browser its size never depends on children */
}
.action-bar.is-empty .ab-top,
.action-bar.is-empty .ab-bottom { visibility: hidden; }
.action-bar .ab-empty-msg {
  display: none;
  color: var(--c-text-faint);
  font-size: var(--t-sm);
  text-align: center;
}
.action-bar.is-empty .ab-empty-msg {
  display: block;
  position: absolute; left: 0; right: 0; top: 50%;
  transform: translateY(-50%);
}
.action-bar .ab-top {
  flex-shrink: 0;
  min-height: 0;
}
.action-bar .ab-bottom {
  flex-shrink: 0;
  min-height: 0;
}
.ab-top { display: flex; gap: var(--s-6); align-items: center; flex-wrap: wrap; }
.ab-row-info {
  font-weight: 600; font-size: var(--t-md); color: var(--c-text);
  display: inline-flex; align-items: center; gap: var(--s-3);
}
.ab-row-info .row-num {
  font-family: var(--f-mono);
  background: var(--c-primary-soft);
  color: var(--c-primary-text);
  padding: 2px var(--s-4);
  border-radius: var(--r-sm);
  font-size: var(--t-base); font-weight: 700;
}
.ab-row-info .section { color: var(--c-text-soft); font-weight: 400; font-size: var(--t-sm); }
.ab-flags {
  font-size: var(--t-sm); color: var(--c-warn-hover);
  background: var(--c-warn-soft); padding: 2px var(--s-4);
  border-radius: var(--r-sm);
}
.ab-spacer { flex: 1; }

/* === Verdict segmented control ============================== */
.verdict-control {
  display: inline-flex;
  background: var(--c-surface-2);
  border: 1px solid var(--c-border);
  border-radius: var(--r-lg);
  padding: 3px;
  gap: 2px;
}
.verdict-control .ab-btn {
  border: 1px solid transparent;
  background: transparent;
  color: var(--c-text-muted);
  font-size: var(--t-sm);
  font-weight: 500;
  padding: 0 var(--s-5);
  height: 28px;
  border-radius: var(--r-md);
  display: inline-flex; align-items: center; gap: var(--s-2);
  white-space: nowrap;
  transition: background var(--d-fast) var(--ease-std),
              color var(--d-fast) var(--ease-std),
              box-shadow var(--d-fast) var(--ease-std),
              transform var(--d-fast) var(--ease-out);
  cursor: pointer;
}
.verdict-control .ab-btn .vc-label { font-weight: 500; }
.verdict-control .ab-btn:hover { background: var(--c-surface); color: var(--c-text); }
.verdict-control .ab-btn:active { transform: translateY(1px); }
.verdict-control .ab-btn .kbd {
  background: var(--c-surface-3); color: var(--c-text-faint);
  border: 1px solid var(--c-border);
  padding: 0 4px; border-radius: var(--r-sm);
  font-family: var(--f-mono); font-size: var(--t-xs);
  margin-left: var(--s-2);
}

/* Active state per kind — fills the slot with semantic color */
.verdict-control .ab-btn.pass.active {
  background: var(--c-success); color: white;
  box-shadow: 0 1px 2px rgba(16,185,129,0.3);
}
.verdict-control .ab-btn.fail.active {
  background: var(--c-danger); color: white;
  box-shadow: 0 1px 2px rgba(239,68,68,0.3);
}
.verdict-control .ab-btn.fix.active {
  background: var(--c-warn); color: white;
  box-shadow: 0 1px 2px rgba(245,158,11,0.3);
}
.verdict-control .ab-btn.skip.active {
  background: var(--c-text-soft); color: white;
  box-shadow: 0 1px 2px rgba(100,116,139,0.3);
}
.verdict-control .ab-btn.active .kbd {
  background: rgba(255,255,255,0.18); color: rgba(255,255,255,0.85);
  border-color: rgba(255,255,255,0.25);
}
.verdict-control .ab-btn.active .ico { stroke-width: 2.5; }

/* === Secondary actions (Auto, Mark) — visually subordinate === */
.ab-secondary { display: inline-flex; gap: 4px; align-items: center; }
.ab-secondary .ab-btn {
  display: inline-flex; align-items: center; gap: var(--s-2);
  padding: 0 var(--s-5);
  height: var(--btn-h-sm);
  border: 1px solid var(--c-border-strong);
  background: var(--c-surface);
  color: var(--c-text-muted);
  border-radius: var(--r-md);
  font-size: var(--t-sm); font-weight: 500;
  transition: background var(--d-fast) var(--ease-std),
              color var(--d-fast) var(--ease-std),
              border-color var(--d-fast) var(--ease-std);
  cursor: pointer;
}
.ab-secondary .ab-btn:hover { background: var(--c-surface-2); color: var(--c-text); }
.ab-secondary .ab-btn:active { transform: translateY(1px); }
.ab-secondary .ab-btn.auto { color: var(--c-purple-text); }
.ab-secondary .ab-btn.auto:hover { background: var(--c-purple-soft); border-color: var(--c-purple); }
.ab-secondary .ab-btn.mark { color: var(--c-teal-text); }
.ab-secondary .ab-btn.mark:hover { background: var(--c-teal-soft); border-color: var(--c-teal); }
.ab-secondary .ab-btn.mark.commitment {
  background: var(--c-warn-soft); border-color: var(--c-warn); color: var(--c-warn-text);
  animation: pulse-warn 2s ease-in-out infinite;
}

/* iOS-style toggle (replaces tiny checkbox) */
.ab-toggle {
  display: inline-flex; align-items: center; gap: var(--s-3);
  cursor: pointer; user-select: none;
  font-size: var(--t-sm); color: var(--c-text-muted);
  margin-left: var(--s-3);
}
.ab-toggle input { position: absolute; opacity: 0; pointer-events: none; }
.ab-toggle .ab-toggle-track {
  width: 30px; height: 18px;
  background: var(--c-border-strong);
  border-radius: 999px;
  position: relative;
  transition: background var(--d-fast) var(--ease-std);
  flex-shrink: 0;
}
.ab-toggle .ab-toggle-track::after {
  content: '';
  position: absolute; top: 2px; left: 2px;
  width: 14px; height: 14px;
  background: white;
  border-radius: 50%;
  box-shadow: 0 1px 2px rgba(0,0,0,0.2);
  transition: transform var(--d-base) var(--ease-out);
}
.ab-toggle input:checked + .ab-toggle-track { background: var(--c-primary); }
.ab-toggle input:checked + .ab-toggle-track::after { transform: translateX(12px); }
.ab-toggle input:focus-visible + .ab-toggle-track {
  outline: 2px solid var(--c-focus);
  outline-offset: 2px;
}

.ab-bottom {
  display: flex; gap: var(--s-4); align-items: center;
  margin-top: 0;
  height: 28px;
}
.ab-bottom textarea {
  flex: 1; resize: none;
  height: 28px; min-height: 28px; max-height: 28px;
  padding: 4px var(--s-5);
  font: inherit; font-size: var(--t-base);
  line-height: 18px;
  border: 1px solid var(--c-border-strong); border-radius: var(--r-md);
  background: var(--c-surface);
  white-space: nowrap; overflow: hidden;
}
.ab-bottom .reset-btn {
  font-size: var(--t-sm); color: var(--c-text-soft);
  background: none; border: none;
  padding: 4px var(--s-4);
  border-radius: var(--r-sm);
  transition: color var(--d-fast) var(--ease-std), background var(--d-fast) var(--ease-std);
}
.ab-bottom .reset-btn:hover { color: var(--c-text); background: var(--c-surface-2); }

/* ── Floating actions (Phase A7: only visible in embedded mode) ──
   The activity rail already provides Settings + Help in normal use.
   FAB only kicks in when running inside Claude Preview MCP / iframe
   where the topbar can be covered by host UI. */
.floating-actions {
  position: fixed;
  bottom: calc(var(--action-bar-h) + var(--s-5));
  left: var(--s-5);
  display: none;          /* hidden by default — rail covers normal case */
  flex-direction: column; gap: var(--s-3);
  z-index: 90;
}
body[data-embedded="1"] .floating-actions { display: flex; }
.fab-btn {
  width: 40px; height: 40px;
  border-radius: 50%;
  border: 1px solid var(--c-border);
  background: var(--c-surface);
  box-shadow: var(--e-2);
  color: var(--c-text-muted);
  display: inline-flex; align-items: center; justify-content: center;
  cursor: pointer;
  transition: all var(--d-fast) var(--ease-std);
}
.fab-btn:hover {
  background: var(--c-surface-2);
  color: var(--c-text);
  border-color: var(--c-border-strong);
  transform: translateY(-1px);
  box-shadow: var(--e-3);
}
.fab-btn:active { transform: translateY(0); }
.fab-btn .ico { width: 18px; height: 18px; }

/* On mobile (action-bar takes less room) shrink + tighten the FAB stack */
@media (max-width: 700px) {
  .floating-actions {
    bottom: calc(var(--action-bar-h) + var(--s-3));
    left: var(--s-3);
    gap: var(--s-2);
  }
  .fab-btn { width: 36px; height: 36px; }
}

/* ── kbd-help (Phase A7: collapsed badge — expands on hover) ──────
   Live above the action bar, bottom-right. Collapses to a single "?"
   so it doesn't compete with the catalog/AI panes. Hover to reveal
   the full shortcut row. Hidden in embedded mode (less screen real
   estate) — users press / for the help modal there. */
.kbd-help {
  position: fixed;
  bottom: calc(var(--action-bar-h) + var(--statusbar-h) + var(--s-5));
  right: var(--s-5);
  font-size: var(--t-xs); color: var(--c-text-soft);
  background: var(--c-surface);
  padding: 4px 10px;
  border-radius: 999px;
  border: 1px solid var(--c-border);
  box-shadow: var(--e-1);
  display: inline-flex; gap: var(--s-4); align-items: center;
  opacity: 0.55;
  z-index: 50;
  cursor: default;
  max-width: 38px; overflow: hidden;
  white-space: nowrap;
  transition: max-width 200ms var(--ease-out), opacity 120ms var(--ease-std);
}
.kbd-help::before {
  content: '⌨';
  font-size: 14px; line-height: 1; color: var(--c-text);
  flex-shrink: 0;
}
.kbd-help:hover { opacity: 1; max-width: 600px; }
.kbd-help .kbd-group {
  display: inline-flex; gap: 3px; align-items: center;
}
.kbd-help .kbd-text { color: var(--c-text-faint); margin-left: 2px; }
body[data-embedded="1"] .kbd-help { display: none; }

/* ── Sync banner ────────────────────────────────────────────────── */
.sync-banner {
  position: fixed; top: 0; left: 0; right: 0;
  z-index: 150;
  padding: var(--s-3) var(--s-7);
  font-size: var(--t-base); text-align: center;
  display: none;
  border-bottom: 1px solid;
}
.sync-banner.show { display: block; animation: slide-down var(--d-base) var(--ease-out); }
.sync-banner.warn   { background: var(--c-warn-soft); color: var(--c-warn-text); border-color: var(--c-warn); }
.sync-banner.danger { background: var(--c-danger-soft); color: var(--c-danger-text); border-color: var(--c-danger); }
.sync-banner button {
  margin-left: var(--s-6);
  padding: 3px var(--s-5);
  border: 1px solid currentColor; background: var(--c-surface);
  border-radius: var(--r-sm); cursor: pointer; font-weight: 600;
  color: inherit;
}
.sync-banner button:hover { background: rgba(0,0,0,0.05); }
.sync-banner .close-x {
  float: right; cursor: pointer;
  padding: 0 var(--s-4); opacity: 0.7;
  transition: opacity var(--d-fast) var(--ease-std);
}
.sync-banner .close-x:hover { opacity: 1; }

/* Versions toolbar buttons */
.versions-btn {
  font-size: var(--t-sm); padding: 4px var(--s-5);
  border: 1px solid var(--c-border);
  background: var(--c-surface);
  border-radius: var(--r-md);
  cursor: pointer;
  text-transform: none; letter-spacing: 0; font-weight: 600;
  color: var(--c-text-muted);
  display: inline-flex; align-items: center; gap: 4px;
  transition: all var(--d-fast) var(--ease-std);
}
.versions-btn:hover {
  background: var(--c-surface-2); color: var(--c-text);
  border-color: var(--c-border-strong);
}
.versions-btn.warn   { border-color: var(--c-warn); background: var(--c-warn-soft); color: var(--c-warn-text); }
.versions-btn.danger {
  border-color: var(--c-danger);
  background: var(--c-danger-soft);
  color: var(--c-danger-text);
  animation: pulse-warn 2s ease-in-out infinite;
}
@keyframes pulse-warn {
  0%, 100% { box-shadow: 0 0 0 0 rgba(239,68,68,0.4); }
  50%      { box-shadow: 0 0 0 4px rgba(239,68,68,0); }
}

/* Sync badge */
.sync-badge {
  display: inline-flex; align-items: center; gap: var(--s-3);
  padding: 4px var(--s-5);
  border-radius: var(--r-pill);
  font-size: var(--t-sm); font-weight: 600;
  margin: var(--s-3) 0 var(--s-5) 0;
}
.sync-badge.in-sync       { background: var(--c-success-soft); color: var(--c-success-text); }
.sync-badge.working-ahead { background: var(--c-info-soft); color: var(--c-info-text); }
.sync-badge.working-behind,
.sync-badge.divergent,
.sync-badge.incomplete-local { background: var(--c-danger-soft); color: var(--c-danger-text); }
.sync-badge.no-snapshots  { background: var(--c-surface-3); color: var(--c-text-soft); }
.sync-badge .latest-id {
  font-family: var(--f-mono); font-weight: 400;
  font-size: var(--t-xs); opacity: 0.85;
}

/* ── Modals ─────────────────────────────────────────────────────── */
.modal-bg {
  position: fixed; inset: 0;
  background: var(--c-overlay);
  display: none; align-items: center; justify-content: center;
  z-index: 200;
  padding: var(--s-7);
  backdrop-filter: blur(4px);
  -webkit-backdrop-filter: blur(4px);
}
.modal-bg[style*="display: flex"], .modal-bg.show {
  display: flex !important;
  animation: modal-bg-in var(--d-base) var(--ease-std);
}
@keyframes modal-bg-in { from { opacity: 0; } to { opacity: 1; } }

.modal {
  background: var(--c-surface);
  padding: var(--s-7) var(--s-8);
  border-radius: var(--r-xl);
  min-width: 480px; max-width: 640px;
  max-height: 85vh; overflow: auto;
  box-shadow: var(--e-5);
  border: 1px solid var(--c-border);
  animation: modal-in var(--d-slow) var(--ease-out);
}
@keyframes modal-in {
  from { opacity: 0; transform: translateY(8px) scale(0.985); }
  to   { opacity: 1; transform: translateY(0) scale(1); }
}
.modal h3 {
  margin: 0 0 var(--s-7);
  font-size: var(--t-xl); font-weight: 700;
  color: var(--c-text);
  display: flex; align-items: center;
  gap: var(--s-4);
  border-bottom: 1px solid var(--c-divider);
  padding-bottom: var(--s-5);
}
.modal h3 > .ico { color: var(--c-primary); flex-shrink: 0; }
.modal h3 > span { flex: 1; }
.modal h3 > .close { margin-left: auto; }
.modal h3 .close {
  background: none; border: none; cursor: pointer;
  font-size: var(--t-xl); color: var(--c-text-soft);
  width: 32px; height: 32px;
  display: inline-flex; align-items: center; justify-content: center;
  border-radius: var(--r-md);
  transition: background var(--d-fast) var(--ease-std), color var(--d-fast) var(--ease-std);
}
.modal h3 .close:hover { background: var(--c-surface-2); color: var(--c-text); }

/* History modal */
.history-item {
  padding: var(--s-4) var(--s-5);
  border-radius: var(--r-md);
  margin-bottom: var(--s-2);
  background: var(--c-surface-2);
  display: flex; gap: var(--s-4); align-items: center;
  transition: background var(--d-fast) var(--ease-std);
}
.history-item:hover { background: var(--c-surface-3); }
.history-item .ts {
  font-family: var(--f-mono); font-size: var(--t-sm);
  color: var(--c-text-muted); flex: 1;
}
.history-item .size { font-size: var(--t-xs); color: var(--c-text-faint); }
.history-item button {
  font-size: var(--t-sm); padding: 4px var(--s-4);
  border: 1px solid var(--c-border-strong);
  background: var(--c-surface);
  color: var(--c-text);
  border-radius: var(--r-sm);
}
.history-item button:hover { background: var(--c-surface-3); }

/* ── Toasts ─────────────────────────────────────────────────────── */
.toast-stack {
  position: fixed;
  top: calc(var(--topbar-h) + var(--s-3));
  right: var(--s-7);
  display: flex; flex-direction: column; gap: var(--s-4);
  z-index: 300; pointer-events: none;
  max-width: 380px;
}
@media (max-width: 700px) {
  .toast-stack { top: 56px; }
}
.toast {
  pointer-events: auto;
  background: var(--c-surface);
  border-radius: var(--r-lg);
  padding: var(--s-5) var(--s-6);
  font-size: var(--t-base);
  box-shadow: var(--e-3);
  border-left: 4px solid var(--c-success);
  border-top: 1px solid var(--c-border);
  border-right: 1px solid var(--c-border);
  border-bottom: 1px solid var(--c-border);
  animation: toast-in var(--d-slow) var(--ease-out);
  max-width: 380px;
  position: relative;
  overflow: hidden;
}
.toast.warn  { border-left-color: var(--c-warn); }
.toast.error { border-left-color: var(--c-danger); }
.toast.info  { border-left-color: var(--c-info); }
.toast.learn { border-left-color: var(--c-purple); }
.toast.learn::before {
  content: ''; position: absolute; inset: 0;
  background: linear-gradient(135deg, var(--c-purple-soft) 0%, transparent 60%);
  pointer-events: none;
}
.toast .title {
  font-weight: 700; margin-bottom: 2px;
  color: var(--c-text);
  position: relative;
}
.toast .body {
  color: var(--c-text-muted);
  position: relative;
}
.toast .close {
  float: right;
  background: transparent; border: none;
  cursor: pointer; opacity: 0.5;
  padding: 2px;
  color: var(--c-text-muted);
  border-radius: var(--r-sm);
  transition: opacity var(--d-fast) var(--ease-std), background var(--d-fast) var(--ease-std);
  position: relative;
  display: inline-flex; align-items: center; justify-content: center;
}
.toast .close:hover { opacity: 1; background: var(--c-surface-2); }
@keyframes toast-in {
  from { transform: translateX(120%); opacity: 0; }
  to   { transform: translateX(0); opacity: 1; }
}
.toast.fading {
  transition: opacity 0.4s var(--ease-std), transform 0.4s var(--ease-std);
  opacity: 0; transform: translateX(80%);
}

/* ── Spinner ────────────────────────────────────────────────────── */
.spinner {
  display: inline-block;
  width: 24px; height: 24px;
  border: 3px solid var(--c-border);
  border-top-color: var(--c-primary);
  border-radius: 50%;
  animation: spin 0.8s linear infinite;
  margin-right: var(--s-4); vertical-align: middle;
}
@keyframes spin { to { transform: rotate(360deg); } }

/* ── Versions modal ─────────────────────────────────────────────── */
.versions-modal-body { min-width: 720px; max-width: 920px; }
.versions-snap-form {
  display: flex; gap: var(--s-3); align-items: center;
  margin-bottom: var(--s-5);
  padding: var(--s-4) var(--s-5);
  background: var(--c-surface-2);
  border: 1px solid var(--c-border);
  border-radius: var(--r-md);
}
.versions-snap-form input[type=text] {
  flex: 1; padding: 6px var(--s-5);
  border: 1px solid var(--c-border-strong);
  background: var(--c-surface);
  border-radius: var(--r-sm);
  font-size: var(--t-base);
}
.versions-snap-form .snap-quick { background: var(--c-info-soft); color: var(--c-info-text); border-color: var(--c-info-soft); }
.versions-snap-form .snap-quick:hover { background: var(--c-primary-soft); border-color: var(--c-primary); }
.versions-snap-form .snap-full  { background: var(--c-warn-soft); color: var(--c-warn-text); border-color: var(--c-warn-soft); }
.versions-snap-form .snap-full:hover { border-color: var(--c-warn); }
.versions-snap-form .auto-snap  { background: var(--c-success-soft); color: var(--c-success-text); border-color: var(--c-success-soft); }
.versions-snap-form .auto-snap:hover { border-color: var(--c-success); }

.versions-actions {
  display: flex; gap: var(--s-3); align-items: center;
  margin-bottom: var(--s-4);
  font-size: var(--t-sm); color: var(--c-text-soft);
}
.versions-actions .diff-helper { color: var(--c-text-faint); }

.versions-list {
  max-height: 50vh; overflow-y: auto;
  border: 1px solid var(--c-border); border-radius: var(--r-md);
  background: var(--c-surface);
}
.version-item {
  padding: var(--s-4) var(--s-5);
  border-bottom: 1px solid var(--c-divider);
  display: grid;
  grid-template-columns: 28px 1fr auto;
  gap: var(--s-4); align-items: center;
  transition: background var(--d-fast) var(--ease-std);
}
.version-item:last-child { border-bottom: none; }
.version-item:hover { background: var(--c-surface-2); }
.version-item.diff-selected { background: var(--c-warn-soft); }
.version-item .v-check { cursor: pointer; }
.version-item .v-tag {
  font-weight: 600; color: var(--c-text); font-size: var(--t-base);
  margin-bottom: 2px;
  display: flex; align-items: center; gap: var(--s-3);
}
.version-item .v-tag .untagged {
  color: var(--c-text-faint); font-style: italic; font-weight: 400;
}
.version-item .v-meta {
  font-size: var(--t-sm); color: var(--c-text-soft);
  font-family: var(--f-mono);
}
.version-item .v-kind {
  display: inline-block; padding: 1px var(--s-4);
  border-radius: var(--r-sm); margin-right: var(--s-3);
  font-size: 9px; font-weight: 700; text-transform: uppercase;
  letter-spacing: 0.04em;
}
.version-item .v-kind.quick { background: var(--c-info-soft); color: var(--c-info-text); }
.version-item .v-kind.full  { background: var(--c-warn-soft); color: var(--c-warn-text); }
.version-item .v-actions { display: flex; gap: var(--s-2); }
.version-item .v-actions button {
  font-size: var(--t-xs); padding: 4px var(--s-4);
  border: 1px solid var(--c-border-strong);
  background: var(--c-surface);
  color: var(--c-text);
  border-radius: var(--r-sm);
  transition: background var(--d-fast) var(--ease-std);
}
.version-item .v-actions button:hover { background: var(--c-surface-2); }
.version-item .v-actions .restore { color: var(--c-warn-text); border-color: var(--c-warn); }
.version-item .v-actions .restore-full { color: var(--c-danger-text); border-color: var(--c-danger); }

.versions-busy {
  padding: var(--s-7); text-align: center; color: var(--c-text-muted);
}
.versions-output {
  margin-top: var(--s-5);
  padding: var(--s-5);
  background: #111827; color: #e5e7eb;
  border-radius: var(--r-md);
  font-size: var(--t-sm);
  font-family: var(--f-mono);
  max-height: 300px; overflow: auto;
  white-space: pre-wrap; word-break: break-all;
  border: 1px solid #1f2937;
}

/* ── Learning modal ─────────────────────────────────────────────── */
.learn-pattern-list {
  max-height: 45vh; overflow: auto;
  border: 1px solid var(--c-border); border-radius: var(--r-md);
  background: var(--c-surface);
}
.learn-pattern-row {
  padding: var(--s-3) var(--s-5);
  border-bottom: 1px solid var(--c-divider);
  display: grid;
  grid-template-columns: 110px 140px 1fr 70px auto;
  gap: var(--s-4); align-items: center;
  font-size: var(--t-sm);
  transition: background var(--d-fast) var(--ease-std);
}
.learn-pattern-row:last-child { border-bottom: none; }
.learn-pattern-row:hover { background: var(--c-surface-2); }
.learn-pattern-row.disabled { opacity: 0.5; }
.learn-pattern-row .pt {
  display: inline-block; padding: 2px var(--s-4);
  border-radius: var(--r-sm);
  font-size: 9px; font-weight: 700; text-transform: uppercase;
  letter-spacing: 0.04em;
}
.learn-pattern-row .pt.filename_brand { background: var(--c-warn-soft); color: var(--c-warn-text); }
.learn-pattern-row .pt.section_vendor { background: var(--c-info-soft); color: var(--c-info-text); }
.learn-pattern-row .pt.row_format_d   { background: var(--c-success-soft); color: var(--c-success-text); }
.learn-pattern-row .trigger {
  font-family: var(--f-mono); font-size: var(--t-xs);
  background: var(--c-surface-3); padding: 2px var(--s-3);
  border-radius: var(--r-sm); color: var(--c-text-muted);
}
.learn-pattern-row .arrow { color: var(--c-text-faint); }
.learn-pattern-row .output { font-weight: 600; color: var(--c-text); }
.learn-pattern-row .conf {
  font-size: var(--t-xs); color: var(--c-text-soft);
  font-family: var(--f-mono);
}
.learn-pattern-row .conf .high { color: var(--c-success-text); font-weight: 700; }
.learn-pattern-row .conf .low { color: var(--c-danger-text); }

/* Confidence badge for plans */
.conf-badge {
  display: inline-block; padding: 2px var(--s-5);
  border-radius: var(--r-pill);
  font-size: var(--t-sm); font-weight: 700;
  margin-left: var(--s-3);
}
.conf-badge.high { background: var(--c-success-soft); color: var(--c-success-text); }
.conf-badge.med  { background: var(--c-warn-soft); color: var(--c-warn-text); }
.conf-badge.low  { background: var(--c-danger-soft); color: var(--c-danger-text); }
.conf-badge .src { font-size: 9px; opacity: 0.8; margin-left: 4px; }

/* ── Audit / DB modal ───────────────────────────────────────────── */
.audit-stats {
  display: grid; grid-template-columns: repeat(4, 1fr);
  gap: var(--s-4); margin-bottom: var(--s-6);
}
.audit-stats .stat-card {
  padding: var(--s-5);
  background: var(--c-surface-2);
  border-radius: var(--r-md);
  border: 1px solid var(--c-border);
  transition: transform var(--d-fast) var(--ease-out);
}
.audit-stats .stat-card:hover { transform: translateY(-1px); border-color: var(--c-border-strong); }
.audit-stats .stat-card .v {
  font-size: var(--t-2xl); font-weight: 700;
  color: var(--c-text); line-height: 1;
}
.audit-stats .stat-card .l {
  font-size: var(--t-xs); color: var(--c-text-soft);
  text-transform: uppercase; letter-spacing: 0.04em;
  margin-top: 2px;
}
.audit-search-row {
  display: flex; gap: var(--s-3); margin-bottom: var(--s-4);
}
.audit-search-row input {
  flex: 1;
}
.audit-search-row select {
  width: auto; min-width: 140px;
  padding: 6px var(--s-5);
}
.audit-list {
  max-height: 50vh; overflow: auto;
  border: 1px solid var(--c-border); border-radius: var(--r-md);
  background: var(--c-surface);
}
.audit-row {
  padding: var(--s-3) var(--s-5);
  border-bottom: 1px solid var(--c-divider);
  display: grid;
  grid-template-columns: 140px 130px 70px 1fr;
  gap: var(--s-4); align-items: center;
  font-size: var(--t-sm);
  transition: background var(--d-fast) var(--ease-std);
}
.audit-row:last-child { border-bottom: none; }
.audit-row:hover { background: var(--c-surface-2); }
.audit-row .ts {
  font-family: var(--f-mono); color: var(--c-text-soft); font-size: var(--t-xs);
}
.audit-row .ac { font-weight: 600; }
.audit-row .ac.status_change       { color: var(--c-success-text); }
.audit-row .ac.notes_update         { color: var(--c-info-text); }
.audit-row .ac.auto_annotate_apply  { color: var(--c-purple-text); }
.audit-row .ac.pdf_edit             { color: var(--c-warn-text); }
.audit-row .ac.snapshot             { color: var(--c-info-text); }
.audit-row .ac.restore              { color: var(--c-danger-text); }
.audit-row .tg {
  font-family: var(--f-mono); font-size: var(--t-xs);
  color: var(--c-text-muted); cursor: pointer;
}
.audit-row .tg:hover { color: var(--c-text); text-decoration: underline; }
.audit-row .ba { color: var(--c-text-soft); font-size: var(--t-xs); }
.audit-row .ba .arrow { color: var(--c-text-faint); margin: 0 var(--s-2); }
.audit-row .ba .before { color: var(--c-danger-text); }
.audit-row .ba .after  { color: var(--c-success-text); }
.audit-search-results .ar-row {
  padding: var(--s-3) var(--s-5);
  border-bottom: 1px solid var(--c-divider);
  font-size: var(--t-sm); cursor: pointer;
  transition: background var(--d-fast) var(--ease-std);
}
.audit-search-results .ar-row:hover { background: var(--c-warn-soft); }
.audit-search-results mark {
  background: var(--c-warn-soft); color: var(--c-warn-text);
  font-weight: 600; padding: 0 2px; border-radius: 2px;
}

/* ── Auto-annotate modal ────────────────────────────────────────── */
.auto-block { margin-bottom: var(--s-5); }
.auto-label {
  font-size: var(--t-xs); font-weight: 700;
  color: var(--c-text-soft);
  text-transform: uppercase; margin-bottom: 4px;
  letter-spacing: 0.06em;
}
.auto-pre {
  margin: 0; padding: var(--s-4) var(--s-5);
  background: var(--c-surface-2);
  border: 1px solid var(--c-border);
  border-radius: var(--r-md);
  font-size: var(--t-base);
  max-height: 100px; overflow: auto;
  white-space: pre-wrap; word-break: break-word;
  font-family: inherit;
  color: var(--c-text);
}
.auto-pre.empty { color: var(--c-text-faint); font-style: italic; }
.auto-batch-list {
  max-height: 50vh; overflow: auto;
  border: 1px solid var(--c-border); border-radius: var(--r-md);
}
.auto-batch-row {
  display: grid; grid-template-columns: 28px 56px 110px 1fr auto;
  gap: var(--s-3); align-items: center;
  padding: var(--s-3) var(--s-4);
  border-bottom: 1px solid var(--c-divider);
  font-size: var(--t-sm);
}
.auto-batch-row:last-child { border-bottom: none; }
.auto-batch-row .b-row { font-family: var(--f-mono); color: var(--c-text-soft); }
.auto-batch-row .b-role {
  display: inline-block; padding: 1px var(--s-3);
  border-radius: var(--r-sm);
  font-size: 9px; font-weight: 700; text-transform: uppercase;
  letter-spacing: 0.04em;
}
.auto-batch-row .b-role.section_header { background: var(--c-info-soft); color: var(--c-info-text); }
.auto-batch-row .b-role.item            { background: var(--c-warn-soft); color: var(--c-warn-text); }
.auto-batch-row .b-role.sub_item        { background: var(--c-success-soft); color: var(--c-success-text); }
.auto-batch-row .b-role.unknown         { background: var(--c-surface-3); color: var(--c-text-soft); }
.auto-batch-row .b-d {
  font-size: var(--t-xs); color: var(--c-text-muted);
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.auto-batch-row.fail { background: var(--c-danger-soft); }
.auto-batch-row.warn { background: var(--c-warn-soft); }

/* ── Reduced motion ─────────────────────────────────────────────── */
@media (prefers-reduced-motion: reduce) {
  *, *::before, *::after {
    animation-duration: 0.01ms !important;
    animation-iteration-count: 1 !important;
    transition-duration: 0.01ms !important;
    scroll-behavior: auto !important;
  }
}

/* ── Responsive ─────────────────────────────────────────────────── */
@media (max-width: 1280px) {
  #app {
    grid-template-columns: minmax(var(--tree-w-md), 280px) 1fr;
    grid-template-areas:
      "safetop safetop"
      "topbar  topbar"
      "tree    center"
      "action  action";
  }
  .pdf-pane { display: none; }
  .pdf-pane.expand { display: flex; grid-column: 2; grid-area: center; }
}

@media (max-width: 900px) {
  #app {
    grid-template-columns: 1fr;
    grid-template-rows: var(--safe-top) var(--topbar-h) 240px 1fr auto;
    grid-template-areas:
      "safetop"
      "topbar"
      "tree"
      "center"
      "action";
  }
  .tree-pane { border-right: none; border-bottom: 1px solid var(--c-border); }
}

@media (max-width: 700px) {
  .topbar { display: none; }
  .mobile-tabs { display: flex; }
  #app {
    grid-template-columns: 1fr;
    grid-template-rows: var(--safe-top) 44px 1fr auto;
    grid-template-areas:
      "safetop"
      "topbar"
      "tree"
      "action";
  }
  #app > .pane { display: none; grid-area: tree; }
  #app > .pane.mobile-active { display: flex; }
  .ab-top { flex-wrap: wrap; gap: var(--s-3); }
  .ab-secondary, .verdict-control { flex-wrap: wrap; }
  .ab-btn { padding: 5px var(--s-4); font-size: var(--t-sm); }
  .modal { min-width: auto !important; max-width: calc(100vw - 24px); padding: var(--s-6); }
  .audit-stats { grid-template-columns: repeat(2, 1fr) !important; }
  .versions-modal-body { min-width: auto !important; }
  .kbd-help { display: none; }
}

@media (max-width: 480px) {
  .audit-stats { grid-template-columns: 1fr !important; }
  .audit-row { grid-template-columns: 1fr !important; gap: 2px !important; }
  .topbar-actions .stats-pill { display: none; }
}

/* ============================================================
 * Sprint UI-2 / UI-3 additions (production polish, motion, a11y)
 * ============================================================ */

/* Hover lift + press feedback on interactive elements */
.btn:not(:disabled):not(.is-disabled),
.topbar-btn,
.versions-btn,
.canvas-toolbar button,
.history-item,
.version-item,
.learn-pattern-row,
.audit-row,
.col-d-menu button,
.wiz-pdf-menu button.pdf-opt,
.tree-row {
  transition:
    transform var(--d-fast) var(--ease-out),
    box-shadow var(--d-fast) var(--ease-std),
    background var(--d-fast) var(--ease-std),
    color var(--d-fast) var(--ease-std),
    border-color var(--d-fast) var(--ease-std);
}
.btn:not(:disabled):not(.is-disabled):active,
.topbar-btn:active,
.versions-btn:active,
.canvas-toolbar button:active {
  transform: translateY(1px);
}
/* Cards lift on hover */
.history-item:hover,
.version-item:hover,
.learn-pattern-row:hover {
  transform: translateY(-1px);
}
.audit-stats .stat-card:hover {
  transform: translateY(-2px);
  box-shadow: var(--e-2);
}

/* Modal slot system (Sprint UI-2.2) — for new modals; old ones still work */
.modal__header {
  display: flex; align-items: center; gap: var(--s-4);
  padding-bottom: var(--s-5);
  border-bottom: 1px solid var(--c-divider);
  margin: calc(-1 * var(--s-7)) calc(-1 * var(--s-8)) var(--s-6);
  padding: var(--s-6) var(--s-8) var(--s-5);
}
.modal__header .ico { color: var(--c-primary); flex-shrink: 0; }
.modal__title { flex: 1; font-size: var(--t-xl); font-weight: 700; }
.modal__body { padding: 0; max-height: 65vh; overflow: auto; }
.modal__footer {
  position: sticky; bottom: 0;
  display: flex; gap: var(--s-3);
  padding: var(--s-5) var(--s-8);
  margin: var(--s-6) calc(-1 * var(--s-8)) calc(-1 * var(--s-7));
  background: var(--c-surface);
  border-top: 1px solid var(--c-divider);
  border-radius: 0 0 var(--r-xl) var(--r-xl);
}
.modal__footer .spacer { flex: 1; }

/* Empty state (Sprint UI-2.3) */
.empty-state {
  padding: var(--s-12) var(--s-7);
  text-align: center;
  display: flex; flex-direction: column; align-items: center;
  gap: var(--s-5);
  color: var(--c-text-soft);
}
.empty-state .es-icon {
  width: 64px; height: 64px;
  border-radius: 50%;
  display: inline-flex; align-items: center; justify-content: center;
  background: var(--c-surface-2);
  border: 1px solid var(--c-border);
  color: var(--c-text-faint);
}
.empty-state .es-icon .ico { width: 28px; height: 28px; stroke-width: 1.5; }
.empty-state .es-title {
  font-size: var(--t-lg); font-weight: 600; color: var(--c-text);
  margin: 0;
}
.empty-state .es-body {
  font-size: var(--t-base); color: var(--c-text-muted);
  max-width: 320px;
}
.empty-state .es-cta {
  display: inline-flex; align-items: center; gap: var(--s-3);
  padding: 6px var(--s-5);
  background: var(--c-primary); color: white;
  border: 0; border-radius: var(--r-md);
  font-size: var(--t-base); font-weight: 600;
  cursor: pointer;
  transition: background var(--d-fast) var(--ease-std);
}
.empty-state .es-cta:hover { background: var(--c-primary-hover); }

/* Error state */
.error-state .es-icon { background: var(--c-danger-soft); border-color: var(--c-danger); color: var(--c-danger); }
.error-state .es-title { color: var(--c-danger-text); }

/* Counter animation (Sprint UI-2.4) */
.count-animate { display: inline-block; }
.count-animate.pop { animation: count-pop var(--d-base) var(--ease-out); }
@keyframes count-pop {
  0%   { transform: scale(1); }
  50%  { transform: scale(1.15); color: var(--c-primary); }
  100% { transform: scale(1); }
}

/* List item enter / exit */
.list-enter { animation: list-enter var(--d-base) var(--ease-out); }
@keyframes list-enter {
  from { opacity: 0; transform: translateY(-4px); }
  to   { opacity: 1; transform: translateY(0); }
}

/* High-contrast theme (Sprint UI-3.2) */
:root[data-theme="hi-contrast"] {
  --c-bg: #000000;
  --c-bg-subtle: #0a0a0a;
  --c-surface: #0a0a0a;
  --c-surface-2: #141414;
  --c-surface-3: #1c1c1c;
  --c-border: #555;
  --c-border-strong: #777;
  --c-divider: #2a2a2a;
  --c-text: #ffffff;
  --c-text-muted: #e8e8e8;
  --c-text-soft: #c8c8c8;
  --c-text-faint: #a0a0a0;
  --c-primary: #4d9bff;
  --c-primary-soft: #002850;
  --c-primary-text: #b8d6ff;
  --c-success: #2bff7a;
  --c-success-soft: #003015;
  --c-success-text: #80ffb0;
  --c-danger: #ff5050;
  --c-danger-soft: #380000;
  --c-danger-text: #ffaaaa;
  --c-warn: #ffaa00;
  --c-warn-soft: #2a1a00;
  --c-warn-text: #ffd980;
  color-scheme: dark;
}
:root[data-theme="hi-contrast"] :focus-visible {
  outline: 3px solid #ffff00 !important;
  outline-offset: 2px !important;
}

/* Mobile (Sprint UI-3.4) — action-bar already docked, just enlarge tap targets */
@media (max-width: 700px) {
  .action-bar {
    padding: var(--s-3) var(--s-4);
    box-shadow: 0 -8px 24px rgba(0,0,0,0.12);
  }
  .ab-btn, .verdict-control .ab-btn {
    min-height: 44px;       /* Apple-HIG tap target */
    padding: 0 var(--s-5);
  }
  .topbar-btn { min-height: 44px; min-width: 44px; }
  .canvas-toolbar button { min-height: 36px; min-width: 36px; }
}

/* Command palette (Sprint UI-2.6) */
.cmdk-bg {
  position: fixed; inset: 0;
  background: var(--c-overlay);
  display: none; align-items: flex-start; justify-content: center;
  z-index: 400;
  padding-top: 12vh;
  backdrop-filter: blur(6px);
  -webkit-backdrop-filter: blur(6px);
  animation: modal-bg-in var(--d-base) var(--ease-std);
}
.cmdk-bg.show { display: flex; }
.cmdk {
  background: var(--c-surface);
  border: 1px solid var(--c-border);
  border-radius: var(--r-xl);
  width: 640px; max-width: calc(100vw - var(--s-8));
  box-shadow: var(--e-5);
  overflow: hidden;
  animation: modal-in var(--d-slow) var(--ease-out);
}
.cmdk-input-wrap {
  display: flex; align-items: center; gap: var(--s-3);
  padding: var(--s-5) var(--s-6);
  border-bottom: 1px solid var(--c-divider);
}
.cmdk-input-wrap .ico { color: var(--c-text-faint); flex-shrink: 0; }
.cmdk-input {
  flex: 1; border: 0; background: transparent;
  font-size: var(--t-lg);
  color: var(--c-text);
  outline: none;
  padding: 0;
  height: 28px;
}
.cmdk-input::placeholder { color: var(--c-text-faint); }
.cmdk-hint {
  font-size: var(--t-xs); color: var(--c-text-faint);
  display: inline-flex; gap: var(--s-2); align-items: center;
}
.cmdk-list {
  max-height: 50vh; overflow: auto;
  padding: var(--s-2);
}
.cmdk-section {
  font-size: var(--t-xs); font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.06em;
  color: var(--c-text-faint);
  padding: var(--s-4) var(--s-5) var(--s-2);
}
.cmdk-item {
  display: flex; align-items: center; gap: var(--s-4);
  padding: var(--s-3) var(--s-5);
  cursor: pointer;
  border-radius: var(--r-md);
  font-size: var(--t-base);
  color: var(--c-text);
  transition: background var(--d-fast) var(--ease-std);
}
.cmdk-item:hover { background: var(--c-surface-2); }
.cmdk-item.cmdk-active {
  background: var(--c-primary-soft); color: var(--c-primary-text);
}
.cmdk-item .ico { color: var(--c-text-soft); }
.cmdk-item.cmdk-active .ico { color: var(--c-primary); }
.cmdk-item-main { flex: 1; min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.cmdk-item-meta { font-size: var(--t-xs); color: var(--c-text-faint); font-family: var(--f-mono); }
.cmdk-item-shortcut { display: inline-flex; gap: 2px; }
.cmdk-empty {
  padding: var(--s-9) var(--s-7);
  text-align: center;
  color: var(--c-text-faint);
  font-size: var(--t-base);
}

/* Toast action button (Undo) */
.toast-action {
  display: inline-block; margin-top: var(--s-3);
  background: transparent; border: 1px solid var(--c-border);
  color: var(--c-primary);
  padding: 3px var(--s-4);
  border-radius: var(--r-sm);
  font-size: var(--t-sm); font-weight: 600;
  cursor: pointer;
  transition: all var(--d-fast) var(--ease-std);
}
.toast-action:hover { background: var(--c-primary-soft); border-color: var(--c-primary); }

/* Onboarding overlay (Sprint UI-3.1) */
.onboard-bg {
  position: fixed; inset: 0;
  background: rgba(15,23,42,0.65);
  z-index: 500;
  display: none; align-items: center; justify-content: center;
  backdrop-filter: blur(2px);
}
.onboard-bg.show { display: flex; animation: modal-bg-in var(--d-base) var(--ease-std); }
.onboard-card {
  background: var(--c-surface);
  border-radius: var(--r-xl);
  padding: var(--s-9);
  max-width: 480px;
  box-shadow: var(--e-5);
  border: 1px solid var(--c-border);
  animation: modal-in var(--d-slow) var(--ease-out);
}
.onboard-card h2 {
  margin: 0 0 var(--s-3); font-size: var(--t-xl); font-weight: 700;
}
.onboard-card p {
  margin: var(--s-3) 0;
  color: var(--c-text-muted);
  line-height: var(--lh-loose);
}
.onboard-shortcuts {
  display: grid; grid-template-columns: 1fr 1fr; gap: var(--s-3);
  margin: var(--s-5) 0;
  font-size: var(--t-sm);
}
.onboard-shortcuts li {
  display: flex; gap: var(--s-3); align-items: center;
  list-style: none;
}
.onboard-shortcuts .kbd { flex-shrink: 0; }
.onboard-actions {
  display: flex; justify-content: flex-end; gap: var(--s-3);
  margin-top: var(--s-6);
}

/* Topbar Settings dropdown (Sprint UI-2.5) — position: fixed escapes
   the topbar's stacking context (z-index: 10 was clipped behind the
   docked action-bar at z=60). JS positions it relative to the gear
   button + clamps inside the viewport with extra clearance for
   browser chrome / overlay extensions at the top. */
.topbar-menu-btn { position: relative; }
.topbar-menu {
  position: fixed;
  min-width: 240px; max-width: calc(100vw - 16px);
  max-height: calc(100vh - 80px);   /* never overlap browser top bar */
  overflow-y: auto;                  /* scroll if too tall */
  background: var(--c-surface);
  border: 1px solid var(--c-border);
  border-radius: var(--r-lg);
  box-shadow: var(--e-4);
  padding: 4px;
  z-index: 350;       /* above modals (200), action-bar (60), sync-banner (150) */
  animation: menu-pop var(--d-fast) var(--ease-out);
  display: none;
}
.topbar-menu.show { display: block; }
.topbar-menu button {
  display: flex; width: 100%; align-items: center; gap: var(--s-3);
  padding: var(--s-3) var(--s-5);
  background: transparent; border: 0;
  text-align: left;
  color: var(--c-text);
  font-size: var(--t-base);
  border-radius: var(--r-sm);
  cursor: pointer;
}
.topbar-menu button:hover { background: var(--c-surface-2); }
.topbar-menu button .ico { color: var(--c-text-soft); }
.topbar-menu .menu-sep { height: 1px; background: var(--c-divider); margin: 4px 0; }

/* Diff toast (Sprint S2 — before/after Col D) */
.toast.diff .body { font-family: var(--f-mono); font-size: var(--t-xs); }
.toast.diff .diff-old {
  display: block; color: var(--c-danger-text);
  text-decoration: line-through;
  opacity: 0.85; margin-bottom: 2px;
}
.toast.diff .diff-new {
  display: block; color: var(--c-success-text);
  font-weight: 600;
}

/* Apply to siblings dialog */
.siblings-dialog {
  background: var(--c-success-soft);
  border: 1px solid var(--c-success);
  border-radius: var(--r-md);
  padding: var(--s-5);
  margin: var(--s-5) 0;
  font-size: var(--t-sm);
  color: var(--c-success-text);
}
.siblings-dialog strong { color: var(--c-success-text); font-weight: 700; }
.siblings-dialog ul {
  margin: var(--s-3) 0;
  padding-left: var(--s-7);
}
.siblings-dialog .sibling-actions {
  display: flex; gap: var(--s-3); margin-top: var(--s-4);
}

/* Wizard skill-md tip (Sprint S3.6) */
.wiz-skill-tip {
  background: var(--c-info-soft);
  border-left: 3px solid var(--c-info);
  padding: var(--s-3) var(--s-5);
  border-radius: 0 var(--r-sm) var(--r-sm) 0;
  font-size: var(--t-xs);
  color: var(--c-info-text);
  margin-top: var(--s-3);
}
.wiz-skill-tip strong { font-weight: 700; }

/* ── Print (subtle) ─────────────────────────────────────────────── */
@media print {
  .topbar, .action-bar, .mobile-tabs, .toast-stack, .kbd-help, .canvas-toolbar { display: none !important; }
  #app { display: block; height: auto; }
  .pane { border: none; height: auto; }
  body { background: white; }
}
</style>
</head>
<body>

<!-- ============================================================
     Inline SVG icon sprite (Lucide-derived, stroke 2 / 24x24).
     Use via <svg class="ico"><use href="#i-..."/></svg> or ico('name').
     ============================================================ -->
<svg width="0" height="0" style="position:absolute;width:0;height:0;overflow:hidden" aria-hidden="true">
  <defs>
    <symbol id="i-check" viewBox="0 0 24 24"><path d="M20 6L9 17l-5-5"/></symbol>
    <symbol id="i-x" viewBox="0 0 24 24"><path d="M18 6L6 18M6 6l12 12"/></symbol>
    <symbol id="i-alert" viewBox="0 0 24 24"><path d="M12 9v4M12 17h.01M10.29 3.86L1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"/></symbol>
    <symbol id="i-skip" viewBox="0 0 24 24"><path d="M5 4l10 8-10 8V4zM19 5v14"/></symbol>
    <symbol id="i-pin" viewBox="0 0 24 24"><path d="M21 10c0 7-9 13-9 13s-9-6-9-13a9 9 0 0 1 18 0z"/><circle cx="12" cy="10" r="3"/></symbol>
    <symbol id="i-sparkles" viewBox="0 0 24 24"><path d="M12 3l2 5 5 2-5 2-2 5-2-5-5-2 5-2zM19 14l1 2 2 1-2 1-1 2-1-2-2-1 2-1zM5 16l1 2 2 1-2 1-1 2-1-2-2-1 2-1z"/></symbol>
    <symbol id="i-refresh" viewBox="0 0 24 24"><path d="M23 4v6h-6M1 20v-6h6"/><path d="M3.51 9a9 9 0 0 1 14.85-3.36L23 10M20.49 15a9 9 0 0 1-14.85 3.36L1 14"/></symbol>
    <symbol id="i-pencil" viewBox="0 0 24 24"><path d="M11 4H4a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2v-7"/><path d="M18.5 2.5a2.121 2.121 0 0 1 3 3L12 15l-4 1 1-4 9.5-9.5z"/></symbol>
    <symbol id="i-trash" viewBox="0 0 24 24"><path d="M3 6h18M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"/></symbol>
    <symbol id="i-camera" viewBox="0 0 24 24"><path d="M23 19a2 2 0 0 1-2 2H3a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h4l2-3h6l2 3h4a2 2 0 0 1 2 2z"/><circle cx="12" cy="13" r="4"/></symbol>
    <symbol id="i-eye" viewBox="0 0 24 24"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"/><circle cx="12" cy="12" r="3"/></symbol>
    <symbol id="i-file" viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><path d="M14 2v6h6M16 13H8M16 17H8M10 9H8"/></symbol>
    <symbol id="i-book" viewBox="0 0 24 24"><path d="M2 3h6a4 4 0 0 1 4 4v14a3 3 0 0 0-3-3H2zM22 3h-6a4 4 0 0 0-4 4v14a3 3 0 0 1 3-3h7z"/></symbol>
    <symbol id="i-chart" viewBox="0 0 24 24"><path d="M18 20V10M12 20V4M6 20v-6"/></symbol>
    <symbol id="i-brain" viewBox="0 0 24 24"><path d="M12 5a3 3 0 0 0-5.99-.16A3 3 0 0 0 4 8a3 3 0 0 0 0 6 3 3 0 0 0 2 3 3 3 0 0 0 6 0V5zM12 5a3 3 0 0 1 5.99-.16A3 3 0 0 1 20 8a3 3 0 0 1 0 6 3 3 0 0 1-2 3 3 3 0 0 1-6 0"/></symbol>
    <symbol id="i-search" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.35-4.35"/></symbol>
    <symbol id="i-settings" viewBox="0 0 24 24"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09a1.65 1.65 0 0 0-1-1.51 1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09a1.65 1.65 0 0 0 1.51-1 1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 2.83-2.83l.06.06a1.65 1.65 0 0 0 1.82.33h.04a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82v.04a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1z"/></symbol>
    <symbol id="i-sun" viewBox="0 0 24 24"><circle cx="12" cy="12" r="4"/><path d="M12 2v2M12 20v2M4.93 4.93l1.41 1.41M17.66 17.66l1.41 1.41M2 12h2M20 12h2M4.93 19.07l1.41-1.41M17.66 6.34l1.41-1.41"/></symbol>
    <symbol id="i-moon" viewBox="0 0 24 24"><path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z"/></symbol>
    <symbol id="i-chevron-right" viewBox="0 0 24 24"><path d="M9 18l6-6-6-6"/></symbol>
    <symbol id="i-chevron-down" viewBox="0 0 24 24"><path d="M6 9l6 6 6-6"/></symbol>
    <symbol id="i-arrow-left" viewBox="0 0 24 24"><path d="M19 12H5M12 19l-7-7 7-7"/></symbol>
    <symbol id="i-arrow-right" viewBox="0 0 24 24"><path d="M5 12h14M12 5l7 7-7 7"/></symbol>
    <symbol id="i-plus" viewBox="0 0 24 24"><path d="M12 5v14M5 12h14"/></symbol>
    <symbol id="i-minus" viewBox="0 0 24 24"><path d="M5 12h14"/></symbol>
    <symbol id="i-undo" viewBox="0 0 24 24"><path d="M3 7v6h6"/><path d="M21 17a9 9 0 0 0-15-6.7L3 13"/></symbol>
    <symbol id="i-redo" viewBox="0 0 24 24"><path d="M21 7v6h-6"/><path d="M3 17a9 9 0 0 1 15-6.7L21 13"/></symbol>
    <symbol id="i-save" viewBox="0 0 24 24"><path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2z"/><path d="M17 21v-8H7v8M7 3v5h8"/></symbol>
    <symbol id="i-help" viewBox="0 0 24 24"><circle cx="12" cy="12" r="10"/><path d="M9.09 9a3 3 0 0 1 5.83 1c0 2-3 3-3 3"/><path d="M12 17h.01"/></symbol>
    <symbol id="i-folder" viewBox="0 0 24 24"><path d="M22 19a2 2 0 0 1-2 2H4a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h5l2 3h9a2 2 0 0 1 2 2z"/></symbol>
    <symbol id="i-clock" viewBox="0 0 24 24"><circle cx="12" cy="12" r="10"/><path d="M12 6v6l4 2"/></symbol>
    <symbol id="i-zap" viewBox="0 0 24 24"><path d="M13 2L3 14h9l-1 8 10-12h-9l1-8z"/></symbol>
    <symbol id="i-package" viewBox="0 0 24 24"><path d="M21 8.5l-9-5.5-9 5.5L12 14zM3 8.5v7l9 5.5 9-5.5v-7M12 14v7"/></symbol>
    <symbol id="i-target" viewBox="0 0 24 24"><circle cx="12" cy="12" r="10"/><circle cx="12" cy="12" r="6"/><circle cx="12" cy="12" r="2"/></symbol>
    <symbol id="i-external" viewBox="0 0 24 24"><path d="M18 13v6a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V8a2 2 0 0 1 2-2h6"/><path d="M15 3h6v6M10 14L21 3"/></symbol>
    <symbol id="i-rotate" viewBox="0 0 24 24"><path d="M1 4v6h6"/><path d="M3.51 15a9 9 0 1 0 2.13-9.36L1 10"/></symbol>
    <symbol id="i-square" viewBox="0 0 24 24"><rect x="3" y="3" width="18" height="18" rx="2"/></symbol>
    <symbol id="i-cursor" viewBox="0 0 24 24"><path d="M3 3l7.5 18 2.5-7 7-2.5z"/></symbol>
    <symbol id="i-text" viewBox="0 0 24 24"><path d="M4 7V5h16v2M9 19h6M12 5v14"/></symbol>
    <symbol id="i-copy" viewBox="0 0 24 24"><rect x="9" y="9" width="13" height="13" rx="2"/><path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"/></symbol>
  </defs>
</svg>

<!-- Sync banner (shown when working dir differs from latest snapshot) -->
<div id="sync-banner" class="sync-banner" role="status" aria-live="polite">
  <span id="sync-banner-msg"></span>
  <span id="sync-banner-actions"></span>
  <span class="close-x" onclick="dismissSyncBanner()" role="button" tabindex="0" aria-label="ปิด banner">✕</span>
</div>

<div id="app">

  <!-- ───── TOPBAR ──────────────────────────────────────────────── -->
  <header class="topbar" role="banner">
    <div class="topbar-brand">
      <span class="logo" aria-hidden="true"><svg class="ico" viewBox="0 0 24 24"><path d="M3 7l9-4 9 4-9 4-9-4z" fill="currentColor" stroke="none"/><path d="M3 12l9 4 9-4M3 17l9 4 9-4" opacity="0.5"/></svg></span>
      <span>Comply <span class="sub">Smart Plant 1</span></span>
    </div>
    <!-- Current row context (clickable → focuses tree) -->
    <button class="topbar-context" id="topbar-context" onclick="focusSelectedRow()" title="คลิกเพื่อโฟกัส row ปัจจุบันใน tree" style="display:none">
      <span class="topbar-context-label" id="topbar-context-label"></span>
    </button>
    <div class="topbar-spacer"></div>
    <div class="topbar-actions">
      <button class="topbar-btn" onclick="openCmdK()" title="Command palette (⌘K)" aria-label="open command palette"><svg class="ico" aria-hidden="true"><use href="#i-search"/></svg><span class="kbd" style="margin-left:4px">⌘K</span></button>
      <span class="stats-pill" id="stats-pill" role="status" aria-label="overall progress">
        <span id="stats-pill-progress">—</span>
      </span>
      <button class="topbar-btn icon-only" onclick="toggleTheme()" title="สลับธีม light/dark" aria-label="สลับธีม light/dark" id="theme-toggle"><svg class="ico" aria-hidden="true"><use href="#i-moon"/></svg></button>
      <button class="topbar-btn" onclick="showLearning()" title="HITL learning"><svg class="ico" aria-hidden="true"><use href="#i-brain"/></svg><span>Learn</span></button>
      <button class="topbar-btn" onclick="showSettings()" title="Settings"><svg class="ico" aria-hidden="true"><use href="#i-settings"/></svg><span>Settings</span></button>
      <div class="topbar-menu-btn">
        <button class="topbar-btn icon-only" onclick="toggleTopbarMenu(event)" title="More" aria-label="more actions" aria-haspopup="true">
          <svg class="ico" aria-hidden="true" viewBox="0 0 24 24" stroke-width="2.4">
            <circle cx="12" cy="5" r="1" fill="currentColor"/>
            <circle cx="12" cy="12" r="1" fill="currentColor"/>
            <circle cx="12" cy="19" r="1" fill="currentColor"/>
          </svg>
        </button>
        <div class="topbar-menu" id="topbar-menu" role="menu">
          <button onclick="closeTopbarMenu();showAudit()" role="menuitem"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-chart"/></svg><span>Database &amp; Audit</span></button>
          <button onclick="closeTopbarMenu();showVersions()" role="menuitem"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-package"/></svg><span>Project versions</span></button>
          <div class="menu-sep"></div>
          <button onclick="closeTopbarMenu();showSettings()" role="menuitem"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-settings"/></svg><span>Settings</span></button>
          <button onclick="closeTopbarMenu();showOnboarding(true)" role="menuitem"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-help"/></svg><span>Help &amp; shortcuts</span></button>
        </div>
      </div>
    </div>
  </header>

  <!-- Mobile tab navigation (replaces topbar at <700px) -->
  <nav class="mobile-tabs" role="tablist" aria-label="pane navigation">
    <button class="active" data-tab="tree"   onclick="setMobileTab('tree')"   role="tab" aria-selected="true"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-folder"/></svg><span>Tree</span></button>
    <button              data-tab="center" onclick="setMobileTab('center')" role="tab" aria-selected="false"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-file"/></svg><span>TOR/xlsx</span></button>
    <button              data-tab="pdf"    onclick="setMobileTab('pdf')"    role="tab" aria-selected="false"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-book"/></svg><span>Catalog</span></button>
  </nav>

  <!-- ───── LEFT: Activity Rail ─────────────────────────────────── -->
  <aside class="activity-rail" role="navigation" aria-label="primary navigation">
    <button class="rail-btn active" data-panel="tree" onclick="setRailPanel('tree')" title="Row tree" aria-label="row tree"><svg class="ico" aria-hidden="true"><use href="#i-folder"/></svg></button>
    <button class="rail-btn" data-panel="search" onclick="openCmdK()" title="Search · ⌘K" aria-label="search"><svg class="ico" aria-hidden="true"><use href="#i-search"/></svg></button>
    <button class="rail-btn" data-panel="catalogs" onclick="openCatalogBrowser()" title="Catalog library" aria-label="catalogs"><svg class="ico" aria-hidden="true"><use href="#i-book"/></svg></button>
    <button class="rail-btn" data-panel="learn" onclick="setRailPanel('learn')" title="Learning patterns" aria-label="learn"><svg class="ico" aria-hidden="true"><use href="#i-brain"/></svg></button>
    <button class="rail-btn" data-panel="versions" onclick="setRailPanel('versions')" title="Project versions" aria-label="versions"><svg class="ico" aria-hidden="true"><use href="#i-package"/></svg></button>
    <button class="rail-btn" data-panel="audit" onclick="setRailPanel('audit')" title="Database & audit" aria-label="audit"><svg class="ico" aria-hidden="true"><use href="#i-chart"/></svg></button>
    <span class="rail-spacer"></span>
    <button class="rail-btn" onclick="toggleAiPane()" title="Toggle AI pane" aria-label="AI pane"><svg class="ico" aria-hidden="true"><use href="#i-sparkles"/></svg></button>
    <span class="rail-sep"></span>
    <button class="rail-btn" onclick="toggleTheme()" title="Toggle theme" aria-label="theme" id="rail-theme"><svg class="ico" aria-hidden="true"><use href="#i-moon"/></svg></button>
    <button class="rail-btn" onclick="showSettings()" title="Settings" aria-label="settings"><svg class="ico" aria-hidden="true"><use href="#i-settings"/></svg></button>
    <button class="rail-btn" onclick="showOnboarding(true)" title="Help · ?" aria-label="help"><svg class="ico" aria-hidden="true"><use href="#i-help"/></svg></button>
  </aside>

  <!-- Rail panel host — content swaps based on data-rail-panel -->
  <div class="rail-panel" role="region" aria-label="rail panel">
    <div class="rail-panel-head">
      <span class="panel-title" id="rail-panel-title">Panel</span>
      <button class="panel-close" onclick="setRailPanel('tree')" aria-label="close"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button>
    </div>
    <div class="rail-panel-body" id="rail-panel-body"></div>
  </div>

  <!-- ───── Context Ribbon (mode tabs + sub-toolbar) ─────────────── -->
  <div class="context-ribbon" role="toolbar" aria-label="mode and tools">
    <div class="ribbon-tabs" role="tablist" aria-label="modes">
      <button class="ribbon-tab active" data-mode="verify" onclick="setMode('verify')" role="tab" aria-selected="true"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-eye"/></svg><span>Verify</span></button>
      <button class="ribbon-tab" data-mode="edit" onclick="setMode('edit')" role="tab" aria-selected="false"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-pencil"/></svg><span>Edit</span></button>
      <button class="ribbon-tab" data-mode="reannotate" onclick="setMode('reannotate')" role="tab" aria-selected="false"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-refresh"/></svg><span>Re-annotate</span></button>
      <button class="ribbon-tab" data-mode="apply" onclick="setMode('apply')" role="tab" aria-selected="false"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-sparkles"/></svg><span>Apply Auto</span></button>
    </div>
    <!-- Verify mode: navigation + highlight + DPI -->
    <div class="ribbon-mode-bar mode-verify">
      <div class="rb-group">
        <button onclick="pdfPrev()" title="Catalog page prev · [" aria-label="prev page"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-arrow-left"/></svg></button>
        <span class="rb-info" id="ribbon-pdf-page">— / —</span>
        <button onclick="pdfNext()" title="Catalog page next · ]" aria-label="next page"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-arrow-right"/></svg></button>
      </div>
      <div class="rb-sep"></div>
      <div class="rb-group">
        <button onclick="pdfZoom(-1)" title="Zoom out" aria-label="zoom out"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-minus"/></svg></button>
        <button onclick="pdfZoom(1)" title="Zoom in" aria-label="zoom in"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-plus"/></svg></button>
      </div>
      <div class="rb-sep"></div>
      <label title="Highlight matched text"><input type="checkbox" id="ribbon-hl-toggle" checked onchange="renderPdf()"> Highlight</label>
      <span class="rb-spacer"></span>
      <span class="rb-info">บริบท ±<span id="ribbon-ctx-radius">6</span> rows</span>
      <div class="rb-group">
        <button onclick="ctxRadius(-2)" title="Reduce context">−2</button>
        <button onclick="ctxRadius(2)"  title="More context">+2</button>
      </div>
    </div>
    <!-- Edit mode: tools -->
    <div class="ribbon-mode-bar mode-edit">
      <div class="rb-group">
        <button class="active" data-tool="select" onclick="setTool('select')" title="Select / move · V"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-cursor"/></svg></button>
        <button data-tool="drawRect" onclick="setTool('drawRect')" title="Draw rect · R"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-square"/></svg></button>
        <button data-tool="addText" onclick="setTool('addText')" title="Add text · T"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-text"/></svg></button>
        <button onclick="deleteSelected()" title="Delete selected · Del"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-trash"/></svg></button>
      </div>
      <div class="rb-sep"></div>
      <div class="rb-group">
        <button id="ribbon-undo-btn" onclick="undo()" disabled title="Undo · ⌘Z"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-undo"/></svg></button>
        <button id="ribbon-redo-btn" onclick="redo()" disabled title="Redo · ⇧⌘Z"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-redo"/></svg></button>
      </div>
      <div class="rb-sep"></div>
      <button onclick="showHistory()" title="Catalog edit history"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-clock"/></svg><span>History</span></button>
      <span class="rb-spacer"></span>
      <span class="rb-info" id="ribbon-dirty-ind"></span>
      <button id="ribbon-save-btn" onclick="saveEdits()" disabled title="Save · ⌘S" style="background:var(--c-success);color:white;border-color:var(--c-success);font-weight:600"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-save"/></svg><span>Save</span></button>
    </div>
    <!-- Re-annotate mode: starts wizard if not already in it -->
    <div class="ribbon-mode-bar mode-reannotate">
      <span class="rb-info">Re-annotate the selected row's catalog. Click <strong>Start</strong> to launch the wizard.</span>
      <span class="rb-spacer"></span>
      <button onclick="startReannotate()" title="Start wizard"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-refresh"/></svg><span>Start wizard</span></button>
    </div>
    <!-- Apply Auto mode: AI proposal control -->
    <div class="ribbon-mode-bar mode-apply">
      <span class="rb-info" id="ribbon-llm-summary">Claude · checking…</span>
      <span class="rb-spacer"></span>
      <button onclick="showAutoAnnotate()" title="Preview AI proposal"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-sparkles"/></svg><span>Preview proposal</span></button>
      <button onclick="aiPaneRefresh()" title="Refresh AI pane"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-refresh"/></svg></button>
    </div>
  </div>

  <!-- ───── RIGHT: AI Pane ──────────────────────────────────────── -->
  <aside class="ai-pane" id="ai-pane" role="complementary" aria-label="AI assistant">
    <div class="ai-pane-head" id="ai-pane-head" data-status="offline">
      <span class="ai-status-dot"></span>
      <div class="ai-title">
        <span class="name">Claude</span>
        <span class="sub" id="ai-pane-sub">offline · paste API key in Settings</span>
      </div>
      <button class="ai-collapse" onclick="toggleAiPane()" title="Hide AI pane" aria-label="hide AI pane"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button>
    </div>
    <div class="ai-pane-body" id="ai-pane-body">
      <div class="ai-empty">
        <svg class="ico ico-xl" aria-hidden="true" style="opacity:0.3;margin-bottom:var(--s-3)"><use href="#i-sparkles"/></svg>
        <div>เลือก row เพื่อเริ่ม</div>
      </div>
    </div>
  </aside>

  <!-- ───── BOTTOM: Status Bar ──────────────────────────────────── -->
  <footer class="status-bar" role="status">
    <span class="sb-section sb-row-info" id="sb-row-info">No row selected</span>
    <span class="sb-sep"></span>
    <!-- Phase A6: verdict pills moved here from action bar -->
    <div class="sb-verdict" id="sb-verdict" role="radiogroup" aria-label="verdict (1-4)">
      <button class="sb-vbtn pass" role="radio" aria-checked="false" data-verdict="pass" type="button" onclick="setStatus('pass')" title="ผ่าน · 1" aria-label="ผ่าน">
        <svg class="ico ico-sm" aria-hidden="true"><use href="#i-check"/></svg><span>ผ่าน</span><span class="kbd" aria-hidden="true">1</span>
      </button>
      <button class="sb-vbtn fail" role="radio" aria-checked="false" data-verdict="fail" type="button" onclick="setStatus('fail')" title="ไม่ผ่าน · 2" aria-label="ไม่ผ่าน">
        <svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg><span>ไม่ผ่าน</span><span class="kbd" aria-hidden="true">2</span>
      </button>
      <button class="sb-vbtn fix" role="radio" aria-checked="false" data-verdict="need_fix" type="button" onclick="setStatus('need_fix')" title="ต้องแก้ · 3" aria-label="ต้องแก้">
        <svg class="ico ico-sm" aria-hidden="true"><use href="#i-alert"/></svg><span>แก้</span><span class="kbd" aria-hidden="true">3</span>
      </button>
      <button class="sb-vbtn skip" role="radio" aria-checked="false" data-verdict="skip" type="button" onclick="setStatus('skip')" title="ข้าม · 4" aria-label="ข้าม">
        <svg class="ico ico-sm" aria-hidden="true"><use href="#i-skip"/></svg><span>ข้าม</span><span class="kbd" aria-hidden="true">4</span>
      </button>
      <button class="sb-vbtn reset" type="button" onclick="setStatus('unverified')" title="reset verdict" aria-label="reset verdict">
        <svg class="ico ico-sm" aria-hidden="true"><use href="#i-rotate"/></svg>
      </button>
    </div>
    <span class="sb-sep"></span>
    <span class="sb-section sb-progress" id="sb-progress">
      <strong id="sb-done">0</strong>/<strong id="sb-total">0</strong>
      <span class="bar"><span id="sb-bar" style="width:0"></span></span>
    </span>
    <span class="sb-spacer"></span>
    <span class="sb-section sb-claude offline" id="sb-claude">
      <span class="dot"></span>
      <span id="sb-claude-text">Claude offline</span>
    </span>
    <span class="sb-sep"></span>
    <span class="sb-section">
      <span id="sb-save-state" title="auto-save state">●</span>
      <span style="color:var(--c-text-faint)">saved</span>
    </span>
  </footer>

  <!-- ───── LEFT: Tree ──────────────────────────────────────────── -->
  <section class="pane tree-pane" aria-label="row tree">
    <h2>
      <span class="pane-title"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-folder"/></svg><span>โครงสร้าง</span> <span id="tree-count" style="font-weight:400;color:var(--c-text-faint)"></span></span>
    </h2>
    <div class="stats-bar" id="stats" aria-live="polite"></div>
    <div class="toolbar" role="search">
      <div class="search-wrap">
        <svg class="search-ico ico ico-sm" aria-hidden="true"><use href="#i-search"/></svg>
        <input type="search" id="tree-search" placeholder="ค้น section / row / spec…" aria-label="ค้นหา row">
      </div>
      <div class="filter-row">
        <select id="filter-status" aria-label="กรอง status">
          <option value="">ทุกสถานะ</option>
          <option value="unverified">ยังไม่ตรวจ</option>
          <option value="pass">ผ่าน</option>
          <option value="fail">ไม่ผ่าน</option>
          <option value="need_fix">ต้องแก้</option>
          <option value="skip">ข้าม</option>
        </select>
        <select id="filter-vendor" aria-label="กรอง vendor">
          <option value="">ทุก vendor</option>
          <option value="SMART">SMART</option>
          <option value="TRIO">TRIO</option>
          <option value="SR">SR</option>
        </select>
      </div>
      <div class="filter-row">
        <label><input type="checkbox" id="filter-flags"> มี flag</label>
        <label><input type="checkbox" id="filter-haspdf"> มี catalog</label>
        <button class="mini-btn" onclick="expandAll()" title="แตกทุก section" aria-label="expand all">⊞</button>
        <button class="mini-btn" onclick="collapseAll()" title="หุบทุก section" aria-label="collapse all">⊟</button>
      </div>
    </div>
    <div class="tree-scroll" id="tree" role="tree" aria-label="rows"></div>
  </section>

  <!-- ───── CENTER: TOR top + xlsx bot ──────────────────────────── -->
  <section class="pane center-pane" aria-label="TOR and spreadsheet">
    <!-- Top: TOR -->
    <div class="split-top">
      <h2>
        <span class="pane-title"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-file"/></svg><span>TOR</span> <span style="font-weight:400;color:var(--c-text-faint)" id="tor-info"></span></span>
      </h2>
      <div class="canvas-toolbar" role="toolbar" aria-label="TOR navigation">
        <div class="tb-group">
          <button onclick="torPrev()" title="หน้าก่อน (,)" aria-label="หน้าก่อน"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-arrow-left"/></svg></button>
          <span class="page-ind" id="tor-page-info">— / —</span>
          <button onclick="torNext()" title="หน้าถัดไป (.)" aria-label="หน้าถัดไป"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-arrow-right"/></svg></button>
        </div>
        <div class="tb-sep"></div>
        <button onclick="torJumpToMatch()" title="ไปยังหน้าที่เจอ" aria-label="jump to match"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-target"/></svg><span>Match</span></button>
        <span class="info" id="tor-status">เลือก row เพื่อดู</span>
        <span class="tb-spacer"></span>
        <div class="tb-group">
          <button onclick="torZoom(-1)" title="ย่อ" aria-label="ย่อ"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-minus"/></svg></button>
          <button onclick="torZoom(1)" title="ขยาย" aria-label="ขยาย"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-plus"/></svg></button>
        </div>
      </div>
      <div class="tor-canvas" id="tor-canvas">
        <div class="empty-state">
          <div class="es-icon"><svg class="ico" aria-hidden="true"><use href="#i-file"/></svg></div>
          <div class="es-title">No row selected</div>
          <div class="es-body">เลือก row จากต้นไม้ด้านซ้ายเพื่อดูเนื้อหา TOR ที่เกี่ยวข้อง</div>
          <button class="es-cta" onclick="openCmdK()"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-search"/></svg><span>Browse rows (⌘K)</span></button>
        </div>
      </div>
    </div>

    <!-- Bottom: xlsx preview -->
    <div class="split-bot">
      <h2>
        <span class="pane-title"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-chart"/></svg><span>Comply.xlsx</span> <span style="font-weight:400;color:var(--c-text-faint)" id="xlsx-info"></span></span>
      </h2>
      <div class="canvas-toolbar" role="toolbar" aria-label="spreadsheet context">
        <span class="info">บริบท ±<span id="ctx-radius">6</span> rows</span>
        <div class="tb-group">
          <button onclick="ctxRadius(-2)" title="ลด context (−2 rows)" aria-label="ลด context">−2</button>
          <button onclick="ctxRadius(2)" title="เพิ่ม context (+2 rows)" aria-label="เพิ่ม context">+2</button>
        </div>
        <span class="tb-spacer"></span>
        <span class="info" style="font-style:italic">double-click Col D เพื่อแก้</span>
      </div>
      <div class="xlsx-wrap" id="xlsx-wrap">
        <div class="empty-state">
          <div class="es-icon"><svg class="ico" aria-hidden="true"><use href="#i-chart"/></svg></div>
          <div class="es-title">Comply spec context</div>
          <div class="es-body">เลือก row จากต้นไม้ด้านซ้าย จะแสดง ±6 rows รอบ row ที่เลือก</div>
        </div>
      </div>
    </div>
  </section>

  <!-- ───── RIGHT: Catalog PDF ──────────────────────────────────── -->
  <section class="pane pdf-pane" aria-label="catalog PDF">
    <h2>
      <span class="pane-title"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-book"/></svg><span>Catalog</span></span>
      <span class="filename" id="pdf-filename">(ไม่มี)</span>
      <button class="edit-toggle-btn" id="edit-toggle-btn" onclick="toggleEditMode()" title="Edit annotations" aria-pressed="false"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-pencil"/></svg><span>Edit</span></button>
    </h2>
    <div class="canvas-toolbar" role="toolbar" aria-label="catalog navigation">
      <div class="tb-group">
        <button onclick="pdfPrev()" title="หน้าก่อน ([)" aria-label="หน้าก่อน"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-arrow-left"/></svg></button>
        <span class="page-ind" id="pdf-page-info">— / —</span>
        <button onclick="pdfNext()" title="หน้าถัดไป (])" aria-label="หน้าถัดไป"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-arrow-right"/></svg></button>
      </div>
      <span class="tb-spacer"></span>
      <div class="tb-group">
        <button onclick="pdfZoom(-1)" title="ย่อ (−)" aria-label="ย่อ"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-minus"/></svg></button>
        <button onclick="pdfZoom(1)" title="ขยาย (+)" aria-label="ขยาย"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-plus"/></svg></button>
      </div>
      <div class="tb-sep"></div>
      <label title="ไฮไลต์เนื้อหาที่ตรงกับ row"><input type="checkbox" id="hl-toggle" checked onchange="renderPdf()"> highlight</label>
      <button onclick="openInBrowser()" title="เปิดใน browser tab ใหม่" aria-label="เปิดใน browser"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-external"/></svg></button>
    </div>
    <!-- Edit toolbar (visible only in edit mode) -->
    <div class="edit-toolbar canvas-toolbar" id="edit-toolbar" style="display:none;">
      <div class="tb-group" role="group" aria-label="tools">
        <button class="tool active" data-tool="select" onclick="setTool('select')" title="Select / move · V" aria-label="select tool"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-cursor"/></svg></button>
        <button class="tool" data-tool="drawRect" onclick="setTool('drawRect')" title="Draw rectangle · R" aria-label="draw rectangle"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-square"/></svg></button>
        <button class="tool" data-tool="addText" onclick="setTool('addText')" title="Add text · T" aria-label="add text"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-text"/></svg></button>
        <button onclick="deleteSelected()" title="Delete selected · Del" aria-label="delete"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-trash"/></svg></button>
      </div>
      <div class="tb-sep"></div>
      <div class="tb-group" role="group" aria-label="undo/redo">
        <button onclick="undo()" id="undo-btn" disabled title="Undo · ⌘Z" aria-label="undo"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-undo"/></svg></button>
        <button onclick="redo()" id="redo-btn" disabled title="Redo · ⇧⌘Z" aria-label="redo"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-redo"/></svg></button>
      </div>
      <div class="tb-sep"></div>
      <button onclick="showHistory()" title="Version history (per-PDF snapshots)"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-clock"/></svg><span>History</span></button>
      <span class="tb-spacer"></span>
      <span class="dirty-indicator" id="dirty-ind"></span>
      <button onclick="saveEdits()" id="save-btn" class="save-btn" disabled title="Save · ⌘S"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-save"/></svg><span>Save</span></button>
    </div>
    <!-- Re-annotate / Manual-annotate wizard banner -->
    <div class="manual-mode-banner" id="manual-banner" role="dialog" aria-label="annotate wizard">
      <div class="wiz-progress" id="wiz-progress" aria-hidden="true"></div>
      <div class="target-info" id="manual-target-info">—</div>
      <div class="wiz-actions">
        <button class="wiz-pdf-btn" id="wiz-pdf-btn" onclick="toggleWizPdfPicker(event)" title="เปลี่ยน catalog PDF">
          <svg class="ico ico-sm wiz-pdf-icon" aria-hidden="true"><use href="#i-file"/></svg>
          <span class="wiz-pdf-label" id="wiz-pdf-current">…</span>
          <svg class="ico ico-xs wiz-pdf-caret" aria-hidden="true"><use href="#i-chevron-down"/></svg>
        </button>
        <button class="back" id="wiz-back-btn" onclick="wizBack()" disabled><svg class="ico ico-sm" aria-hidden="true"><use href="#i-arrow-left"/></svg><span>Back</span></button>
        <button class="cancel" onclick="cancelManualAnnotate()"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg><span>Cancel</span></button>
        <button class="save" id="manual-save-btn" onclick="saveManualAnnotate()" disabled><svg class="ico ico-sm" aria-hidden="true"><use href="#i-check"/></svg><span>Save</span></button>
      </div>
    </div>
    <div class="pdf-canvas" id="pdf-canvas">
      <div class="empty-state">
        <div class="es-icon"><svg class="ico" aria-hidden="true"><use href="#i-book"/></svg></div>
        <div class="es-title">Catalog preview</div>
        <div class="es-body">เลือก row ที่อ้างอิง catalog ใน Col D — ระบบจะ highlight ตำแหน่งให้ตรงกับเนื้อหา</div>
      </div>
    </div>
    <div class="pdf-annots" id="pdf-annots" aria-label="annotations on current page"></div>
  </section>
</div>

<!-- Per-PDF history modal (catalog edits) -->
<div class="modal-bg" id="history-modal" role="dialog" aria-modal="true" aria-labelledby="history-modal-title" onclick="if(event.target.id==='history-modal') closeHistory()">
  <div class="modal">
    <h3 id="history-modal-title"><svg class="ico" aria-hidden="true"><use href="#i-clock"/></svg><span>Catalog edit history</span><button class="close" onclick="closeHistory()" aria-label="ปิด"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button></h3>
    <div id="history-info" style="font-size:var(--t-sm);color:var(--c-text-soft);margin-bottom:var(--s-4);"></div>
    <ul id="history-list" style="list-style:none;padding:0;margin:0;"></ul>
    <div style="margin-top:var(--s-6);display:flex;gap:var(--s-4);">
      <button onclick="manualSnapshot()" class="btn"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-camera"/></svg><span>Snapshot now</span></button>
      <button onclick="closeHistory()" class="btn">Close</button>
    </div>
  </div>
</div>

<!-- Project-level versions modal (snap, restore, diff via version.py) -->
<div class="modal-bg" id="versions-modal" role="dialog" aria-modal="true" aria-labelledby="versions-modal-title" onclick="if(event.target.id==='versions-modal') closeVersions()">
  <div class="modal versions-modal-body">
    <h3 id="versions-modal-title"><svg class="ico" aria-hidden="true"><use href="#i-package"/></svg><span>Project versions</span>
      <button class="close" onclick="closeVersions()" aria-label="ปิด"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button>
    </h3>
    <div id="versions-sync-badge"></div>
    <div id="versions-info" style="font-size:var(--t-sm);color:var(--c-text-soft);margin-bottom:var(--s-5);"></div>

    <div class="versions-snap-form">
      <input type="text" id="versions-tag" placeholder='tag (เช่น "before-rcbo-edit")' aria-label="snapshot tag">
      <button class="btn snap-quick" onclick="takeProjectSnap(false)"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-camera"/></svg><span>Quick snap</span></button>
      <button class="btn snap-full"  onclick="takeProjectSnap(true)" title="รวม output/ ทั้งหมด (~200 MB)"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-package"/></svg><span>Full snap</span></button>
      <button class="btn auto-snap" onclick="autoSnapNow()" title="snap เฉพาะถ้า xlsx เปลี่ยน"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-zap"/></svg><span>Auto</span></button>
    </div>

    <div class="versions-actions" id="versions-bulk-actions">
      <button class="btn" onclick="pruneVersions()"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-trash"/></svg><span>Prune (keep 10)</span></button>
      <button class="btn" onclick="loadVersions()"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-refresh"/></svg><span>Refresh</span></button>
      <span style="flex:1"></span>
      <span class="diff-helper" id="diff-helper">เลือก 2 snapshots เพื่อ diff</span>
    </div>

    <div class="versions-list" id="versions-list">
      <div style="color:var(--c-text-faint);padding:var(--s-8);text-align:center;">Loading…</div>
    </div>

    <div id="versions-busy" class="versions-busy" style="display:none;">
      <div class="spinner"></div>
      <div id="versions-busy-msg">working…</div>
    </div>

    <pre id="versions-output" class="versions-output" style="display:none;"></pre>
  </div>
</div>

<!-- HITL Learning modal -->
<div class="modal-bg" id="learn-modal" role="dialog" aria-modal="true" aria-labelledby="learn-modal-title" onclick="if(event.target.id==='learn-modal') closeLearning()">
  <div class="modal" style="min-width:780px;max-width:1000px;">
    <h3 id="learn-modal-title"><svg class="ico" aria-hidden="true"><use href="#i-brain"/></svg><span>HITL Learning</span>
      <button class="close" onclick="closeLearning()" aria-label="ปิด"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button>
    </h3>
    <div style="font-size:var(--t-sm);color:var(--c-text-soft);margin-bottom:var(--s-5);">
      Core proposes → user does visual proof → corrections become rules.
      Patterns ที่ user แก้ซ้ำ ≥ 2 ครั้ง จะถูก promote เป็น rule อัตโนมัติ.
    </div>

    <div class="audit-stats" id="learn-stats"></div>

    <div style="display:flex;gap:var(--s-4);align-items:center;margin:var(--s-6) 0 var(--s-3);">
      <strong style="font-size:var(--t-base);">Learned patterns</strong>
      <select id="learn-pattern-filter" onchange="loadLearnPatterns()" style="width:auto;min-width:140px;">
        <option value="">[all types]</option>
        <option value="filename_brand">filename_brand</option>
        <option value="section_vendor">section_vendor</option>
        <option value="row_format_d">row_format_d</option>
      </select>
      <span style="flex:1"></span>
      <button class="btn save-btn" onclick="runRetrain()"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-refresh"/></svg><span>Retrain now</span></button>
    </div>

    <div class="learn-pattern-list" id="learn-pattern-list"></div>

    <div id="learn-llm-row" style="margin-top:var(--s-6);font-size:var(--t-sm);color:var(--c-text-soft);">
      <strong>LLM provider:</strong> <span id="learn-llm-name">off</span>
      — เปิดใช้ทาง env var <code>COMPLY_LLM</code> (Anthropic / OpenAI / local Ollama)
    </div>

    <pre id="learn-output" class="versions-output" style="display:none;margin-top:var(--s-5);"></pre>
  </div>
</div>

<!-- Audit log + DB stats modal -->
<div class="modal-bg" id="audit-modal" role="dialog" aria-modal="true" aria-labelledby="audit-modal-title" onclick="if(event.target.id==='audit-modal') closeAudit()">
  <div class="modal" style="min-width:780px;max-width:1000px;">
    <h3 id="audit-modal-title"><svg class="ico" aria-hidden="true"><use href="#i-chart"/></svg><span>Database &amp; Audit</span>
      <button class="close" onclick="closeAudit()" aria-label="ปิด"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button>
    </h3>

    <!-- DB stats summary -->
    <div id="audit-stats" class="audit-stats"></div>

    <!-- Search bar (FTS5 over rows) -->
    <div class="audit-search-row">
      <input type="search" id="audit-search" placeholder="ค้นหาข้าม Col A/B/C/D/E (FTS5)…" oninput="onAuditSearch()" aria-label="full-text search">
      <select id="audit-action-filter" onchange="loadAudit()" aria-label="กรอง action">
        <option value="">[all actions]</option>
        <option value="status_change">status_change</option>
        <option value="notes_update">notes_update</option>
        <option value="auto_annotate_apply">auto_annotate_apply</option>
        <option value="pdf_edit">pdf_edit</option>
        <option value="snapshot">snapshot</option>
        <option value="restore">restore</option>
        <option value="refresh">refresh</option>
      </select>
    </div>

    <!-- FTS results (only when searching) -->
    <div id="audit-search-results" style="display:none;"></div>

    <!-- Audit log timeline -->
    <div id="audit-list-wrap">
      <h4 style="margin: var(--s-6) 0 var(--s-3); font-size: var(--t-base); color: var(--c-text-muted);">Recent activity</h4>
      <div class="audit-list" id="audit-list"></div>
    </div>
  </div>
</div>

<!-- Auto-annotate modal -->
<div class="modal-bg" id="auto-modal" role="dialog" aria-modal="true" aria-labelledby="auto-modal-title" onclick="if(event.target.id==='auto-modal') closeAuto()">
  <div class="modal" style="min-width:640px;max-width:800px;">
    <h3 id="auto-modal-title"><svg class="ico" aria-hidden="true"><use href="#i-sparkles"/></svg><span>Auto-annotate</span> <span id="auto-title" style="font-weight:400;color:var(--c-text-soft);font-size:var(--t-base);"></span>
      <button class="close" onclick="closeAuto()" aria-label="ปิด"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button>
    </h3>

    <!-- Per-row preview -->
    <div id="auto-single">
      <div id="auto-meta" style="font-size:var(--t-sm);color:var(--c-text-soft);margin-bottom:var(--s-4);"></div>
      <div id="auto-warn" class="versions-output" style="display:none;background:var(--c-warn-soft);color:var(--c-warn-text);"></div>

      <div class="auto-block">
        <div class="auto-label">Col C (proposed)</div>
        <pre class="auto-pre" id="auto-c"></pre>
      </div>
      <div class="auto-block">
        <div class="auto-label">Col D (proposed)</div>
        <pre class="auto-pre" id="auto-d"></pre>
      </div>
      <div class="auto-block">
        <div class="auto-label">PDF annotations <span id="auto-ann-count"></span></div>
        <pre class="auto-pre" id="auto-ann"></pre>
      </div>

      <div style="margin-top:var(--s-6);display:flex;gap:var(--s-4);align-items:center;flex-wrap:wrap;">
        <label style="font-size:var(--t-sm);"><input type="checkbox" id="auto-write-xlsx" checked> เขียน Col C/D ลง xlsx</label>
        <label style="font-size:var(--t-sm);"><input type="checkbox" id="auto-write-pdf" checked> เขียน rect+label ลง PDF</label>
        <span style="flex:1"></span>
        <button class="btn" onclick="closeAuto()">Cancel</button>
        <button class="btn" onclick="showBatchAuto()"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-package"/></svg><span>Batch mode…</span></button>
        <button class="btn save-btn" onclick="applyAutoSingle()">✓ Apply</button>
      </div>
    </div>

    <!-- Batch preview -->
    <div id="auto-batch" style="display:none;">
      <div id="auto-batch-info" style="font-size:var(--t-sm);color:var(--c-text-soft);margin-bottom:var(--s-3);"></div>
      <div class="auto-batch-list" id="auto-batch-list"></div>
      <div style="margin-top:var(--s-6);display:flex;gap:var(--s-4);align-items:center;flex-wrap:wrap;">
        <button class="btn" onclick="showSingleAuto()">← back to single</button>
        <span style="flex:1"></span>
        <button class="btn" onclick="closeAuto()">Cancel</button>
        <button class="btn save-btn" onclick="applyAutoBatch()" id="auto-batch-apply">✓ Apply selected</button>
      </div>
    </div>

    <pre id="auto-output" class="versions-output" style="display:none;margin-top:var(--s-5);"></pre>
  </div>
</div>

<!-- ─── Command palette (⌘K) — fuzzy row search + actions ─── -->
<div class="cmdk-bg" id="cmdk-bg" onclick="if(event.target.id==='cmdk-bg') closeCmdK()">
  <div class="cmdk" role="dialog" aria-modal="true" aria-labelledby="cmdk-input">
    <div class="cmdk-input-wrap">
      <svg class="ico" aria-hidden="true"><use href="#i-search"/></svg>
      <input type="text" id="cmdk-input" class="cmdk-input" placeholder="ค้น row, section, หรือ action…" autocomplete="off" spellcheck="false">
      <span class="cmdk-hint"><span class="kbd">esc</span></span>
    </div>
    <div class="cmdk-list" id="cmdk-list" role="listbox"></div>
  </div>
</div>

<!-- ─── Onboarding card (first launch) ─── -->
<div class="onboard-bg" id="onboard-bg" onclick="if(event.target.id==='onboard-bg') closeOnboarding()">
  <div class="onboard-card" role="dialog" aria-labelledby="onboard-title">
    <h2 id="onboard-title">Welcome to Comply Verify</h2>
    <p>คลิกแถวจาก tree ด้านซ้าย → ดู TOR + xlsx + catalog ในที่เดียว → ตัดสิน verdict ด้วย <span class="kbd">1</span>–<span class="kbd">4</span> หรือคลิกใน action bar ด้านล่าง</p>
    <ul class="onboard-shortcuts">
      <li><span class="kbd">⌘K</span><span>Command palette / ค้นทุกอย่าง</span></li>
      <li><span class="kbd">J</span><span class="kbd">K</span><span>row ถัดไป / ก่อนหน้า</span></li>
      <li><span class="kbd">N</span><span>row uncertain ถัดไป</span></li>
      <li><span class="kbd">1</span>–<span class="kbd">4</span><span>verdict pass/fail/fix/skip</span></li>
      <li><span class="kbd">[</span><span class="kbd">]</span><span>หน้า PDF</span></li>
      <li><span class="kbd">,</span><span class="kbd">.</span><span>หน้า TOR</span></li>
      <li><span class="kbd">?</span><span>เปิด help นี้</span></li>
      <li><span class="kbd">⌘S</span><span>save edits</span></li>
    </ul>
    <p style="font-size:var(--t-sm);color:var(--c-text-soft)">คลิก Col D เพื่อ Re-annotate / Auto-annotate / Edit. กด Settings ด้านขวาบนเพื่อปรับ theme + language</p>
    <div class="onboard-actions">
      <button class="btn" onclick="closeOnboarding()">Skip</button>
      <button class="btn btn-primary" onclick="closeOnboarding();dontShowOnboardingAgain()">Got it</button>
    </div>
  </div>
</div>

<!-- ─── Settings panel ─── -->
<div class="modal-bg" id="settings-modal" role="dialog" aria-modal="true" aria-labelledby="settings-modal-title" onclick="if(event.target.id==='settings-modal') closeSettings()">
  <div class="modal" style="min-width:480px;max-width:560px;">
    <h3 id="settings-modal-title"><svg class="ico" aria-hidden="true"><use href="#i-settings"/></svg><span>Settings</span><button class="close" onclick="closeSettings()" aria-label="ปิด"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button></h3>
    <div style="display:flex;flex-direction:column;gap:var(--s-5)">
      <label style="display:flex;justify-content:space-between;align-items:center;gap:var(--s-4)">
        <span><strong>Theme</strong> <span style="color:var(--c-text-soft);font-size:var(--t-sm);display:block">Light / Dark / System / High contrast</span></span>
        <select id="settings-theme" onchange="setThemeFromSettings(this.value)" style="width:auto;min-width:140px">
          <option value="">System</option>
          <option value="light">Light</option>
          <option value="dark">Dark</option>
          <option value="hi-contrast">High contrast</option>
        </select>
      </label>
      <div>
        <strong>LLM provider — Claude (Code or API)</strong>
        <span style="color:var(--c-text-soft);font-size:var(--t-sm);display:block;margin-bottom:var(--s-3)">Phase 1: <strong>Claude Code</strong> via OAuth uses your <strong>Claude Max</strong> subscription (no metered API charges). API key path remains as fallback.</span>
        <div id="settings-llm-card" style="background:var(--c-surface-2);border:1px solid var(--c-border);border-radius:var(--r-md);padding:var(--s-5);font-size:var(--t-sm);display:flex;flex-direction:column;gap:var(--s-4)">
          <div id="settings-llm-status" style="font-family:var(--f-mono)">loading…</div>
          <details style="font-size:var(--t-xs);color:var(--c-text-soft)">
            <summary style="cursor:pointer;color:var(--c-text-muted);font-weight:600">API key fallback (only if Claude Max not available)</summary>
          <div style="display:flex;flex-direction:column;gap:var(--s-2);margin-top:var(--s-3)">
            <label style="font-size:var(--t-sm);color:var(--c-text-muted);font-weight:600">API key</label>
            <div style="display:flex;gap:var(--s-2)">
              <input type="password" id="settings-api-key" placeholder="sk-ant-…" style="flex:1;font-family:var(--f-mono);font-size:var(--t-sm)" autocomplete="off" spellcheck="false">
              <button class="btn" type="button" onclick="toggleApiKeyVisibility()" id="settings-api-key-toggle" title="show/hide" style="padding:6px var(--s-4)"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-eye"/></svg></button>
            </div>
            <span style="font-size:var(--t-xs);color:var(--c-text-faint)">รับจาก <a href="https://console.anthropic.com/settings/keys" target="_blank" rel="noopener" style="color:var(--c-primary)">console.anthropic.com</a> · ห้ามแชร์ใน chat</span>
          </div>
          <div style="display:flex;gap:var(--s-3)">
            <label style="flex:1;display:flex;flex-direction:column;gap:var(--s-2)">
              <span style="font-size:var(--t-sm);color:var(--c-text-muted);font-weight:600">Model</span>
              <select id="settings-model" style="font-family:var(--f-mono);font-size:var(--t-sm)">
                <option value="claude-sonnet-4-5">Sonnet 4.5 (recommended, $3/M in)</option>
                <option value="claude-opus-4-5">Opus 4.5 ($15/M in, edge cases)</option>
                <option value="claude-haiku-4-5">Haiku 4.5 ($0.80/M in, fastest)</option>
              </select>
            </label>
            <label style="width:120px;display:flex;flex-direction:column;gap:var(--s-2)">
              <span style="font-size:var(--t-sm);color:var(--c-text-muted);font-weight:600">$/day cap</span>
              <input type="number" id="settings-budget" min="0" step="0.5" value="5" style="font-family:var(--f-mono);font-size:var(--t-sm)">
            </label>
          </div>
          <div style="display:flex;justify-content:flex-end;gap:var(--s-3);margin-top:var(--s-2)">
            <button class="btn btn-danger" type="button" onclick="clearApiKey()" id="settings-clear-key" style="font-size:var(--t-sm)">Clear key</button>
            <button class="btn btn-primary" type="button" onclick="saveApiKey()" id="settings-save-key" style="font-size:var(--t-sm)"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-save"/></svg><span>Save &amp; activate</span></button>
          </div>
          </details>
        </div>
      </div>
      <label style="display:flex;justify-content:space-between;align-items:center;gap:var(--s-4)">
        <span><strong>Reduce motion</strong> <span style="color:var(--c-text-soft);font-size:var(--t-sm);display:block">Disable animations &amp; transitions</span></span>
        <input type="checkbox" id="settings-reduce-motion" onchange="setReduceMotion(this.checked)" style="width:auto">
      </label>
      <label style="display:flex;justify-content:space-between;align-items:center;gap:var(--s-4)">
        <span><strong>Show keyboard hints</strong> <span style="color:var(--c-text-soft);font-size:var(--t-sm);display:block">Bottom-right shortcut overlay</span></span>
        <input type="checkbox" id="settings-show-kbd" onchange="setShowKbd(this.checked)" style="width:auto" checked>
      </label>
      <label style="display:flex;justify-content:space-between;align-items:center;gap:var(--s-4)">
        <span><strong>Embedded mode</strong> <span style="color:var(--c-text-soft);font-size:var(--t-sm);display:block">Push topbar down 56px (Claude Preview, iframe, browser ext)</span></span>
        <input type="checkbox" id="settings-embedded" onchange="setEmbeddedMode(this.checked)" style="width:auto">
      </label>
      <label style="display:flex;justify-content:space-between;align-items:center;gap:var(--s-4)">
        <span><strong>Top inset (custom)</strong> <span style="color:var(--c-text-soft);font-size:var(--t-sm);display:block">Override pixel offset above topbar — adjust until your host's toolbar stops covering buttons</span></span>
        <span style="display:inline-flex;align-items:center;gap:var(--s-3)">
          <input type="range" id="settings-safe-top" min="0" max="120" step="4" value="0" oninput="setSafeTop(this.value)" style="width:140px">
          <span id="settings-safe-top-val" style="font-family:var(--f-mono);font-size:var(--t-sm);width:40px;text-align:right">0px</span>
        </span>
      </label>
      <div style="border-top:1px solid var(--c-divider);padding-top:var(--s-5);font-size:var(--t-sm);color:var(--c-text-soft)">
        <strong>Keyboard shortcuts</strong><br>
        <code style="font-family:var(--f-mono);font-size:var(--t-xs)">⌘K</code> command palette ·
        <code style="font-family:var(--f-mono);font-size:var(--t-xs)">?</code> this help ·
        <code style="font-family:var(--f-mono);font-size:var(--t-xs)">J/K</code> nav ·
        <code style="font-family:var(--f-mono);font-size:var(--t-xs)">1–4</code> verdict
      </div>
    </div>
  </div>
</div>

<!-- ─── Bulk re-annotate modal (Sprint 3) ─── -->
<div class="modal-bg" id="bulk-modal" role="dialog" aria-modal="true" aria-labelledby="bulk-modal-title" onclick="if(event.target.id==='bulk-modal') closeBulk()">
  <div class="modal" style="min-width:560px;max-width:760px;">
    <h3 id="bulk-modal-title"><svg class="ico" aria-hidden="true"><use href="#i-refresh"/></svg><span>Apply pattern to similar rows</span><button class="close" onclick="closeBulk()" aria-label="ปิด"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-x"/></svg></button></h3>
    <div id="bulk-info" style="font-size:var(--t-sm);color:var(--c-text-muted);margin-bottom:var(--s-5)"></div>
    <div id="bulk-list" style="max-height:50vh;overflow:auto;border:1px solid var(--c-border);border-radius:var(--r-md)"></div>
    <div style="display:flex;gap:var(--s-3);margin-top:var(--s-5);align-items:center">
      <span style="flex:1;font-size:var(--t-sm);color:var(--c-text-soft)" id="bulk-summary"></span>
      <button class="btn" onclick="closeBulk()">Cancel</button>
      <button class="btn btn-primary" id="bulk-apply" onclick="applyBulk()">Apply selected</button>
    </div>
  </div>
</div>

<!-- Docked action bar — always visible, content swaps based on selection -->
<!-- Phase A6: verdict pills moved to status bar; action bar focuses on
     Auto/Mark/auto-next + free-form notes. -->
<div class="action-bar is-empty" id="action-bar" role="toolbar" aria-label="row actions">
  <div class="ab-empty-msg">เลือก row เพื่อทำงาน — verdict อยู่ที่ status bar (1–4)</div>
  <div class="ab-top">
    <div class="ab-row-info" id="ab-row-info">—</div>
    <div class="ab-flags" id="ab-flags"></div>
    <span class="ab-spacer"></span>

    <!-- Secondary actions (Auto, Mark) -->
    <div class="ab-secondary">
      <button class="ab-btn auto" onclick="showAutoAnnotate()" title="Auto-annotate (preview)">
        <svg class="ico ico-sm" aria-hidden="true"><use href="#i-sparkles"/></svg><span>Auto</span>
      </button>
      <button class="ab-btn mark" id="ab-mark-btn" onclick="startManualAnnotate()" title="ลาก rect ใน catalog เพื่อแก้ Col D">
        <svg class="ico ico-sm" aria-hidden="true"><use href="#i-pin"/></svg><span>Mark</span>
      </button>
    </div>

    <label class="ab-toggle" title="ไป row ถัดไปอัตโนมัติหลัง verdict">
      <input type="checkbox" id="auto-advance" onchange="toggleAutoAdvance()" checked>
      <span class="ab-toggle-track"></span>
      <span class="ab-toggle-label">auto-next</span>
    </label>
  </div>
  <div class="ab-bottom">
    <textarea id="ab-notes" placeholder="บันทึก / Notes…" oninput="saveNotesDebounced()" aria-label="row notes"></textarea>
  </div>
</div>

<!-- ──────────── Phase 2: Catalog browser modal ──────────── -->
<div class="modal-bg" id="catalog-modal" role="dialog" aria-modal="true" aria-labelledby="catalog-modal-title" onclick="if(event.target.id==='catalog-modal') closeCatalogBrowser()">
  <div class="modal" id="catalog-modal-inner" style="max-width:1100px;width:96vw;height:88vh;display:flex;flex-direction:column">
    <div class="modal-head">
      <h3 id="catalog-modal-title" style="margin:0;display:flex;align-items:center;gap:8px">
        <svg class="ico" aria-hidden="true"><use href="#i-book"/></svg>
        <span>Catalog Library</span>
        <span id="catalog-stats-pill" style="margin-left:6px;font-weight:500;font-size:var(--t-xs);color:var(--c-text-soft);background:var(--c-bg-soft);padding:2px 8px;border-radius:999px">loading…</span>
      </h3>
      <button class="modal-close" onclick="closeCatalogBrowser()" aria-label="close">✕</button>
    </div>
    <div style="display:flex;gap:var(--s-3);padding:var(--s-3) var(--s-5);border-bottom:1px solid var(--c-divider);flex-shrink:0">
      <input type="search" id="catalog-search" placeholder="ค้น brand / model / section / filename…" style="flex:1;padding:6px 10px;border:1px solid var(--c-border);border-radius:var(--r-md);font-size:var(--t-sm)" oninput="catalogSearchDebounced()">
      <select id="catalog-section-filter" onchange="catalogReload()" style="padding:6px 10px;border:1px solid var(--c-border);border-radius:var(--r-md);font-size:var(--t-sm)">
        <option value="">All sections</option>
      </select>
      <button class="btn" onclick="catalogReingest()" title="Re-scan output/ for new PDFs"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-refresh"/></svg><span>Re-scan</span></button>
    </div>
    <div style="flex:1;overflow:hidden;display:grid;grid-template-columns: 380px 1fr;gap:0;min-height:0">
      <!-- LEFT: list -->
      <div id="catalog-list" style="overflow-y:auto;border-right:1px solid var(--c-divider);padding:var(--s-2);min-height:0">
        <div style="padding:20px;color:var(--c-text-faint);text-align:center">loading…</div>
      </div>
      <!-- RIGHT: detail -->
      <div id="catalog-detail" style="overflow-y:auto;padding:var(--s-5);min-height:0">
        <div style="padding:40px;color:var(--c-text-faint);text-align:center;font-style:italic">เลือก catalog จาก list ด้านซ้าย</div>
      </div>
    </div>
  </div>
</div>

<!-- Toast notifications (top-right) -->
<div id="toasts" class="toast-stack" role="region" aria-live="polite" aria-label="notifications"></div>

<!-- Phase A5: floating annotation toolbar (mounted on body to escape
     stacking contexts; visibility + position are driven by JS).      -->
<div id="float-annot-toolbar" class="float-annot-toolbar" role="toolbar" aria-label="annotation actions">
  <span class="fat-type" id="fat-type"><svg class="ico ico-sm" aria-hidden="true"><use href="#i-square"/></svg><span id="fat-type-label">Square</span></span>
  <span class="fat-sep"></span>
  <span class="fat-swatch" id="fat-color-swatch" title="Stroke color (red)"></span>
  <span class="fat-meta" id="fat-meta">1pt</span>
  <span class="fat-sep"></span>
  <button class="fat-btn" id="fat-duplicate" type="button" onclick="duplicateSelected()" title="Duplicate (D)" aria-label="duplicate annotation">
    <svg class="ico ico-sm" aria-hidden="true"><use href="#i-copy"/></svg><span>Duplicate</span>
  </button>
  <button class="fat-btn danger" id="fat-delete" type="button" onclick="deleteSelected()" title="Delete (Del)" aria-label="delete annotation">
    <svg class="ico ico-sm" aria-hidden="true"><use href="#i-trash"/></svg><span>Delete</span>
  </button>
  <span class="fat-arrow" aria-hidden="true"></span>
</div>

<!-- Always-on floating Settings access — survives any top-overlay (browser
     extension / Claude Preview toolbar / etc.) covering the topbar. -->
<div class="floating-actions" id="floating-actions">
  <button class="fab-btn" onclick="openCmdK()" title="Command palette · ⌘K" aria-label="open command palette">
    <svg class="ico" aria-hidden="true"><use href="#i-search"/></svg>
  </button>
  <button class="fab-btn" onclick="showSettings()" title="Settings" aria-label="settings">
    <svg class="ico" aria-hidden="true"><use href="#i-settings"/></svg>
  </button>
  <button class="fab-btn" onclick="showOnboarding(true)" title="Help · ?" aria-label="help">
    <svg class="ico" aria-hidden="true"><use href="#i-help"/></svg>
  </button>
</div>

<aside class="kbd-help" aria-label="keyboard shortcuts">
  <span class="kbd-group"><span class="kbd">J</span><span class="kbd">K</span><span class="kbd-text">rows</span></span>
  <span class="kbd-group"><span class="kbd">N</span><span class="kbd-text">uncertain</span></span>
  <span class="kbd-group"><span class="kbd">1</span>–<span class="kbd">4</span><span class="kbd-text">verdict</span></span>
  <span class="kbd-group"><span class="kbd">[</span><span class="kbd">]</span><span class="kbd-text">PDF</span></span>
  <span class="kbd-group"><span class="kbd">,</span><span class="kbd">.</span><span class="kbd-text">TOR</span></span>
  <span class="kbd-group"><span class="kbd">?</span><span class="kbd-text">help</span></span>
</aside>

<script>
// ── Icon helper ───────────────────────────────────────────────
// Returns an inline <svg><use href="#i-{name}"/></svg> string for use
// in template literals. Centralises icon rendering so every glyph in
// the UI comes from the same SVG sprite (no emoji mixing).
function ico(name, size = 16, extraCls = '') {
  let cls = 'ico';
  if (size <= 12) cls += ' ico-xs';
  else if (size === 14) cls += ' ico-sm';
  else if (size === 20) cls += ' ico-lg';
  else if (size >= 24) cls += ' ico-xl';
  if (extraCls) cls += ' ' + extraCls;
  return `<svg class="${cls}" aria-hidden="true"><use href="#i-${name}"/></svg>`;
}

let DATA = null;
let SELECTED_ROW = null;
let ROWS_BY_NUM = {};
let TREE_ROOT = null;
let VISIBLE_ROWS = new Set();          // rows passing filter
let EXPANDED = new Set();              // tree node keys expanded
let CTX_RADIUS = 6;
let TOR_DPI = 110, PDF_DPI = 130;
let TOR_PAGE = 1, TOR_TARGET_PAGE = 1;
let TOR_PAGES = 0, TOR_HITS = 0;
let CURRENT_PDF = null, PDF_PAGE = 1, CURRENT_HIGHLIGHT = null;

async function init() {
  const r = await fetch('/api/index');
  DATA = await r.json();
  ROWS_BY_NUM = Object.fromEntries(DATA.rows.map(r => [r.row, r]));
  TREE_ROOT = DATA.tree;
  buildRowBlobs();

  // expand top-level by default
  for (const c of TREE_ROOT.children) EXPANDED.add(c.key);

  applyFilters();
  renderStats();
  renderTree();

  // Reflect "always work on latest" status on the main UI
  if (DATA.version_sync) refreshSyncIndicators(DATA.version_sync);

  // Periodically refresh sync status (every 60s) so divergence detected
  // by other tools (e.g. Google Drive sync conflicts) shows up.
  setInterval(async () => {
    try {
      const r = await fetch('/api/versions/sync');
      const sync = await r.json();
      refreshSyncIndicators(sync);
    } catch (e) {}
  }, 60000);

  const last = parseInt(localStorage.getItem('lastRow'));
  if (last && ROWS_BY_NUM[last]) selectRow(last, false);
  else if (DATA.rows.length) selectRow(DATA.rows[0].row, false);
}

// ── Search normalization ───────────────────────────────────────
/* Match xlsx/comply text regardless of:
   - precomposed vs decomposed Thai SARA AM (ำ ↔ ํา)
   - case
   - Whitespace differences
   Also support multi-token AND matching ("RCBO Schneider" → both must appear).
*/
function normalizeForSearch(s) {
  if (s == null) return '';
  return String(s).toLowerCase()
    .replace(/ํา/g, 'ำ')   // decomposed → precomposed SARA AM
    .replace(/\s+/g, ' ');
}
function tokenizeQuery(q) {
  // split on whitespace; quote support: "a b" treated as one token
  const tokens = [];
  const re = /"([^"]+)"|(\S+)/g;
  let m;
  while ((m = re.exec(q)) !== null) {
    const t = (m[1] || m[2] || '').trim();
    if (t) tokens.push(normalizeForSearch(t));
  }
  return tokens;
}
const SECTION_QUERY = /^\d+(\.\d+){0,4}\.?$/;

// Per-row search blob cache (rebuilt on init, not on every filter)
let ROW_BLOBS = new Map();
function buildRowBlobs() {
  ROW_BLOBS.clear();
  for (const r of DATA.rows) {
    const parts = [
      `r${r.row}`, r.A||'', r.B||'', r.C||'', r.D||'', r.E||'', r.F||'',
      r.section||'', (r.parsed && r.parsed.brand)||'', (r.parsed && r.parsed.model)||'',
    ];
    ROW_BLOBS.set(r.row, normalizeForSearch(parts.join(' ')));
  }
}

// ── Filtering ──────────────────────────────────────────────────
function applyFilters() {
  const rawQ = document.getElementById('tree-search').value.trim();
  const tokens = tokenizeQuery(rawQ);
  const fStat = document.getElementById('filter-status').value;
  const fVen  = document.getElementById('filter-vendor').value;
  const fFlag = document.getElementById('filter-flags').checked;
  const fPdf  = document.getElementById('filter-haspdf').checked;
  const status = DATA.status || {};

  // Detect "pure section number" query (e.g. "5.1.2") and expand its match
  // semantics: match rows whose section IS or BEGINS WITH this prefix.
  let sectionPrefix = null;
  if (tokens.length === 1 && SECTION_QUERY.test(tokens[0])) {
    sectionPrefix = tokens[0].replace(/\.$/, '');  // strip trailing dot
  }

  VISIBLE_ROWS.clear();
  for (const r of DATA.rows) {
    const st = (status[r.row] && status[r.row].status) || 'unverified';
    if (fStat && st !== fStat) continue;
    if (fVen && r.E !== fVen) continue;
    if (fFlag && (!r.auto_flags || !r.auto_flags.length)) continue;
    if (fPdf && !r.pdf_rel) continue;

    if (tokens.length) {
      const blob = ROW_BLOBS.get(r.row) || '';
      if (sectionPrefix) {
        // Section query: row's section must start with the prefix
        const sec = (r.section || '').toString();
        if (sec !== sectionPrefix && !sec.startsWith(sectionPrefix + '.')) continue;
      } else {
        // All tokens must appear (AND)
        let ok = true;
        for (const t of tokens) { if (!blob.includes(t)) { ok = false; break; } }
        if (!ok) continue;
      }
    }
    VISIBLE_ROWS.add(r.row);
  }

  // Auto-expand path to all matched rows when a search is active
  if (tokens.length) {
    for (const rn of VISIBLE_ROWS) expandPathToRow(rn);
  }
  renderTree();
}

// ── Tree rendering ─────────────────────────────────────────────
function visibleRowsInNode(node) {
  // count rows in this node + descendants that pass filter
  let count = 0;
  for (const r of node.rows) if (VISIBLE_ROWS.has(r)) count++;
  for (const c of node.children) count += visibleRowsInNode(c);
  return count;
}
function nodeHasMatch(node) { return visibleRowsInNode(node) > 0; }

// Tree render is debounced on rapid filter changes (Sprint S3.1)
let _TREE_RENDER_PENDING = false;
function renderTree() {
  if (_TREE_RENDER_PENDING) return;
  _TREE_RENDER_PENDING = true;
  // Use rAF so multiple renderTree() calls within a frame coalesce
  requestAnimationFrame(() => {
    _TREE_RENDER_PENDING = false;
    const out = [];
    for (const c of TREE_ROOT.children) renderNode(c, 0, out);
    const treeEl = document.getElementById('tree');
    // innerHTML in one shot is faster than DocumentFragment for static markup
    treeEl.innerHTML = out.join('');
    document.getElementById('tree-count').textContent = `(${VISIBLE_ROWS.size})`;
  });
}

function renderNode(node, depth, out) {
  if (!nodeHasMatch(node)) return;
  const expanded = EXPANDED.has(node.key);
  const hasKids = node.children.length > 0 || node.rows.length > 0;

  // Render the node row(s):
  // Section/item/sub nodes show their representative row (first one) + children
  const repRow = node.rows[0];
  const r = repRow ? ROWS_BY_NUM[repRow] : null;
  const status = (DATA.status && DATA.status[repRow]) ? DATA.status[repRow].status : 'unverified';

  const flagCount = r && r.auto_flags ? r.auto_flags.length : 0;
  const flagHtml = flagCount
    ? `<span class="tree-flags" title="${escapeHtml(r.auto_flags.map(f=>f.msg).join(' / '))}">${ico('alert',12)}${flagCount}</span>`
    : '';

  const icon = ({section:'§', item:'›', sub:'·', root:''})[node.type] || '';
  const iconCls = node.type;
  const indent = depth * 12;

  let label = node.label;
  // Append context for items/subs
  if (r) {
    const preview = (r.B || r.C || '').toString().trim().replace(/\s+/g,' ').slice(0, 60);
    if (node.type === 'section' && r.D && /ยี่ห้อ/.test(r.D)) {
      const bm = r.D.match(/ยี่ห้อ\s+([^รุ่น]+?)\s+รุ่น\s+(.+)/);
      if (bm) label = `${node.label} — ${bm[1].trim()} ${bm[2].trim()}`.slice(0, 60);
    } else if (preview && (node.type === 'item' || node.type === 'sub')) {
      label = `${node.label} ${preview}`.slice(0, 70);
    }
  }
  const rowAttr = repRow ? `data-row="${repRow}"` : '';
  const sel = (SELECTED_ROW === repRow) ? ' selected' : '';
  const stAttr = (status && status !== 'unverified') ? ` data-status="${status}"` : '';

  out.push(`<div class="tree-node${expanded?' expanded':''}" data-key="${escapeHtml(node.key)}" style="padding-left:${indent}px">`);
  out.push(`<div class="tree-row${sel}"${stAttr} ${rowAttr} onclick="onTreeRowClick(event, '${escapeHtml(node.key)}', ${repRow||'null'})">`);
  if (hasKids && node.children.length) {
    out.push(`<span class="tree-chev" onclick="event.stopPropagation();toggleNode('${escapeHtml(node.key)}')">${expanded?ico('chevron-down',10):ico('chevron-right',10)}</span>`);
  } else {
    out.push(`<span class="tree-chev empty">·</span>`);
  }
  out.push(`<span class="tree-icon ${iconCls}">${icon}</span>`);
  out.push(`<span class="tree-label">${escapeHtml(label)}</span>`);
  // Confidence dot — at-a-glance for which rows still need review
  out.push(confDotHtml(r));
  out.push(flagHtml);
  out.push('</div>');

  // Additional rows in this node beyond the first → render as siblings
  for (let i = 1; i < node.rows.length; i++) {
    const rn = node.rows[i];
    if (!VISIBLE_ROWS.has(rn)) continue;
    const rr = ROWS_BY_NUM[rn];
    if (!rr) continue;
    const sel2 = (SELECTED_ROW === rn) ? ' selected' : '';
    const st2 = (DATA.status && DATA.status[rn]) ? DATA.status[rn].status : 'unverified';
    const stAttr2 = (st2 && st2 !== 'unverified') ? ` data-status="${st2}"` : '';
    const fc2 = rr.auto_flags ? rr.auto_flags.length : 0;
    const fh2 = fc2 ? `<span class="tree-flags" title="${escapeHtml(rr.auto_flags.map(f=>f.msg).join(' / '))}">${ico('alert',12)}${fc2}</span>` : '';
    const lab2 = (rr.B || rr.C || '').toString().trim().replace(/\s+/g,' ').slice(0, 70);
    out.push(`<div class="tree-row${sel2}"${stAttr2} data-row="${rn}" onclick="onTreeRowClick(event, null, ${rn})" style="padding-left:${(depth+1)*12+18}px">`);
    out.push(`<span class="tree-chev empty">·</span><span class="tree-icon">·</span>`);
    out.push(`<span class="tree-label">R${rn} ${escapeHtml(lab2)}</span>`);
    out.push(fh2);
    out.push('</div>');
  }

  out.push('<div class="tree-children">');
  for (const c of node.children) renderNode(c, depth + 1, out);
  out.push('</div></div>');
}

function toggleNode(key) {
  if (EXPANDED.has(key)) EXPANDED.delete(key);
  else EXPANDED.add(key);
  renderTree();
}
function expandAll() {
  function walk(n){ EXPANDED.add(n.key); for (const c of n.children) walk(c); }
  for (const c of TREE_ROOT.children) walk(c);
  renderTree();
}
function collapseAll() {
  EXPANDED.clear();
  for (const c of TREE_ROOT.children) EXPANDED.add(c.key);
  renderTree();
}

function onTreeRowClick(e, nodeKey, rowNum) {
  if (rowNum != null) selectRow(rowNum);
  // also expand the clicked node so its children are visible
  if (nodeKey) {
    EXPANDED.add(nodeKey);
    renderTree();
  }
}

// ── Selection: drives all 4 panes ──────────────────────────────
async function selectRow(rowNum, scroll = true) {
  SELECTED_ROW = rowNum;
  localStorage.setItem('lastRow', rowNum);
  const r = ROWS_BY_NUM[rowNum];
  if (!r) return;

  // 0) topbar context chip — show R# · section · brand/model summary
  updateTopbarContext(r);

  // 1) action bar
  renderActionBar(r);

  // 2) TOR
  loadTOR(rowNum);

  // 3) xlsx preview
  loadXlsx(rowNum);

  // 4) catalog PDF
  if (r.pdf_rel) {
    loadPdf(r.pdf_rel, r.parsed.page || 1, computeHighlight(r));
  } else {
    CURRENT_PDF = null;
    document.getElementById('pdf-canvas').innerHTML =
      `<div class="empty-canvas">Row นี้ไม่มี catalog อ้างอิง<br>(${r.parsed.type || 'empty'})</div>`;
    document.getElementById('pdf-filename').textContent = '(ไม่มี)';
    document.getElementById('pdf-page-info').textContent = '— / —';
    document.getElementById('pdf-annots').innerHTML = '';
  }

  // 5) update tree highlight + auto-expand path + scroll into view
  expandPathToRow(rowNum);
  renderTree();
  if (scroll) {
    requestAnimationFrame(() => {
      const el = document.querySelector(`.tree-row[data-row="${rowNum}"]`);
      if (el) el.scrollIntoView({ block: 'nearest', behavior: 'smooth' });
    });
  }
}

// ── Topbar current-row context chip ───────────────────────────
function updateTopbarContext(r) {
  const chip = document.getElementById('topbar-context');
  const lbl  = document.getElementById('topbar-context-label');
  if (!chip || !lbl || !r) { if (chip) chip.style.display = 'none'; return; }
  const sec = r.section || '?';
  let meta = '';
  if (r.parsed && r.parsed.brand && r.parsed.model) {
    meta = `${r.parsed.brand} ${r.parsed.model}`;
  } else if (r.parsed && r.parsed.brand) {
    meta = r.parsed.brand;
  } else if (r.B) {
    meta = (r.B || '').toString().trim().replace(/\s+/g, ' ').slice(0, 50);
  }
  lbl.innerHTML =
    `<span class="row-num">R${r.row}</span>` +
    `<span class="ctx-section">${escapeHtml(sec)}</span>` +
    (meta ? `<span class="sep">·</span><span class="ctx-meta">${escapeHtml(meta)}</span>` : '');
  chip.style.display = '';
}
function focusSelectedRow() {
  if (!SELECTED_ROW) return;
  const el = document.querySelector(`.tree-row[data-row="${SELECTED_ROW}"]`);
  if (el) el.scrollIntoView({block: 'nearest', behavior: 'smooth'});
}

/* Walk the tree, find any node containing the row, expand all ancestors. */
function expandPathToRow(rowNum) {
  const path = [];
  function walk(node, ancestors) {
    if (node.rows && node.rows.indexOf(rowNum) >= 0) {
      for (const a of ancestors) path.push(a.key);
      return true;
    }
    for (const c of (node.children||[])) {
      if (walk(c, ancestors.concat([node]))) return true;
    }
    return false;
  }
  walk(TREE_ROOT, []);
  for (const k of path) if (k && k !== 'ROOT') EXPANDED.add(k);
}

// ── Action bar ─────────────────────────────────────────────────
function renderActionBar(r) {
  // Toggle is-empty class (no display flip → no animation re-trigger,
  // bar stays in the same grid slot regardless of selection state).
  const bar = document.getElementById('action-bar');
  bar.classList.remove('is-empty');
  const cur = (DATA.status && DATA.status[r.row]) || {};
  const st = cur.status || 'unverified';

  let secLabel = r.section || '';
  if (r.parsed.brand || r.parsed.model) {
    secLabel += ` · ${r.parsed.brand||''} ${r.parsed.model||''}`.trim();
  }
  document.getElementById('ab-row-info').innerHTML =
    `<span class="row-num">R${r.row}</span> <span class="section">${escapeHtml(secLabel)}</span> ${r.E ? `<span class="vendor-tag ${r.E}">${escapeHtml(r.E)}</span>` : ''}`;

  const flags = r.auto_flags || [];
  document.getElementById('ab-flags').innerHTML = flags.length
    ? flags.map(f => `<span title="${escapeHtml(f.msg)}">⚠ ${escapeHtml(f.msg.slice(0,60))}</span>`).join(' · ')
    : '';

  for (const cls of ['pass','fail','fix','skip']) {
    const btn = document.querySelector('.ab-btn.' + cls);
    if (btn) btn.classList.toggle('active', st === (cls==='fix'?'need_fix':cls));
  }
  document.getElementById('ab-notes').value = cur.notes || '';

  // Pulse the 📍 Mark button when Col D is "ยินดีปฏิบัติ" (commitment) —
  // the most common case where the user might want to manually locate
  // the spec the AI failed to find.
  const markBtn = document.getElementById('ab-mark-btn');
  if (markBtn) {
    const dCommit = (r.D || '').trim().startsWith('ยินดีปฏิบัติ');
    markBtn.classList.toggle('commitment', dCommit && !!r.pdf_rel);
    markBtn.disabled = !r.pdf_rel;
    markBtn.title = !r.pdf_rel
      ? 'ไม่มี catalog PDF ให้ annotate'
      : (dCommit ? '⚠ Col D เป็น commitment — คลิกเพื่อหาเนื้อหาใน catalog เอง'
                 : 'ลาก rect ใน catalog เพื่อ override Col D');
  }
}

let notesTimer = null;
function saveNotesDebounced() { clearTimeout(notesTimer); notesTimer = setTimeout(saveNotes, 400); }
async function saveNotes() {
  if (!SELECTED_ROW) return;
  const notes = document.getElementById('ab-notes').value;
  await fetch('/api/status', {method: 'POST', headers: {'Content-Type':'application/json'},
    body: JSON.stringify({row: SELECTED_ROW, notes})});
  if (!DATA.status) DATA.status = {};
  DATA.status[SELECTED_ROW] = DATA.status[SELECTED_ROW] || {};
  DATA.status[SELECTED_ROW].notes = notes;
}
async function setStatus(status) {
  if (!SELECTED_ROW) return;
  const notes = document.getElementById('ab-notes').value || '';
  // Optimistic UI: set verdict button + tree row indicator before round-trip
  const treeEl = document.querySelector(`.tree-row[data-row="${SELECTED_ROW}"]`);
  if (treeEl) {
    if (status && status !== 'unverified') treeEl.setAttribute('data-status', status);
    else treeEl.removeAttribute('data-status');
  }
  // Update verdict-control radio state
  for (const cls of ['pass','fail','fix','skip']) {
    const btn = document.querySelector('.verdict-control .ab-btn.' + cls);
    if (btn) {
      const target = (cls === 'fix') ? 'need_fix' : cls;
      const active = (status === target);
      btn.classList.toggle('active', active);
      btn.setAttribute('aria-checked', active ? 'true' : 'false');
    }
  }
  const r = await fetch('/api/status', {method: 'POST', headers: {'Content-Type':'application/json'},
    body: JSON.stringify({row: SELECTED_ROW, status, notes})});
  const j = await r.json();
  if (!DATA.status) DATA.status = {};
  DATA.status[SELECTED_ROW] = j.entry;
  renderStats();
  renderTree();
  const row = ROWS_BY_NUM[SELECTED_ROW];
  if (row) renderActionBar(row);
  // Verdict counts as a learning signal — accumulate; auto-retrain at threshold
  if (status && status !== 'unverified') {
    if (typeof tickRetrain === 'function') tickRetrain('status');
  }
}

// ── Skeleton helpers ───────────────────────────────────────────
function skeletonImage() {
  return `<div class="skel-stack">
    <div class="skel skel-img"></div>
  </div>`;
}
function skeletonTable() {
  return `<div class="skel-stack">
    <div class="skel skel-line tall" style="width:60%"></div>
    <div class="skel skel-line" style="width:90%"></div>
    <div class="skel skel-line" style="width:95%"></div>
    <div class="skel skel-line" style="width:80%"></div>
    <div class="skel skel-line" style="width:88%"></div>
    <div class="skel skel-line" style="width:70%"></div>
  </div>`;
}

// ── TOR pane ───────────────────────────────────────────────────
async function loadTOR(rowNum) {
  const url = `/api/tor_page?row=${rowNum}&dpi=${TOR_DPI}`;
  const c = document.getElementById('tor-canvas');
  c.innerHTML = skeletonImage();
  try {
    const r = await fetch(url);
    if (!r.ok) {
      c.innerHTML = '<div class="empty-canvas">โหลด TOR ไม่ได้</div>';
      return;
    }
    TOR_TARGET_PAGE = parseInt(r.headers.get('X-TOR-Page') || '1');
    TOR_HITS = parseInt(r.headers.get('X-TOR-Hits') || '0');
    const hY0 = parseFloat(r.headers.get('X-Highlight-Y0'));
    const hY1 = parseFloat(r.headers.get('X-Highlight-Y1'));
    const torSection = r.headers.get('X-Section') || '';
    const torRange = r.headers.get('X-Section-Range') || '';
    TOR_PAGE = TOR_TARGET_PAGE;
    const blob = await r.blob();
    const imgUrl = URL.createObjectURL(blob);
    showImageWithScroll(c, imgUrl, hY0, hY1, `TOR p${TOR_PAGE}`);
    if (!TOR_PAGES) {
      const m = await fetch('/api/tor_meta').then(r => r.json());
      TOR_PAGES = m.pages || 0;
      document.getElementById('tor-info').textContent = m.filename || '';
    }
    document.getElementById('tor-page-info').textContent = `${TOR_PAGE} / ${TOR_PAGES}`;
    let status = '';
    if (TOR_HITS > 0) {
      status = `✓ ${TOR_HITS} hit${TOR_HITS>1?'s':''}`;
    } else if (torRange) {
      status = `ไม่เจอข้อความ — แสดงต้น section ${torSection}`;
    } else {
      status = 'ไม่เจอข้อความ';
    }
    if (torSection && torRange) status += `  ·  section ${torSection} P${torRange}`;
    document.getElementById('tor-status').textContent = status;
  } catch (e) {
    c.innerHTML = '<div class="empty-canvas">โหลด TOR error</div>';
  }
}
async function torReload() {
  if (!SELECTED_ROW) return;
  const c = document.getElementById('tor-canvas');
  const url = `/api/tor_page?row=${SELECTED_ROW}&page=${TOR_PAGE}&dpi=${TOR_DPI}`;
  const r = await fetch(url);
  const hY0 = parseFloat(r.headers.get('X-Highlight-Y0'));
  const hY1 = parseFloat(r.headers.get('X-Highlight-Y1'));
  const blob = await r.blob();
  showImageWithScroll(c, URL.createObjectURL(blob), hY0, hY1, `TOR p${TOR_PAGE}`);
  document.getElementById('tor-page-info').textContent = `${TOR_PAGE} / ${TOR_PAGES}`;
}

/* Render an image into a scrollable container and auto-scroll so that the
   highlight Y range (in source-pixel coords) sits ~25% from the top of the
   visible viewport. Falls back to top if no Y given. */
function showImageWithScroll(container, src, y0, y1, alt) {
  container.innerHTML = '';
  const img = document.createElement('img');
  if (alt) img.alt = alt;
  container.appendChild(img);
  const onReady = () => {
    if (!isFinite(y0)) { container.scrollTop = 0; return; }
    const scale = img.clientWidth / (img.naturalWidth || 1);
    const targetMid = ((y0 + (isFinite(y1) ? y1 : y0)) / 2) * scale;
    const offset = container.clientHeight * 0.25;
    container.scrollTop = Math.max(0, targetMid - offset);
  };
  img.onload = onReady;
  img.src = src;
  if (img.complete && (img.naturalWidth > 0 || img.naturalHeight > 0)) {
    onReady();
  }
}
function torPrev() { if (TOR_PAGE > 1) { TOR_PAGE--; torReload(); } }
function torNext() { if (TOR_PAGE < TOR_PAGES) { TOR_PAGE++; torReload(); } }
function torZoom(d) { TOR_DPI = Math.max(70, Math.min(200, TOR_DPI + d*20)); torReload(); }
function torJumpToMatch() { TOR_PAGE = TOR_TARGET_PAGE; torReload(); }

// ── XLSX preview ───────────────────────────────────────────────
async function loadXlsx(rowNum) {
  const w = document.getElementById('xlsx-wrap');
  w.innerHTML = skeletonTable();
  const r = await fetch(`/api/row_context?row=${rowNum}&radius=${CTX_RADIUS}`);
  const j = await r.json();
  document.getElementById('ctx-radius').textContent = CTX_RADIUS;
  let html = `<table class="xlsx-table">
    <thead><tr>
      <th class="row-num">#</th>
      <th class="col-A">A</th>
      <th class="col-B">B คุณลักษณะที่ต้องการ (TOR)</th>
      <th class="col-C">C คุณลักษณะที่เสนอ</th>
      <th class="col-D">D เอกสารอ้างอิง</th>
      <th class="col-E">E</th>
      <th class="col-F">F สถานะ TOR</th>
    </tr></thead><tbody>`;
  for (const row of j.rows) {
    const cls = row.row === j.target ? 'target' : '';
    html += `<tr class="${cls}" data-row="${row.row}" onclick="selectRow(${row.row})">`;
    html += `<td class="row-num">${row.row}</td>`;
    html += `<td class="col-A">${escapeHtml(row.A||'')}</td>`;
    html += `<td class="col-B">${escapeHtml(row.B||'')}</td>`;
    html += `<td class="col-C">${escapeHtml(row.C||'')}</td>`;
    // Col D — RIGHT-click opens the actions menu (annotate / revert / edit / auto);
    // double-click triggers inline editing; left-click selects the row (default).
    const dVal = row.D || '';
    const dCommit = dVal.trim().startsWith('ยินดีปฏิบัติ');
    const dCls = 'col-D editable' + (dCommit ? ' commitment' : ' has-ref');
    html += `<td class="${dCls}"
             oncontextmenu="onColDContextMenu(event, ${row.row})"
             ondblclick="editColD(event, ${row.row})"
             title="คลิกขวา → เปิดเมนู · ดับเบิลคลิก → แก้ไขในช่อง">`;
    html += `<span class="d-text">${escapeHtml(dVal)}</span>`;
    html += `</td>`;
    html += `<td class="col-E">${row.E ? `<span class="vendor-tag ${row.E}">${escapeHtml(row.E)}</span>` : ''}</td>`;
    html += `<td class="col-F">${escapeHtml(row.F||'')}</td>`;
    html += `</tr>`;
  }
  html += '</tbody></table>';
  w.innerHTML = html;
  // scroll target into view
  setTimeout(() => {
    const t = w.querySelector('tr.target');
    if (t) t.scrollIntoView({block: 'center'});
  }, 0);
  document.getElementById('xlsx-info').textContent = `R${j.target}`;
}
function ctxRadius(d) {
  CTX_RADIUS = Math.max(2, Math.min(30, CTX_RADIUS + d));
  if (SELECTED_ROW) loadXlsx(SELECTED_ROW);
}

// Inline Col D editing — double-click in xlsx preview → editable textarea →
// save → records HITL feedback
function editColD(e, rowNum) {
  e.stopPropagation();
  const td = e.currentTarget;
  const original = td.textContent;
  td.contentEditable = 'plaintext-only';
  td.classList.add('editing');
  td.focus();
  const sel = window.getSelection();
  sel.selectAllChildren(td);

  // Phase B3: live autocomplete dropdown
  const ac = _colDAcOpen(td, rowNum);

  const finish = async (commit) => {
    _colDAcClose();
    td.contentEditable = 'false';
    td.classList.remove('editing');
    const newVal = td.textContent.trim();
    if (!commit || newVal === original.trim()) {
      td.textContent = original;
      return;
    }
    // Send to server
    try {
      const r = await fetch('/api/row/col_d', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({row: rowNum, col_d: newVal, original}),
      });
      const j = await r.json();
      if (j.ok) {
        toast(`✓ R${rowNum} Col D updated`,
              `${original.slice(0, 40)} → ${newVal.slice(0, 40)}`,
              'learn', 4000);
        // Trigger auto-retrain since this is a strong correction signal
        if (typeof tickRetrain === 'function') tickRetrain('inline');
        // Reload xlsx + tree
        if (SELECTED_ROW) loadXlsx(SELECTED_ROW);
        // Refresh in-memory rows snapshot
        const idx = await fetch('/api/index').then(r => r.json());
        DATA = idx;
        ROWS_BY_NUM = Object.fromEntries(idx.rows.map(r => [r.row, r]));
        renderTree();
      } else {
        toast('Save failed', j.error || 'unknown', 'error', 5000);
        td.textContent = original;
      }
    } catch (err) {
      toast('Save failed', err.message, 'error', 5000);
      td.textContent = original;
    }
  };

  td.addEventListener('blur', () => {
    // Defer so a click on a suggestion can be processed first
    setTimeout(() => { if (!_COL_D_AC || !_COL_D_AC._wasClicked) finish(true); }, 120);
  }, {once: true});
  td.addEventListener('keydown', (ev) => {
    // Phase B3: autocomplete keyboard nav (intercept before commit/cancel)
    if (_COL_D_AC && _COL_D_AC.visible) {
      if (ev.key === 'ArrowDown') { ev.preventDefault(); _colDAcMove(1); return; }
      if (ev.key === 'ArrowUp')   { ev.preventDefault(); _colDAcMove(-1); return; }
      if (ev.key === 'Tab')       { ev.preventDefault(); _colDAcAcceptCurrent(td); return; }
      if (ev.key === 'Enter' && _COL_D_AC.activeIdx >= 0 && !ev.shiftKey) {
        ev.preventDefault(); _colDAcAcceptCurrent(td); return;
      }
    }
    if (ev.key === 'Enter' && !ev.shiftKey) { ev.preventDefault(); td.blur(); }
    if (ev.key === 'Escape') {
      if (_COL_D_AC && _COL_D_AC.visible) { ev.preventDefault(); _colDAcClose(); return; }
      ev.preventDefault(); finish(false); td.blur();
    }
  });
  td.addEventListener('input', () => _colDAcQueueRefresh(td, rowNum));
}

// ── Phase B3: Col D live autocomplete ─────────────────────────
// Single-instance dropdown panel below the editing td. Fetches from
// /api/row/col_d/suggest, debounced. Keyboard nav works through the
// editColD keydown handler above.
let _COL_D_AC = null;
let _COL_D_AC_TIMER = 0;

function _colDAcOpen(td, rowNum) {
  _colDAcClose();
  const panel = document.createElement('div');
  panel.className = 'col-d-ac-panel';
  panel.setAttribute('role', 'listbox');
  panel.setAttribute('aria-label', 'Col D suggestions');
  document.body.appendChild(panel);
  _COL_D_AC = {panel, td, rowNum, items: [], activeIdx: -1, visible: false, _wasClicked: false};
  _colDAcPosition();
  // Reposition on resize / scroll inside the table
  window.addEventListener('scroll', _colDAcPosition, true);
  window.addEventListener('resize', _colDAcPosition);
  // First fetch (no query)
  _colDAcQueueRefresh(td, rowNum, 0);
  return _COL_D_AC;
}
function _colDAcClose() {
  if (!_COL_D_AC) return;
  try { _COL_D_AC.panel.remove(); } catch (e) {}
  window.removeEventListener('scroll', _colDAcPosition, true);
  window.removeEventListener('resize', _colDAcPosition);
  _COL_D_AC = null;
  if (_COL_D_AC_TIMER) { clearTimeout(_COL_D_AC_TIMER); _COL_D_AC_TIMER = 0; }
}
function _colDAcPosition() {
  if (!_COL_D_AC) return;
  const r = _COL_D_AC.td.getBoundingClientRect();
  const p = _COL_D_AC.panel;
  p.style.left = Math.round(r.left) + 'px';
  p.style.top  = Math.round(r.bottom + 4) + 'px';
  p.style.minWidth = Math.round(r.width) + 'px';
}
function _colDAcQueueRefresh(td, rowNum, delay) {
  if (delay === undefined) delay = 250;
  if (_COL_D_AC_TIMER) clearTimeout(_COL_D_AC_TIMER);
  _COL_D_AC_TIMER = setTimeout(() => _colDAcRefresh(td, rowNum), delay);
}
async function _colDAcRefresh(td, rowNum) {
  if (!_COL_D_AC) return;
  const q = (td.textContent || '').trim();
  try {
    const r = await fetch(`/api/row/col_d/suggest?row=${rowNum}&q=${encodeURIComponent(q)}`);
    const j = await r.json();
    if (!_COL_D_AC) return;  // closed during request
    if (!j.ok) { _COL_D_AC.panel.style.display = 'none'; _COL_D_AC.visible = false; return; }
    _colDAcRender(j.suggestions || []);
  } catch (e) { /* swallow */ }
}
function _colDAcRender(items) {
  if (!_COL_D_AC) return;
  _COL_D_AC.items = items;
  _COL_D_AC.activeIdx = items.length ? 0 : -1;
  if (!items.length) {
    _COL_D_AC.panel.style.display = 'none';
    _COL_D_AC.visible = false;
    return;
  }
  _COL_D_AC.panel.style.display = '';
  _COL_D_AC.visible = true;
  _COL_D_AC.panel.innerHTML = items.map((s, i) => {
    const conf = s.confidence ? `<span class="ac-conf">${Math.round(s.confidence*100)}%</span>` : '';
    const cls = i === _COL_D_AC.activeIdx ? 'ac-item active' : 'ac-item';
    const kindLbl = ({ai:'AI', neighbor:'Neighbor', shape:'Shape'})[s.kind] || s.kind;
    return `<div class="${cls}" data-i="${i}" role="option" aria-selected="${i === _COL_D_AC.activeIdx}">
      <span class="ac-kind ${s.kind}">${escapeHtml(kindLbl)}</span>
      <span class="ac-text">${escapeHtml(s.text)}</span>
      <span class="ac-meta">${escapeHtml(s.label || '')}${conf}</span>
    </div>`;
  }).join('');
  // Mouse interactions
  _COL_D_AC.panel.querySelectorAll('.ac-item').forEach(el => {
    el.addEventListener('mousedown', (ev) => {
      ev.preventDefault();
      _COL_D_AC._wasClicked = true;
      const i = parseInt(el.dataset.i);
      _COL_D_AC.activeIdx = i;
      _colDAcAcceptCurrent(_COL_D_AC.td);
    });
    el.addEventListener('mouseenter', () => {
      _COL_D_AC.activeIdx = parseInt(el.dataset.i);
      _colDAcRefreshActive();
    });
  });
}
function _colDAcRefreshActive() {
  if (!_COL_D_AC) return;
  _COL_D_AC.panel.querySelectorAll('.ac-item').forEach((el, i) => {
    el.classList.toggle('active', i === _COL_D_AC.activeIdx);
    el.setAttribute('aria-selected', i === _COL_D_AC.activeIdx ? 'true' : 'false');
  });
}
function _colDAcMove(delta) {
  if (!_COL_D_AC || !_COL_D_AC.items.length) return;
  const n = _COL_D_AC.items.length;
  _COL_D_AC.activeIdx = ((_COL_D_AC.activeIdx + delta) % n + n) % n;
  _colDAcRefreshActive();
}
function _colDAcAcceptCurrent(td) {
  if (!_COL_D_AC || _COL_D_AC.activeIdx < 0) return;
  const item = _COL_D_AC.items[_COL_D_AC.activeIdx];
  if (!item) return;
  td.textContent = item.text;
  // Place caret at end
  const range = document.createRange();
  range.selectNodeContents(td);
  range.collapse(false);
  const sel2 = window.getSelection();
  sel2.removeAllRanges();
  sel2.addRange(range);
  _colDAcClose();
  // Fire input event so any listener (including our debounced refresh)
  // sees the new value, but next blur/Enter will commit it.
  td.dispatchEvent(new Event('input'));
}

// ── Col D context dropdown menu ───────────────────────────────
// Right-click (contextmenu) on a Col D cell shows a dropdown of actions
// tailored to the row's current state:
//   • commitment ("ยินดีปฏิบัติ") → Mark / Auto / Re-annotate / Edit
//   • has reference                → Re-annotate / Auto / Mark / Edit / Revert
//
// Left-click selects the row (same as clicking elsewhere on the row).
// Double-click still triggers inline editing (editColD) for power users.

let _COL_D_MENU = null;

function closeColDMenu() {
  if (_COL_D_MENU) { _COL_D_MENU.remove(); _COL_D_MENU = null; }
  document.removeEventListener('click', _onDocClickForColDMenu, true);
  document.removeEventListener('keydown', _onEscForColDMenu, true);
}
function _onDocClickForColDMenu(e) {
  if (!_COL_D_MENU) return;
  if (!_COL_D_MENU.contains(e.target)) closeColDMenu();
}
function _onEscForColDMenu(e) {
  if (e.key === 'Escape') closeColDMenu();
}

function onColDContextMenu(e, rowNum) {
  // Right-click: suppress browser context menu, show our actions menu.
  e.preventDefault();
  e.stopPropagation();
  if (e.target.closest('td').classList.contains('editing')) return;  // inline-edit mode, skip
  selectRow(rowNum);     // make sure the row context is loaded
  showColDMenu(e, rowNum);
}
// Back-compat: old call-sites still wired to onColDClick
function onColDClick(e, rowNum) { return onColDContextMenu(e, rowNum); }

function showColDMenu(e, rowNum) {
  closeColDMenu();
  const row = ROWS_BY_NUM[rowNum];
  if (!row) return;
  const dVal = (row.D || '').trim();
  const isCommit = dVal.startsWith('ยินดีปฏิบัติ');
  const hasPdf = !!row.pdf_rel;
  const sec = row.section || '?';
  const parsed = row.parsed || {};
  const pType = parsed.type || 'empty';
  const isBrandModel = (pType === 'brand_model');
  const role = row.role || 'unknown';

  // Re-annotate label changes by role for clarity in the menu
  let reannLabel = 'Re-annotate (rect + label)';
  let reannHint  = 'ลบของเก่า แล้ววาดใหม่';
  if (isBrandModel) {
    reannLabel = 'Re-annotate ยี่ห้อ + รุ่น';
    reannHint  = '2 steps · brand → model';
  } else if (role === 'item') {
    reannLabel = `Re-annotate ข้อ ${parsed.item || '?'})`;
  } else if (role === 'sub_item') {
    reannLabel = `Re-annotate ข้อย่อย ${parsed.subitem || '?'}.`;
  } else if (role === 'section_header') {
    reannLabel = 'Re-annotate section';
  }

  const menu = document.createElement('div');
  menu.className = 'col-d-menu';
  let typeBadge = isBrandModel ? '· brand_model'
                : (isCommit ? '· commitment'
                : (pType !== 'empty' ? '· '+pType : ''));
  let html = `<div class="menu-header">R${rowNum} · ${escapeHtml(sec)} ${typeBadge}</div>`;

  if (isCommit) {
    // → Switch FROM commitment TO real annotation
    html += `<button class="primary" data-act="mark" ${!hasPdf?'disabled':''}>
      <span class="icon">${ico('pin',16)}</span>
      <span class="label">Mark in catalog → annotate</span>
      <span class="hint">${hasPdf?'รับ rect + label':'no PDF'}</span>
    </button>`;
    html += `<button data-act="reannotate">
      <span class="icon">${ico('refresh',16)}</span>
      <span class="label">${escapeHtml(reannLabel)}</span>
      <span class="hint">เลือก PDF ได้</span>
    </button>`;
    html += `<button data-act="auto" ${!hasPdf?'disabled':''}>
      <span class="icon">${ico('sparkles',16)}</span>
      <span class="label">Auto-annotate (AI tries again)</span>
      <span class="hint">preview ก่อน apply</span>
    </button>`;
    html += `<div class="sep"></div>`;
    html += `<button data-act="edit">
      <span class="icon">${ico('pencil',16)}</span>
      <span class="label">Edit Col D manually</span>
      <span class="hint">double-click ก็ได้</span>
    </button>`;
  } else {
    // → Already has a reference — re-annotate is the headline action
    html += `<button class="primary" data-act="reannotate">
      <span class="icon">${ico('refresh',16)}</span>
      <span class="label">${escapeHtml(reannLabel)}</span>
      <span class="hint">${escapeHtml(reannHint)}</span>
    </button>`;
    html += `<button data-act="auto">
      <span class="icon">${ico('sparkles',16)}</span>
      <span class="label">Auto-annotate (AI proposal)</span>
      <span class="hint">preview ก่อน apply</span>
    </button>`;
    html += `<button data-act="mark" ${!hasPdf?'disabled':''}>
      <span class="icon">${ico('pin',16)}</span>
      <span class="label">Add extra annotation</span>
      <span class="hint">ไม่ลบของเก่า</span>
    </button>`;
    html += `<button data-act="edit">
      <span class="icon">${ico('pencil',16)}</span>
      <span class="label">Edit Col D manually</span>
    </button>`;
    html += `<div class="sep"></div>`;
    html += `<button class="danger" data-act="revert">
      <span class="icon">${ico('rotate',16)}</span>
      <span class="label">Revert to "ยินดีปฏิบัติตามข้อกำหนด"</span>
      <span class="hint">บันทึกประวัติก่อน</span>
    </button>`;
  }

  menu.innerHTML = html;
  document.body.appendChild(menu);
  _COL_D_MENU = menu;

  // Position near the click, but keep on-screen
  const W = menu.offsetWidth, H = menu.offsetHeight;
  let x = e.clientX + 4, y = e.clientY + 4;
  if (x + W > window.innerWidth - 8)  x = window.innerWidth - W - 8;
  if (y + H > window.innerHeight - 8) y = e.clientY - H - 4;
  menu.style.left = `${Math.max(8, x)}px`;
  menu.style.top  = `${Math.max(8, y)}px`;

  // Wire button actions
  menu.querySelectorAll('button[data-act]').forEach(btn => {
    btn.addEventListener('click', (ev) => {
      ev.stopPropagation();
      const act = btn.dataset.act;
      closeColDMenu();
      handleColDAction(act, rowNum);
    });
  });

  setTimeout(() => {
    document.addEventListener('click', _onDocClickForColDMenu, true);
    document.addEventListener('keydown', _onEscForColDMenu, true);
  }, 0);
}

async function handleColDAction(action, rowNum) {
  const row = ROWS_BY_NUM[rowNum];
  if (!row) return;
  switch (action) {
    case 'mark':
      startManualAnnotate();
      break;
    case 'reannotate':
      startReannotate();
      break;
    case 'auto':
      showAutoAnnotate();
      break;
    case 'edit':
      // Programmatically focus the Col D cell and put it in edit mode
      const td = document.querySelector(`tr[data-row="${rowNum}"] td.col-D`);
      if (td) {
        const fakeEvent = {currentTarget: td, stopPropagation: () => {}};
        editColD(fakeEvent, rowNum);
      }
      break;
    case 'revert':
      await revertColDToCommitment(rowNum);
      break;
  }
}

async function revertColDToCommitment(rowNum) {
  const row = ROWS_BY_NUM[rowNum];
  if (!row) return;
  const old = row.D || '';
  if (old.trim().startsWith('ยินดีปฏิบัติ')) {
    toast('Already commitment', 'Col D เป็น "ยินดีปฏิบัติ" อยู่แล้ว', 'info', 2500);
    return;
  }
  if (!confirm(`Revert Col D ของ R${rowNum} เป็น "ยินดีปฏิบัติตามข้อกำหนด"?\n\n` +
               `เดิม: ${old.slice(0, 100)}\n\n` +
               `(ระบบ snapshot ก่อน — annotations ใน PDF ไม่ลบ)`)) return;
  try {
    const r = await fetch('/api/row/col_d', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        row: rowNum,
        col_d: 'ยินดีปฏิบัติตามข้อกำหนด',
        original: old,
      }),
    });
    const j = await r.json();
    if (j.ok) {
      toast(`↩ R${rowNum} reverted`, 'Col D = ยินดีปฏิบัติตามข้อกำหนด', 'learn', 4000);
      if (typeof tickRetrain === 'function') tickRetrain('revert');
      // Reload xlsx + tree
      if (SELECTED_ROW) loadXlsx(SELECTED_ROW);
      const idx = await fetch('/api/index').then(r => r.json());
      DATA = idx;
      ROWS_BY_NUM = Object.fromEntries(idx.rows.map(r => [r.row, r]));
      renderTree();
    } else {
      toast('Revert failed', j.error || 'unknown', 'error', 5000);
    }
  } catch (err) {
    toast('Revert failed', err.message, 'error', 5000);
  }
}

// ── Catalog PDF ────────────────────────────────────────────────
function computeHighlight(r) {
  const p = r.parsed;
  if (p.type === 'brand_model') return 'ยี่ห้อ|รุ่น';
  // Trailing period is what real annotation labels use ("ข้อย่อย 1.").
  // It's redundant for the structured matcher but safer if older annotations
  // ever fall back to substring matching.
  if (p.item != null && p.subitem != null) return `ข้อ ${p.item}) ข้อย่อย ${p.subitem}.`;
  if (p.subitem != null) return `ข้อย่อย ${p.subitem}.`;
  if (p.item != null) return `ข้อ ${p.item})`;
  return null;
}
async function loadPdf(rel, page, highlight) {
  CURRENT_PDF = {rel, meta: null};
  CURRENT_HIGHLIGHT = highlight;
  PDF_PAGE = page;
  document.getElementById('pdf-filename').textContent = rel;
  document.getElementById('pdf-canvas').innerHTML = skeletonImage();
  const r = await fetch('/api/pdf_meta?rel=' + encodeURIComponent(rel));
  if (!r.ok) {
    document.getElementById('pdf-canvas').innerHTML = '<div class="empty-canvas">โหลด PDF ไม่ได้</div>';
    return;
  }
  CURRENT_PDF.meta = await r.json();
  if (PDF_PAGE > CURRENT_PDF.meta.pages) PDF_PAGE = CURRENT_PDF.meta.pages;
  if (PDF_PAGE < 1) PDF_PAGE = 1;
  // If edit mode is on, re-initialize the working set with the new PDF's annots
  if (EDIT_MODE) {
    initEditAnnots();
  } else {
    EDIT_ANNOTS = [];
    ORIGINAL_ANNOTS_BY_ID = {};
    UNDO_STACK = [];
    REDO_STACK = [];
    SELECTED_ANN_ID = null;
  }
  renderPdf();
}
async function renderPdf() {
  if (!CURRENT_PDF) return;
  const useHl = !EDIT_MODE && document.getElementById('hl-toggle').checked ? CURRENT_HIGHLIGHT : null;
  let url = '/api/pdf_page?rel=' + encodeURIComponent(CURRENT_PDF.rel)
          + '&page=' + PDF_PAGE + '&dpi=' + PDF_DPI;
  if (EDIT_MODE) url += '&edit=1&_t=' + Date.now();
  if (useHl) url += '&highlight=' + encodeURIComponent(useHl);
  const c = document.getElementById('pdf-canvas');
  try {
    const r = await fetch(url);
    if (!r.ok) {
      c.innerHTML = '<div class="empty-canvas">โหลด PDF ไม่ได้</div>';
      return;
    }
    const hY0 = parseFloat(r.headers.get('X-Highlight-Y0'));
    const hY1 = parseFloat(r.headers.get('X-Highlight-Y1'));
    const blob = await r.blob();
    const imgUrl = URL.createObjectURL(blob);
    if (EDIT_MODE) {
      showPdfPageWithOverlay(c, imgUrl, hY0, hY1);
    } else {
      showImageWithScroll(c, imgUrl, hY0, hY1, `page ${PDF_PAGE}`);
    }
  } catch (e) {
    c.innerHTML = '<div class="empty-canvas">PDF render error</div>';
    return;
  }
  document.getElementById('pdf-page-info').textContent = `${PDF_PAGE} / ${CURRENT_PDF.meta.pages}`;
  renderAnnots();
}
// Structured matcher mirroring the Python _match_annot_label (handles
// "ข้อย่อย 1" ≠ "ข้อย่อย 10" and other digit-boundary cases).
const _LBL_ITEM_RE    = /ข้อ\s*(\d+)\s*\)/;
const _LBL_SUBITEM_RE = /ข้อย่อย\s*(\d+)/;
const _LBL_LITERALS   = new Set(['ยี่ห้อ', 'รุ่น']);
function _parseLabel(s) {
  if (!s) return {};
  s = s.trim();
  if (_LBL_LITERALS.has(s)) return {kind: 'literal', literal: s};
  const out = {};
  const i = s.match(_LBL_ITEM_RE);
  const sub = s.match(_LBL_SUBITEM_RE);
  if (i) out.item = parseInt(i[1]);
  if (sub) out.subitem = parseInt(sub[1]);
  if (Object.keys(out).length) out.kind = 'section_label';
  return out;
}
function matchAnnotLabel(query, content) {
  if (!query || !content) return false;
  const qf = _parseLabel(query);
  const cf = _parseLabel(content);
  if (qf.kind === 'literal') return cf.kind === 'literal' && cf.literal === qf.literal;
  if (qf.kind !== 'section_label') {
    if (query === content) return true;
    // word-boundary fallback: query not followed by another digit
    const re = new RegExp(query.replace(/[.*+?^${}()|[\]\\]/g, '\\$&') + '(?!\\d)');
    return re.test(content);
  }
  if (cf.kind !== 'section_label') return false;
  if (qf.subitem != null) {
    if (cf.subitem == null) {
      return qf.item != null && cf.item === qf.item;
    }
    if (cf.subitem !== qf.subitem) return false;
    if (qf.item != null && cf.item != null && cf.item !== qf.item) return false;
    return true;
  }
  if (qf.item != null) {
    return cf.item === qf.item && cf.subitem == null;
  }
  return false;
}

function renderAnnots() {
  const a = document.getElementById('pdf-annots');
  if (!CURRENT_PDF || !CURRENT_PDF.meta) { a.innerHTML = ''; return; }
  const annots = EDIT_MODE ? EDIT_ANNOTS.filter(x => !x._deleted) : CURRENT_PDF.meta.annots;
  const cur = annots.filter(x => x.page === PDF_PAGE);
  if (!cur.length) { a.innerHTML = '<div style="color:#9ca3af">(ไม่มี annotation บนหน้านี้)</div>'; return; }
  const queries = (CURRENT_HIGHLIGHT || '').split('|').map(s=>s.trim()).filter(Boolean);
  a.innerHTML = cur.map(x => {
    const m = queries.some(q => matchAnnotLabel(q, x.contents || ''));
    const sel = (x._id != null && x._id === SELECTED_ANN_ID) ? ' selected' : '';
    return `<div class="ann-row ${m?'matched':''}${sel}"><b>${x.type}</b> ${x.contents ? escapeHtml(x.contents) : '<i style="color:#9ca3af">(empty)</i>'}</div>`;
  }).join('');
}
function pdfPrev() { if (CURRENT_PDF && PDF_PAGE > 1) { PDF_PAGE--; renderPdf(); } }
function pdfNext() { if (CURRENT_PDF && PDF_PAGE < CURRENT_PDF.meta.pages) { PDF_PAGE++; renderPdf(); } }
function pdfZoom(d) { PDF_DPI = Math.max(70, Math.min(260, PDF_DPI + d*25)); renderPdf(); }
function openInBrowser() {
  if (!CURRENT_PDF) return;
  window.open('/api/raw_pdf?rel=' + encodeURIComponent(CURRENT_PDF.rel) + '#page=' + PDF_PAGE, '_blank');
}

// ── Edit-mode state + SVG overlay ──────────────────────────────
let EDIT_MODE = false;
let EDIT_TOOL = 'select';
let EDIT_ANNOTS = [];                  // working copy with _id, _deleted, _isNew flags
let ORIGINAL_ANNOTS_BY_ID = {};        // _id → original snapshot (for diff at save)
let SELECTED_ANN_ID = null;
let UNDO_STACK = [];
let REDO_STACK = [];
let DIRTY = false;
let NEW_ID_COUNTER = 0;
let DRAG_STATE = null;                 // {mode, ann, start, origRect}

function newClientId() { NEW_ID_COUNTER++; return 'new-' + NEW_ID_COUNTER; }

function snapshotAnnots() {
  return EDIT_ANNOTS.map(a => ({...a}));
}
function commitChange() {
  UNDO_STACK.push(snapshotAnnots());
  if (UNDO_STACK.length > 200) UNDO_STACK.shift();
  REDO_STACK = [];
  setDirty(true);
  refreshOverlay();
  refreshUndoRedoButtons();
}
function undo() {
  if (UNDO_STACK.length === 0) return;
  REDO_STACK.push(snapshotAnnots());
  EDIT_ANNOTS = UNDO_STACK.pop();
  if (!EDIT_ANNOTS.find(a => a._id === SELECTED_ANN_ID)) SELECTED_ANN_ID = null;
  setDirty(hasUnsavedChanges());
  refreshOverlay();
  refreshUndoRedoButtons();
}
function redo() {
  if (REDO_STACK.length === 0) return;
  UNDO_STACK.push(snapshotAnnots());
  EDIT_ANNOTS = REDO_STACK.pop();
  setDirty(hasUnsavedChanges());
  refreshOverlay();
  refreshUndoRedoButtons();
}
function hasUnsavedChanges() {
  for (const a of EDIT_ANNOTS) {
    if (a._isNew && !a._deleted) return true;
    if (a._deleted && !a._isNew) return true;
    if (!a._isNew) {
      const o = ORIGINAL_ANNOTS_BY_ID[a._id];
      if (!o) continue;
      if (a.contents !== o.contents) return true;
      const r1 = a.rect, r2 = o.rect;
      if (Math.abs(r1[0]-r2[0])>0.5 || Math.abs(r1[1]-r2[1])>0.5 ||
          Math.abs(r1[2]-r2[2])>0.5 || Math.abs(r1[3]-r2[3])>0.5) return true;
    }
  }
  return false;
}
function setDirty(v) {
  DIRTY = v;
  document.getElementById('save-btn').disabled = !v;
  document.getElementById('dirty-ind').textContent = v ? '● ยังไม่บันทึก' : '';
}
function refreshUndoRedoButtons() {
  document.getElementById('undo-btn').disabled = UNDO_STACK.length === 0;
  document.getElementById('redo-btn').disabled = REDO_STACK.length === 0;
}

function toggleEditMode() {
  if (!CURRENT_PDF) return;
  if (EDIT_MODE && DIRTY) {
    if (!confirm('มีการแก้ไขที่ยังไม่บันทึก — ออกจาก edit mode จะทิ้งงาน?')) return;
  }
  EDIT_MODE = !EDIT_MODE;
  document.getElementById('edit-toggle-btn').classList.toggle('active', EDIT_MODE);
  // Update edit/view toggle button label (preserves SVG icon + text span)
  const etb = document.getElementById('edit-toggle-btn');
  etb.innerHTML = EDIT_MODE
    ? `${ico('eye', 14)}<span>View</span>`
    : `${ico('pencil', 14)}<span>Edit</span>`;
  etb.setAttribute('aria-pressed', EDIT_MODE ? 'true' : 'false');
  document.getElementById('edit-toolbar').style.display = EDIT_MODE ? '' : 'none';
  const c = document.getElementById('pdf-canvas');
  c.classList.toggle('edit-mode', EDIT_MODE);
  if (EDIT_MODE) {
    initEditAnnots();
    setTool('select');
  } else {
    closeTextEditor();
    SELECTED_ANN_ID = null;
  }
  renderPdf();
}

function initEditAnnots() {
  EDIT_ANNOTS = [];
  ORIGINAL_ANNOTS_BY_ID = {};
  let inlineSeq = 0;
  for (const a of (CURRENT_PDF.meta.annots || [])) {
    // Inline annots all have xref=0 — use their inline_index to give each a
    // unique client _id, so click+select+delete works per-annot.
    let id;
    if (a.xref) {
      id = 'x' + a.xref;
    } else {
      const idx = (typeof a.inline_index === 'number') ? a.inline_index : inlineSeq++;
      id = 'inl-p' + a.page + '-' + idx;
    }
    const copy = {
      _id: id,
      xref: a.xref,
      _inline: !a.xref,                      // mark for delete-path routing
      _inline_index: a.inline_index,
      page: a.page,
      type: a.type,
      rect: a.rect.slice(),
      contents: a.contents,
    };
    EDIT_ANNOTS.push(copy);
    ORIGINAL_ANNOTS_BY_ID[id] = {
      contents: a.contents,
      rect: a.rect.slice(),
    };
  }
  UNDO_STACK = [];
  REDO_STACK = [];
  SELECTED_ANN_ID = null;
  setDirty(false);
  refreshUndoRedoButtons();
}

function setTool(name) {
  EDIT_TOOL = name;
  for (const b of document.querySelectorAll('.edit-toolbar button.tool')) {
    b.classList.toggle('active', b.dataset.tool === name);
  }
  const c = document.getElementById('pdf-canvas');
  c.classList.remove('tool-drawRect', 'tool-addText');
  if (name === 'drawRect') c.classList.add('tool-drawRect');
  else if (name === 'addText') c.classList.add('tool-addText');
}

// SVG overlay rendering
let OVERLAY_HOST = null;
let OVERLAY_SVG = null;

function showPdfPageWithOverlay(container, src, y0, y1) {
  container.innerHTML = '';
  const host = document.createElement('div');
  host.className = 'pdf-page-host';
  const img = document.createElement('img');
  img.className = 'pdf-page-img';
  host.appendChild(img);
  const svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
  svg.classList.add('pdf-overlay', 'editable');
  host.appendChild(svg);
  container.appendChild(host);
  OVERLAY_HOST = host;
  OVERLAY_SVG = svg;
  // Drag-to-create / click handlers on the SVG (attach before any paint)
  svg.addEventListener('pointerdown', onOverlayPointerDown);

  const onReady = () => {
    const pageSize = (CURRENT_PDF && CURRENT_PDF.meta && CURRENT_PDF.meta.page_sizes[PDF_PAGE - 1])
                     || [img.naturalWidth || 595, img.naturalHeight || 842];
    svg.setAttribute('viewBox', `0 0 ${pageSize[0]} ${pageSize[1]}`);
    svg.style.width = '100%';
    svg.style.height = '100%';
    refreshOverlay();
    // auto-scroll
    if (isFinite(y0)) {
      const scale = img.clientWidth / (img.naturalWidth || 1);
      const targetMid = ((y0 + (isFinite(y1) ? y1 : y0)) / 2) * scale;
      container.scrollTop = Math.max(0, targetMid - container.clientHeight * 0.25);
    }
  };
  // Bind handlers BEFORE setting src so a synchronous blob-URL load doesn't
  // race past the assignment (caused existing rects to be invisible after
  // toggling Edit mode — refreshOverlay() never fired).
  img.onload = onReady;
  img.onerror = () => {
    container.innerHTML = '<div class="empty-canvas">โหลด PDF page ไม่ได้</div>';
  };
  img.src = src;
  // Defensive: if the image is already complete (cached / blob loaded
  // synchronously), trigger the handler manually.
  if (img.complete && (img.naturalWidth > 0 || img.naturalHeight > 0)) {
    onReady();
  }
}

function refreshOverlay() {
  if (!OVERLAY_SVG) return;
  while (OVERLAY_SVG.firstChild) OVERLAY_SVG.removeChild(OVERLAY_SVG.firstChild);
  const onPage = EDIT_ANNOTS.filter(a => a.page === PDF_PAGE && !a._deleted);
  for (const a of onPage) {
    OVERLAY_SVG.appendChild(buildAnnotNode(a));
  }
  if (SELECTED_ANN_ID != null) {
    const sel = onPage.find(a => a._id === SELECTED_ANN_ID);
    if (sel) OVERLAY_SVG.appendChild(buildHandles(sel));
  }
  renderAnnots();
  // Phase A5: keep floating toolbar in sync with selection + annot pos
  if (typeof updateFloatingToolbar === 'function') updateFloatingToolbar();
}

// Compute the FreeText fontsize from rect height — must mirror the
// backend's apply_pdf_edits formula so edit-mode preview === saved PDF.
function freetextFontSize(rect) {
  const h = Math.max(0, rect[3] - rect[1]);
  return Math.max(6, Math.min(14, h * 0.65));
}

function buildAnnotNode(a) {
  const NS = 'http://www.w3.org/2000/svg';
  const g = document.createElementNS(NS, 'g');
  g.classList.add('annot');
  g.dataset.id = a._id;
  const isSelected = (a._id === SELECTED_ANN_ID);
  const isNew = !!a._isNew;
  if (isSelected) g.classList.add('selected');
  if (isNew) g.classList.add('is-new');
  const [x0, y0, x1, y1] = a.rect;
  const w = Math.max(0.5, x1 - x0);
  const h = Math.max(0.5, y1 - y0);

  // The page render NOW bakes in all existing annots (WYSIWYG with preview),
  // so the SVG only needs to:
  //   • show visible rect/text for NEWLY-drawn annots that aren't on disk yet
  //     OR for the currently selected annot (drag/resize feedback)
  //   • always provide an invisible hit area so click/hover/select work
  //
  // For existing annots in static view, ann-rect is rendered transparent
  // (no double-rect with the baked image).

  // Visible outline rect (transparent unless new/selected)
  const rect = document.createElementNS(NS, 'rect');
  rect.classList.add('ann-rect');
  if (a.type === 'FreeText') rect.classList.add('freetext');
  rect.setAttribute('x', x0); rect.setAttribute('y', y0);
  rect.setAttribute('width', w);
  rect.setAttribute('height', h);
  g.appendChild(rect);

  // Hit area for click/hover/select
  const hit = document.createElementNS(NS, 'rect');
  hit.classList.add('ann-hit');
  hit.setAttribute('x', x0); hit.setAttribute('y', y0);
  hit.setAttribute('width', w);
  hit.setAttribute('height', h);
  g.appendChild(hit);

  // Visible text only for newly-drawn FreeText (baked image already has
  // existing FreeText text; drawing again would show double).
  if (isNew && a.type === 'FreeText' && a.contents) {
    const fontSize = freetextFontSize(a.rect);
    const lines = a.contents.split('\n');
    const PAD_X = 2;
    const PAD_Y = 2;
    const lineH = fontSize * 1.15;
    for (let i = 0; i < lines.length; i++) {
      const t = document.createElementNS(NS, 'text');
      t.classList.add('ann-text');
      t.setAttribute('x', x0 + PAD_X);
      t.setAttribute('y', y0 + PAD_Y + i * lineH);
      t.setAttribute('font-size', fontSize);
      t.setAttribute('dominant-baseline', 'hanging');
      t.textContent = lines[i];
      g.appendChild(t);
    }
  }
  return g;
}

function buildHandles(a) {
  const NS = 'http://www.w3.org/2000/svg';
  const g = document.createElementNS(NS, 'g');
  g.classList.add('handles');
  const [x0, y0, x1, y1] = a.rect;
  const cx = (x0 + x1) / 2, cy = (y0 + y1) / 2;
  const handles = [
    ['nw', x0, y0], ['n', cx, y0], ['ne', x1, y0],
    ['w', x0, cy], ['e', x1, cy],
    ['sw', x0, y1], ['s', cx, y1], ['se', x1, y1],
  ];
  const HSIZE = 4;  // PDF-pt; will be smaller on screen due to viewBox scaling
  for (const [name, hx, hy] of handles) {
    const h = document.createElementNS(NS, 'rect');
    h.classList.add('handle', 'h-' + name);
    h.setAttribute('x', hx - HSIZE/2);
    h.setAttribute('y', hy - HSIZE/2);
    h.setAttribute('width', HSIZE);
    h.setAttribute('height', HSIZE);
    h.dataset.handle = name;
    g.appendChild(h);
  }
  return g;
}

// Pointer event coords: screen → PDF-pt
function eventToPdfPt(e) {
  const svg = OVERLAY_SVG;
  const pt = svg.createSVGPoint();
  pt.x = e.clientX; pt.y = e.clientY;
  const ctm = svg.getScreenCTM();
  if (!ctm) return [0,0];
  const inv = ctm.inverse();
  const p = pt.matrixTransform(inv);
  return [p.x, p.y];
}

function onOverlayPointerDown(e) {
  if (!EDIT_MODE) return;
  closeTextEditor();
  // Resize handle?
  if (e.target.classList.contains('handle')) {
    const sel = EDIT_ANNOTS.find(a => a._id === SELECTED_ANN_ID);
    if (!sel) return;
    // Snapshot BEFORE the drag starts so undo can restore it
    UNDO_STACK.push(snapshotAnnots());
    if (UNDO_STACK.length > 200) UNDO_STACK.shift();
    REDO_STACK = [];
    DRAG_STATE = {
      mode: 'resize',
      handle: e.target.dataset.handle,
      ann: sel,
      origRect: sel.rect.slice(),
      start: eventToPdfPt(e),
    };
    e.target.setPointerCapture(e.pointerId);
    e.preventDefault(); e.stopPropagation();
    return;
  }
  // Click on annotation?
  const annNode = e.target.closest('g.annot');
  if (annNode) {
    const id = annNode.dataset.id;
    SELECTED_ANN_ID = id;
    const sel = EDIT_ANNOTS.find(a => a._id === id);
    if (!sel) return;
    if (EDIT_TOOL === 'select') {
      // Snapshot before move starts
      UNDO_STACK.push(snapshotAnnots());
      if (UNDO_STACK.length > 200) UNDO_STACK.shift();
      REDO_STACK = [];
      DRAG_STATE = {
        mode: 'move',
        ann: sel,
        origRect: sel.rect.slice(),
        start: eventToPdfPt(e),
      };
      OVERLAY_SVG.setPointerCapture(e.pointerId);
    }
    refreshOverlay();
    e.preventDefault(); e.stopPropagation();
    return;
  }
  // Click on empty area
  if (EDIT_TOOL === 'drawRect') {
    const [x, y] = eventToPdfPt(e);
    DRAG_STATE = {
      mode: 'drawRect',
      start: [x, y],
      cur: [x, y],
    };
    OVERLAY_SVG.setPointerCapture(e.pointerId);
    drawPreviewRect();
  } else if (EDIT_TOOL === 'addText') {
    const [x, y] = eventToPdfPt(e);
    addTextAt(x, y);
  } else {
    SELECTED_ANN_ID = null;
    refreshOverlay();
  }
  e.preventDefault();
}

function drawPreviewRect() {
  if (!DRAG_STATE || DRAG_STATE.mode !== 'drawRect') return;
  // Remove previous preview
  const prev = OVERLAY_SVG.querySelector('rect.draw-preview');
  if (prev) prev.remove();
  const [x0, y0] = DRAG_STATE.start, [x1, y1] = DRAG_STATE.cur;
  const r = document.createElementNS('http://www.w3.org/2000/svg', 'rect');
  r.classList.add('draw-preview');
  r.setAttribute('x', Math.min(x0, x1));
  r.setAttribute('y', Math.min(y0, y1));
  r.setAttribute('width', Math.abs(x1 - x0));
  r.setAttribute('height', Math.abs(y1 - y0));
  OVERLAY_SVG.appendChild(r);
}

document.addEventListener('pointermove', e => {
  if (!DRAG_STATE) return;
  const [px, py] = eventToPdfPt(e);
  if (DRAG_STATE.mode === 'move') {
    const dx = px - DRAG_STATE.start[0], dy = py - DRAG_STATE.start[1];
    const o = DRAG_STATE.origRect;
    DRAG_STATE.ann.rect = [o[0]+dx, o[1]+dy, o[2]+dx, o[3]+dy];
    refreshOverlay();
  } else if (DRAG_STATE.mode === 'resize') {
    const dx = px - DRAG_STATE.start[0], dy = py - DRAG_STATE.start[1];
    const o = DRAG_STATE.origRect;
    let r = o.slice();
    const h = DRAG_STATE.handle;
    if (h.includes('n')) r[1] = o[1] + dy;
    if (h.includes('s')) r[3] = o[3] + dy;
    if (h.includes('w')) r[0] = o[0] + dx;
    if (h.includes('e')) r[2] = o[2] + dx;
    // ensure x0 < x1, y0 < y1
    if (r[0] > r[2]) [r[0], r[2]] = [r[2], r[0]];
    if (r[1] > r[3]) [r[1], r[3]] = [r[3], r[1]];
    DRAG_STATE.ann.rect = r;
    refreshOverlay();
  } else if (DRAG_STATE.mode === 'drawRect') {
    DRAG_STATE.cur = [px, py];
    drawPreviewRect();
  }
});
document.addEventListener('pointerup', e => {
  if (!DRAG_STATE) return;
  if (DRAG_STATE.mode === 'move' || DRAG_STATE.mode === 'resize') {
    const ann = DRAG_STATE.ann;
    const orig = DRAG_STATE.origRect;
    const moved = ann.rect.some((v,i) => Math.abs(v - orig[i]) > 0.1);
    if (moved) {
      setDirty(true);
      refreshUndoRedoButtons();
    } else {
      // No real change — discard the speculative undo we pushed at pointerdown
      UNDO_STACK.pop();
      refreshUndoRedoButtons();
    }
  } else if (DRAG_STATE.mode === 'drawRect') {
    const [x0, y0] = DRAG_STATE.start, [x1, y1] = DRAG_STATE.cur;
    const w = Math.abs(x1 - x0), h = Math.abs(y1 - y0);
    if (w >= 4 && h >= 4) {
      addRectAt(Math.min(x0,x1), Math.min(y0,y1), Math.max(x0,x1), Math.max(y0,y1));
    }
    const prev = OVERLAY_SVG.querySelector('rect.draw-preview');
    if (prev) prev.remove();
  }
  DRAG_STATE = null;
});

// Helper: push current state to undo stack (BEFORE applying a change).
function _commitBeforeChange() {
  UNDO_STACK.push(snapshotAnnots());
  if (UNDO_STACK.length > 200) UNDO_STACK.shift();
  REDO_STACK = [];
}

// ── Add rect / text / delete ───────────────────────────────────
function addRectAt(x0, y0, x1, y1) {
  _commitBeforeChange();
  const id = newClientId();
  const ann = {
    _id: id, _isNew: true,
    xref: null,
    page: PDF_PAGE,
    type: 'Square',
    rect: [x0, y0, x1, y1],
    contents: '',
  };
  EDIT_ANNOTS.push(ann);
  SELECTED_ANN_ID = id;
  setDirty(true);
  refreshOverlay();
  refreshUndoRedoButtons();
  setTool('select');
}
function addTextAt(x, y, defaultText) {
  _commitBeforeChange();
  const id = newClientId();
  const w = 130, h = 14;
  const ann = {
    _id: id, _isNew: true,
    xref: null,
    page: PDF_PAGE,
    type: 'FreeText',
    rect: [x, y, x + w, y + h],
    contents: defaultText || 'text',
  };
  EDIT_ANNOTS.push(ann);
  SELECTED_ANN_ID = id;
  setDirty(true);
  refreshOverlay();
  refreshUndoRedoButtons();
  // Open inline editor immediately
  setTool('select');
  setTimeout(() => openTextEditor(ann), 50);
}
function deleteSelected() {
  if (!SELECTED_ANN_ID) return;
  const idx = EDIT_ANNOTS.findIndex(a => a._id === SELECTED_ANN_ID);
  if (idx < 0) return;
  _commitBeforeChange();
  const a = EDIT_ANNOTS[idx];
  if (a._isNew) {
    EDIT_ANNOTS.splice(idx, 1);
  } else {
    a._deleted = true;
  }
  SELECTED_ANN_ID = null;
  setDirty(true);
  refreshOverlay();
  refreshUndoRedoButtons();
}

// Phase A5: clone the selected annot with a small offset so the user
// sees both original and copy. The clone is a NEW annot (no xref);
// when the user saves, apply_pdf_edits mints it as a fresh PDF annot.
function duplicateSelected() {
  if (!SELECTED_ANN_ID) return;
  const orig = EDIT_ANNOTS.find(a => a._id === SELECTED_ANN_ID);
  if (!orig) return;
  _commitBeforeChange();
  const OFF = 12;  // PDF-pt offset
  const id = newClientId();
  const r = orig.rect;
  // Clamp the offset so the duplicate doesn't fall off the page
  const pageSize = (CURRENT_PDF && CURRENT_PDF.meta && CURRENT_PDF.meta.page_sizes[PDF_PAGE - 1]) || [595, 842];
  const w = r[2] - r[0], h = r[3] - r[1];
  const dx = (r[2] + OFF <= pageSize[0]) ? OFF : -OFF;
  const dy = (r[3] + OFF <= pageSize[1]) ? OFF : -OFF;
  const clone = {
    _id: id, _isNew: true,
    xref: null,
    page: orig.page,
    type: orig.type,
    rect: [r[0] + dx, r[1] + dy, r[0] + dx + w, r[1] + dy + h],
    contents: orig.contents || '',
  };
  EDIT_ANNOTS.push(clone);
  SELECTED_ANN_ID = id;
  setDirty(true);
  refreshOverlay();
  refreshUndoRedoButtons();
}

// ── Phase A5: floating annotation toolbar ──────────────────────
// Mounted on <body> to escape stacking contexts. Driven by:
//   • SELECTED_ANN_ID  — show/hide
//   • EDIT_MODE        — only visible in edit/reannotate
//   • refreshOverlay() — repositions after any visible change
//   • catalog scroll/zoom/page change — repositions
let _FAT_RAF = 0;
function _fatEl() { return document.getElementById('float-annot-toolbar'); }
function updateFloatingToolbar() {
  // rAF coalesce — if multiple events fire in one frame, do one paint
  if (_FAT_RAF) return;
  _FAT_RAF = requestAnimationFrame(() => {
    _FAT_RAF = 0;
    _updateFloatingToolbarNow();
  });
}
function _updateFloatingToolbarNow() {
  const tb = _fatEl();
  if (!tb) return;
  const ann = SELECTED_ANN_ID
    ? EDIT_ANNOTS.find(a => a._id === SELECTED_ANN_ID && !a._deleted)
    : null;
  if (!EDIT_MODE || !ann || !OVERLAY_SVG) {
    tb.classList.remove('visible');
    return;
  }
  // Find the rendered SVG group for this annot
  const node = OVERLAY_SVG.querySelector('g.annot[data-id="' + (CSS.escape ? CSS.escape(ann._id) : ann._id) + '"]');
  if (!node) {
    tb.classList.remove('visible');
    return;
  }
  // Clip-aware: hide if the annot scrolled out of the catalog pane
  const rect = node.getBoundingClientRect();
  if (rect.width < 1 || rect.height < 1) {
    tb.classList.remove('visible');
    return;
  }
  // Catalog viewport bounds (the .pdf-canvas / pdf column that owns the SVG)
  const pdfWrap = OVERLAY_SVG.closest('.pdf-canvas') || OVERLAY_SVG.closest('.pdf-pane') || OVERLAY_SVG.parentElement;
  if (pdfWrap) {
    const wrapR = pdfWrap.getBoundingClientRect();
    const intersects =
      rect.right >= wrapR.left && rect.left <= wrapR.right &&
      rect.bottom >= wrapR.top && rect.top <= wrapR.bottom;
    if (!intersects) {
      tb.classList.remove('visible');
      return;
    }
  }
  // Update labels
  const typeLbl = document.getElementById('fat-type-label');
  const typeIco = tb.querySelector('.fat-type use');
  if (typeLbl) typeLbl.textContent = ann.type === 'FreeText' ? 'FreeText' : 'Square';
  if (typeIco) typeIco.setAttribute('href', ann.type === 'FreeText' ? '#i-text' : '#i-square');
  const swatch = document.getElementById('fat-color-swatch');
  if (swatch) swatch.style.background = 'rgb(255,0,0)';
  const meta = document.getElementById('fat-meta');
  if (meta) meta.textContent = ann.type === 'FreeText' ? 'red text' : '1pt red';

  // Position above (or below if no room)
  tb.classList.add('visible');
  // Force a reflow so width is measured with the new content
  const tbW = tb.offsetWidth || 220;
  const tbH = tb.offsetHeight || 40;
  const GAP = 8;
  const vw = window.innerWidth, vh = window.innerHeight;

  let top = rect.top - tbH - GAP;
  let placeBelow = false;
  if (top < 8) {
    top = rect.bottom + GAP;
    placeBelow = true;
  }
  // Center on the annot but clamp to viewport
  let left = rect.left + rect.width / 2 - tbW / 2;
  left = Math.max(8, Math.min(left, vw - tbW - 8));
  if (top + tbH > vh - 8) top = vh - tbH - 8;

  tb.style.left = Math.round(left) + 'px';
  tb.style.top = Math.round(top) + 'px';
  tb.classList.toggle('below', placeBelow);

  // Position the arrow pointing back at the annot
  const arrow = tb.querySelector('.fat-arrow');
  if (arrow) {
    const annCenterX = rect.left + rect.width / 2;
    let arrowX = annCenterX - left - 5;  // arrow is 10px wide
    arrowX = Math.max(8, Math.min(arrowX, tbW - 18));
    arrow.style.left = Math.round(arrowX) + 'px';
  }
}
// Repaint on common reposition triggers
window.addEventListener('scroll', updateFloatingToolbar, true);
window.addEventListener('resize', updateFloatingToolbar);

// ── Inline text editor ─────────────────────────────────────────
let TEXT_EDITOR = null;
function openTextEditor(a) {
  closeTextEditor();
  if (!OVERLAY_HOST || a.type !== 'FreeText') return;
  const img = OVERLAY_HOST.querySelector('img.pdf-page-img');
  if (!img) return;
  const pageSize = CURRENT_PDF.meta.page_sizes[PDF_PAGE - 1];
  const scaleX = img.clientWidth / pageSize[0];
  const scaleY = img.clientHeight / pageSize[1];
  const ed = document.createElement('textarea');
  ed.className = 'text-editor';
  ed.value = a.contents || '';
  ed.style.left = (a.rect[0] * scaleX) + 'px';
  ed.style.top = (a.rect[1] * scaleY) + 'px';
  ed.style.width = ((a.rect[2] - a.rect[0]) * scaleX) + 'px';
  ed.style.height = Math.max(20, (a.rect[3] - a.rect[1]) * scaleY) + 'px';
  ed.style.fontSize = Math.max(8, (a.rect[3] - a.rect[1]) * scaleY * 0.7) + 'px';
  OVERLAY_HOST.appendChild(ed);
  ed.focus();
  ed.select();
  TEXT_EDITOR = {ed, ann: a, originalContents: a.contents};
  ed.addEventListener('blur', commitTextEditor);
  ed.addEventListener('keydown', e => {
    if (e.key === 'Escape') { e.preventDefault(); cancelTextEditor(); }
    else if (e.key === 'Enter' && (e.metaKey || e.ctrlKey)) { e.preventDefault(); commitTextEditor(); }
  });
}
function commitTextEditor() {
  if (!TEXT_EDITOR) return;
  const newVal = TEXT_EDITOR.ed.value;
  const a = TEXT_EDITOR.ann;
  if (newVal !== TEXT_EDITOR.originalContents) {
    _commitBeforeChange();
    a.contents = newVal;
    setDirty(true);
    refreshUndoRedoButtons();
  }
  TEXT_EDITOR.ed.remove();
  TEXT_EDITOR = null;
  refreshOverlay();
}
function cancelTextEditor() {
  if (!TEXT_EDITOR) return;
  TEXT_EDITOR.ed.remove();
  TEXT_EDITOR = null;
}
function closeTextEditor() { commitTextEditor(); }

// Double-click annotation → open text editor (FreeText only)
document.addEventListener('dblclick', e => {
  if (!EDIT_MODE) return;
  const annNode = e.target.closest('g.annot');
  if (!annNode) return;
  const id = annNode.dataset.id;
  const a = EDIT_ANNOTS.find(x => x._id === id);
  if (a && a.type === 'FreeText') {
    SELECTED_ANN_ID = id;
    openTextEditor(a);
  }
});

// ── Save ───────────────────────────────────────────────────────
function computeEditDiff() {
  const edits = [];
  for (const a of EDIT_ANNOTS) {
    if (a._isNew && !a._deleted) {
      const e = {action:'create', client_id: a._id, page: a.page,
                 type: a.type, rect: a.rect, contents: a.contents};
      // Pin fontsize for FreeText so the saved PDF text size matches the
      // SVG overlay the user just saw. Same formula as freetextFontSize().
      if (a.type === 'FreeText') e.fontsize = freetextFontSize(a.rect);
      edits.push(e);
    } else if (!a._isNew && a._deleted) {
      if (a._inline || !a.xref) {
        // Inline annot — backend rewrites the page's /Annots array
        edits.push({action: 'delete_inline',
                    page: a.page,
                    inline_index: a._inline_index,
                    type: a.type,
                    rect: a.rect,
                    contents: a.contents});
      } else {
        edits.push({action: 'delete', xref: a.xref});
      }
    } else if (!a._isNew && !a._deleted) {
      const o = ORIGINAL_ANNOTS_BY_ID[a._id];
      if (!o) continue;
      const u = {action: 'update', xref: a.xref};
      let dirty = false;
      if (a.contents !== o.contents) { u.contents = a.contents; dirty = true; }
      const r1 = a.rect, r2 = o.rect;
      if (Math.abs(r1[0]-r2[0])>0.5 || Math.abs(r1[1]-r2[1])>0.5 ||
          Math.abs(r1[2]-r2[2])>0.5 || Math.abs(r1[3]-r2[3])>0.5) {
        u.rect = a.rect; dirty = true;
      }
      if (dirty) edits.push(u);
    }
  }
  return edits;
}
async function saveEdits() {
  if (!CURRENT_PDF || !DIRTY) return;
  closeTextEditor();
  const edits = computeEditDiff();
  if (!edits.length) { setDirty(false); return; }
  const r = await fetch('/api/pdf_save', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({rel: CURRENT_PDF.rel, edits}),
  });
  const j = await r.json();
  if (!j.ok) {
    alert('Save failed: ' + (j.error || JSON.stringify(j)));
    return;
  }
  // Reload PDF + meta to get fresh xrefs
  const meta = await fetch('/api/pdf_meta?rel=' + encodeURIComponent(CURRENT_PDF.rel)).then(r => r.json());
  CURRENT_PDF.meta = meta;
  initEditAnnots();
  setDirty(false);
  renderPdf();
  console.log(`✓ saved: ${j.applied} applied, ${j.errors} errors, snapshot: ${j.snapshot}`);
}

// ── History panel ──────────────────────────────────────────────
async function showHistory() {
  if (!CURRENT_PDF) return;
  const r = await fetch('/api/pdf_history?rel=' + encodeURIComponent(CURRENT_PDF.rel));
  const j = await r.json();
  const list = document.getElementById('history-list');
  list.innerHTML = '';
  document.getElementById('history-info').textContent =
    `${j.snapshots.length} snapshots — ${CURRENT_PDF.rel}`;
  for (const s of j.snapshots) {
    const ts = new Date(s.mtime * 1000).toLocaleString();
    const li = document.createElement('li');
    li.className = 'history-item';
    li.innerHTML = `
      <span class="ts">${escapeHtml(s.id)}</span>
      <span class="size">${(s.size/1024).toFixed(1)} KB · ${escapeHtml(ts)}</span>
      <button onclick="restoreSnapshot('${escapeHtml(s.id)}')">↺ Restore</button>`;
    list.appendChild(li);
  }
  if (!j.snapshots.length) {
    list.innerHTML = '<div style="padding:10px;color:#9ca3af;">ยังไม่มี snapshot</div>';
  }
  document.getElementById('history-modal').classList.add('show');
}
function closeHistory() {
  document.getElementById('history-modal').classList.remove('show');
}
async function restoreSnapshot(snapshotId) {
  if (!confirm(`Restore PDF จาก snapshot "${snapshotId}"?\nสถานะปัจจุบันจะถูก snapshot ก่อน restore`)) return;
  const r = await fetch('/api/pdf_restore', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({rel: CURRENT_PDF.rel, snapshot: snapshotId}),
  });
  const j = await r.json();
  if (!j.ok) { alert('Restore failed: ' + (j.error || '?')); return; }
  closeHistory();
  // reload
  const meta = await fetch('/api/pdf_meta?rel=' + encodeURIComponent(CURRENT_PDF.rel)).then(r => r.json());
  CURRENT_PDF.meta = meta;
  if (EDIT_MODE) initEditAnnots();
  renderPdf();
}
async function manualSnapshot() {
  if (!CURRENT_PDF) return;
  await fetch('/api/pdf_snapshot', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({rel: CURRENT_PDF.rel, tag: 'manual'}),
  });
  showHistory();  // refresh list
}

// ── Project versions modal (wraps scripts/version.py) ─────────
let DIFF_SELECTION = [];   // up to 2 ids selected for diff

function showVersions() {
  document.getElementById('versions-modal').classList.add('show');
  loadVersions();
}
function closeVersions() {
  document.getElementById('versions-modal').classList.remove('show');
  hideVersionsBusy();
  document.getElementById('versions-output').style.display = 'none';
  DIFF_SELECTION = [];
}
function showVersionsBusy(msg) {
  document.getElementById('versions-busy-msg').textContent = msg || 'working…';
  document.getElementById('versions-busy').style.display = '';
}
function hideVersionsBusy() {
  document.getElementById('versions-busy').style.display = 'none';
}
function showVersionsOutput(text) {
  const pre = document.getElementById('versions-output');
  pre.textContent = text || '(no output)';
  pre.style.display = '';
}

async function loadVersions() {
  const list = document.getElementById('versions-list');
  list.innerHTML = '<div style="color:#9ca3af;padding:20px;text-align:center;">Loading…</div>';
  try {
    const [rv, rs] = await Promise.all([
      fetch('/api/versions'),
      fetch('/api/versions/sync'),
    ]);
    const j = await rv.json();
    const sync = await rs.json();
    document.getElementById('versions-info').innerHTML =
      j.available
        ? `${ico('folder',14)} ${escapeHtml(j.root)} · <strong>${j.snapshots.length}</strong> snapshot(s)`
        : '<span style="color:#b91c1c">scripts/version.py ไม่พบ — ระบบ versioning ไม่พร้อม</span>';
    renderSyncBadge(sync);
    refreshSyncIndicators(sync);
    if (!j.snapshots.length) {
      list.innerHTML = '<div style="color:#9ca3af;padding:24px;text-align:center;">ยังไม่มี snapshot — กด Quick snap เพื่อเริ่ม</div>';
      return;
    }
    // Mark the current latest visually so users see at a glance which is "live"
    const latestId = sync && sync.latest && sync.latest.id;
    list.innerHTML = j.snapshots.map(s => renderVersionItem(s, s.id === latestId)).join('');
    updateDiffHelper();
  } catch (e) {
    list.innerHTML = `<div style="color:#b91c1c;padding:20px;">โหลดไม่ได้: ${escapeHtml(e.message)}</div>`;
  }
}

const _SYNC_LABELS = {
  in_sync:           {label: '✓ ตรงกับ snapshot ล่าสุด',         cls: 'in-sync'},
  working_ahead:     {label: '⬆ working dir ใหม่กว่า snapshot',  cls: 'working-ahead'},
  working_behind:    {label: '⬇ working dir เก่ากว่า snapshot',  cls: 'working-behind'},
  divergent:         {label: '⚠ ต่างจาก snapshot ล่าสุด',         cls: 'divergent'},
  incomplete_local:  {label: '⚠ ไฟล์หายจาก working dir',          cls: 'incomplete-local'},
  no_snapshots:      {label: 'ยังไม่มี snapshot',                  cls: 'no-snapshots'},
};

function renderSyncBadge(sync) {
  const host = document.getElementById('versions-sync-badge');
  if (!sync || !host) { if (host) host.innerHTML = ''; return; }
  const meta = _SYNC_LABELS[sync.state] || _SYNC_LABELS.no_snapshots;
  const latest = sync.latest;
  let detail = '';
  if (latest) {
    detail = `<span class="latest-id">↳ ${escapeHtml(latest.id)}` +
             `${latest.tag ? ' · ' + escapeHtml(latest.tag) : ''}</span>`;
  }
  host.innerHTML = `<div class="sync-badge ${meta.cls}">${escapeHtml(meta.label)} ${detail}</div>`;
}

function refreshSyncIndicators(sync) {
  // 1. Top-level Versions button — change colour for warn / danger states
  const btn = document.querySelector('.versions-btn');
  if (btn) {
    btn.classList.remove('warn', 'danger');
    if (sync && (sync.state === 'working_behind' || sync.state === 'divergent' ||
                 sync.state === 'incomplete_local')) {
      btn.classList.add('danger');
    } else if (sync && sync.state === 'working_ahead') {
      btn.classList.add('warn');
    }
  }
  // 2. Top banner — only for behind / divergent / incomplete (these need user
  //    decision; "ahead" is auto-handled at boot so no banner needed)
  const banner = document.getElementById('sync-banner');
  const app = document.getElementById('app');
  if (!sync || !banner) return;
  if (window._SYNC_BANNER_DISMISSED) {
    banner.classList.remove('show');
    if (app) app.classList.remove('has-sync-banner');
    return;
  }
  let bannerCls = '', msg = '', actions = '';
  if (sync.state === 'working_behind') {
    bannerCls = 'danger';
    msg = `⚠ Working dir เก่ากว่า snapshot ล่าสุด (${escapeHtml(sync.latest.id)}). โหลดเวอร์ชันล่าสุด?`;
    actions = `<button onclick="quickRestoreLatest(false)">↺ Restore xlsx</button>`
            + `<button onclick="quickRestoreLatest(true)">↻ Restore full</button>`;
  } else if (sync.state === 'divergent') {
    bannerCls = 'danger';
    msg = `⚠ ต่างจาก snapshot ล่าสุด (${escapeHtml(sync.latest.id)})`;
    actions = `<button onclick="showVersions()">${ico('package',14)} ตรวจสอบ</button>`;
  } else if (sync.state === 'incomplete_local') {
    bannerCls = 'danger';
    msg = `⚠ ไฟล์บางส่วนหาย — เทียบกับ snapshot ${escapeHtml(sync.latest.id)}`;
    actions = `<button onclick="showVersions()">${ico('package',14)} ตรวจสอบ</button>`;
  }
  if (msg) {
    banner.className = 'sync-banner show ' + bannerCls;
    document.getElementById('sync-banner-msg').textContent = '';
    document.getElementById('sync-banner-msg').textContent = msg;
    document.getElementById('sync-banner-actions').innerHTML = actions;
    if (app) app.classList.add('has-sync-banner');
  } else {
    banner.classList.remove('show');
    if (app) app.classList.remove('has-sync-banner');
  }
}

function dismissSyncBanner() {
  window._SYNC_BANNER_DISMISSED = true;
  document.getElementById('sync-banner').classList.remove('show');
  const app = document.getElementById('app');
  if (app) app.classList.remove('has-sync-banner');
}

async function quickRestoreLatest(full) {
  const r = await fetch('/api/versions/sync');
  const sync = await r.json();
  if (!sync.latest) { alert('ไม่มี snapshot ให้ restore'); return; }
  await restoreVersion(sync.latest.id, full);
}

function renderVersionItem(s, isLatest) {
  const sizeMB = (s.size / (1024 * 1024)).toFixed(1);
  const ts = s.timestamp ? new Date(s.timestamp).toLocaleString() : '';
  const diffSel = DIFF_SELECTION.includes(s.id) ? ' diff-selected' : '';
  const latestBadge = isLatest ? ' <span style="background:#10b981;color:white;padding:1px 6px;border-radius:3px;font-size:9px;font-weight:600;">LATEST</span>' : '';
  return `
    <div class="version-item${diffSel}" style="${isLatest ? 'border-left: 3px solid #10b981;' : ''}">
      <input type="checkbox" class="v-check" ${DIFF_SELECTION.includes(s.id) ? 'checked' : ''}
             onchange="toggleDiffSelect('${escapeHtml(s.id)}', this.checked)"
             title="เลือกเพื่อ diff">
      <div>
        <div class="v-tag">
          <span class="v-kind ${s.kind}">${s.kind}</span>
          ${s.tag ? escapeHtml(s.tag) : '<span class="untagged">(no tag)</span>'}
          ${latestBadge}
        </div>
        <div class="v-meta">${escapeHtml(s.id)} · ${sizeMB} MB · ${s.n_output} files · ${escapeHtml(ts)}</div>
      </div>
      <div class="v-actions">
        <button onclick="diffOne('${escapeHtml(s.id)}')" title="Diff กับ snapshot ล่าสุด">diff</button>
        <button class="restore" onclick="restoreVersion('${escapeHtml(s.id)}', false)">↺ restore</button>
        ${s.has_tarball ? `<button class="restore-full" onclick="restoreVersion('${escapeHtml(s.id)}', true)" title="restore output/ ทั้งหมด">↻ full</button>` : ''}
      </div>
    </div>`;
}

function toggleDiffSelect(id, checked) {
  if (checked) {
    if (!DIFF_SELECTION.includes(id)) DIFF_SELECTION.push(id);
    if (DIFF_SELECTION.length > 2) DIFF_SELECTION.shift();
  } else {
    DIFF_SELECTION = DIFF_SELECTION.filter(x => x !== id);
  }
  loadVersions();
  if (DIFF_SELECTION.length === 2) {
    runDiff(DIFF_SELECTION[0], DIFF_SELECTION[1]);
  }
}

function updateDiffHelper() {
  const helper = document.getElementById('diff-helper');
  if (DIFF_SELECTION.length === 0) helper.textContent = 'เลือก 2 snapshots เพื่อ diff';
  else if (DIFF_SELECTION.length === 1) helper.textContent = 'เลือกอีก 1 snapshot';
  else helper.textContent = `diff: ${DIFF_SELECTION[0]} ↔ ${DIFF_SELECTION[1]}`;
}

async function runDiff(id1, id2) {
  showVersionsBusy('comparing snapshots…');
  try {
    const r = await fetch(`/api/versions/diff?id1=${encodeURIComponent(id1)}&id2=${encodeURIComponent(id2)}`);
    const j = await r.json();
    showVersionsOutput(j.stdout || j.error || '(empty)');
  } finally {
    hideVersionsBusy();
  }
}
async function diffOne(id) {
  // diff between this snapshot and the latest one
  const r = await fetch('/api/versions');
  const j = await r.json();
  const latest = j.snapshots[0];
  if (!latest || latest.id === id) {
    showVersionsOutput('(this is already the latest snapshot)');
    return;
  }
  await runDiff(id, latest.id);
}

async function takeProjectSnap(full) {
  const tag = document.getElementById('versions-tag').value.trim();
  if (full) {
    if (!confirm('Full snapshot จะ tar.gz ทั้ง output/ (~200 MB) ใช้เวลาประมาณ 30-60 วินาที — ดำเนินการต่อ?')) return;
  }
  showVersionsBusy(full ? 'compressing output/ (อาจใช้เวลา 30-60s)…' : 'snapshotting…');
  try {
    const r = await fetch('/api/versions/snap', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({tag, full}),
    });
    const j = await r.json();
    showVersionsOutput(j.stdout || j.stderr || j.error || '(no output)');
    if (j.ok) document.getElementById('versions-tag').value = '';
    await loadVersions();
  } finally {
    hideVersionsBusy();
  }
}

async function autoSnapNow() {
  showVersionsBusy('checking xlsx for changes…');
  try {
    const r = await fetch('/api/versions/auto-snap', {method: 'POST'});
    const j = await r.json();
    showVersionsOutput(j.stdout || j.error || '(no output)');
    await loadVersions();
  } finally {
    hideVersionsBusy();
  }
}

async function pruneVersions() {
  const keep = parseInt(prompt('เก็บ N snapshots ล่าสุด (ลบที่เหลือ):', '10'));
  if (!keep || keep < 1) return;
  showVersionsBusy(`pruning to ${keep}…`);
  try {
    const r = await fetch('/api/versions/prune', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({keep}),
    });
    const j = await r.json();
    showVersionsOutput(j.stdout || '(no output)');
    await loadVersions();
  } finally {
    hideVersionsBusy();
  }
}

async function restoreVersion(id, full) {
  const msg = full
    ? `⚠️ FULL RESTORE จะแทนที่ทั้ง output/ ของคุณด้วย snapshot ${id}\n` +
      `(ระบบจะ snapshot สถานะปัจจุบันก่อนเสมอ)\n\nดำเนินการต่อ?`
    : `Restore xlsx + SKILL.md จาก ${id}?\n` +
      `(ระบบจะ snapshot สถานะปัจจุบันก่อน)`;
  if (!confirm(msg)) return;
  showVersionsBusy(full ? 'restoring full output/ (อาจใช้เวลา)…' : 'restoring…');
  try {
    const r = await fetch('/api/versions/restore', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({id, full}),
    });
    const j = await r.json();
    showVersionsOutput(j.stdout || j.stderr || j.error || '(no output)');
    if (j.ok) {
      await loadVersions();
      alert('✓ Restore เสร็จแล้ว — กำลัง reload หน้าเพื่อ reset state');
      setTimeout(() => location.reload(), 500);
    }
  } finally {
    hideVersionsBusy();
  }
}

// ── Auto-annotate modal ───────────────────────────────────────
let _AUTO_PLAN = null;
let _AUTO_BATCH_PLANS = [];

function showAutoAnnotate() {
  if (!SELECTED_ROW) { alert('เลือก row ก่อน'); return; }
  document.getElementById('auto-modal').classList.add('show');
  document.getElementById('auto-single').style.display = '';
  document.getElementById('auto-batch').style.display = 'none';
  document.getElementById('auto-output').style.display = 'none';
  loadAutoPreview(SELECTED_ROW);
}
function closeAuto() {
  document.getElementById('auto-modal').classList.remove('show');
}

async function loadAutoPreview(rowNum) {
  const r = await fetch(`/api/auto_annotate/preview?row=${rowNum}`);
  const plan = await r.json();
  _AUTO_PLAN = plan;
  const role = (plan.role && plan.role.role) || 'unknown';
  document.getElementById('auto-title').textContent =
    `R${rowNum} · section ${plan.section || '?'} · role: ${role}`;
  // Confidence badge — high/med/low based on threshold
  const cf = plan.confidence ?? 0;
  const cfCls = cf >= 0.85 ? 'high' : cf >= 0.65 ? 'med' : 'low';
  const genTag = plan.generator
    ? `<span class="src">${escapeHtml(plan.generator)}</span>` : '';
  const cfBadge = `<span class="conf-badge ${cfCls}">${(cf*100).toFixed(0)}% ${genTag}</span>`;
  let provLine = '';
  if (plan.provenance && Object.keys(plan.provenance).length) {
    provLine = `<br>🧠 <em>learned</em>: ` +
      Object.entries(plan.provenance).map(([k, v]) =>
        `${k} ← <code>${escapeHtml(v.trigger || '')}</code> (${((v.confidence||0)*100).toFixed(0)}%, ${v.samples} samples)`
      ).join(' · ');
  }
  document.getElementById('auto-meta').innerHTML =
    `<strong>Catalog:</strong> ${escapeHtml((plan.pdf_rel || '').split('/').pop() || '(none)')}` +
    (plan.match ? ` · matched <strong>${plan.match.score}</strong> tokens on page <strong>${plan.match.page}</strong>` : '') +
    ` ${cfBadge}${provLine}`;

  const warn = document.getElementById('auto-warn');
  if (plan.warnings && plan.warnings.length) {
    warn.style.display = '';
    warn.textContent = plan.warnings.map(w => '⚠ ' + w).join('\n');
  } else {
    warn.style.display = 'none';
  }

  const setPre = (id, text) => {
    const el = document.getElementById(id);
    if (text) { el.classList.remove('empty'); el.textContent = text; }
    else { el.classList.add('empty'); el.textContent = '(empty)'; }
  };
  setPre('auto-c', plan.proposed_c || '');
  setPre('auto-d', plan.proposed_d || '');
  const ann = plan.annotations || [];
  document.getElementById('auto-ann-count').textContent = `(${ann.length})`;
  setPre('auto-ann', ann.length
    ? ann.map(a => `[${a.type}] page ${a.page} rect=${JSON.stringify(a.rect)} contents=${JSON.stringify(a.contents)}`).join('\n')
    : '(no annotations)');
}

async function applyAutoSingle() {
  if (!_AUTO_PLAN) return;
  if (!_AUTO_PLAN.ok) {
    if (!confirm('Plan has warnings — apply anyway?')) return;
  }
  const writeXlsx = document.getElementById('auto-write-xlsx').checked;
  const writePdf = document.getElementById('auto-write-pdf').checked;
  const r = await fetch('/api/auto_annotate/apply', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({row: _AUTO_PLAN.row, write_xlsx: writeXlsx, write_pdf: writePdf}),
  });
  const j = await r.json();
  const out = document.getElementById('auto-output');
  out.style.display = '';
  out.textContent = JSON.stringify(j, null, 2);
  if (j.ok) {
    alert('✓ Applied');
    closeAuto();
    location.reload();
  }
}

async function showBatchAuto() {
  document.getElementById('auto-single').style.display = 'none';
  document.getElementById('auto-batch').style.display = '';
  const list = document.getElementById('auto-batch-list');
  list.innerHTML = '<div style="padding:20px;color:#9ca3af;text-align:center;">Computing plans for all rows…</div>';
  const r = await fetch('/api/auto_annotate/batch_preview');
  const j = await r.json();
  _AUTO_BATCH_PLANS = j.plans || [];
  document.getElementById('auto-batch-info').innerHTML =
    `<strong>${j.count}</strong> rows have plans · ` +
    Object.entries(j.by_role).filter(([k,v]) => v).map(([k,v]) => `${k}: ${v}`).join(' · ');
  if (!_AUTO_BATCH_PLANS.length) {
    list.innerHTML = '<div style="padding:20px;color:#9ca3af;text-align:center;">ไม่มี row ที่ต้อง auto-annotate</div>';
    return;
  }
  list.innerHTML = _AUTO_BATCH_PLANS.map(p => renderBatchRow(p)).join('');
}
function showSingleAuto() {
  document.getElementById('auto-single').style.display = '';
  document.getElementById('auto-batch').style.display = 'none';
}

function renderBatchRow(plan) {
  const role = (plan.role && plan.role.role) || 'unknown';
  const cls = (plan.warnings && plan.warnings.length) ? 'warn' : '';
  const okBox = plan.ok && !(plan.warnings && plan.warnings.length);
  return `
    <div class="auto-batch-row ${cls}">
      <input type="checkbox" class="b-check" data-row="${plan.row}" ${okBox ? 'checked' : ''}>
      <span class="b-row">R${plan.row}</span>
      <span class="b-role ${role}">${role}</span>
      <span class="b-d" title="${escapeHtml(plan.proposed_d || '')}">${escapeHtml(plan.proposed_d || '(empty)').slice(0, 80)}</span>
      <button onclick="loadAutoPreviewFromBatch(${plan.row})" style="font-size:var(--t-xs);padding:2px 6px;">${ico('eye',12)} view</button>
    </div>`;
}
function loadAutoPreviewFromBatch(rowNum) {
  showSingleAuto();
  loadAutoPreview(rowNum);
}

async function applyAutoBatch() {
  const checked = [...document.querySelectorAll('.b-check:checked')]
    .map(el => parseInt(el.dataset.row));
  if (!checked.length) { alert('เลือก row อย่างน้อย 1 อัน'); return; }
  if (!confirm(`Apply auto-annotate ${checked.length} rows?\n(ระบบจะ snap ก่อน)`)) return;
  const list = document.getElementById('auto-batch-list');
  list.innerHTML = `<div style="padding:20px;color:#4b5563;text-align:center;"><div class="spinner"></div>กำลัง apply ${checked.length} rows…</div>`;
  const r = await fetch('/api/auto_annotate/batch_apply', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({rows: checked}),
  });
  const j = await r.json();
  const out = document.getElementById('auto-output');
  out.style.display = '';
  out.textContent = `Applied ${j.applied_ok}/${j.count} rows\n\n` +
    JSON.stringify(j.results.slice(0, 20), null, 2) +
    (j.results.length > 20 ? `\n... and ${j.results.length - 20} more` : '');
  alert(`✓ ${j.applied_ok}/${j.count} applied`);
  closeAuto();
  location.reload();
}

// ── Manual annotate workflow (when AI fell back to "ยินดีปฏิบัติ") ──
// State machine:
//   1. user clicks 📍 Mark on a row that has a catalog
//   2. we load context (suggested label) + force edit mode + drawRect tool
//   3. user click-drags to create a Square; we auto-add a paired FreeText
//      label nearby with the SKILL.md content
//   4. user can drag/resize either rect
//   5. user clicks ✓ Save → POST /api/manual_annotate/save which writes
//      both annotations into the PDF AND auto-fills Col D in xlsx based on
//      the page where the rect sits

// ── Re-annotate Wizard ────────────────────────────────────────
// Unified state machine for the manual-annotate (📍 Mark on commitment)
// AND the brand_model / item / sub_item re-annotate flow. Both run
// through the same banner UI and the same drawing hook; they differ
// only in:
//   • how many (rect, label) pairs the wizard expects (steps[])
//   • whether existing annots with matching label prefixes should be
//     deleted before writing the new ones (delete_existing)
//   • which save endpoint we POST to (manual_annotate vs reannotate)
let MANUAL_MODE = false;          // true while wizard is active
let MANUAL_TARGET_ROW = null;
let WIZ = null;                   // wizard state, see _newWizState()

function _newWizState() {
  return {
    mode: 'manual',           // 'manual' | 'reannotate'
    row: null,
    pdf_rel: null,
    pdf_meta: null,
    candidates: [],
    steps: [],                // [{label, hint, kind}]
    drawn: [],                // [{square_id, label_id, content_rect, label_rect, label_text}]
    current_step: 0,
    delete_existing: false,
    delete_prefixes: [],
    brand_hint: '',
    model_hint: '',
    section: '',
    col_b: '',
  };
}

// Public entry points ─────────────────────────────────────────
async function startManualAnnotate() {
  // Used by the floating action bar's 📍 Mark button — single-step flow
  // for commitment rows that have no PDF yet (or the user wants to add
  // a fresh annotation).
  await _startWizard('manual');
}

async function startReannotate() {
  // Used by the Col D dropdown's "🔁 Re-annotate" entry — multi-step flow
  // that also deletes prior annotations matching this row's label prefix.
  await _startWizard('reannotate');
}

async function _startWizard(mode) {
  if (!SELECTED_ROW) { toast('No row selected', 'เลือก row ก่อน', 'warn', 2500); return; }

  const cr = await fetch(`/api/reannotate/context?row=${SELECTED_ROW}`);
  const cx = await cr.json();
  if (!cx.ok) {
    toast('Wizard error', 'โหลด context ไม่ได้: ' + (cx.error || ''), 'error', 4000);
    return;
  }

  // Pick PDF — prefer the row's current; otherwise first candidate
  let chosen_pdf_rel = cx.pdf_rel;
  if (!chosen_pdf_rel && cx.candidates && cx.candidates.length) {
    chosen_pdf_rel = cx.candidates[0].rel;
  }
  if (!chosen_pdf_rel) {
    toast('No catalog PDF', 'ไม่มี catalog PDF สำหรับ row นี้ — ลอง section อื่น', 'warn', 4000);
    return;
  }

  WIZ = _newWizState();
  WIZ.mode = mode;
  WIZ.row = cx.row;
  WIZ.pdf_rel = chosen_pdf_rel;
  WIZ.candidates = cx.candidates || [];
  WIZ.steps = cx.steps || [{label: '', hint: 'ลาก rect รอบเนื้อหา', kind: 'row'}];
  WIZ.delete_existing = (mode === 'reannotate');
  WIZ.delete_prefixes = cx.delete_label_prefixes || [];
  WIZ.brand_hint = cx.brand_hint || '';
  WIZ.model_hint = cx.model_hint || '';
  WIZ.section = cx.section || '';
  WIZ.col_b = cx.col_b || '';
  WIZ.current_step = 0;

  MANUAL_MODE = true;
  MANUAL_TARGET_ROW = cx.row;

  // Load chosen PDF if it differs from currently displayed
  if (!CURRENT_PDF || CURRENT_PDF.rel !== chosen_pdf_rel) {
    await loadPdf(chosen_pdf_rel, 1, null);
  }
  WIZ.pdf_meta = CURRENT_PDF && CURRENT_PDF.meta;

  // Show banner and render initial state
  document.getElementById('manual-banner').classList.add('show');
  _renderWizBanner();

  if (!EDIT_MODE) toggleEditMode();
  setTool('drawRect');
}

// Render banner UI for the current wizard step ────────────────
function _renderWizBanner() {
  if (!WIZ) return;
  const totalSteps = WIZ.steps.length;
  const i = WIZ.current_step;
  const active = WIZ.steps[Math.min(i, totalSteps - 1)] || {};
  const isDone = i >= totalSteps;

  // Stepper chips
  const prog = document.getElementById('wiz-progress');
  const chips = WIZ.steps.map((st, idx) => {
    let cls = 'step';
    if (idx < i) cls += ' done';
    else if (idx === i && !isDone) cls += ' active';
    const mark = idx < i
      ? `<span class="check">${ico('check',10)}</span>`
      : `<span class="num">${idx+1}</span>`;
    return `<span class="${cls}">${mark} ${escapeHtml(st.label || '·')}</span>`;
  }).join(`<span class="arrow">${ico('chevron-right',10)}</span>`);
  prog.innerHTML = chips;

  // Target info / hint
  const ti = document.getElementById('manual-target-info');
  const headIcon = WIZ.mode === 'reannotate' ? ico('refresh',14) : ico('pin',14);
  const headLabel = WIZ.mode === 'reannotate' ? 'Re-annotate' : 'Mark';
  const head = `${headIcon} <strong>${headLabel} R${WIZ.row}</strong> · section <code>${escapeHtml(WIZ.section)}</code>`;
  let body;
  if (isDone) {
    body = `<span style="font-size:var(--t-sm);color:var(--c-success-text)">${ico('check',12)} ครบทุก step — กด <strong>Save</strong> เพื่อบันทึก</span>`;
  } else {
    body = `<span style="font-size:var(--t-sm)">Step ${i+1}/${totalSteps} · label: <code>${escapeHtml(active.label||'')}</code> — ${active.hint||''}</span>`;
  }
  ti.innerHTML = head + '<br>' + body;

  // PDF picker label
  const pdfLabel = document.getElementById('wiz-pdf-current');
  if (pdfLabel) {
    const cur = (WIZ.candidates.find(c => c.rel === WIZ.pdf_rel) || {}).name
                 || (WIZ.pdf_rel||'').split('/').pop() || WIZ.pdf_rel;
    pdfLabel.textContent = cur;
    pdfLabel.title = WIZ.pdf_rel;
  }

  // Buttons
  document.getElementById('wiz-back-btn').disabled = (WIZ.drawn.length === 0);
  document.getElementById('manual-save-btn').disabled = !isDone;
}

// Cancel / back / save ────────────────────────────────────────
function cancelManualAnnotate() {
  if (MANUAL_MODE && DIRTY) {
    if (!confirm('ยกเลิกการ annotate? การเปลี่ยนแปลงที่ยังไม่ save จะหาย')) return;
  }
  MANUAL_MODE = false;
  MANUAL_TARGET_ROW = null;
  WIZ = null;
  closeWizPdfPicker();
  document.getElementById('manual-banner').classList.remove('show');
  if (EDIT_MODE) {
    if (DIRTY) UNDO_STACK = [];
    setDirty(false);
    EDIT_ANNOTS = [];
    SELECTED_ANN_ID = null;
    if (EDIT_MODE) toggleEditMode();
  }
}

function wizBack() {
  if (!WIZ || WIZ.drawn.length === 0) return;
  // Pop the last drawn pair, remove its annots from EDIT_ANNOTS
  const last = WIZ.drawn.pop();
  EDIT_ANNOTS = EDIT_ANNOTS.filter(a => a._id !== last.square_id && a._id !== last.label_id);
  SELECTED_ANN_ID = null;
  WIZ.current_step = WIZ.drawn.length;
  setDirty(EDIT_ANNOTS.some(a => a._isNew));
  refreshOverlay();
  setTool('drawRect');
  _renderWizBanner();
}

async function saveManualAnnotate() {
  if (!WIZ || !MANUAL_TARGET_ROW) return;
  if (WIZ.drawn.length < WIZ.steps.length) {
    toast('Wizard incomplete', `ต้องวาด ${WIZ.steps.length} rect ก่อน — ตอนนี้วาดแล้ว ${WIZ.drawn.length}`, 'warn', 3000);
    return;
  }

  const btn = document.getElementById('manual-save-btn');
  btn.disabled = true;
  const origLabel = btn.textContent;
  btn.textContent = 'Saving…';
  try {
    let url, payload;
    // Read latest rects from EDIT_ANNOTS (they may have been moved/resized)
    const stepsPayload = WIZ.drawn.map(d => {
      const sq = EDIT_ANNOTS.find(a => a._id === d.square_id);
      const lb = EDIT_ANNOTS.find(a => a._id === d.label_id);
      return {
        content_rect: sq ? sq.rect : d.content_rect,
        label_rect:   lb ? lb.rect : d.label_rect,
        label_text:   lb ? (lb.contents || '') : (d.label_text || ''),
      };
    });

    if (WIZ.mode === 'reannotate' || WIZ.steps.length > 1 || WIZ.delete_existing) {
      url = '/api/reannotate/save';
      payload = {
        row: WIZ.row,
        pdf_rel: CURRENT_PDF.rel,
        page: PDF_PAGE,                 // all current steps share the visible page
        steps: stepsPayload,
        delete_existing: !!WIZ.delete_existing,
        delete_prefixes: WIZ.delete_prefixes || [],
        // brand/model hints flow into Col D for section_header brand_model rows
        brand: WIZ.brand_hint,
        model: WIZ.model_hint,
      };
    } else {
      // Single-step manual flow keeps the legacy endpoint for back-compat
      url = '/api/manual_annotate/save';
      const s = stepsPayload[0];
      payload = {
        row: WIZ.row,
        page: PDF_PAGE,
        content_rect: s.content_rect,
        label_rect: s.label_rect,
        label_text: s.label_text,
        pdf_rel: CURRENT_PDF.rel,
      };
    }

    const r = await fetch(url, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload),
    });
    const j = await r.json();
    if (!j.ok) {
      toast('Save failed', j.error || JSON.stringify(j), 'error', 6000);
      btn.disabled = false; btn.textContent = origLabel;
      return;
    }
    toast(`✓ R${WIZ.row} saved`,
          `Col D → ${(j.new_d||'').slice(0, 80)}${(j.new_d||'').length > 80 ? '…' : ''}`,
          'learn', 5000);
    if (typeof tickRetrain === 'function') tickRetrain('reannotate');
    MANUAL_MODE = false;
    WIZ = null;
    document.getElementById('manual-banner').classList.remove('show');
    setTimeout(() => location.reload(), 600);
  } catch (e) {
    toast('Save error', e.message, 'error', 5000);
    btn.disabled = false; btn.textContent = origLabel;
  }
}

// Drawing hook ─────────────────────────────────────────────────
// When wizard is active, each new Square the user draws becomes one
// step. We auto-place a paired FreeText label and advance the cursor.
const _origAddRectAt = addRectAt;
addRectAt = function(x0, y0, x1, y1) {
  _origAddRectAt(x0, y0, x1, y1);
  if (!MANUAL_MODE || !WIZ) return;
  // If user already finished all steps but draws another rect, treat as
  // free-form (don't auto-label).
  if (WIZ.current_step >= WIZ.steps.length) {
    _renderWizBanner();
    return;
  }

  const sq = EDIT_ANNOTS[EDIT_ANNOTS.length - 1];
  if (!sq) return;
  const stepDef = WIZ.steps[WIZ.current_step];
  const labelText = stepDef.label || '';

  // Auto-place a label rect — prefer right of the square, then below, then above
  const pageSize = (CURRENT_PDF.meta && CURRENT_PDF.meta.page_sizes[PDF_PAGE - 1]) || [595, 842];
  const labelW = Math.min(180, Math.max(60, labelText.length * 7 + 16));
  const labelH = 14;
  let lx0 = x1 + 5, ly0 = (y0 + y1) / 2 - labelH / 2;
  if (lx0 + labelW > pageSize[0] - 5) {
    // place below
    lx0 = x0; ly0 = y1 + 4;
    if (ly0 + labelH > pageSize[1] - 5) {
      // place above
      lx0 = x0; ly0 = y0 - labelH - 4;
    }
  }
  const lid = newClientId();
  const label = {
    _id: lid, _isNew: true, xref: null,
    page: PDF_PAGE, type: 'FreeText',
    rect: [lx0, ly0, lx0 + labelW, ly0 + labelH],
    contents: labelText,
  };
  EDIT_ANNOTS.push(label);
  WIZ.drawn.push({
    square_id: sq._id, label_id: lid,
    content_rect: sq.rect.slice(),
    label_rect: label.rect.slice(),
    label_text: labelText,
  });
  WIZ.current_step += 1;
  SELECTED_ANN_ID = lid;
  setDirty(true);
  refreshOverlay();

  // Decide next tool based on whether more steps remain
  if (WIZ.current_step < WIZ.steps.length) {
    // Stay in drawRect for the next step but flash a hint
    setTool('drawRect');
  } else {
    setTool('select');
  }
  _renderWizBanner();
};

// PDF picker dropdown ─────────────────────────────────────────
let _WIZ_PDF_MENU = null;
function toggleWizPdfPicker(e) {
  if (e) e.stopPropagation();
  if (_WIZ_PDF_MENU) { closeWizPdfPicker(); return; }
  if (!WIZ) return;

  const menu = document.createElement('div');
  menu.className = 'wiz-pdf-menu';
  let html = `<div class="menu-header">เปลี่ยน catalog · section ${escapeHtml(WIZ.section)}</div>`;
  if (!WIZ.candidates.length) {
    html += `<div style="padding:var(--s-5);color:var(--c-text-faint);text-align:center;">ไม่มี candidate</div>`;
  } else {
    html += WIZ.candidates.map(c => {
      const cur = (c.rel === WIZ.pdf_rel);
      return `<button class="pdf-opt ${cur?'is-current':''}" data-rel="${escapeHtml(c.rel)}">
        <span class="pdf-name">${cur?'✓ ':''}${escapeHtml(c.name)}</span>
        <span class="pdf-folder">${escapeHtml(c.folder)}</span>
      </button>`;
    }).join('');
  }
  menu.innerHTML = html;
  document.body.appendChild(menu);
  _WIZ_PDF_MENU = menu;

  // Position below the trigger button
  const btn = document.getElementById('wiz-pdf-btn');
  const r = btn.getBoundingClientRect();
  let x = r.left, y = r.bottom + 4;
  const W = menu.offsetWidth, H = menu.offsetHeight;
  if (x + W > window.innerWidth - 8) x = window.innerWidth - W - 8;
  if (y + H > window.innerHeight - 8) y = r.top - H - 4;
  menu.style.left = `${Math.max(8, x)}px`;
  menu.style.top  = `${Math.max(8, y)}px`;

  menu.querySelectorAll('button.pdf-opt').forEach(b => {
    b.addEventListener('click', async (ev) => {
      ev.stopPropagation();
      const rel = b.dataset.rel;
      closeWizPdfPicker();
      await _wizSwitchPdf(rel);
    });
  });
  setTimeout(() => {
    document.addEventListener('click', _onDocClickForWizPdfMenu, true);
    document.addEventListener('keydown', _onEscForWizPdfMenu, true);
  }, 0);
}
function closeWizPdfPicker() {
  if (_WIZ_PDF_MENU) { _WIZ_PDF_MENU.remove(); _WIZ_PDF_MENU = null; }
  document.removeEventListener('click', _onDocClickForWizPdfMenu, true);
  document.removeEventListener('keydown', _onEscForWizPdfMenu, true);
}
function _onDocClickForWizPdfMenu(e) {
  if (_WIZ_PDF_MENU && !_WIZ_PDF_MENU.contains(e.target) &&
      e.target.id !== 'wiz-pdf-btn' && !e.target.closest('#wiz-pdf-btn')) {
    closeWizPdfPicker();
  }
}
function _onEscForWizPdfMenu(e) {
  if (e.key === 'Escape') { e.stopPropagation(); closeWizPdfPicker(); }
}

async function _wizSwitchPdf(rel) {
  if (!WIZ || !rel || rel === WIZ.pdf_rel) return;
  if (WIZ.drawn.length > 0) {
    if (!confirm('เปลี่ยน PDF จะลบ rect ที่วาดไว้ในหน้านี้ — ต่อไหม?')) return;
  }
  // Drop in-progress drawings
  for (const d of WIZ.drawn) {
    EDIT_ANNOTS = EDIT_ANNOTS.filter(a => a._id !== d.square_id && a._id !== d.label_id);
  }
  WIZ.drawn = [];
  WIZ.current_step = 0;
  WIZ.pdf_rel = rel;
  // Update brand/model hint from the new filename
  const stem = rel.split('/').pop().replace(/\.pdf$/i, '');
  // Use a quick parse mirroring parse_brand_model_from_filename (ASCII first token)
  const stripped = stem.replace(/^\d+(?:\.\d+){1,3}\.?\s+/, '').replace(/\([^)]*\)/g, '').replace(/\s+/g, ' ').trim();
  const m = stripped.match(/[A-Za-z]{2,}\S*/);
  if (m) {
    WIZ.brand_hint = m[0][0].toUpperCase() + m[0].slice(1);
    WIZ.model_hint = stripped.slice(stripped.indexOf(m[0]) + m[0].length).trim();
  }
  await loadPdf(rel, 1, null);
  WIZ.pdf_meta = CURRENT_PDF && CURRENT_PDF.meta;
  setTool('drawRect');
  setDirty(false);
  EDIT_ANNOTS = [];
  SELECTED_ANN_ID = null;
  refreshOverlay();
  _renderWizBanner();
  toast('PDF switched', rel.split('/').pop(), 'info', 2500);
}

// ── Toast notifications ───────────────────────────────────────
function toast(title, body, kind = 'info', timeout = 5000) {
  const wrap = document.getElementById('toasts');
  if (!wrap) return;
  const el = document.createElement('div');
  el.className = 'toast ' + kind;
  el.setAttribute('role', kind === 'error' ? 'alert' : 'status');
  el.setAttribute('aria-live', kind === 'error' ? 'assertive' : 'polite');
  el.innerHTML = `<button class="close" type="button" aria-label="ปิด" onclick="this.parentElement.remove()">${ico('x',12)}</button>` +
                 `<div class="title">${escapeHtml(title)}</div>` +
                 (body ? `<div class="body">${escapeHtml(body)}</div>` : '');
  // Keep stack reasonable
  while (wrap.children.length >= 6) wrap.firstChild.remove();
  wrap.appendChild(el);
  if (timeout) {
    setTimeout(() => {
      el.classList.add('fading');
      setTimeout(() => el.remove(), 400);
    }, timeout);
  }
}

// ── Theme toggle (light/dark) ─────────────────────────────────
function toggleTheme() {
  const root = document.documentElement;
  const cur = root.getAttribute('data-theme');
  let next;
  if (cur === 'dark') next = 'light';
  else if (cur === 'light') next = 'dark';
  else {
    // No explicit preference yet → flip from system
    const isDark = window.matchMedia('(prefers-color-scheme: dark)').matches;
    next = isDark ? 'light' : 'dark';
  }
  root.setAttribute('data-theme', next);
  try { localStorage.setItem('comply-theme', next); } catch (e) {}
  const btn = document.getElementById('theme-toggle');
  if (btn) btn.innerHTML = ico(next === 'dark' ? 'sun' : 'moon', 16);
}
(function initTheme() {
  try {
    const saved = localStorage.getItem('comply-theme');
    if (saved === 'dark' || saved === 'light') {
      document.documentElement.setAttribute('data-theme', saved);
      const btn = document.getElementById('theme-toggle');
      if (btn) btn.innerHTML = `<svg class="ico" aria-hidden="true"><use href="#i-${saved === 'dark' ? 'sun' : 'moon'}"/></svg>`;
    }
  } catch (e) {}
})();

// ── Modal infrastructure: ESC to close + focus trap ────────────
const MODAL_CLOSERS = {
  'history-modal':  () => closeHistory(),
  'versions-modal': () => closeVersions(),
  'learn-modal':    () => closeLearning(),
  'audit-modal':    () => closeAudit(),
  'auto-modal':     () => closeAuto(),
};
function _topMostOpenModal() {
  const ids = Object.keys(MODAL_CLOSERS);
  for (const id of ids.reverse()) {
    const el = document.getElementById(id);
    if (el && (el.classList.contains('show') || el.style.display === 'flex')) return el;
  }
  return null;
}
document.addEventListener('keydown', (e) => {
  if (e.key !== 'Escape') return;
  const m = _topMostOpenModal();
  if (!m) return;
  // Don't intercept if user is in the middle of editing a Col D cell or text editor
  if (typeof TEXT_EDITOR !== 'undefined' && TEXT_EDITOR) return;
  const closer = MODAL_CLOSERS[m.id];
  if (closer) { e.preventDefault(); closer(); }
}, true);
// Simple focus trap: when a modal opens, focus its first focusable element
const _modalObserver = new MutationObserver((muts) => {
  for (const m of muts) {
    if (m.type !== 'attributes') continue;
    const el = m.target;
    if (!el.classList || !el.classList.contains('modal-bg')) continue;
    const open = el.classList.contains('show') || el.style.display === 'flex';
    if (open) {
      const first = el.querySelector('input, select, textarea, button, [tabindex="0"]');
      if (first) setTimeout(() => first.focus({preventScroll: true}), 50);
    }
  }
});
document.querySelectorAll('.modal-bg').forEach(m => {
  _modalObserver.observe(m, {attributes: true, attributeFilter: ['class', 'style']});
});

// ── Auto-retrain (HITL "system gets smarter") ─────────────────
// Track feedback events client-side; whenever we cross a threshold,
// trigger a background retrain and show a toast with what was learned.
let _PENDING_RETRAIN_COUNT = 0;
const RETRAIN_THRESHOLD = 5;
async function tickRetrain(reason = 'feedback') {
  _PENDING_RETRAIN_COUNT++;
  if (_PENDING_RETRAIN_COUNT < RETRAIN_THRESHOLD) return;
  _PENDING_RETRAIN_COUNT = 0;
  try {
    const r = await fetch('/api/learn/retrain', {method: 'POST'});
    const j = await r.json();
    const total = (j.promoted || 0) + (j.updated || 0);
    if (total > 0) {
      toast(`🧠 Learned from your edits`,
            `${j.promoted} new pattern${j.promoted!==1?'s':''}, ${j.updated} updated. Future suggestions will use these.`,
            'learn', 7000);
    }
  } catch (e) {}
}

// ── Smart navigation: jump to next "uncertain" row ────────────
// "Uncertain" = unverified + has flags + has PDF.  Navigates within the
// currently-visible (filtered) tree only.
function nextUncertainRow() {
  const status = DATA.status || {};
  const visible = [...document.querySelectorAll('.tree-row[data-row]')]
    .map(el => parseInt(el.dataset.row));
  const start = visible.indexOf(SELECTED_ROW);
  // Score each candidate: lower = more interesting to inspect
  function score(rn) {
    const row = ROWS_BY_NUM[rn];
    if (!row) return 1e9;
    const st = (status[rn] && status[rn].status) || 'unverified';
    if (st !== 'unverified') return 1e9;        // skip already-verified
    if (!row.pdf_rel) return 1e6;                // no PDF → low priority
    let s = 0;
    if (row.auto_flags && row.auto_flags.length) s -= 100;  // flags = high priority
    if (row.parsed && row.parsed.type === 'commitment') s += 50;
    if (row.needs_col_d) s -= 50;
    return s;
  }
  // Look forward from the selected row first
  const ahead = visible.slice(start + 1).map(rn => [score(rn), rn]);
  const before = visible.slice(0, start + 1).map(rn => [score(rn), rn]);
  ahead.sort((a, b) => a[0] - b[0]);
  before.sort((a, b) => a[0] - b[0]);
  const target = (ahead[0] && ahead[0][0] < 1e6 ? ahead[0][1]
                 : before[0] && before[0][0] < 1e6 ? before[0][1] : null);
  if (target) {
    selectRow(target);
    toast('▸ Next uncertain', `R${target}`, 'info', 1500);
  } else {
    toast('No uncertain rows left', 'ตรวจครบทุก row ที่มี catalog แล้ว 🎉', 'info', 3000);
  }
}

// ── Mobile tab navigation ─────────────────────────────────────
function setMobileTab(tab) {
  for (const b of document.querySelectorAll('.mobile-tabs button')) {
    const active = b.dataset.tab === tab;
    b.classList.toggle('active', active);
    b.setAttribute('aria-selected', active ? 'true' : 'false');
  }
  const map = {tree: '.tree-pane', center: '.center-pane', pdf: '.pdf-pane'};
  for (const [k, sel] of Object.entries(map)) {
    const el = document.querySelector(sel);
    if (el) el.classList.toggle('mobile-active', k === tab);
  }
}

// Auto-show "center" tab on row select for mobile
const _origSelectRow = selectRow;
selectRow = async function(rowNum, scroll) {
  await _origSelectRow(rowNum, scroll);
  if (window.innerWidth <= 700) {
    // After picking a row on mobile, switch to the work area
    setMobileTab('center');
  }
};

// ── HITL Learning modal ───────────────────────────────────────
function showLearning() {
  document.getElementById('learn-modal').classList.add('show');
  loadLearnStats();
  loadLearnPatterns();
}
function closeLearning() {
  document.getElementById('learn-modal').classList.remove('show');
  document.getElementById('learn-output').style.display = 'none';
}

async function loadLearnStats() {
  const r = await fetch('/api/learn/stats');
  const j = await r.json();
  const acc = (j.accuracy * 100).toFixed(0);
  const cards = [
    ['total_feedbacks', `feedbacks (${j.window_days}d)`],
    ['accepted', 'accepted'],
    ['edited', 'edited'],
    ['rejected', 'rejected'],
    ['accuracy_pct', `accuracy ${acc}%`],
    ['patterns_total', 'patterns'],
    ['patterns_enabled', 'enabled'],
  ];
  const data = {...j, accuracy_pct: `${acc}%`};
  document.getElementById('learn-stats').innerHTML = cards.map(([k, l]) =>
    `<div class="stat-card"><div class="v">${data[k] ?? 0}</div><div class="l">${l}</div></div>`
  ).join('');
  document.getElementById('learn-llm-name').textContent =
    j.llm.available ? `${j.llm.name} ✓` : 'off (rule-only)';
}

async function loadLearnPatterns() {
  const t = document.getElementById('learn-pattern-filter').value;
  const url = '/api/learn/patterns' + (t ? `?type=${encodeURIComponent(t)}` : '');
  const r = await fetch(url);
  const j = await r.json();
  const list = document.getElementById('learn-pattern-list');
  if (!j.patterns.length) {
    list.innerHTML = '<div style="padding:20px;color:#9ca3af;text-align:center;">ยังไม่มี learned patterns — แก้ไข Col D หลายๆ row แล้วกด Retrain</div>';
    return;
  }
  list.innerHTML = j.patterns.map(p => {
    const cls = p.confidence >= 0.8 ? 'high' : p.confidence < 0.5 ? 'low' : '';
    const dis = p.enabled ? '' : ' disabled';
    return `<div class="learn-pattern-row${dis}" data-id="${p.id}">
      <span class="pt ${p.type}">${p.type}</span>
      <span class="trigger">${escapeHtml(p.trigger_key)}</span>
      <span><span class="arrow">→</span> <span class="output">${escapeHtml(p.output)}</span></span>
      <span class="conf"><span class="${cls}">${(p.confidence*100).toFixed(0)}%</span> · ${p.samples_total}</span>
      <label style="font-size:10px;display:flex;gap:3px;align-items:center;">
        <input type="checkbox" ${p.enabled?'checked':''} onchange="togglePattern(${p.id}, this.checked)"> on
      </label>
    </div>`;
  }).join('');
}

async function togglePattern(id, enabled) {
  await fetch(`/api/learn/pattern/${id}`, {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({enabled}),
  });
  loadLearnPatterns();
}

async function runRetrain() {
  const out = document.getElementById('learn-output');
  out.style.display = '';
  out.textContent = 'Retraining…';
  const r = await fetch('/api/learn/retrain', {method: 'POST'});
  const j = await r.json();
  out.textContent = `✓ Retrain complete\n  promoted: ${j.promoted}\n  updated:  ${j.updated}\n  patterns_examined: ${j.patterns_examined}`;
  loadLearnStats();
  loadLearnPatterns();
}

// ── Audit + DB modal ──────────────────────────────────────────
function showAudit() {
  document.getElementById('audit-modal').classList.add('show');
  loadAuditStats();
  loadAudit();
  document.getElementById('audit-search').value = '';
  document.getElementById('audit-search-results').style.display = 'none';
  document.getElementById('audit-list-wrap').style.display = '';
}
function closeAudit() {
  document.getElementById('audit-modal').classList.remove('show');
}

// ── Phase 2: Catalog browser ──────────────────────────────────────
let _CAT_SEL = null;        // selected catalog in detail pane
let _CAT_TIMER = 0;
let _CAT_LIST = [];

function openCatalogBrowser() {
  document.getElementById('catalog-modal').classList.add('show');
  catalogReload();
  catalogPopulateSectionFilter();
}
function closeCatalogBrowser() {
  document.getElementById('catalog-modal').classList.remove('show');
}
function catalogSearchDebounced() {
  if (_CAT_TIMER) clearTimeout(_CAT_TIMER);
  _CAT_TIMER = setTimeout(catalogReload, 200);
}
async function catalogReload() {
  const list = document.getElementById('catalog-list');
  if (!list) return;
  list.innerHTML = '<div style="padding:20px;color:var(--c-text-faint);text-align:center">loading…</div>';
  const q = document.getElementById('catalog-search').value.trim();
  const sec = document.getElementById('catalog-section-filter').value;
  const params = new URLSearchParams();
  if (q) params.set('q', q);
  if (sec) params.set('section', sec);
  params.set('limit', '300');
  try {
    const r = await fetch(`/api/catalogs?${params.toString()}`);
    const j = await r.json();
    _CAT_LIST = j.items || [];
    catalogRenderList();
    catalogLoadStats();
  } catch (e) {
    list.innerHTML = `<div style="padding:20px;color:var(--c-danger-text);text-align:center">load error: ${escapeHtml(e.message)}</div>`;
  }
}
async function catalogLoadStats() {
  try {
    const r = await fetch('/api/catalogs/stats');
    const j = await r.json();
    const pill = document.getElementById('catalog-stats-pill');
    if (pill) pill.textContent = `${j.catalogs} catalogs · ${j.companies} co · ${j.projects} proj · ${j.row_links} bound`;
  } catch (e) {}
}
async function catalogPopulateSectionFilter() {
  const sel = document.getElementById('catalog-section-filter');
  if (!sel || sel.children.length > 1) return;  // already populated
  // Cheap: reuse data we already have if possible
  try {
    const r = await fetch('/api/catalogs?limit=500');
    const j = await r.json();
    const sections = [...new Set((j.items || []).map(c => c.section_hint).filter(Boolean))].sort();
    for (const s of sections) {
      const opt = document.createElement('option');
      opt.value = s; opt.textContent = s;
      sel.appendChild(opt);
    }
  } catch (e) {}
}
function catalogRenderList() {
  const list = document.getElementById('catalog-list');
  if (!_CAT_LIST.length) {
    list.innerHTML = '<div style="padding:20px;color:var(--c-text-faint);text-align:center">No catalogs match.</div>';
    return;
  }
  list.innerHTML = _CAT_LIST.map(c => {
    const active = (_CAT_SEL && c.catalog_id === _CAT_SEL) ? ' active' : '';
    const filename = (c.pdf_rel || '').split('/').pop();
    return `<div class="cat-item${active}" data-id="${c.catalog_id}" onclick="catalogSelect(${c.catalog_id})">
      <div class="cat-head">
        ${c.section_hint ? `<span class="cat-section">${escapeHtml(c.section_hint)}</span>` : ''}
        <span class="cat-brand">${escapeHtml(c.brand || '—')}</span>
        <span class="cat-model">${escapeHtml(c.model || '')}</span>
        <span class="cat-pages">p.${c.pages || '?'}</span>
      </div>
      <div class="cat-rel">${escapeHtml(filename)}</div>
    </div>`;
  }).join('');
}
async function catalogSelect(catalog_id) {
  _CAT_SEL = catalog_id;
  document.querySelectorAll('#catalog-list .cat-item').forEach(el => {
    el.classList.toggle('active', parseInt(el.dataset.id) === catalog_id);
  });
  const detail = document.getElementById('catalog-detail');
  detail.innerHTML = '<div style="padding:20px;color:var(--c-text-faint)">loading…</div>';
  try {
    const r = await fetch(`/api/catalogs/${catalog_id}`);
    const j = await r.json();
    if (!j.ok) {
      detail.innerHTML = `<div style="color:var(--c-danger-text)">${escapeHtml(j.error || 'load failed')}</div>`;
      return;
    }
    const links_r = await fetch(`/api/catalogs/${catalog_id}/links`);
    const links_j = await links_r.json();
    catalogRenderDetail(j.catalog, links_j.links || []);
  } catch (e) {
    detail.innerHTML = `<div style="color:var(--c-danger-text)">error: ${escapeHtml(e.message)}</div>`;
  }
}
function catalogRenderDetail(c, links) {
  const detail = document.getElementById('catalog-detail');
  const filename = (c.pdf_rel || '').split('/').pop();
  detail.innerHTML = `
    <h3>
      ${c.section_hint ? `<span class="cat-section-pill">${escapeHtml(c.section_hint)}</span>` : ''}
      <span style="font-family:var(--f-mono);font-size:var(--t-sm);color:var(--c-text-soft)">${escapeHtml(filename)}</span>
    </h3>
    <div style="font-size:var(--t-xs);color:var(--c-text-faint);margin-bottom:var(--s-3);font-family:var(--f-mono);word-break:break-all">${escapeHtml(c.pdf_rel || '')}</div>

    <div class="meta-grid">
      <label>Brand</label>
      <input id="cat-edit-brand" value="${escapeHtml(c.brand || '')}" placeholder="e.g. Lenovo">
      <label>Model</label>
      <input id="cat-edit-model" value="${escapeHtml(c.model || '')}" placeholder="e.g. ThinkSystem SR630">
      <label>Category</label>
      <input id="cat-edit-category" value="${escapeHtml(c.category || '')}" placeholder="e.g. Server / Switch / Rack">
      <label>Section</label>
      <input id="cat-edit-section" value="${escapeHtml(c.section_hint || '')}" placeholder="e.g. 5.1.1.2">
      <label>Description</label>
      <textarea id="cat-edit-desc" placeholder="vendor specs / notes">${escapeHtml(c.description || '')}</textarea>
    </div>

    <div class="detail-actions">
      <button class="primary" onclick="catalogSaveMeta(${c.catalog_id})">${ico('save', 13)}<span>Save metadata</span></button>
      ${SELECTED_ROW ? `<button class="success" onclick="catalogApplyToRow(${c.catalog_id})">${ico('check', 13)}<span>Apply to R${SELECTED_ROW}</span></button>` : `<button disabled title="select a row first">${ico('check', 13)}<span>Apply to row…</span></button>`}
      <button onclick="catalogOpenPdf('${escapeHtml(c.pdf_rel || '')}')">${ico('eye', 13)}<span>Open PDF</span></button>
    </div>

    <div class="detail-section">
      <h4>Annotations (DB-stored, per page)</h4>
      ${(c.annotations || []).length ? c.annotations.map(a => `
        <div class="annot-row">
          <span class="pg">p.${a.page}</span>
          <span class="type">${escapeHtml(a.type)}</span>
          <span class="contents" title="${escapeHtml(a.contents || '')}">${escapeHtml(a.contents || '(no text)')}</span>
          <button onclick="catalogDeleteAnnot(${c.catalog_id}, ${a.annot_id})" title="Delete">×</button>
        </div>`).join('') : '<div style="font-size:var(--t-xs);color:var(--c-text-faint);font-style:italic;padding:4px 8px">No DB-stored annotations yet. (Existing annotations are baked into the PDF and edited via the row edit flow.)</div>'}
    </div>

    <div class="detail-section">
      <h4>Used by (${links.length} row${links.length === 1 ? '' : 's'})</h4>
      ${links.length ? links.map(l => `
        <div class="links-row">
          <span class="row-num" onclick="closeCatalogBrowser();selectRow(${l.row_num})" title="Jump to row">R${l.row_num}</span>
          <span style="color:var(--c-text-soft)">${escapeHtml(l.project_name || '')}</span>
          <span style="color:var(--c-text-faint);font-family:var(--f-mono);overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${escapeHtml((l.col_d_text || '').slice(0, 80))}</span>
        </div>`).join('') : '<div style="font-size:var(--t-xs);color:var(--c-text-faint);font-style:italic;padding:4px 8px">Not yet bound to any row.</div>'}
    </div>`;
}
async function catalogSaveMeta(catalog_id) {
  const body = {
    brand: document.getElementById('cat-edit-brand').value.trim() || null,
    model: document.getElementById('cat-edit-model').value.trim() || null,
    category: document.getElementById('cat-edit-category').value.trim() || null,
    section_hint: document.getElementById('cat-edit-section').value.trim() || null,
    description: document.getElementById('cat-edit-desc').value.trim() || null,
  };
  try {
    const r = await fetch(`/api/catalogs/${catalog_id}`, {
      method: 'PATCH', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(body),
    });
    const j = await r.json();
    if (j.ok) {
      toast('✓ Saved', `Catalog #${catalog_id} metadata updated`, 'learn', 2500);
      catalogReload();         // refresh list (filter / sort may change)
      catalogSelect(catalog_id);
    } else {
      toast('Save failed', j.error || 'unknown', 'error', 4000);
    }
  } catch (e) { toast('Save error', e.message, 'error', 4000); }
}
async function catalogApplyToRow(catalog_id) {
  if (!SELECTED_ROW) { toast('Pick a row first', '', 'warn', 2500); return; }
  const page = prompt(`หน้าใน catalog ที่ row R${SELECTED_ROW} อ้างอิง? (เว้นว่าง = ไม่ระบุหน้า)`);
  // user cancelled
  if (page === null) return;
  const pageN = page.trim() ? parseInt(page) : null;
  try {
    const r = await fetch('/api/row/apply_catalog', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({row: SELECTED_ROW, catalog_id, page: pageN}),
    });
    const j = await r.json();
    if (j.ok) {
      toast(`✓ Applied`, `R${SELECTED_ROW} → ${(j.col_d || '').slice(0, 40)}`, 'learn', 4000);
      closeCatalogBrowser();
      // Refresh xlsx + tree to show new Col D
      if (typeof loadXlsx === 'function' && SELECTED_ROW) loadXlsx(SELECTED_ROW);
      const idx = await fetch('/api/index').then(r => r.json());
      DATA = idx;
      ROWS_BY_NUM = Object.fromEntries(idx.rows.map(r => [r.row, r]));
      if (typeof renderTree === 'function') renderTree();
      if (typeof tickRetrain === 'function') tickRetrain('apply_catalog');
    } else {
      toast('Apply failed', j.error || 'unknown', 'error', 4500);
    }
  } catch (e) { toast('Apply error', e.message, 'error', 4500); }
}
async function catalogDeleteAnnot(catalog_id, annot_id) {
  if (!confirm(`Delete annotation #${annot_id}?`)) return;
  try {
    await fetch(`/api/catalogs/${catalog_id}/annotations/${annot_id}`,
                {method: 'DELETE'});
    catalogSelect(catalog_id);
  } catch (e) { toast('Delete error', e.message, 'error', 3000); }
}
function catalogOpenPdf(pdf_rel) {
  // Open the PDF in a new tab (uses existing /api/raw_pdf)
  if (!pdf_rel) return;
  window.open(`/api/raw_pdf?rel=${encodeURIComponent(pdf_rel)}`, '_blank');
}
async function catalogReingest() {
  toast('Re-scanning…', 'reading output/ for new PDFs', 'info', 2000);
  try {
    const r = await fetch('/api/catalogs/reingest', {method: 'POST'});
    const j = await r.json();
    if (j.ok) {
      toast('✓ Re-scan done',
            `${j.scanned} PDFs (${j.inserted} new, ${j.updated} updated, ${j.skipped} unchanged)`,
            'learn', 4000);
      catalogReload();
    } else {
      toast('Re-scan failed', j.error || 'unknown', 'error', 4000);
    }
  } catch (e) { toast('Re-scan error', e.message, 'error', 4000); }
}

async function loadAuditStats() {
  const r = await fetch('/api/db/stats');
  const j = await r.json();
  const cards = [
    ['rows_total', 'rows'],
    ['rows_with_pdf', 'with PDF'],
    ['pdfs', 'catalog PDFs'],
    ['annotations', 'annotations'],
    ['snapshots', 'snapshots'],
    ['audit_entries', 'audit entries'],
    ['auto_annotates_applied', 'auto-annotates'],
    ['rows_needs_col_d', 'need Col D'],
  ];
  document.getElementById('audit-stats').innerHTML = cards.map(([k, l]) =>
    `<div class="stat-card"><div class="v">${j[k] ?? 0}</div><div class="l">${l}</div></div>`
  ).join('');
}

async function loadAudit() {
  const action = document.getElementById('audit-action-filter').value;
  let url = '/api/db/audit?limit=100';
  if (action) url += '&action=' + encodeURIComponent(action);
  const r = await fetch(url);
  const j = await r.json();
  const list = document.getElementById('audit-list');
  if (!j.entries.length) {
    list.innerHTML = '<div style="padding:20px;color:#9ca3af;text-align:center;">ยังไม่มี audit log</div>';
    return;
  }
  list.innerHTML = j.entries.map(e => {
    const tsHuman = e.ts ? e.ts.replace('T', ' ').slice(5, 19) : '';
    let body = '';
    if (e.before != null || e.after != null) {
      body = `<span class="ba"><span class="before">${escapeHtml(e.before || '∅')}</span><span class="arrow">→</span><span class="after">${escapeHtml(e.after || '∅')}</span></span>`;
    } else if (e.details) {
      body = `<span class="ba">${escapeHtml(JSON.stringify(e.details).slice(0, 80))}</span>`;
    }
    const targetClick = (e.target_type === 'row')
      ? `onclick="closeAudit();selectRow(${parseInt(e.target_id)||0})"` : '';
    return `<div class="audit-row">
      <span class="ts">${escapeHtml(tsHuman)}</span>
      <span class="ac ${e.action}">${escapeHtml(e.action)}</span>
      <span class="tg" ${targetClick} title="${escapeHtml(e.target_type||'')}">${escapeHtml((e.target_type||'')+(e.target_id ? ' '+e.target_id : ''))}</span>
      ${body}
    </div>`;
  }).join('');
}

let _AUDIT_SEARCH_TIMER = null;
function onAuditSearch() {
  clearTimeout(_AUDIT_SEARCH_TIMER);
  _AUDIT_SEARCH_TIMER = setTimeout(runAuditSearch, 250);
}
async function runAuditSearch() {
  const q = document.getElementById('audit-search').value.trim();
  const wrap = document.getElementById('audit-search-results');
  if (!q) {
    wrap.style.display = 'none';
    document.getElementById('audit-list-wrap').style.display = '';
    return;
  }
  document.getElementById('audit-list-wrap').style.display = 'none';
  wrap.style.display = '';
  wrap.innerHTML = '<div style="padding:14px;color:#9ca3af;">กำลังค้น…</div>';
  try {
    const r = await fetch('/api/db/search?q=' + encodeURIComponent(q));
    const j = await r.json();
    if (!j.results.length) {
      wrap.innerHTML = '<div style="padding:14px;color:#9ca3af;">ไม่พบ</div>';
      return;
    }
    wrap.innerHTML = `<h4 style="margin:8px 0 4px;font-size:12px;color:#4b5563;">FTS Results (${j.results.length})</h4>` +
      j.results.map(h =>
        `<div class="ar-row" onclick="closeAudit();selectRow(${h.row})">
          <strong>R${h.row}</strong> <span style="color:#6b7280">${escapeHtml(h.section || '')}</span>
          <div style="margin-top:2px">${h.snippet_b || ''}</div>
          ${h.snippet_d ? `<div style="margin-top:2px;color:#6b7280">D: ${h.snippet_d}</div>` : ''}
        </div>`
      ).join('');
  } catch (e) {
    wrap.innerHTML = `<div style="padding:14px;color:#b91c1c;">error: ${escapeHtml(e.message)}</div>`;
  }
}

// ── Stats ──────────────────────────────────────────────────────
function renderStats() {
  const st = DATA.stats, status = DATA.status || {};
  const c = {pass:0, fail:0, need_fix:0, skip:0};
  for (const k in status) if (c[status[k].status] !== undefined) c[status[k].status]++;
  const done = c.pass + c.fail + c.need_fix + c.skip;
  const remain = st.total - done;
  const pct = st.total ? Math.round((done / st.total) * 100) : 0;
  document.getElementById('stats').innerHTML = `
    <span><strong>${st.total}</strong> rows</span>
    <span><strong style="color:var(--c-success-text)">${c.pass}</strong> ✓</span>
    <span><strong style="color:var(--c-danger-text)">${c.fail}</strong> ✗</span>
    <span><strong style="color:var(--c-warn-text)">${c.need_fix}</strong> ⚠</span>
    <span><strong style="color:var(--c-text-muted)">${c.skip}</strong> ⏭</span>
    <span><strong>${remain}</strong> เหลือ</span>`;
  // Topbar stats-pill — compact summary
  const pill = document.getElementById('stats-pill-progress');
  if (pill) {
    pill.innerHTML =
      `<strong>${pct}%</strong> <span class="sep">·</span> ` +
      `<span style="color:var(--c-success-text)">${c.pass} ✓</span> <span class="sep">·</span> ` +
      `<span style="color:var(--c-danger-text)">${c.fail} ✗</span> <span class="sep">·</span> ` +
      `<span>${remain} เหลือ</span>`;
  }
}

// ── helpers ────────────────────────────────────────────────────
function escapeHtml(s) {
  return (s == null ? '' : String(s)).replace(/[&<>"']/g, c => ({
    '&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'
  }[c]));
}

/** Confidence-dot for a row: green = looks OK, amber = commitment but
 *  has a catalog (might need manual mark), red = no PDF / parse error,
 *  gray = empty. */
function confDotHtml(r) {
  if (!r) return '';
  let cls = 'none', tip = 'no data';
  if (r.is_commitment === undefined) {
    r.is_commitment = (r.D || '').trim().startsWith('ยินดีปฏิบัติ');
  }
  if (r.parsed && r.parsed.type === 'commitment' && r.pdf_rel) {
    cls = 'med'; tip = 'commitment but catalog exists — consider Mark';
  } else if (r.pdf_rel && r.D) {
    cls = 'high'; tip = 'has catalog + Col D';
  } else if (!r.pdf_rel && (r.D || '').trim()) {
    cls = 'low'; tip = 'Col D set but no PDF resolved';
  } else if (r.D) {
    cls = 'med'; tip = 'has Col D, no catalog';
  }
  return `<span class="conf-dot ${cls}" title="${escapeHtml(tip)}"></span>`;
}

// ── filter wiring ──────────────────────────────────────────────
['tree-search','filter-status','filter-vendor','filter-flags','filter-haspdf'].forEach(id => {
  document.getElementById(id).addEventListener('input', applyFilters);
  document.getElementById(id).addEventListener('change', applyFilters);
});

// ── Keyboard ───────────────────────────────────────────────────
document.addEventListener('keydown', e => {
  const tag = e.target.tagName;
  const inText = (tag === 'INPUT' || tag === 'TEXTAREA' || tag === 'SELECT');
  // Cmd/Ctrl-S — save (works even from input)
  if ((e.metaKey || e.ctrlKey) && e.key.toLowerCase() === 's') {
    e.preventDefault();
    if (EDIT_MODE) saveEdits();
    return;
  }
  // Cmd/Ctrl-Z / Shift+Cmd+Z — undo/redo (only when in edit mode and not editing text)
  if (EDIT_MODE && !inText && (e.metaKey || e.ctrlKey) && e.key.toLowerCase() === 'z') {
    e.preventDefault();
    if (e.shiftKey) redo(); else undo();
    return;
  }
  if (EDIT_MODE && !inText && (e.metaKey || e.ctrlKey) && e.key.toLowerCase() === 'y') {
    e.preventDefault(); redo(); return;
  }
  if (inText) return;
  // Edit-mode tool shortcuts
  if (EDIT_MODE) {
    if (e.key === 'v' || e.key === 'V') { setTool('select'); e.preventDefault(); return; }
    if (e.key === 'r' || e.key === 'R') { setTool('drawRect'); e.preventDefault(); return; }
    if (e.key === 't' || e.key === 'T') { setTool('addText'); e.preventDefault(); return; }
    if ((e.key === 'Delete' || e.key === 'Backspace') && SELECTED_ANN_ID) {
      e.preventDefault(); deleteSelected(); return;
    }
    // Phase A5: D = duplicate selected annotation (Acrobat-style)
    if ((e.key === 'd' || e.key === 'D') && !e.metaKey && !e.ctrlKey && SELECTED_ANN_ID) {
      e.preventDefault(); duplicateSelected(); return;
    }
    if (e.key === 'Escape') {
      if (TEXT_EDITOR) cancelTextEditor();
      else if (DIRTY) toggleEditMode();
      else { EDIT_MODE = false; toggleEditMode(); toggleEditMode(); /* clean exit */ }
      e.preventDefault(); return;
    }
    // Arrow keys nudge selected annotation (when not editing text)
    if (SELECTED_ANN_ID && (e.key === 'ArrowUp' || e.key === 'ArrowDown' ||
                            e.key === 'ArrowLeft' || e.key === 'ArrowRight')) {
      const sel = EDIT_ANNOTS.find(a => a._id === SELECTED_ANN_ID);
      if (sel) {
        const step = e.shiftKey ? 5 : 1;
        const dx = e.key === 'ArrowLeft' ? -step : e.key === 'ArrowRight' ? step : 0;
        const dy = e.key === 'ArrowUp' ? -step : e.key === 'ArrowDown' ? step : 0;
        _commitBeforeChange();
        sel.rect = [sel.rect[0]+dx, sel.rect[1]+dy, sel.rect[2]+dx, sel.rect[3]+dy];
        setDirty(true);
        refreshOverlay();
        refreshUndoRedoButtons();
        e.preventDefault();
        return;
      }
    }
  }
  // Global navigation
  if (e.key === 'j' || e.key === 'ArrowDown')  { moveRow(1);  e.preventDefault(); }
  else if (e.key === 'k' || e.key === 'ArrowUp') { moveRow(-1); e.preventDefault(); }
  else if (e.key === 'n' || e.key === 'N') { nextUncertainRow(); e.preventDefault(); }
  else if (e.key === '1') setStatus('pass');
  else if (e.key === '2') setStatus('fail');
  else if (e.key === '3') setStatus('need_fix');
  else if (e.key === '4') setStatus('skip');
  else if (e.key === '[') pdfPrev();
  else if (e.key === ']') pdfNext();
  else if (e.key === ',') torPrev();
  else if (e.key === '.') torNext();
  else if (e.key === '+' || e.key === '=') pdfZoom(1);
  else if (e.key === '-') pdfZoom(-1);
});
function moveRow(delta) {
  // walk through visible rows in tree order
  const visible = [...document.querySelectorAll('.tree-row[data-row]')]
    .map(el => parseInt(el.dataset.row));
  if (!visible.length) return;
  let idx = visible.indexOf(SELECTED_ROW);
  if (idx < 0) idx = 0;
  else idx = Math.max(0, Math.min(visible.length - 1, idx + delta));
  selectRow(visible[idx]);
}

// Auto-advance after verdict (toggleable)
let AUTO_ADVANCE = (localStorage.getItem('autoAdvance') !== '0');
function toggleAutoAdvance() {
  AUTO_ADVANCE = !AUTO_ADVANCE;
  localStorage.setItem('autoAdvance', AUTO_ADVANCE ? '1' : '0');
  const cb = document.getElementById('auto-advance');
  if (cb) cb.checked = AUTO_ADVANCE;
}
// patch setStatus to optionally advance
const _origSetStatus = setStatus;
setStatus = async function(status) {
  await _origSetStatus(status);
  if (AUTO_ADVANCE && status !== 'unverified') moveRow(1);
};

// ============================================================
// Sprint UI-2 / UI-3 / S1 / S2 / S3 — JS additions
// ============================================================

// ── Partial refresh helper (Sprint S1.1 — kills location.reload) ──
async function partialRefresh(opts = {}) {
  // Re-fetch /api/index, rebuild ROWS_BY_NUM, refresh tree + xlsx + stats
  // without losing scroll/expanded state.
  try {
    const j = await fetch('/api/index').then(r => r.json());
    DATA = j;
    ROWS_BY_NUM = Object.fromEntries(j.rows.map(r => [r.row, r]));
    TREE_ROOT = j.tree;
    buildRowBlobs();
    applyFilters();
    renderStats();
    if (SELECTED_ROW) {
      // Re-render xlsx + action bar for the still-selected row
      const r = ROWS_BY_NUM[SELECTED_ROW];
      if (r) {
        loadXlsx(SELECTED_ROW);
        renderActionBar(r);
      }
    }
    if (opts.reloadPdf && CURRENT_PDF) {
      // Force PDF re-fetch (annotations may have changed)
      await loadPdf(CURRENT_PDF.rel, PDF_PAGE, CURRENT_HIGHLIGHT);
    }
  } catch (e) {
    toast('Refresh failed', e.message, 'error', 4000);
  }
}

// ── Toast with action button (Sprint S1.2 — Undo support) ──────
function toastWithAction(title, body, actionLabel, actionFn, kind = 'info', timeout = 6000) {
  const wrap = document.getElementById('toasts');
  if (!wrap) return null;
  const el = document.createElement('div');
  el.className = 'toast ' + kind;
  el.setAttribute('role', 'status');
  el.setAttribute('aria-live', 'polite');
  el.innerHTML =
    `<button class="close" type="button" aria-label="ปิด" onclick="this.parentElement.remove()">${ico('x',12)}</button>` +
    `<div class="title">${escapeHtml(title)}</div>` +
    (body ? `<div class="body">${body}</div>` : '') +
    `<button class="toast-action">${escapeHtml(actionLabel)}</button>`;
  while (wrap.children.length >= 6) wrap.firstChild.remove();
  wrap.appendChild(el);
  el.querySelector('.toast-action').addEventListener('click', () => {
    actionFn();
    el.remove();
  });
  if (timeout) {
    setTimeout(() => {
      if (!el.parentElement) return;
      el.classList.add('fading');
      setTimeout(() => el.remove(), 400);
    }, timeout);
  }
  return el;
}

// Diff toast (Sprint S2 — before/after)
function toastDiff(title, oldStr, newStr, undoFn) {
  const body =
    `<span class="diff-old">${escapeHtml(oldStr || '(empty)')}</span>` +
    `<span class="diff-new">${escapeHtml(newStr || '(empty)')}</span>`;
  const wrap = document.getElementById('toasts');
  if (!wrap) return;
  const el = document.createElement('div');
  el.className = 'toast diff learn';
  el.setAttribute('role', 'status');
  el.innerHTML =
    `<button class="close" type="button" aria-label="ปิด" onclick="this.parentElement.remove()">${ico('x',12)}</button>` +
    `<div class="title">${escapeHtml(title)}</div>` +
    `<div class="body">${body}</div>` +
    (undoFn ? `<button class="toast-action">Undo</button>` : '');
  while (wrap.children.length >= 6) wrap.firstChild.remove();
  wrap.appendChild(el);
  if (undoFn) {
    el.querySelector('.toast-action').addEventListener('click', () => {
      undoFn();
      el.remove();
    });
  }
  setTimeout(() => {
    if (!el.parentElement) return;
    el.classList.add('fading');
    setTimeout(() => el.remove(), 400);
  }, 8000);
}

// ── setStatus undo support (Sprint S1.2) ───────────────────────
let _LAST_STATUS_UNDO = null;
const _origSetStatusForUndo = setStatus;
setStatus = async function(status) {
  if (!SELECTED_ROW) return;
  const prev = (DATA.status && DATA.status[SELECTED_ROW]) ? DATA.status[SELECTED_ROW].status : 'unverified';
  await _origSetStatusForUndo(status);
  if (status !== 'unverified' && prev !== status) {
    const rowToRevert = SELECTED_ROW;
    toastWithAction(
      `Marked ${labelFor(status)}`,
      `<code>R${rowToRevert}</code>${prev !== 'unverified' ? ` (was ${labelFor(prev)})` : ''}`,
      'Undo',
      async () => {
        SELECTED_ROW = rowToRevert;
        await _origSetStatusForUndo(prev);
      },
      kindFor(status),
      5000,
    );
  }
};
function labelFor(s) {
  return ({pass: '✓ Pass', fail: '✗ Fail', need_fix: '⚠ Fix', skip: '⏭ Skip', unverified: 'unverified'})[s] || s;
}
function kindFor(s) {
  return ({pass: 'info', fail: 'warn', need_fix: 'warn', skip: 'info'})[s] || 'info';
}

// ── Replace location.reload calls with partial refresh ─────────
// (Sprint S1.1) — patch the wizard's saveManualAnnotate
const _origSaveManualAnnotate = saveManualAnnotate;
saveManualAnnotate = async function() {
  if (!WIZ || !MANUAL_TARGET_ROW) return;
  if (WIZ.drawn.length < WIZ.steps.length) {
    toast('Wizard incomplete', `ต้องวาด ${WIZ.steps.length} rect ก่อน — ตอนนี้วาดแล้ว ${WIZ.drawn.length}`, 'warn', 3000);
    return;
  }
  const btn = document.getElementById('manual-save-btn');
  btn.disabled = true;
  const origLabel = btn.innerHTML;
  btn.innerHTML = `${ico('refresh',14)}<span>Saving…</span>`;
  try {
    const stepsPayload = WIZ.drawn.map(d => {
      const sq = EDIT_ANNOTS.find(a => a._id === d.square_id);
      const lb = EDIT_ANNOTS.find(a => a._id === d.label_id);
      return {
        content_rect: sq ? sq.rect : d.content_rect,
        label_rect:   lb ? lb.rect : d.label_rect,
        label_text:   lb ? (lb.contents || '') : (d.label_text || ''),
      };
    });
    let url, payload;
    if (WIZ.mode === 'reannotate' || WIZ.steps.length > 1 || WIZ.delete_existing) {
      url = '/api/reannotate/save';
      payload = {
        row: WIZ.row,
        pdf_rel: CURRENT_PDF.rel,
        page: PDF_PAGE,
        steps: stepsPayload,
        delete_existing: !!WIZ.delete_existing,
        delete_prefixes: WIZ.delete_prefixes || [],
        brand: WIZ.brand_hint,
        model: WIZ.model_hint,
      };
    } else {
      url = '/api/manual_annotate/save';
      const s = stepsPayload[0];
      payload = {row: WIZ.row, page: PDF_PAGE, content_rect: s.content_rect, label_rect: s.label_rect, label_text: s.label_text, pdf_rel: CURRENT_PDF.rel};
    }
    const r = await fetch(url, {method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify(payload)});
    const j = await r.json();
    if (!j.ok) {
      toast('Save failed', j.error || JSON.stringify(j), 'error', 6000);
      btn.disabled = false; btn.innerHTML = origLabel;
      return;
    }
    // Surface PDF-edit errors (Sprint S1.2 — bug #4.1 fix)
    if (j.pdf_result && j.pdf_result.errors > 0) {
      toast(`PDF edits had ${j.pdf_result.errors} errors`,
            `Applied: ${j.pdf_result.applied}, Errors: ${j.pdf_result.errors}. Check audit log.`,
            'warn', 8000);
    }
    // Diff toast with Undo option
    const oldD = j.old_d || '';
    const newD = j.new_d || '';
    const undoRow = WIZ.row;
    const undoFn = oldD ? async () => {
      await fetch('/api/row/col_d', {method: 'POST', headers: {'Content-Type':'application/json'},
        body: JSON.stringify({row: undoRow, col_d: oldD, original: newD})});
      await partialRefresh({reloadPdf: true});
      toast('Reverted', `R${undoRow} · Col D restored`, 'info', 3000);
    } : null;
    toastDiff('Re-annotated · Col D updated', oldD, newD, undoFn);
    if (typeof tickRetrain === 'function') tickRetrain('reannotate');
    MANUAL_MODE = false;
    const wasReannotate = (WIZ.mode === 'reannotate');
    const savedRow = WIZ.row;
    const wasBM = WIZ.steps.length === 2 && WIZ.steps[0].kind === 'brand';
    WIZ = null;
    document.getElementById('manual-banner').classList.remove('show');
    if (EDIT_MODE) toggleEditMode();
    await partialRefresh({reloadPdf: true});
    // Apply-to-siblings prompt for brand_model (Sprint S1.2)
    if (wasBM && wasReannotate) {
      offerApplyToSiblings(savedRow);
    }
  } catch (e) {
    toast('Save error', e.message, 'error', 5000);
    btn.disabled = false; btn.innerHTML = origLabel;
  }
};

// ── Apply-to-siblings (Sprint S1.2) ────────────────────────────
async function offerApplyToSiblings(savedRow) {
  const r = ROWS_BY_NUM[savedRow];
  if (!r) return;
  const sec = r.section;
  if (!sec) return;
  // Find sibling rows: same section root (5.1.1), brand_model type, NOT yet
  // matching the saved Col D shape
  const sectionRoot = sec.split('.').slice(0, 2).join('.');
  const siblings = (DATA.rows || []).filter(x => {
    if (x.row === savedRow) return false;
    if (!x.section) return false;
    if (!x.section.startsWith(sectionRoot)) return false;
    const t = (x.parsed && x.parsed.type) || '';
    return t === 'brand_model' || x.role === 'section_header';
  }).slice(0, 8);
  if (siblings.length === 0) return;
  toastWithAction(
    `Apply same pattern to ${siblings.length} similar rows?`,
    `<span style="font-size:var(--t-xs);color:var(--c-text-muted)">Sibling brand_model rows in section ${escapeHtml(sectionRoot)}.x</span>`,
    'Review',
    () => openBulkDialog(savedRow, siblings),
    'learn', 8000,
  );
}

function openBulkDialog(sourceRow, siblings) {
  const m = document.getElementById('bulk-modal');
  document.getElementById('bulk-info').innerHTML =
    `Source: <code>R${sourceRow}</code> — applying brand_model pattern (filename-based) to siblings.`;
  const list = document.getElementById('bulk-list');
  list.innerHTML = siblings.map(r => {
    const dPreview = (r.D || '').slice(0, 60);
    return `<label class="version-item" style="cursor:pointer">
      <input type="checkbox" class="bulk-row-cb" data-row="${r.row}" checked>
      <div>
        <div class="v-tag">R${r.row} · ${escapeHtml(r.section || '?')}</div>
        <div class="v-meta">${escapeHtml(dPreview)}${(r.D||'').length > 60 ? '…' : ''}</div>
      </div>
      <div></div>
    </label>`;
  }).join('');
  document.getElementById('bulk-summary').textContent =
    `${siblings.length} rows queued · brand/model auto-derived from each row's catalog filename`;
  m.classList.add('show');
}
function closeBulk() { document.getElementById('bulk-modal').classList.remove('show'); }
async function applyBulk() {
  const checked = [...document.querySelectorAll('.bulk-row-cb:checked')]
    .map(cb => parseInt(cb.dataset.row));
  if (!checked.length) { toast('Nothing selected', '', 'warn', 2500); closeBulk(); return; }
  const btn = document.getElementById('bulk-apply');
  btn.disabled = true;
  btn.innerHTML = `${ico('refresh',14)} Applying…`;
  let ok = 0, fail = 0;
  for (const rn of checked) {
    try {
      const r = await fetch(`/api/auto_annotate/preview?row=${rn}`).then(r=>r.json());
      if (!r.ok) { fail++; continue; }
      const ap = await fetch('/api/auto_annotate/apply', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({row: rn, write_xlsx: true, write_pdf: true})});
      const aj = await ap.json();
      if (aj.ok) ok++; else fail++;
    } catch (e) { fail++; }
  }
  closeBulk();
  toast(`Bulk apply: ${ok} ok / ${fail} fail`, '', ok && !fail ? 'info' : 'warn', 5000);
  await partialRefresh();
}

// ── Force-learn / Pin pattern (Sprint S2.1) ────────────────────
async function pinAsTemplate(rowNum) {
  try {
    const r = await fetch('/api/learn/pin_pattern', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({row: rowNum})
    });
    const j = await r.json();
    if (j.ok) {
      toast('🧠 Pattern pinned', `Pattern จาก R${rowNum} จะถูกใช้เป็น template`, 'learn', 5000);
    } else {
      toast('Pin failed', j.error || 'unknown', 'error', 4000);
    }
  } catch (e) {
    toast('Pin failed', e.message, 'error', 4000);
  }
}

// ── Command palette (Sprint UI-2.6) ────────────────────────────
let _CMDK_OPEN = false;
let _CMDK_RESULTS = [];
let _CMDK_INDEX = 0;
const CMDK_ACTIONS = [
  {id:'showAuto',  label:'Auto-annotate row',          icon:'sparkles', shortcut:'A', kind:'action', fn: () => SELECTED_ROW && showAutoAnnotate()},
  {id:'reann',     label:'Re-annotate row',            icon:'refresh',  shortcut:'R', kind:'action', fn: () => SELECTED_ROW && startReannotate()},
  {id:'mark',      label:'Mark in catalog (manual)',   icon:'pin',      shortcut:'M', kind:'action', fn: () => SELECTED_ROW && startManualAnnotate()},
  {id:'pin',       label:'Pin row as template',        icon:'sparkles', shortcut:'P', kind:'action', fn: () => SELECTED_ROW && pinAsTemplate(SELECTED_ROW)},
  {id:'next-uncertain', label:'Next uncertain row',    icon:'arrow-right', shortcut:'N', kind:'action', fn: () => nextUncertainRow()},
  {id:'pass',      label:'Mark as Pass',               icon:'check',    shortcut:'1', kind:'action', fn: () => SELECTED_ROW && setStatus('pass')},
  {id:'fail',      label:'Mark as Fail',               icon:'x',        shortcut:'2', kind:'action', fn: () => SELECTED_ROW && setStatus('fail')},
  {id:'fix',       label:'Mark as Need fix',           icon:'alert',    shortcut:'3', kind:'action', fn: () => SELECTED_ROW && setStatus('need_fix')},
  {id:'skip',      label:'Mark as Skip',               icon:'skip',     shortcut:'4', kind:'action', fn: () => SELECTED_ROW && setStatus('skip')},
  {id:'showLearning', label:'Open Learning panel',     icon:'brain',    kind:'action', fn: () => showLearning()},
  {id:'showAudit',    label:'Open Database & Audit',   icon:'chart',    kind:'action', fn: () => showAudit()},
  {id:'showVersions', label:'Open Project versions',   icon:'package',  kind:'action', fn: () => showVersions()},
  {id:'showSettings', label:'Open Settings',           icon:'settings', kind:'action', fn: () => showSettings()},
  {id:'snap',         label:'Quick snapshot',          icon:'camera',   kind:'action', fn: () => takeProjectSnap(false)},
  {id:'theme',        label:'Toggle theme (light/dark)', icon:'moon',   kind:'action', fn: () => toggleTheme()},
  {id:'help',         label:'Show keyboard shortcuts', icon:'help',     kind:'action', fn: () => showOnboarding(true)},
];
function openCmdK() {
  document.getElementById('cmdk-bg').classList.add('show');
  const inp = document.getElementById('cmdk-input');
  inp.value = '';
  setTimeout(() => inp.focus(), 30);
  _CMDK_OPEN = true;
  renderCmdK('');
}
function closeCmdK() {
  document.getElementById('cmdk-bg').classList.remove('show');
  _CMDK_OPEN = false;
}
function renderCmdK(q) {
  const list = document.getElementById('cmdk-list');
  q = (q || '').trim().toLowerCase();
  const items = [];
  // Actions filtered
  for (const a of CMDK_ACTIONS) {
    const score = q ? (a.label.toLowerCase().includes(q) ? 1 : (a.id.includes(q) ? 0.5 : 0)) : 1;
    if (score > 0) items.push({...a, score});
  }
  // Row search (when q non-empty)
  if (q && DATA && DATA.rows) {
    const qNorm = (typeof normalizeForSearch === 'function') ? normalizeForSearch(q) : q;
    let matched = 0;
    for (const r of DATA.rows) {
      if (matched >= 20) break;
      const blob = ROW_BLOBS.get(r.row) || '';
      if (!blob.includes(qNorm)) continue;
      items.push({
        id: 'row-' + r.row,
        kind: 'row',
        rowNum: r.row,
        label: `R${r.row} · ${(r.section || '?')} · ${(r.B || '').toString().slice(0, 60).trim()}`,
        meta: r.D ? r.D.slice(0, 40) : '',
        icon: 'arrow-right',
        score: 0.4,
        fn: () => { closeCmdK(); selectRow(r.row); },
      });
      matched++;
    }
  }
  items.sort((a, b) => b.score - a.score);
  _CMDK_RESULTS = items;
  _CMDK_INDEX = 0;
  if (!items.length) {
    list.innerHTML = '<div class="cmdk-empty">No results · ลองค้นชื่อ section หรือ keyword</div>';
    return;
  }
  // Render with section labels
  let html = '';
  let lastKind = null;
  for (let i = 0; i < items.length; i++) {
    const it = items[i];
    if (it.kind !== lastKind) {
      html += `<div class="cmdk-section">${it.kind === 'row' ? 'Rows' : 'Actions'}</div>`;
      lastKind = it.kind;
    }
    html += `<div class="cmdk-item${i === 0 ? ' cmdk-active' : ''}" data-i="${i}" role="option">
      ${ico(it.icon || 'arrow-right',16)}
      <div class="cmdk-item-main">${escapeHtml(it.label)}</div>
      ${it.meta ? `<div class="cmdk-item-meta">${escapeHtml(it.meta)}</div>` : ''}
      ${it.shortcut ? `<span class="cmdk-item-shortcut"><span class="kbd">${it.shortcut}</span></span>` : ''}
    </div>`;
  }
  list.innerHTML = html;
  list.querySelectorAll('.cmdk-item').forEach(el => {
    el.addEventListener('mouseenter', () => { _setCmdKIndex(parseInt(el.dataset.i)); });
    el.addEventListener('click', () => { _runCmdKItem(parseInt(el.dataset.i)); });
  });
}
function _setCmdKIndex(i) {
  _CMDK_INDEX = Math.max(0, Math.min(i, _CMDK_RESULTS.length - 1));
  const els = document.querySelectorAll('.cmdk-item');
  els.forEach((el, idx) => el.classList.toggle('cmdk-active', idx === _CMDK_INDEX));
  const active = els[_CMDK_INDEX];
  if (active) active.scrollIntoView({block: 'nearest'});
}
function _runCmdKItem(i) {
  const it = _CMDK_RESULTS[i];
  if (!it) return;
  closeCmdK();
  if (typeof it.fn === 'function') it.fn();
}

// ── Settings panel (Sprint UI-3.3) ─────────────────────────────
function showSettings() {
  const m = document.getElementById('settings-modal');
  // Sync current values
  const t = document.documentElement.getAttribute('data-theme') || '';
  document.getElementById('settings-theme').value = t;
  // LLM status — full Claude detail
  fetch('/api/learn/llm_status').then(r => r.json()).then(j => {
    const el = document.getElementById('settings-llm-status');
    if (!el) return;
    if (!j.available || j.name === 'off') {
      el.innerHTML = `<span style="color:var(--c-text-faint)">OFF — set ANTHROPIC_API_KEY + COMPLY_LLM=anthropic in .env to enable</span>`;
      return;
    }
    const spentPct = j.budget_usd_per_day
      ? Math.round((j.spent_today_usd || 0) / j.budget_usd_per_day * 100)
      : 0;
    const barColor = spentPct >= 90 ? 'var(--c-danger)'
                   : spentPct >= 70 ? 'var(--c-warn)'
                   : 'var(--c-success)';
    el.innerHTML = `
      <div style="display:flex;justify-content:space-between;margin-bottom:var(--s-2)">
        <span style="color:var(--c-success-text);font-weight:600">● ${escapeHtml(j.model || j.name)}</span>
        <span style="color:var(--c-text-soft)">${j.calls_today || 0} calls today</span>
      </div>
      <div style="margin-bottom:var(--s-2)">
        <span style="color:var(--c-text-muted)">Tokens:</span>
        in ${(j.tokens_in_today || 0).toLocaleString()} ·
        out ${(j.tokens_out_today || 0).toLocaleString()} ·
        cache ${(j.tokens_cache_read_today || 0).toLocaleString()}
      </div>
      <div style="margin-bottom:var(--s-2)">
        <span style="color:var(--c-text-muted)">Budget:</span>
        $${(j.spent_today_usd || 0).toFixed(3)} / $${(j.budget_usd_per_day || 0).toFixed(2)}
        (${spentPct}%)
      </div>
      <div style="background:var(--c-surface-3);height:4px;border-radius:2px;overflow:hidden">
        <div style="background:${barColor};height:100%;width:${Math.min(100, spentPct)}%;transition:width 0.3s"></div>
      </div>`;
  }).catch(() => {
    const el = document.getElementById('settings-llm-status');
    if (el) el.textContent = 'failed to load';
  });
  // Reduce motion
  const rm = localStorage.getItem('comply-reduce-motion') === '1';
  document.getElementById('settings-reduce-motion').checked = rm;
  // Show kbd
  const showKbd = localStorage.getItem('comply-show-kbd') !== '0';
  document.getElementById('settings-show-kbd').checked = showKbd;
  // Embedded mode (sync to current body attribute)
  const embCb = document.getElementById('settings-embedded');
  if (embCb) embCb.checked = document.body.getAttribute('data-embedded') === '1';
  // Top inset slider
  const stCur = document.body.style.getPropertyValue('--safe-top') ||
                getComputedStyle(document.documentElement).getPropertyValue('--safe-top') || '0';
  const stPx = parseInt(stCur) || 0;
  const stInp = document.getElementById('settings-safe-top');
  const stLbl = document.getElementById('settings-safe-top-val');
  if (stInp) stInp.value = stPx;
  if (stLbl) stLbl.textContent = stPx + 'px';
  // Pre-fill model + budget from current llm_status
  fetch('/api/learn/llm_status').then(r => r.json()).then(j => {
    refreshLlmStatusCard(j);
    const mSel = document.getElementById('settings-model');
    const bInp = document.getElementById('settings-budget');
    if (mSel && j.model) {
      // Match model prefix (server returns full id like 'claude-sonnet-4-5-20250929')
      for (const o of mSel.options) {
        if (j.model.startsWith(o.value)) { mSel.value = o.value; break; }
      }
    }
    if (bInp && typeof j.budget_usd_per_day === 'number') {
      bInp.value = j.budget_usd_per_day;
    }
  });
  m.classList.add('show');
}
function closeSettings() { document.getElementById('settings-modal').classList.remove('show'); }
function setThemeFromSettings(t) {
  if (t) {
    document.documentElement.setAttribute('data-theme', t);
    try { localStorage.setItem('comply-theme', t); } catch (e) {}
  } else {
    document.documentElement.removeAttribute('data-theme');
    try { localStorage.removeItem('comply-theme'); } catch (e) {}
  }
  // Update theme toggle icon
  const btn = document.getElementById('theme-toggle');
  const isDark = t === 'dark' || (t === '' && window.matchMedia('(prefers-color-scheme: dark)').matches);
  if (btn) btn.innerHTML = `<svg class="ico" aria-hidden="true"><use href="#i-${isDark ? 'sun' : 'moon'}"/></svg>`;
}
function setReduceMotion(on) {
  document.documentElement.classList.toggle('reduce-motion', on);
  try { localStorage.setItem('comply-reduce-motion', on ? '1' : '0'); } catch (e) {}
}
function setShowKbd(on) {
  const el = document.querySelector('.kbd-help');
  if (el) el.style.display = on ? '' : 'none';
  try { localStorage.setItem('comply-show-kbd', on ? '1' : '0'); } catch (e) {}
}

// ── Onboarding (Sprint UI-3.1) ─────────────────────────────────
function showOnboarding(force = false) {
  if (!force && localStorage.getItem('comply-onboarded') === '1') return;
  document.getElementById('onboard-bg').classList.add('show');
}
function closeOnboarding() { document.getElementById('onboard-bg').classList.remove('show'); }
function dontShowOnboardingAgain() { try { localStorage.setItem('comply-onboarded','1'); } catch(e) {} }

// ── Topbar Settings dropdown (Sprint UI-2.5) ───────────────────
function toggleTopbarMenu(e) {
  if (e) e.stopPropagation();
  const m = document.getElementById('topbar-menu');
  if (m.classList.contains('show')) {
    closeTopbarMenu();
    return;
  }
  // Position the fixed-position menu directly under the trigger button.
  // The trigger is the .topbar-menu-btn or its child icon-button.
  const trigger = document.querySelector('.topbar-menu-btn .topbar-btn')
                || document.querySelector('.topbar-menu-btn');
  if (trigger) {
    const r = trigger.getBoundingClientRect();
    // Show first so we can measure offsetWidth/Height
    m.classList.add('show');
    const W = m.offsetWidth || 240;
    const H = m.offsetHeight || 240;
    // Safe area: leave clearance for browser chrome / overlay bars that
    // some users have at the top (translate bar, password manager bar).
    const SAFE_TOP = 8;
    const SAFE_BOTTOM = 8;
    let left = r.right - W;                 // right-align with trigger
    let top  = r.bottom + 4;                 // default: just below the gear
    // Horizontal clamp
    if (left < 8) left = 8;
    if (left + W > window.innerWidth - 8) left = window.innerWidth - W - 8;
    // Vertical clamp:
    //   if not enough room below → flip above the trigger
    //   if still not enough → clamp to safe-top with internal scroll
    if (top + H > window.innerHeight - SAFE_BOTTOM) {
      const above = r.top - H - 4;
      if (above >= SAFE_TOP) {
        top = above;
      } else {
        // Neither side has room — pin near top, content scrolls
        top = SAFE_TOP;
      }
    }
    if (top < SAFE_TOP) top = SAFE_TOP;
    m.style.left = left + 'px';
    m.style.top  = top  + 'px';
  } else {
    m.classList.add('show');
  }
  setTimeout(() => {
    document.addEventListener('click', _closeTopbarMenuOnDocClick, true);
    document.addEventListener('keydown', _closeTopbarMenuOnEsc, true);
  }, 0);
}
function closeTopbarMenu() {
  const m = document.getElementById('topbar-menu');
  m.classList.remove('show');
  m.style.left = ''; m.style.top = '';
  document.removeEventListener('click', _closeTopbarMenuOnDocClick, true);
  document.removeEventListener('keydown', _closeTopbarMenuOnEsc, true);
}
function _closeTopbarMenuOnDocClick(e) {
  // Don't close when clicking inside the dropdown itself
  if (e.target.closest('#topbar-menu')) return;
  if (!e.target.closest('.topbar-menu-btn')) closeTopbarMenu();
}
function _closeTopbarMenuOnEsc(e) {
  if (e.key === 'Escape') { e.stopPropagation(); closeTopbarMenu(); }
}

// ── Wizard skill.md tip (Sprint S3.6) ──────────────────────────
const SKILL_TIPS = {
  brand_model: 'Brand logo มัก<strong>หน้า 1 บนซ้าย/ขวา</strong>; model number มักในตาราง spec — label ชี้ <code>ยี่ห้อ</code>/<code>รุ่น</code> เท่านั้น',
  item: 'Label ใช้รูป <code>{section} ข้อ N)</code> — rect ครอบเฉพาะข้อความข้อนั้น (ไม่รวม spec ข้างๆ)',
  sub_item: 'Label ใช้รูป <code>{section} ข้อ N) ข้อย่อย M.</code> — rect เฉพาะข้อย่อยนั้น',
  section: 'Section header rect ครอบ block ใหญ่ของ section นั้นใน catalog',
};
const _origRenderWizBanner = (typeof _renderWizBanner === 'function') ? _renderWizBanner : null;
if (_origRenderWizBanner) {
  _renderWizBanner = function() {
    _origRenderWizBanner();
    if (!WIZ) return;
    const ti = document.getElementById('manual-target-info');
    if (!ti) return;
    const stepKind = WIZ.steps[Math.min(WIZ.current_step, WIZ.steps.length - 1)]?.kind;
    let tipKey;
    if (stepKind === 'brand' || stepKind === 'model') tipKey = 'brand_model';
    else tipKey = stepKind || 'item';
    const tip = SKILL_TIPS[tipKey];
    if (tip && !ti.querySelector('.wiz-skill-tip')) {
      ti.insertAdjacentHTML('beforeend',
        `<div class="wiz-skill-tip">${ico('help',12)} <strong>SKILL.md tip:</strong> ${tip}</div>`);
    }
  };
}

// ── Counter animation (Sprint UI-2.4) ──────────────────────────
const _origRenderStats = renderStats;
let _LAST_PCT = null;
renderStats = function() {
  _origRenderStats();
  const pillEl = document.getElementById('stats-pill-progress');
  if (pillEl) {
    const m = pillEl.textContent.match(/(\d+)%/);
    if (m) {
      const pct = parseInt(m[1]);
      if (_LAST_PCT !== null && pct !== _LAST_PCT) {
        const strong = pillEl.querySelector('strong');
        if (strong) {
          strong.classList.remove('pop');
          // Force reflow then add class to retrigger animation
          void strong.offsetWidth;
          strong.classList.add('pop');
        }
      }
      _LAST_PCT = pct;
    }
  }
};

// ── Global keyboard: Cmd+K, ?, Esc ─────────────────────────────
document.addEventListener('keydown', (e) => {
  // Cmd+K / Ctrl+K — open palette
  if ((e.metaKey || e.ctrlKey) && e.key.toLowerCase() === 'k') {
    e.preventDefault();
    if (_CMDK_OPEN) closeCmdK(); else openCmdK();
    return;
  }
  // ? — open help/onboarding
  if (e.key === '?' && !['INPUT','TEXTAREA','SELECT'].includes(e.target.tagName)) {
    e.preventDefault();
    showOnboarding(true);
    return;
  }
  // Cmdk navigation
  if (_CMDK_OPEN) {
    if (e.key === 'Escape') { e.preventDefault(); closeCmdK(); return; }
    if (e.key === 'ArrowDown') { e.preventDefault(); _setCmdKIndex(_CMDK_INDEX + 1); return; }
    if (e.key === 'ArrowUp')   { e.preventDefault(); _setCmdKIndex(_CMDK_INDEX - 1); return; }
    if (e.key === 'Enter')     { e.preventDefault(); _runCmdKItem(_CMDK_INDEX); return; }
  }
}, true);
document.addEventListener('input', (e) => {
  if (e.target && e.target.id === 'cmdk-input') {
    renderCmdK(e.target.value);
  }
}, true);

// ── Init: show onboarding on first launch ──────────────────────
setTimeout(() => {
  if (localStorage.getItem('comply-onboarded') !== '1') {
    showOnboarding(false);
  }
}, 600);
// Restore reduce-motion + show-kbd preferences
if (localStorage.getItem('comply-reduce-motion') === '1') {
  document.documentElement.classList.add('reduce-motion');
}
if (localStorage.getItem('comply-show-kbd') === '0') {
  const el = document.querySelector('.kbd-help');
  if (el) el.style.display = 'none';
}

// ── Embedded mode (auto-detect iframe / Claude Preview MCP overlay) ──
// When the app runs inside an iframe (Claude Preview, embedded view,
// etc.), there's typically a host toolbar overlapping the top ~48px.
// We carve out --safe-top space so the topbar isn't covered.
(function applyEmbeddedMode() {
  let embedded;
  const explicit = localStorage.getItem('comply-embedded');
  if (explicit === '1' || explicit === '0') {
    embedded = explicit === '1';
  } else {
    // Auto: in an iframe → assume yes
    try {
      embedded = (window.self !== window.top);
    } catch (e) {
      embedded = true;       // cross-origin iframe access throws
    }
  }
  if (embedded) document.body.setAttribute('data-embedded', '1');
})();
function setEmbeddedMode(on) {
  if (on) {
    document.body.setAttribute('data-embedded', '1');
    try { localStorage.setItem('comply-embedded', '1'); } catch (e) {}
  } else {
    document.body.removeAttribute('data-embedded');
    try { localStorage.setItem('comply-embedded', '0'); } catch (e) {}
  }
}

// Top safe-area pixel inset (overrides --safe-top via inline style on body)
function setSafeTop(px) {
  const v = Math.max(0, Math.min(200, parseInt(px) || 0));
  if (v === 0) {
    document.body.style.removeProperty('--safe-top');
    try { localStorage.removeItem('comply-safe-top'); } catch (e) {}
  } else {
    document.body.style.setProperty('--safe-top', v + 'px');
    try { localStorage.setItem('comply-safe-top', String(v)); } catch (e) {}
  }
  const lbl = document.getElementById('settings-safe-top-val');
  if (lbl) lbl.textContent = v + 'px';
}
// Restore on boot
(function _initSafeTop() {
  try {
    const saved = parseInt(localStorage.getItem('comply-safe-top'));
    if (saved > 0) document.body.style.setProperty('--safe-top', saved + 'px');
  } catch (e) {}
})();

// API key — show/hide + save + clear
function toggleApiKeyVisibility() {
  const inp = document.getElementById('settings-api-key');
  if (!inp) return;
  inp.type = inp.type === 'password' ? 'text' : 'password';
}
async function saveApiKey() {
  const inp = document.getElementById('settings-api-key');
  const model = document.getElementById('settings-model').value;
  const budget = parseFloat(document.getElementById('settings-budget').value);
  const key = inp.value.trim();
  if (!key) {
    toast('Empty API key', 'พิมพ์ key ก่อน หรือกด Clear ถ้าต้องการล้าง', 'warn', 3000);
    return;
  }
  if (!key.startsWith('sk-ant-')) {
    toast('Invalid format', 'API key ต้องเริ่มด้วย sk-ant-', 'error', 4000);
    return;
  }
  const btn = document.getElementById('settings-save-key');
  const orig = btn.innerHTML;
  btn.disabled = true;
  btn.innerHTML = `${ico('refresh',14)}<span>Saving…</span>`;
  try {
    const r = await fetch('/api/settings/api_key', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({api_key: key, model, budget_usd_per_day: budget}),
    });
    const j = await r.json();
    if (!j.ok) {
      toast('Save failed', j.error || 'unknown', 'error', 5000);
    } else {
      toast(`✓ Claude activated`, `${j.status?.model || 'provider'} · budget $${(j.status?.budget_usd_per_day || 0).toFixed(2)}/day`, 'learn', 5000);
      // Mask the key after save
      inp.value = '••••••••' + key.slice(-4);
      inp.type = 'password';
      // Refresh status display
      if (typeof showSettings === 'function') {
        // Re-fetch status without reopening modal
        fetch('/api/learn/llm_status').then(r => r.json()).then(refreshLlmStatusCard);
      }
    }
  } catch (e) {
    toast('Save error', e.message, 'error', 5000);
  } finally {
    btn.disabled = false;
    btn.innerHTML = orig;
  }
}
async function clearApiKey() {
  if (!confirm('ลบ API key ออกจาก .env? Claude provider จะปิดทันที')) return;
  try {
    const r = await fetch('/api/settings/api_key', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({api_key: ''}),
    });
    const j = await r.json();
    if (j.ok) {
      toast('Key cleared', 'Claude provider ปิดแล้ว — rules-only mode', 'info', 3500);
      const inp = document.getElementById('settings-api-key');
      if (inp) { inp.value = ''; inp.type = 'password'; }
      fetch('/api/learn/llm_status').then(r => r.json()).then(refreshLlmStatusCard);
    } else {
      toast('Clear failed', j.error || 'unknown', 'error', 4000);
    }
  } catch (e) { toast('Clear error', e.message, 'error', 4000); }
}
function refreshLlmStatusCard(j) {
  const el = document.getElementById('settings-llm-status');
  if (!el) return;
  if (!j.available || j.name === 'off' || j.provider_kind === 'off') {
    el.innerHTML = `
      <div style="color:var(--c-text-faint)">OFF</div>
      <div style="font-size:var(--t-xs);color:var(--c-text-muted);margin-top:var(--s-2)">
        ▸ <strong>Recommended</strong>: install Claude Code CLI (<code style="background:var(--c-surface);padding:1px 4px;border-radius:3px">npm i -g @anthropic-ai/claude-code</code>),
        run <code style="background:var(--c-surface);padding:1px 4px;border-radius:3px">claude login</code>, then set <code>COMPLY_LLM=claude_code</code> and restart.<br>
        ▸ <em>Or</em> paste an API key below as fallback.
      </div>`;
    return;
  }
  // Phase 1: Claude Code via Claude Max OAuth
  if (j.provider_kind === 'claude_code') {
    const auth = j.auth_mode || 'unknown';
    const authBadge = auth === 'claude_max'
      ? '<span style="background:var(--c-success-soft);color:var(--c-success-text);padding:2px 8px;border-radius:999px;font-size:11px;font-weight:700">Claude Max OAuth</span>'
      : `<span style="background:var(--c-warn-soft);color:var(--c-warn-text);padding:2px 8px;border-radius:999px;font-size:11px;font-weight:700">${escapeHtml(auth)}</span>`;
    el.innerHTML = `
      <div style="display:flex;justify-content:space-between;margin-bottom:var(--s-2);align-items:center">
        <span style="color:var(--c-success-text);font-weight:600">● ${escapeHtml(j.model || 'claude-code')}</span>
        ${authBadge}
      </div>
      <div style="margin-bottom:var(--s-2);font-size:var(--t-xs);color:var(--c-text-muted)">
        ${j.calls_today || 0} calls today · agentic mode (Read + Grep + custom MCP tools)
      </div>
      <div style="font-size:var(--t-xs);color:var(--c-text-faint)">
        💡 Cost = subscription (no per-call charge). Each row run = 1 agentic loop, may take 10–30 s.
      </div>`;
    return;
  }
  // Anthropic API path (legacy)
  const spentPct = j.budget_usd_per_day
    ? Math.round((j.spent_today_usd || 0) / j.budget_usd_per_day * 100)
    : 0;
  const barColor = spentPct >= 90 ? 'var(--c-danger)'
                 : spentPct >= 70 ? 'var(--c-warn)'
                 : 'var(--c-success)';
  el.innerHTML = `
    <div style="display:flex;justify-content:space-between;margin-bottom:var(--s-2)">
      <span style="color:var(--c-success-text);font-weight:600">● ${escapeHtml(j.model || j.name)}</span>
      <span style="background:var(--c-bg-soft);color:var(--c-text-soft);padding:2px 8px;border-radius:999px;font-size:11px;font-weight:600">API key (metered)</span>
    </div>
    <div style="margin-bottom:var(--s-2);font-size:var(--t-xs)">
      <span style="color:var(--c-text-muted)">Tokens:</span>
      in ${(j.tokens_in_today || 0).toLocaleString()} ·
      out ${(j.tokens_out_today || 0).toLocaleString()} ·
      cache ${(j.tokens_cache_read_today || 0).toLocaleString()}
    </div>
    <div style="margin-bottom:var(--s-2);font-size:var(--t-xs)">
      <span style="color:var(--c-text-muted)">Budget:</span>
      $${(j.spent_today_usd || 0).toFixed(3)} / $${(j.budget_usd_per_day || 0).toFixed(2)}
      (${spentPct}%)
    </div>
    <div style="background:var(--c-surface-3);height:4px;border-radius:2px;overflow:hidden">
      <div style="background:${barColor};height:100%;width:${Math.min(100, spentPct)}%;transition:width 0.3s"></div>
    </div>`;
}

// ── Mobile: ensure topbar still has access via hamburger (UI-3.4) ──
// At <700px the topbar is hidden; mobile-tabs has 3 tabs already.
// Add a 4th tab "More" that toggles a drawer of Settings/Audit/Versions.
(function injectMobileMore() {
  const tabs = document.querySelector('.mobile-tabs');
  if (!tabs) return;
  if (tabs.querySelector('[data-tab="more"]')) return;
  const moreBtn = document.createElement('button');
  moreBtn.dataset.tab = 'more';
  moreBtn.setAttribute('role', 'tab');
  moreBtn.setAttribute('aria-selected', 'false');
  moreBtn.innerHTML = `${ico('settings',14)}<span>More</span>`;
  moreBtn.onclick = () => toggleTopbarMenu();
  tabs.appendChild(moreBtn);
})();

// ============================================================
// Phase A — Acrobat-style layout: rail / ribbon / AI pane / status
// ============================================================

// ── Activity rail panels ──────────────────────────────────────
function setRailPanel(panel) {
  // 'tree' is the default — no rail-panel attribute (keeps the existing
  // tree-pane visible). For other panels (learn / versions / audit /
  // search), open them as MODALS for now (existing UI). Future: render
  // them inside .rail-panel-body for in-place navigation.
  document.querySelectorAll('.activity-rail .rail-btn').forEach(b => {
    b.classList.toggle('active', b.dataset.panel === panel);
  });
  // Tree panel = default state, no override
  if (panel === 'tree' || !panel) {
    document.body.removeAttribute('data-rail-panel');
    return;
  }
  // For MVP, route other rail items to existing modal flows
  if (panel === 'learn') showLearning();
  else if (panel === 'versions') showVersions();
  else if (panel === 'audit') showAudit();
  // After the modal action, return rail to tree
  setTimeout(() => {
    document.querySelector('.activity-rail .rail-btn[data-panel="tree"]')?.classList.add('active');
    document.querySelectorAll('.activity-rail .rail-btn').forEach(b => {
      if (b.dataset.panel !== 'tree') b.classList.remove('active');
    });
  }, 200);
}

// ── Mode tabs (Verify / Edit / Re-annotate / Apply) ───────────
function setMode(mode) {
  document.body.setAttribute('data-mode', mode);
  document.querySelectorAll('.ribbon-tab').forEach(t => {
    const isActive = t.dataset.mode === mode;
    t.classList.toggle('active', isActive);
    t.setAttribute('aria-selected', isActive ? 'true' : 'false');
  });
  // Sync edit mode toggle with the existing toggleEditMode flow
  if (mode === 'edit' && typeof EDIT_MODE !== 'undefined' && !EDIT_MODE) {
    if (typeof toggleEditMode === 'function') toggleEditMode();
  } else if (mode !== 'edit' && typeof EDIT_MODE !== 'undefined' && EDIT_MODE) {
    if (typeof toggleEditMode === 'function') toggleEditMode();
  }
  // Apply mode → load AI proposal eagerly
  if (mode === 'apply' && SELECTED_ROW) aiPaneRefresh();
  // Phase A5: floating toolbar visibility tracks edit mode
  if (typeof updateFloatingToolbar === 'function') updateFloatingToolbar();
}
// Default mode
document.body.setAttribute('data-mode', 'verify');

// ── AI pane toggle + render ───────────────────────────────────
let _AI_PANE_OPEN = (localStorage.getItem('comply-ai-pane') !== '0');
function applyAiPaneState() {
  document.body.setAttribute('data-ai-pane', _AI_PANE_OPEN ? '1' : '0');
}
applyAiPaneState();
function toggleAiPane() {
  _AI_PANE_OPEN = !_AI_PANE_OPEN;
  try { localStorage.setItem('comply-ai-pane', _AI_PANE_OPEN ? '1' : '0'); } catch (e) {}
  applyAiPaneState();
}

// Cache last AI proposal per row to avoid re-fetching
const _AI_CACHE = new Map();

function aiPaneRender(row, plan) {
  const body = document.getElementById('ai-pane-body');
  const head = document.getElementById('ai-pane-head');
  const sub  = document.getElementById('ai-pane-sub');
  if (!body) return;
  if (!row) {
    body.innerHTML = `<div class="ai-empty">
      <svg class="ico ico-xl" aria-hidden="true" style="opacity:0.3;margin-bottom:var(--s-3)"><use href="#i-sparkles"/></svg>
      <div>เลือก row เพื่อเริ่ม</div>
    </div>`;
    if (sub) sub.textContent = 'ready';
    return;
  }
  if (sub) sub.textContent = `R${row.row} · ${row.section || '?'}`;

  // Compose a primary "Proposal" section from auto_annotate_plan
  const conf = plan ? (plan.confidence || 0) : 0;
  const confPct = Math.round(conf * 100);
  const confCls = conf >= 0.85 ? '' : conf >= 0.6 ? 'med' : 'low';
  const proposed = plan ? (plan.proposed_d || '') : '';
  const generator = plan ? (plan.generator || 'rules') : '';
  const usedClaude = generator.startsWith('claude');
  const rationale = plan && plan.llm ? (plan.llm.rationale || '') : '';
  const escalation = plan && plan.llm ? plan.llm.escalation : null;

  // Phase B5: which learned patterns fired? Pull from plan.provenance —
  // this is what auto_annotate_plan attaches when apply_learned_brand /
  // apply_learned_vendor / etc. return a hit. Each entry is keyed by
  // its role (brand, vendor, …) and carries pattern_type, trigger,
  // confidence, samples.
  const provenance = (plan && plan.provenance) || {};
  const triggered = [];
  for (const [role, prov] of Object.entries(provenance)) {
    if (!prov || prov.kind !== 'learned') continue;
    triggered.push({
      role,
      pattern_type: prov.pattern_type || 'pattern',
      trigger: prov.trigger || '',
      confidence: prov.confidence || 0,
      samples: prov.samples || 0,
    });
  }
  const patternsSection = triggered.length ? `
    <div class="ai-patterns" aria-label="patterns triggered">
      <div class="ai-patterns-head">${ico('brain',12)}<span>Patterns triggered</span><span class="ai-patterns-count">${triggered.length}</span></div>
      ${triggered.map(t => `
        <div class="ai-pattern-row" title="role: ${escapeHtml(t.role)}">
          <span class="ai-pattern-type">${escapeHtml(t.pattern_type)}</span>
          <span class="ai-pattern-trigger">${escapeHtml(t.trigger)}</span>
          <span class="ai-pattern-meta">
            <span class="ai-pattern-conf">${Math.round(t.confidence*100)}%</span>
            <span class="ai-pattern-samples">${t.samples} samples</span>
          </span>
        </div>`).join('')}
    </div>` : '';

  const propSection = `
    <div class="ai-section">
      <h4>${ico('sparkles',14)}<span>Proposal</span> <span style="margin-left:auto;font-weight:500;text-transform:none;letter-spacing:0;color:var(--c-text-soft)">${escapeHtml(generator)}</span></h4>
      ${proposed
        ? `<div class="ai-proposal-text">${escapeHtml(proposed)}</div>
           <div class="ai-conf-bar ${confCls}"><span style="width:${confPct}%"></span></div>
           <div class="ai-conf-text">
             <span>Confidence</span>
             <strong>${confPct}%</strong>
           </div>`
        : '<div style="font-style:italic;color:var(--c-text-faint);font-size:var(--t-sm)">No proposal</div>'}
      ${rationale ? `<div class="ai-rationale">${escapeHtml(rationale)}</div>` : ''}
      ${patternsSection}
      ${escalation ? `
        <div style="background:var(--c-warn-soft);border-left:3px solid var(--c-warn);padding:var(--s-3) var(--s-4);border-radius:0 var(--r-sm) var(--r-sm) 0;font-size:var(--t-xs);color:var(--c-warn-text);margin:var(--s-3) 0">
          <strong>Claude is uncertain:</strong><br>${escapeHtml(escalation.question || '')}
        </div>` : ''}
      ${proposed ? `
        <div class="ai-actions">
          <button class="ai-accept" onclick="aiAccept()" title="Apply proposal · then verdict pass">${ico('check',14)}<span>Accept</span></button>
          <button class="ai-edit" onclick="aiEditInline()">${ico('pencil',14)}<span>Edit</span></button>
          <button class="ai-reject" onclick="aiReject()">${ico('x',14)}<span>Reject</span></button>
        </div>` : ''}
    </div>`;

  // Teach-back section (always visible)
  const teachSection = `
    <div class="ai-section ai-teach">
      <h4>${ico('brain',14)}<span>Teach Claude</span></h4>
      <div class="ai-tags" id="ai-tags">
        <span class="ai-tag" data-tag="#wrong-page" onclick="toggleAiTag(this)">#wrong-page</span>
        <span class="ai-tag" data-tag="#brand-wrong" onclick="toggleAiTag(this)">#brand-wrong</span>
        <span class="ai-tag" data-tag="#missing-spec" onclick="toggleAiTag(this)">#missing-spec</span>
        <span class="ai-tag" data-tag="#typo" onclick="toggleAiTag(this)">#typo</span>
        <span class="ai-tag" data-tag="#format" onclick="toggleAiTag(this)">#format</span>
        <span class="ai-tag" data-tag="#commitment" onclick="toggleAiTag(this)">#commitment</span>
      </div>
      <textarea id="ai-teach-text" placeholder="ทำไมต้องแก้? (optional)"></textarea>
      <button class="ai-send" onclick="aiTeachSend()">Send to Claude</button>
    </div>`;

  // Phase 1 (Claude Code as core): Live agent run — user clicks Run, the
  // SSE endpoint streams Claude's reasoning + tool calls + final proposal.
  const liveSection = `
    <div class="ai-section ai-claude-live" id="ai-claude-live-section">
      <h4>${ico('sparkles',14)}<span>Run with Claude Code</span>
        <span style="margin-left:auto;font-size:11px;font-weight:500;color:var(--c-text-faint);text-transform:none;letter-spacing:0">live · agentic</span>
      </h4>
      <div class="ai-claude-actions">
        <button class="ai-claude-run" id="ai-claude-run-btn" onclick="aiClaudeRun()" title="Run Claude with full tool use (Read/Grep + custom tools). Watch reasoning live.">
          ${ico('zap',14)}<span>Run</span>
        </button>
        <button class="ai-claude-stop" id="ai-claude-stop-btn" onclick="aiClaudeStop()" disabled>
          ${ico('x',14)}<span>Stop</span>
        </button>
      </div>
      <div class="ai-claude-log" id="ai-claude-log" aria-label="Claude live reasoning"></div>
    </div>`;

  // Recent corrections (compact)
  const recentSection = `
    <div class="ai-section">
      <h4>${ico('clock',14)}<span>Recent (last 5)</span></h4>
      <div id="ai-recent" style="font-size:var(--t-xs);color:var(--c-text-soft);font-family:var(--f-mono)">loading…</div>
    </div>`;

  body.innerHTML = propSection + liveSection + teachSection + recentSection;
  // Lazy-load recent corrections
  fetch('/api/learn/stats').then(r => r.json()).then(j => {
    const el = document.getElementById('ai-recent');
    if (!el || !j) return;
    const acc = (j.accuracy * 100).toFixed(0);
    el.innerHTML = `${j.total_feedbacks||0} feedbacks · ${acc}% accuracy<br>${j.patterns_total||0} patterns (${j.patterns_enabled||0} enabled)`;
  }).catch(() => {});
}

async function aiPaneRefresh() {
  if (!SELECTED_ROW) { aiPaneRender(null); return; }
  const row = ROWS_BY_NUM[SELECTED_ROW];
  if (!row) return;
  // Show row context immediately, fetch proposal in bg
  aiPaneRender(row, null);
  if (_AI_CACHE.has(SELECTED_ROW)) {
    aiPaneRender(row, _AI_CACHE.get(SELECTED_ROW));
    return;
  }
  try {
    const r = await fetch(`/api/auto_annotate/preview?row=${SELECTED_ROW}`);
    const j = await r.json();
    if (j && j.ok !== false) {
      _AI_CACHE.set(SELECTED_ROW, j);
      aiPaneRender(row, j);
    }
  } catch (e) {
    console.warn('ai-pane fetch failed', e);
  }
}

// AI action handlers
async function aiAccept() {
  if (!SELECTED_ROW) return;
  const plan = _AI_CACHE.get(SELECTED_ROW);
  if (!plan) { toast('No proposal to accept', '', 'warn', 2500); return; }
  try {
    const r = await fetch('/api/auto_annotate/apply', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({row: SELECTED_ROW, write_xlsx: true, write_pdf: true})
    });
    const j = await r.json();
    if (j.ok) {
      toast('✓ Proposal applied', `R${SELECTED_ROW} · Col D updated`, 'learn', 4000);
      _AI_CACHE.delete(SELECTED_ROW);
      await partialRefresh({reloadPdf: true});
    } else {
      toast('Apply failed', j.error || 'unknown', 'error', 4500);
    }
  } catch (e) { toast('Apply error', e.message, 'error', 4500); }
}
function aiEditInline() {
  if (!SELECTED_ROW) return;
  // Open inline Col D editor on the target row
  const td = document.querySelector(`tr[data-row="${SELECTED_ROW}"] td.col-D`);
  if (td && typeof editColD === 'function') {
    editColD({currentTarget: td, stopPropagation: () => {}}, SELECTED_ROW);
  }
}
async function aiReject() {
  if (!SELECTED_ROW) return;
  const plan = _AI_CACHE.get(SELECTED_ROW);
  if (!plan) return;
  // Required reason — focus the teach textarea
  const t = document.getElementById('ai-teach-text');
  if (t) {
    t.placeholder = 'Why was this wrong? (required to record a rejection)';
    t.focus();
  }
  toast('Add reason', 'พิมพ์เหตุผลใน Teach Claude แล้วกด Send', 'warn', 3500);
}
function toggleAiTag(el) {
  el.classList.toggle('active');
  const t = document.getElementById('ai-teach-text');
  if (!t) return;
  const tag = el.dataset.tag;
  if (el.classList.contains('active')) {
    if (!t.value.includes(tag)) t.value = (t.value ? t.value + ' ' : '') + tag;
  } else {
    t.value = t.value.replace(new RegExp('\\\\s*' + tag + '\\\\s*'), ' ').trim();
  }
}
async function aiTeachSend() {
  if (!SELECTED_ROW) return;
  const t = document.getElementById('ai-teach-text');
  const text = (t && t.value || '').trim();
  if (!text) { toast('Empty', 'พิมพ์ข้อความก่อน', 'warn', 2500); return; }
  // Record as feedback with provenance.user_note
  const row = ROWS_BY_NUM[SELECTED_ROW];
  const plan = _AI_CACHE.get(SELECTED_ROW) || {};
  try {
    const r = await fetch('/api/learn/feedback', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({
        row_num: SELECTED_ROW,
        section: row?.section,
        input_b: row?.B || '',
        input_pdf_rel: row?.pdf_rel || null,
        input_role: row?.role || '',
        input_filename: row?.pdf_rel ? row.pdf_rel.split('/').pop() : null,
        suggested_c: plan.proposed_c || '',
        suggested_d: plan.proposed_d || '',
        confidence: plan.confidence || 0,
        generator: plan.generator || 'rules',
        provenance: {user_note: text, llm: plan.llm || null},
        user_action: 'edited',
        final_c: row?.C || '',
        final_d: row?.D || '',
      }),
    });
    const j = await r.json();
    if (j.ok) {
      toast('✓ Sent to Claude', `Feedback recorded for R${SELECTED_ROW}`, 'learn', 3500);
      if (t) t.value = '';
      document.querySelectorAll('#ai-tags .ai-tag.active').forEach(x => x.classList.remove('active'));
      if (typeof tickRetrain === 'function') tickRetrain('teach-back');
    }
  } catch (e) { toast('Send error', e.message, 'error', 4000); }
}

// ── Phase 1: Claude Code live-streaming runner ────────────────
// Opens an EventSource on /api/claude/stream, renders each event as a
// chip in the live log, fills in the proposal at completion.
let _CLAUDE_ES = null;

function aiClaudeRun() {
  if (!SELECTED_ROW) {
    toast('Pick a row first', 'เลือก row จาก tree', 'warn', 2500);
    return;
  }
  if (_CLAUDE_ES) { try { _CLAUDE_ES.close(); } catch(e){} _CLAUDE_ES = null; }
  const log = document.getElementById('ai-claude-log');
  const runBtn = document.getElementById('ai-claude-run-btn');
  const stopBtn = document.getElementById('ai-claude-stop-btn');
  if (!log) return;
  log.innerHTML = '';
  log.classList.add('streaming');
  if (runBtn)  runBtn.disabled = true;
  if (stopBtn) stopBtn.disabled = false;
  _aiClaudeAppend(log, 'sys', `Asking Claude to analyze R${SELECTED_ROW}…`);

  const url = `/api/claude/stream?row=${SELECTED_ROW}`;
  const es = new EventSource(url);
  _CLAUDE_ES = es;

  es.onmessage = (e) => {
    let evt;
    try { evt = JSON.parse(e.data); } catch (err) { return; }
    if (!evt || !evt.type) return;
    switch (evt.type) {
      case 'thinking':
        _aiClaudeAppend(log, 'thinking', evt.text || '');
        break;
      case 'tool_use':
        _aiClaudeAppend(log, 'tool', `${evt.name}(${_aiClaudeShort(evt.input)})`);
        break;
      case 'tool_result':
        _aiClaudeAppend(log, evt.is_error ? 'tool-err' : 'tool-res',
                         (evt.text || '').slice(0, 200));
        break;
      case 'text':
        _aiClaudeAppend(log, 'text', evt.content || '');
        break;
      case 'result':
        _aiClaudeFinish(log, evt);
        es.close(); _CLAUDE_ES = null;
        if (runBtn)  runBtn.disabled = false;
        if (stopBtn) stopBtn.disabled = true;
        log.classList.remove('streaming');
        break;
      case 'error':
        const errStr = String(evt.error || 'unknown error');
        _aiClaudeAppend(log, 'err', errStr);
        // Common helpful hint: not-logged-in
        if (/not logged in|please run .?login|auth login/i.test(errStr) ||
            errStr.includes('returned an error')) {
          const hint = document.createElement('div');
          hint.className = 'aclog sys';
          hint.style.borderLeftColor = 'var(--c-warn)';
          hint.style.background = 'var(--c-warn-soft)';
          hint.style.color = 'var(--c-warn-text)';
          hint.style.fontStyle = 'normal';
          hint.style.fontWeight = '600';
          hint.innerHTML = 'ℹ Run <code style="background:var(--c-surface);padding:1px 4px;border-radius:3px">claude auth login</code> in a terminal once to authenticate (uses Claude Max subscription).';
          log.appendChild(hint);
        }
        es.close(); _CLAUDE_ES = null;
        if (runBtn)  runBtn.disabled = false;
        if (stopBtn) stopBtn.disabled = true;
        log.classList.remove('streaming');
        break;
    }
  };
  es.onerror = () => {
    _aiClaudeAppend(log, 'err', 'Stream disconnected');
    try { es.close(); } catch (e) {}
    _CLAUDE_ES = null;
    if (runBtn)  runBtn.disabled = false;
    if (stopBtn) stopBtn.disabled = true;
    log.classList.remove('streaming');
  };
  es.addEventListener('done', () => {
    try { es.close(); } catch (e) {}
    _CLAUDE_ES = null;
    if (runBtn)  runBtn.disabled = false;
    if (stopBtn) stopBtn.disabled = true;
    log.classList.remove('streaming');
  });
}

function aiClaudeStop() {
  if (_CLAUDE_ES) {
    try { _CLAUDE_ES.close(); } catch(e) {}
    _CLAUDE_ES = null;
  }
  const log = document.getElementById('ai-claude-log');
  if (log) {
    _aiClaudeAppend(log, 'sys', '⏹ Stopped by user');
    log.classList.remove('streaming');
  }
  const runBtn = document.getElementById('ai-claude-run-btn');
  const stopBtn = document.getElementById('ai-claude-stop-btn');
  if (runBtn)  runBtn.disabled = false;
  if (stopBtn) stopBtn.disabled = true;
}

function _aiClaudeAppend(log, kind, text) {
  const el = document.createElement('div');
  el.className = 'aclog ' + kind;
  el.textContent = text;
  log.appendChild(el);
  log.scrollTop = log.scrollHeight;
}

function _aiClaudeShort(input) {
  if (!input) return '';
  try {
    if (typeof input === 'string') return input.slice(0, 60);
    const s = JSON.stringify(input);
    return s.length > 80 ? s.slice(0, 77) + '…' : s;
  } catch (e) { return ''; }
}

function _aiClaudeFinish(log, evt) {
  const prop = evt.proposal;
  const ms   = evt.elapsed_ms || 0;
  const cost = (evt.cost_usd || 0).toFixed(4);
  const wrap = document.createElement('div');
  wrap.className = 'aclog result';
  if (prop && prop.input) {
    const txt = prop.input.col_d_text || prop.input.col_d || prop.input.brand || '(see details)';
    const conf = prop.input.confidence || 0;
    wrap.innerHTML = `
      <div class="aclog-result-head">
        ${ico('check',12)}<strong>${escapeHtml(prop.name)}</strong>
        <span class="ac-meta">${ms}ms · $${cost}</span>
      </div>
      <div class="aclog-result-text">${escapeHtml(txt)}</div>
      <div class="aclog-result-conf">conf: ${(conf*100).toFixed(0)}%</div>
      ${prop.input.rationale ? `<div class="aclog-result-rat">${escapeHtml(prop.input.rationale)}</div>` : ''}
      <div class="aclog-result-actions">
        <button class="ai-accept" onclick='aiClaudeAccept(${JSON.stringify(prop).replace(/'/g, "&apos;")})'>${ico('check',13)}<span>Accept</span></button>
        <button class="ai-reject" onclick='aiClaudeReject(${JSON.stringify(prop).replace(/'/g, "&apos;")})'>${ico('x',13)}<span>Reject</span></button>
      </div>`;
  } else {
    wrap.innerHTML = `
      <div class="aclog-result-head">
        ${ico('alert',12)}<strong>No proposal</strong>
        <span class="ac-meta">${ms}ms</span>
      </div>
      <div class="aclog-result-rat">Claude finished without calling a structured tool. See log above.</div>`;
  }
  log.appendChild(wrap);
  log.scrollTop = log.scrollHeight;
}

async function aiClaudeAccept(proposal) {
  if (!SELECTED_ROW || !proposal || !proposal.input) return;
  const text = proposal.input.col_d_text || '';
  if (!text) { toast('No text to apply', 'Proposal ไม่มี col_d_text', 'warn', 3000); return; }
  try {
    const r = await fetch('/api/row/col_d', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({row: SELECTED_ROW, col_d: text, original: ROWS_BY_NUM[SELECTED_ROW]?.D || ''}),
    });
    const j = await r.json();
    if (j.ok) {
      toast('✓ Accepted', `R${SELECTED_ROW} Col D updated`, 'learn', 3500);
      if (typeof tickRetrain === 'function') tickRetrain('claude_accept');
      if (SELECTED_ROW) loadXlsx(SELECTED_ROW);
    } else { toast('Save failed', j.error || 'unknown', 'error', 4000); }
  } catch (e) { toast('Save error', e.message, 'error', 4000); }
}

async function aiClaudeReject(proposal) {
  // Record as a learning signal — Claude proposed X, user rejected.
  if (!SELECTED_ROW) return;
  const row = ROWS_BY_NUM[SELECTED_ROW];
  try {
    await fetch('/api/learn/feedback', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({
        row_num: SELECTED_ROW,
        section: row?.section,
        input_b: row?.B || '',
        suggested_d: proposal?.input?.col_d_text || '',
        confidence: proposal?.input?.confidence || 0,
        generator: 'claude_code',
        provenance: {tool: proposal?.name},
        user_action: 'rejected',
        final_d: row?.D || '',
      }),
    });
    toast('✓ Rejected', 'Logged as training signal', 'info', 2500);
  } catch (e) { /* swallow */ }
}

// ── Status bar update ─────────────────────────────────────────
function statusBarUpdate() {
  const sbRow = document.getElementById('sb-row-info');
  if (!sbRow) return;
  const r = SELECTED_ROW ? ROWS_BY_NUM[SELECTED_ROW] : null;
  // Phase A6: row-selected attr drives whether verdict pills are clickable
  document.body.setAttribute('data-row-selected', r ? '1' : '0');
  if (r) {
    const sec = r.section || '?';
    const summary = (r.D || r.B || '').toString().slice(0, 60).trim();
    sbRow.innerHTML = `<strong>R${r.row}</strong> · ${escapeHtml(sec)} <span style="color:var(--c-text-faint);margin-left:6px">${escapeHtml(summary)}</span>`;
  } else {
    sbRow.textContent = 'No row selected';
  }
  // Phase A6: highlight the active verdict pill
  const status = (r && DATA?.status?.[SELECTED_ROW]?.status) || 'unverified';
  document.querySelectorAll('#sb-verdict .sb-vbtn[data-verdict]').forEach(btn => {
    const isActive = btn.dataset.verdict === status;
    btn.setAttribute('aria-checked', isActive ? 'true' : 'false');
  });
  // Progress
  if (DATA?.stats) {
    const status = DATA.status || {};
    const done = Object.values(status).filter(s => s.status && s.status !== 'unverified').length;
    const total = DATA.stats.total || 0;
    const pct = total ? Math.round(done/total * 100) : 0;
    document.getElementById('sb-done').textContent = done;
    document.getElementById('sb-total').textContent = total;
    document.getElementById('sb-bar').style.width = pct + '%';
  }
}

// Wire AI pane + status bar update into selectRow
const _origSelectRowForLayout = selectRow;
selectRow = async function(rowNum, scroll) {
  await _origSelectRowForLayout(rowNum, scroll);
  statusBarUpdate();
  if (_AI_PANE_OPEN) aiPaneRefresh();
};
const _origSetStatusForStatusBar = setStatus;
setStatus = async function(status) {
  await _origSetStatusForStatusBar(status);
  statusBarUpdate();
};
const _origRenderStatsForStatusBar = renderStats;
renderStats = function() {
  _origRenderStatsForStatusBar();
  statusBarUpdate();
};

// Status bar: Claude status
async function statusBarClaude() {
  try {
    const r = await fetch('/api/learn/llm_status');
    const j = await r.json();
    const sbC = document.getElementById('sb-claude');
    const txt = document.getElementById('sb-claude-text');
    if (!sbC || !txt) return;
    if (j.available) {
      sbC.classList.add('online'); sbC.classList.remove('offline');
      // Phase 1: Claude Code via Max OAuth → no $/day, just label
      if (j.provider_kind === 'claude_code') {
        const auth = j.auth_mode === 'claude_max' ? 'Max' : (j.auth_mode || '');
        txt.textContent = `${j.model || 'claude-code'} · ${auth} · ${j.calls_today || 0} calls`;
      } else {
        txt.textContent = `${j.model || j.name} · $${(j.spent_today_usd || 0).toFixed(2)}/$${(j.budget_usd_per_day || 5).toFixed(0)}`;
      }
      const h = document.getElementById('ai-pane-head');
      const sub = document.getElementById('ai-pane-sub');
      if (h) h.setAttribute('data-status', 'online');
      if (sub && SELECTED_ROW) sub.textContent = `R${SELECTED_ROW} · ${j.model || ''}`;
    } else {
      sbC.classList.remove('online'); sbC.classList.add('offline');
      txt.textContent = 'Claude offline';
      const h = document.getElementById('ai-pane-head');
      if (h) h.setAttribute('data-status', 'offline');
    }
  } catch (e) {}
}
statusBarClaude();
setInterval(statusBarClaude, 15000);   // refresh every 15s

// Sync ribbon controls to existing state
function _syncRibbonState() {
  // Catalog page indicator
  const pdfPg = document.getElementById('pdf-page-info');
  const ribPg = document.getElementById('ribbon-pdf-page');
  if (pdfPg && ribPg) ribPg.textContent = pdfPg.textContent;
  // Context radius
  const ctx = document.getElementById('ctx-radius');
  const ribCtx = document.getElementById('ribbon-ctx-radius');
  if (ctx && ribCtx) ribCtx.textContent = ctx.textContent;
  // Sync hl-toggle ↔ ribbon-hl-toggle
  const hl = document.getElementById('hl-toggle');
  const rhl = document.getElementById('ribbon-hl-toggle');
  if (hl && rhl) rhl.checked = hl.checked;
  // Edit toolbar dirty + save
  const dirty = document.getElementById('dirty-ind');
  const ribDirty = document.getElementById('ribbon-dirty-ind');
  if (dirty && ribDirty) ribDirty.textContent = dirty.textContent;
  const save = document.getElementById('save-btn');
  const ribSave = document.getElementById('ribbon-save-btn');
  if (save && ribSave) ribSave.disabled = save.disabled;
  const ub = document.getElementById('undo-btn');
  const rub = document.getElementById('ribbon-undo-btn');
  if (ub && rub) rub.disabled = ub.disabled;
  const rb = document.getElementById('redo-btn');
  const rrb = document.getElementById('ribbon-redo-btn');
  if (rb && rrb) rrb.disabled = rb.disabled;
  // Tools active state
  const activeTool = document.querySelector('.edit-toolbar button.tool.active');
  const tname = activeTool ? activeTool.dataset.tool : 'select';
  document.querySelectorAll('.ribbon-mode-bar.mode-edit button[data-tool]').forEach(b => {
    b.classList.toggle('active', b.dataset.tool === tname);
  });
  // Apply mode LLM summary
  const llmSum = document.getElementById('ribbon-llm-summary');
  if (llmSum) {
    fetch('/api/learn/llm_status').then(r=>r.json()).then(j => {
      if (j.available) llmSum.textContent = `Claude · ${j.model || j.name} · $${(j.spent_today_usd || 0).toFixed(2)}/day`;
      else llmSum.textContent = `Claude OFF — paste API key in Settings`;
    }).catch(() => {});
  }
}
setInterval(_syncRibbonState, 800);

// Initial state
setTimeout(() => { statusBarUpdate(); _syncRibbonState(); }, 300);

init();
</script>
</body>
</html>
"""



# ---------------------------------------------------------------------------
# Boot
# ---------------------------------------------------------------------------

def boot() -> None:
    print(f"[boot] root={ROOT}")

    # Bring up the SQLite layer first so subsequent index/sync calls have
    # somewhere to mirror to.
    print(f"[boot] opening database at {DB_PATH.relative_to(ROOT)}…")
    db.init_db(DB_PATH)
    # One-time migration: import legacy verification_status.json if the DB
    # table is empty
    try:
        existing = db.get_all_status()
    except Exception:
        existing = {}
    if not existing and STATUS_PATH.exists():
        n = db.import_status_from_json(STATUS_PATH)
        if n:
            print(f"[boot] migrated {n} verification_status entries → DB")

    print(f"[boot] indexing PDFs in {OUTPUT}…")
    build_pdf_index()
    n_indexed = len({p for v in PDF_INDEX.values() for p in v} |
                    {p for v in SECTION_INDEX.values() for p in v})
    print(f"[boot] {len(PDF_INDEX)} ref-keys + {len(SECTION_INDEX)} section-keys → "
          f"{n_indexed} unique PDFs indexed")
    print(f"[boot] loading {XLSX_PATH.name}…")
    load_rows()
    print(f"[boot] {len(ROWS)} rows")
    collect_extra_refs()
    print("[boot] indexing TOR sections + page text…")
    build_tor_section_index()
    index_tor_text()
    ch5 = sum(1 for s in TOR_SECTION_INDEX if s.startswith("5"))
    print(f"[boot] {len(TOR_SECTION_INDEX)} TOR sections ({ch5} in chapter 5), "
          f"{len(TOR_PAGE_TEXTS)} pages indexed for text-search")

    # Mirror everything into the DB so queries can hit it instead of the
    # in-memory caches.
    sync_db_from_memory()

    # Phase 2: catalog library bootstrap (idempotent)
    # Ensure default company + active project exist so row_catalog_links
    # has a project_id to point at, then ingest output/ PDFs as catalogs.
    try:
        cm_id = catalog.upsert_company(name="Smart Solution", code="SMART")
        proj = catalog.get_active_project()
        if not proj:
            pid = catalog.upsert_project(
                company_id=cm_id, name="Smart Plant 1", code="SP1",
                xlsx_rel=str(XLSX_PATH.relative_to(ROOT)),
                output_rel=str(OUTPUT.relative_to(ROOT)),
            )
            catalog.set_active_project(pid)
            proj = catalog.get_active_project()
        result = catalog.ingest_output_dir(OUTPUT)
        if result.get("ok"):
            print(f"[boot] catalog library: {result['scanned']} PDFs scanned "
                  f"({result['inserted']} new, {result['updated']} updated, "
                  f"{result['skipped']} unchanged) · "
                  f"active project={proj.get('name') if proj else '?'}")
    except Exception as e:
        sys.stderr.write(f"[boot] catalog ingest failed: {e}\n")

    # Load .env file (if present) so ANTHROPIC_API_KEY etc. are available
    _load_dotenv(ROOT / ".env")

    # Install Claude provider — Claude Code (Agent SDK) takes priority over
    # API direct, since Phase 1 uses Claude Max OAuth (no metered API costs).
    # Set COMPLY_LLM=claude_code (default) or =anthropic for the legacy path.
    _llm_mode = (os.environ.get("COMPLY_LLM") or "claude_code").lower()
    if _llm_mode in ("claude_code", "claude-code"):
        try:
            from app import claude_code_provider as cp
            if cp.install_into_learning():
                p = cp.get_provider()
                bs = p.budget_status() if p else {}
                print(f"[boot] Claude Code provider installed: model={bs.get('model')} "
                      f"auth={bs.get('auth_mode')} (Max = unlimited)")
            else:
                if not cp.SDK_AVAILABLE:
                    print("[boot] Claude Code provider OFF (claude-agent-sdk not installed)")
                else:
                    print("[boot] Claude Code provider OFF (set COMPLY_LLM=claude_code to enable)")
        except Exception as e:
            sys.stderr.write(f"[boot] Claude Code install failed: {e}\n")
    elif _llm_mode == "anthropic":
        try:
            from app import anthropic_provider as ap
            if ap.install_into_learning():
                p = ap.get_provider()
                bs = p.budget_status() if p else {}
                print(f"[boot] Anthropic API provider installed: model={bs.get('model')} "
                      f"budget=${bs.get('budget_usd_per_day'):.2f}/day "
                      f"spent_today=${bs.get('spent_today_usd', 0):.2f}")
            else:
                print("[boot] Anthropic API provider OFF (set ANTHROPIC_API_KEY to enable)")
        except Exception as e:
            sys.stderr.write(f"[boot] Anthropic install failed: {e}\n")

    # Always-load-latest invariant: ensure the latest snapshot reflects the
    # current working state, OR surface a divergence the user must resolve.
    if _version_script_available():
        sync = boot_sync_check()
        latest = get_version_sync_status().get("latest")
        latest_id = latest["id"] if latest else "(none)"
        action = sync.get("performed") or "—"
        state = sync.get("status") or "?"
        print(f"[boot] version sync: state={state}, latest={latest_id}, action={action}")
        if state in ("working_behind", "divergent", "incomplete_local"):
            print("[boot] ⚠ working dir differs from latest snapshot — review in 📚 Versions")


def main() -> None:
    boot()
    host = "127.0.0.1"
    port = 5173
    url = f"http://{host}:{port}"
    print("\n  ✦ Comply Verify GUI ✦")
    print(f"  → open {url} in browser\n")

    def _open():
        time.sleep(0.6)
        try:
            webbrowser.open(url)
        except Exception:
            pass
    threading.Thread(target=_open, daemon=True).start()

    app.run(host=host, port=port, debug=False, threaded=True)


if __name__ == "__main__":
    main()
