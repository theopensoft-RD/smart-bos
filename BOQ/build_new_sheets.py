"""
สร้างชีท TRIO ใหม่ / SR ใหม่ / Solution ใหม่ ที่กระจายราคาเป็น 40/30/30
จาก ปร.4 (Master = 83,055,976.68 บาท)

วิธีกระจาย: แต่ละ item ของ ปร.4 จะถูกแบ่งจำนวน (qty) เป็น 40/30/30 ตามอัตราส่วน
โดยใช้ Largest Remainder Method (Hamilton method) เพื่อให้เป็นจำนวนเต็ม

ตัวอย่าง:
- qty 12 → TRIO 5, SR 4, SOL 3 (raw 4.8, 3.6, 3.6 → floor 4,3,3 + remainder 2)
- qty 85 → TRIO 34, SR 26, SOL 25 (raw 34, 25.5, 25.5 → floor 34,25,25 + rem 1)
- qty 1  → TRIO 1, SR 0, SOL 0 (only one vendor gets it; TRIO has highest frac)
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from copy import copy

SRC = '5.ปริมาณงาน Smart Plant P1  23-04-2569 - Copy.xlsx'

RATIOS = (0.40, 0.30, 0.30)  # TRIO, SR, SOL

def split_qty_balanced(n, unit_value, running, targets):
    """Split n into 3 ints summing to n, weighted by RATIOS but assigning the
    remainder unit(s) to whichever vendor is currently most behind their value target.
    Mutates `running` in place to add the assigned values."""
    if n is None or n <= 0:
        return (0, 0, 0)
    raw = [n * r for r in RATIOS]
    flo = [int(x) for x in raw]
    rem = int(round(n - sum(flo)))
    for _ in range(rem):
        # Score = current_progress / target. Lowest = most behind.
        scores = [
            (running[i] + flo[i] * unit_value) / targets[i] if targets[i] > 0 else float('inf')
            for i in range(3)
        ]
        winner = min(range(3), key=lambda i: (scores[i], i))
        flo[winner] += 1
    for i in range(3):
        running[i] += flo[i] * unit_value
    return tuple(flo)

# ALLOC will be filled dynamically in main()
ALLOC = {}

# Section grouping for headers/sub-totals (header_row, items_rows, label)
SECTIONS = [
    ('1', 'ระบบบริหารจัดการน้ำเสียอัจฉริยะ', None),
    ('1.1', 'ระบบควบคุมการบำบัดน้ำเสียอัจฉริยะเพื่อรองรับเครื่องจักร 85 เครื่อง', [6,7,8,9,10,11,12]),
    ('1.2', 'ตู้ควบคุมเครื่องจักรสำหรับควบคุมเครื่องจักรประเภทต่างๆภายในโรงบำบัดน้ำเสีย เมืองพัทยา', [14,15]),
    ('1.3', 'ตู้ควบคุมเครื่องจักรสำหรับควบคุมเครื่องสูบน้ำเสียต่างๆที่อยู่ภายนอกโรงบำบัดน้ำเสีย เมืองพัทยา', [18,19,20]),
    ('1.4', 'ตู้ประมวลผลความต้องการออกซิเจนทางชีวเคมี', [23,24,25,26]),
    ('1.5', 'ตู้ประมวลผลค่าออกซิเจนละลายน้ำ', [29,30,31]),
    ('1.6', 'ตู้ควบคุมสถานีแสดงผลการบำบัดน้ำเสียอัจฉริยะ', [34,35,36,37]),
    ('1.7', 'ค่าพัฒนาระบบบริหารจัดการน้ำเสียอัจฉริยะ', [40,41]),
    ('1.8', 'งานเดินสายเชื่อมสัญญาณจากเครื่องจักรเข้าตู้ควบคุมการแสดงผล', [46,47,48]),
    ('2', 'งานระบบเครือข่ายคอมพิวเตอร์', None),
    ('2.1', 'งานระบบเครือข่ายคอมพิวเตอร์ อาคารสำนักงาน', list(range(52,66))),
    ('2.2', 'งานระบบเครือข่ายคอมพิวเตอร์ บ้านพักข้าราชการ', list(range(68,80)) + [81,82]),
    ('2.3', 'งานระบบเครือข่ายคอมพิวเตอร์ อาคารบำบัดน้ำเสีย', list(range(86,98)) + [99,100]),
    ('2.4', 'งานระบบเครือข่ายคอมพิวเตอร์ อาคารซ่อมบำรุง', list(range(104,116)) + [117,118]),
    ('2.5', 'งานระบบเครือข่ายคอมพิวเตอร์ บ่อสัมผัสคลอรีน', list(range(122,130)) + [131,132]),
    ('2.6', 'งานระบบเครือข่ายคอมพิวเตอร์ ป้อม รปภ.และภายนอกอาคาร', list(range(136,144)) + [145,146]),
]

def eval_simple(v, wb=None, sheet_name=None, _depth=0):
    """Resolve numeric cell value. Handles literal arithmetic and cell references.
    Examples: 12, '=370100*100/107', '=C18', '=BackUp_งานขุด!C4'.
    """
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s.startswith('='):
        try:
            return float(s)
        except Exception:
            return 0
    if _depth > 5 or wb is None:
        return 0
    expr = s[1:].replace(' ', '')
    # Resolve cell refs (with or without sheet prefix)
    import re
    def _ref_repl(m):
        sn = m.group(1)
        col = m.group(2)
        row = m.group(3)
        target_sheet = sn if sn else sheet_name
        if not target_sheet or target_sheet not in wb.sheetnames:
            return '0'
        cell = wb[target_sheet][f"{col}{row}"]
        val = eval_simple(cell.value, wb, target_sheet, _depth+1)
        return repr(val)
    # Pattern: optional 'SheetName!' then column letters and row digits
    expr = re.sub(r"(?:'?([^'!]+)'?!)?([A-Z]+)(\d+)", _ref_repl, expr)
    allowed = set('0123456789+-*/.()')
    if all(ch in allowed for ch in expr):
        try:
            return eval(expr, {'__builtins__': {}}, {})
        except Exception:
            return 0
    return 0

import re
def resolve_name(name_value, src_qty):
    """Resolve dynamic name formulas like ="text" & C63 & "text" using src_qty."""
    if not isinstance(name_value, str) or not name_value.startswith('='):
        return name_value
    # Strip leading '=' and replace C<row> refs with src_qty
    expr = name_value[1:]
    expr = re.sub(r'C\d+', str(src_qty), expr)
    # Eval simple string concat: "..." & 48 & "..."
    parts = []
    for tok in re.split(r'\s*&\s*', expr):
        tok = tok.strip()
        if tok.startswith('"') and tok.endswith('"'):
            parts.append(tok[1:-1])
        else:
            parts.append(tok)
    return ''.join(parts)

def main():
    # Load with formulas preserved
    wb = load_workbook(SRC)
    src_ws = wb['ปร.4']
    src = {}
    # First pass: read all source rows
    all_rows = sorted(set(r for sec in SECTIONS if sec[2] for r in sec[2]))
    for row_num in all_rows:
        r = src_ws[row_num]
        qty_resolved = eval_simple(r[2].value, wb, 'ปร.4')
        qty_int = int(round(qty_resolved)) if isinstance(qty_resolved, (int, float)) else 0
        # Resolve B column too (handle cross-sheet refs like =BackUp_งานขุด!B4)
        name_val = r[1].value
        if isinstance(name_val, str) and name_val.startswith('=') and 'BackUp' in name_val:
            import re
            m = re.match(r"=([^!]+)!([A-Z]+\d+)", name_val)
            if m:
                ref_sheet, ref_cell = m.group(1).strip("'"), m.group(2)
                if ref_sheet in wb.sheetnames:
                    name_val = wb[ref_sheet][ref_cell].value
        src[row_num] = {
            'no':       r[0].value,
            'name':     resolve_name(name_val, qty_int),
            'qty':      qty_resolved,
            'unit':     r[3].value,
            'eq_price': eval_simple(r[4].value, wb, 'ปร.4'),
            'lab_price':eval_simple(r[6].value, wb, 'ปร.4'),
            'src_ref':  r[9].value,
            'qty_int':  qty_int,
        }

    # Second pass: compute deficit-aware allocations.
    # Process items in DESCENDING order of total value so big items lock in first
    # and small items fill the gaps to keep totals near 40/30/30.
    grand_total = sum(s['qty_int'] * (s['eq_price'] + s['lab_price']) for s in src.values())
    targets = [grand_total * r for r in RATIOS]
    running = [0.0, 0.0, 0.0]
    items_by_value = sorted(
        all_rows,
        key=lambda rn: -(src[rn]['qty_int'] * (src[rn]['eq_price'] + src[rn]['lab_price']))
    )
    for row_num in items_by_value:
        s = src[row_num]
        upv = s['eq_price'] + s['lab_price']
        ALLOC[row_num] = split_qty_balanced(s['qty_int'], upv, running, targets)

    # Remove existing new sheets if rerun
    for sn in ['TRIO ใหม่', 'SR ใหม่', 'Solution ใหม่']:
        if sn in wb.sheetnames:
            del wb[sn]

    # Styles
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(name='TH SarabunPSK', size=14, bold=True)
    normal = Font(name='TH SarabunPSK', size=14)
    header_fill = PatternFill('solid', start_color='D9E1F2')
    section_fill = PatternFill('solid', start_color='FFF2CC')
    subtotal_fill = PatternFill('solid', start_color='E2EFDA')
    target_fill = PatternFill('solid', start_color='FCE4D6')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')

    def build_sheet(sheet_name, sheet_idx, target_pct, label):
        ws = wb.create_sheet(sheet_name)
        # Column widths
        widths = {'A':8, 'B':45, 'C':10, 'D':10, 'E':14, 'F':16, 'G':12, 'H':14, 'I':18, 'J':30}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Title row
        ws['A1'] = f'การกระจายราคา BOQ ส่วนของ {label} (เป้าหมาย {target_pct*100:.0f}%)'
        ws['A1'].font = Font(name='TH SarabunPSK', size=18, bold=True)
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = center

        # Column headers (rows 2-3)
        ws.cell(row=2, column=1, value='ลำดับ')
        ws.cell(row=2, column=2, value='รายการ')
        ws.cell(row=2, column=3, value='ปริมาณ')
        ws.merge_cells('C2:D2')
        ws.cell(row=2, column=5, value='ค่าอุปกรณ์ต่อหน่วย')
        ws.merge_cells('E2:F2')
        ws.cell(row=2, column=7, value='ค่าแรงงานต่อหน่วย')
        ws.merge_cells('G2:H2')
        ws.cell(row=2, column=9, value='รวมราคาค่าอุปกรณ์และค่าแรงงาน')
        ws.cell(row=2, column=10, value='ราคากลาง / แหล่งที่มา')
        ws.cell(row=3, column=3, value='จำนวน')
        ws.cell(row=3, column=4, value='หน่วย')
        ws.cell(row=3, column=5, value='ราคา (บาท)')
        ws.cell(row=3, column=6, value='จำนวนเงิน')
        ws.cell(row=3, column=7, value='ราคา (บาท)')
        ws.cell(row=3, column=8, value='จำนวนเงิน')
        for r in (2,3):
            for col in range(1,11):
                cell = ws.cell(row=r, column=col)
                cell.font = bold
                cell.alignment = center
                cell.fill = header_fill
                cell.border = border

        out_row = 4
        section_subtotals = []  # list of (label, formula_range, total_row_address)
        for sec_no, sec_name, sec_rows in SECTIONS:
            if sec_rows is None:
                # Top-level header (1, 2, etc.)
                ws.cell(row=out_row, column=1, value=sec_no)
                ws.cell(row=out_row, column=2, value=sec_name)
                for col in range(1,11):
                    cell = ws.cell(row=out_row, column=col)
                    cell.font = bold
                    cell.fill = section_fill
                    cell.border = border
                    cell.alignment = left if col == 2 else center
                out_row += 1
                continue

            # Filter items where this sheet has qty > 0
            sec_items = []
            for sr in sec_rows:
                qty = ALLOC.get(sr, (0,0,0))[sheet_idx]
                if qty and qty > 0:
                    sec_items.append((sr, qty))
            if not sec_items:
                continue  # Skip section if nothing for this sheet

            # Skip writing a separate section header when the first leaf already
            # carries the same section number (e.g., section 1.2 has leaf "1.2")
            first_no = src[sec_items[0][0]]['no']
            if str(first_no) != str(sec_no):
                ws.cell(row=out_row, column=1, value=sec_no)
                ws.cell(row=out_row, column=2, value=sec_name)
                for col in range(1,11):
                    cell = ws.cell(row=out_row, column=col)
                    cell.font = bold
                    cell.fill = section_fill
                    cell.border = border
                    cell.alignment = left if col == 2 else center
                out_row += 1
            first_item_row = out_row

            for src_row, qty in sec_items:
                s = src[src_row]
                ws.cell(row=out_row, column=1, value=s['no'])
                ws.cell(row=out_row, column=2, value=s['name'])
                ws.cell(row=out_row, column=3, value=qty)
                ws.cell(row=out_row, column=4, value=s['unit'])
                ws.cell(row=out_row, column=5, value=s['eq_price'] or 0)
                ws.cell(row=out_row, column=6, value=f"=C{out_row}*E{out_row}")
                ws.cell(row=out_row, column=7, value=s['lab_price'] or 0)
                ws.cell(row=out_row, column=8, value=f"=C{out_row}*G{out_row}")
                ws.cell(row=out_row, column=9, value=f"=F{out_row}+H{out_row}")
                ws.cell(row=out_row, column=10, value=s['src_ref'])
                # Format
                for col in range(1,11):
                    cell = ws.cell(row=out_row, column=col)
                    cell.font = normal
                    cell.border = border
                    if col in (1,3,4):
                        cell.alignment = center
                    elif col == 2 or col == 10:
                        cell.alignment = left
                    else:
                        cell.alignment = right
                        cell.number_format = '#,##0.00'
                out_row += 1

            # Section sub-total row
            ws.cell(row=out_row, column=2, value=f'สรุปราคา{sec_name}')
            ws.cell(row=out_row, column=9, value=f"=SUM(I{first_item_row}:I{out_row-1})")
            for col in range(1,11):
                cell = ws.cell(row=out_row, column=col)
                cell.font = bold
                cell.fill = subtotal_fill
                cell.border = border
                if col == 2:
                    cell.alignment = left
                elif col == 9:
                    cell.alignment = right
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = center
            section_subtotals.append(out_row)
            out_row += 1

        # Grand total
        out_row += 1
        ws.cell(row=out_row, column=2, value='รวมราคาทั้งสิ้น')
        sum_formula = '=' + '+'.join(f'I{r}' for r in section_subtotals)
        ws.cell(row=out_row, column=9, value=sum_formula)
        for col in range(1,11):
            cell = ws.cell(row=out_row, column=col)
            cell.font = Font(name='TH SarabunPSK', size=16, bold=True)
            cell.fill = target_fill
            cell.border = border
            if col == 2:
                cell.alignment = left
            elif col == 9:
                cell.alignment = right
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = center
        grand_total_row = out_row

        # Target/variance section
        out_row += 2
        ws.cell(row=out_row, column=2, value='ราคารวม ปร.4 (Master)')
        ws.cell(row=out_row, column=9, value="=ปร.4!I149")
        ws.cell(row=out_row, column=9).number_format = '#,##0.00'
        out_row += 1
        ws.cell(row=out_row, column=2, value=f'เป้าหมาย {label} ({target_pct*100:.0f}%)')
        ws.cell(row=out_row, column=9, value=f"=I{out_row-1}*{target_pct}")
        ws.cell(row=out_row, column=9).number_format = '#,##0.00'
        out_row += 1
        ws.cell(row=out_row, column=2, value='เปอร์เซ็นต์จริง')
        ws.cell(row=out_row, column=9, value=f"=I{grand_total_row}/I{out_row-2}")
        ws.cell(row=out_row, column=9).number_format = '0.00%'
        out_row += 1
        ws.cell(row=out_row, column=2, value='ส่วนต่างจากเป้าหมาย')
        ws.cell(row=out_row, column=9, value=f"=I{grand_total_row}-I{out_row-2}")
        ws.cell(row=out_row, column=9).number_format = '#,##0.00;[Red](#,##0.00)'
        # Style summary block
        for r in range(grand_total_row+2, out_row+1):
            for col in (2, 9):
                cell = ws.cell(row=r, column=col)
                cell.font = bold
                cell.border = border
                cell.alignment = right if col == 9 else left
                cell.fill = target_fill

        # Freeze header
        ws.freeze_panes = 'A4'

    build_sheet('TRIO ใหม่', 0, 0.40, 'TRIO')
    build_sheet('SR ใหม่', 1, 0.30, 'SR')
    build_sheet('Solution ใหม่', 2, 0.30, 'Solution')

    out = '5.ปริมาณงาน Smart Plant P1  23-04-2569 - Copy.xlsx'
    wb.save(out)
    print(f"Saved: {out}")

if __name__ == '__main__':
    main()
