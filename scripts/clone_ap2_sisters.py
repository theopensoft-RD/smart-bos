#!/usr/bin/env python3
"""
Clone Access Point Type 2 (EAP660 HD) reference PDF to 3 sister sections
and update xlsx for all 4 sister rows.
"""

import os
import shutil

import fitz
import openpyxl

fitz.TOOLS.mupdf_display_errors(False)

BASE = "/Users/ppxndpxdd/Library/CloudStorage/GoogleDrive-smartsolution.in.th@gmail.com/My Drive/Smart Solution/Project/Pattaya Project/Smart Plant 1/co-work/claude-code"

REF_PDF = os.path.join(
    BASE,
    "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์",
    "5.2.1. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารสำนักงาน",
    "5.2.1.-12",
    "5.2.1.12. อุปกรณ์กระจายสัญญาณไร้สาย (Access Point) แบบที่ 2 TP-Link EAP660 HD.pdf"
)

SOURCE_CATALOG = os.path.join(
    BASE,
    "catalog/left/อุปกรณ์กระจายสัญญาณไร้สาย (Access Point) แบบที่ 2 -t_EAP660 HD.pdf"
)

XLSX_PATH = os.path.join(BASE, "output/Comply spec Smart Plant 1.xlsx")

ITEM_NAME = "อุปกรณ์กระจายสัญญาณไร้สาย (Access Point) แบบที่ 2 TP-Link EAP660 HD"
REF_SECTION = "5.2.1.12"

SISTERS = [
    {
        "section": "5.2.2.7",
        "parent_dir": "5.2.2. งานระบบเครือข่ายคอมพิวเตอร์ ภายในบ้านพักข้าราชการ",
        "folder": "5.2.2.-7",
    },
    {
        "section": "5.2.3.7",
        "parent_dir": "5.2.3. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารบำบัดน้ำเสีย",
        "folder": "5.2.3.-7",
    },
    {
        "section": "5.2.4.7",
        "parent_dir": "5.2.4. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารซ่อมบำรุง",
        "folder": "5.2.4.-7",
    },
]


def clone_pdf(sister):
    section = sister["section"]
    out_dir = os.path.join(
        BASE,
        "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์",
        sister["parent_dir"],
        sister["folder"],
    )
    os.makedirs(out_dir, exist_ok=True)

    out_filename = f"{section}. {ITEM_NAME}.pdf"
    out_path = os.path.join(out_dir, out_filename)

    # Open source catalog
    doc = fitz.open(SOURCE_CATALOG)
    page_count = doc.page_count

    # Open reference to read all annotations
    ref_doc = fitz.open(REF_PDF)

    for pno in range(page_count):
        page = doc[pno]
        ref_page = ref_doc[pno] if pno < ref_doc.page_count else None

        if ref_page is None:
            continue

        # Copy all annotations from reference, substituting section numbers
        for a in ref_page.annots():
            atype = a.type[0]
            content = a.info.get("content", "")
            rect = a.rect
            colors = a.colors
            border = a.border

            # Substitute section number in content
            new_content = content.replace(REF_SECTION, section)

            if atype == 2:  # FreeText
                new_annot = page.add_freetext_annot(
                    rect,
                    new_content,
                    fontsize=14,
                    fontname="helv",
                    text_color=(1, 0, 0),
                    fill_color=None,
                    align=1,  # center
                    border_color=None,
                )
                new_annot.set_border(border)
                new_annot.update()

            elif atype == 4:  # Square/Rect
                new_annot = page.add_rect_annot(rect)
                stroke = colors.get("stroke", [])
                if stroke:
                    new_annot.set_colors(stroke=stroke)
                new_annot.set_border(border)
                new_annot.info["content"] = new_content
                new_annot.update()

    ref_doc.close()

    # Save with O_TRUNC workaround
    tmp_path = out_path + ".tmp"
    doc.save(
        tmp_path,
        garbage=4,
        clean=True,
        deflate=True,
    )
    doc.close()

    # Move tmp to final
    flags = os.O_WRONLY | os.O_CREAT | os.O_TRUNC
    fd = os.open(out_path, flags)
    os.close(fd)
    shutil.move(tmp_path, out_path)

    print(f"  Created: {out_path}")
    return out_path


def update_xlsx():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Trio_SR_Solution"]

    # All 4 sections: ref section 5.2.1.12 at R373, sisters at R432, R495, R558
    sections_rows = [
        ("5.2.1.12", 373),
        ("5.2.2.7", 432),
        ("5.2.3.7", 495),
        ("5.2.4.7", 558),
    ]

    for section, parent_row in sections_rows:
        cat = f"{section} {ITEM_NAME}"

        # Col D values for each row (parent + 8 ข้อย่อย)
        col_d_values = [
            "ยี่ห้อ TP-Link รุ่น EAP660 HD",                                        # parent
            f"เทียบเท่าข้อกำหนด เอกสาร {cat} หน้า 9 ข้อ {section} ข้อย่อย 1.",  # ข้อย่อย 1
            f"เทียบเท่าข้อกำหนด เอกสาร {cat} หน้า 9 ข้อ {section} ข้อย่อย 2.",  # ข้อย่อย 2
            f"เทียบเท่าข้อกำหนด เอกสาร {cat} หน้า 9 ข้อ {section} ข้อย่อย 3.",  # ข้อย่อย 3
            "ยินดีปฏิบัติตามข้อกำหนด",   # ข้อย่อย 4
            "ยินดีปฏิบัติตามข้อกำหนด",   # ข้อย่อย 5
            "ยินดีปฏิบัติตามข้อกำหนด",   # ข้อย่อย 6
            "ยินดีปฏิบัติตามข้อกำหนด",   # ข้อย่อย 7
            "ยินดีปฏิบัติตามข้อกำหนด",   # ข้อย่อย 8
        ]

        for i, d_val in enumerate(col_d_values):
            row = parent_row + i
            ws.cell(row=row, column=4).value = d_val   # Col D
            ws.cell(row=row, column=5).value = "TRIO"  # Col E
            ws.cell(row=row, column=6).value = "รอ user ตรวจสอบ"  # Col F

        print(f"  Updated xlsx rows R{parent_row}-R{parent_row+8} for {section}")

    wb.save(XLSX_PATH)
    print(f"  Saved: {XLSX_PATH}")


if __name__ == "__main__":
    print("=== Cloning AP Type 2 PDFs ===")
    for sister in SISTERS:
        print(f"\nProcessing {sister['section']}...")
        clone_pdf(sister)

    print("\n=== Updating xlsx ===")
    update_xlsx()

    print("\nDone.")
