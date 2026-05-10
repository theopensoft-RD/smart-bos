#!/usr/bin/env python3
"""Clone L3 Switch annotated PDF to 3 sister sections and update xlsx."""

import os
import sys
import tempfile

import fitz  # PyMuPDF
import openpyxl

fitz.TOOLS.mupdf_display_errors(False)

BASE = "/Users/ppxndpxdd/Library/CloudStorage/GoogleDrive-smartsolution.in.th@gmail.com/My Drive/Smart Solution/Project/Pattaya Project/Smart Plant 1/co-work/claude-code"

REF_PDF = os.path.join(
    BASE,
    "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์",
    "5.2.1. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารสำนักงาน",
    "5.2.1.-6",
    "5.2.1.6. อุปกรณ์กระจายสัญญาณ (L3 Switch) ขนาด 24 ช่อง Ruijie RG-CS85-24GT8XS-D.pdf"
)

SOURCE_CATALOG = os.path.join(
    BASE,
    "catalog/left",
    "อุปกรณ์กระจายสัญญาณ (L3 Switch) ขนาด 24 ช่อง - RG-CS85-24GT8XS-D รองรับ 24-Port 1GE RJ45 Layer 3.pdf"
)

XLSX_PATH = os.path.join(BASE, "output/Comply spec Smart Plant 1.xlsx")

PRODUCT_NAME = "อุปกรณ์กระจายสัญญาณ (L3 Switch) ขนาด 24 ช่อง Ruijie RG-CS85-24GT8XS-D"

SISTERS = [
    {
        "section": "5.2.2.2",
        "section_dot": "5.2.2.2.",
        "dir": os.path.join(
            BASE,
            "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์",
            "5.2.2. งานระบบเครือข่ายคอมพิวเตอร์ ภายในบ้านพักข้าราชการ",
            "5.2.2.-2"
        ),
        "parent_row": 390,
        "sub_rows": list(range(391, 400)),  # 391-399
    },
    {
        "section": "5.2.3.2",
        "section_dot": "5.2.3.2.",
        "dir": os.path.join(
            BASE,
            "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์",
            "5.2.3. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารบำบัดน้ำเสีย",
            "5.2.3.-2"
        ),
        "parent_row": 453,
        "sub_rows": list(range(454, 463)),  # 454-462
    },
    {
        "section": "5.2.4.2",
        "section_dot": "5.2.4.2.",
        "dir": os.path.join(
            BASE,
            "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์",
            "5.2.4. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารซ่อมบำรุง",
            "5.2.4.-2"
        ),
        "parent_row": 516,
        "sub_rows": list(range(517, 526)),  # 517-525
    },
]

# Also include 5.2.1.6 for xlsx updates
SECTION_516 = {
    "section": "5.2.1.6",
    "section_dot": "5.2.1.6.",
    "parent_row": 318,
    "sub_rows": list(range(319, 328)),  # 319-327
}


def save_pdf_gdrive(doc, path):
    """Save PDF using O_TRUNC workaround for Google Drive."""
    tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    tmp.close()
    doc.save(tmp.name, garbage=4, clean=True, deflate=True)
    with open(tmp.name, 'rb') as src:
        data = src.read()
    if os.path.exists(path):
        fd = os.open(path, os.O_WRONLY | os.O_TRUNC)
        os.write(fd, data)
        os.close(fd)
    else:
        with open(path, 'wb') as f:
            f.write(data)
    os.unlink(tmp.name)


def collect_annotations(ref_path):
    """Collect all annotations from reference PDF, organized by page index."""
    doc = fitz.open(ref_path)
    pages_annots = []
    for _pi, page in enumerate(doc):
        page_data = []
        for a in page.annots():
            atype = a.type[0]
            content = a.info.get('content', '')
            rect = tuple(a.rect)
            colors = a.colors
            border = a.border
            # Get DA for FreeText
            da = None
            if atype == 2:
                da_raw = doc.xref_get_key(a.xref, 'DA')
                if da_raw and da_raw[0] != 'null':
                    da = da_raw[1]
            page_data.append({
                'type': atype,
                'content': content,
                'rect': rect,
                'colors': colors,
                'border': border,
                'da': da,
            })
        pages_annots.append(page_data)
    doc.close()
    return pages_annots


def clone_pdf(sister, ref_annots):
    """Clone source catalog PDF with adapted annotations for the sister section."""
    section = sister['section']
    section_dot = sister['section_dot']
    out_dir = sister['dir']
    os.makedirs(out_dir, exist_ok=True)

    filename = f"{section_dot} {PRODUCT_NAME}.pdf"
    out_path = os.path.join(out_dir, filename)
    filename_no_pdf = filename.replace('.pdf', '')

    doc = fitz.open(SOURCE_CATALOG)
    num_pages = len(doc)

    for pi, page in enumerate(doc):
        page_annots = ref_annots[pi] if pi < len(ref_annots) else []

        for ann in page_annots:
            atype = ann['type']
            content = ann['content']
            rect = fitz.Rect(ann['rect'])
            colors = ann['colors']
            border = ann['border']
            da = ann['da']

            if atype == 2:  # FreeText
                # Adapt content
                if 'หน้า' in content and section_dot.rstrip('.') not in content:
                    # This is a header from the reference - rebuild it
                    page_num = pi + 1
                    new_content = f"{filename_no_pdf} หน้า {page_num}"
                elif content == '5.2.1.6':
                    new_content = section
                elif content.startswith('5.2.1.6'):
                    new_content = content.replace('5.2.1.6', section)
                else:
                    new_content = content

                a = page.add_freetext_annot(rect, new_content)
                if da:
                    doc.xref_set_key(a.xref, 'DA', f'({da})')
                a.update()

            elif atype == 4:  # Square
                a = page.add_rect_annot(rect)
                stroke = colors.get('stroke', [1, 0, 0])
                if stroke:
                    a.set_colors(stroke=stroke)
                fill = colors.get('fill', [])
                if fill:
                    a.set_colors(fill=fill)
                bw = border.get('width', 0.8) if border else 0.8
                a.set_border(width=bw)
                if ann['content']:
                    a.set_info(content=ann['content'])
                a.update()

    save_pdf_gdrive(doc, out_path)
    doc.close()
    return out_path


def col_d_values(section):
    """Return list of 10 col D values: [parent, sub1..sub9]."""
    # CAT filename without .pdf
    cat = f"{section}. {PRODUCT_NAME}"

    vals = [
        # parent
        "ยี่ห้อ Ruijie รุ่น RG-CS85-24GT8XS-D",
        # sub1 - Layer 3
        "ยินดีปฏิบัติตามข้อกำหนด",
        # sub2 - RIPv2 OSPF
        "ยินดีปฏิบัติตามข้อกำหนด",
        # sub3 - 24x 10/100/1000
        f"เทียบเท่าข้อกำหนด เอกสาร {cat} หน้า 5 ข้อ {section} ข้อย่อย 3.",
        # sub4 - SFP/SFP+
        "ยินดีปฏิบัติตามข้อกำหนด",
        # sub5 - LED
        "ยินดีปฏิบัติตามข้อกำหนด",
        # sub6 - 32K MAC
        f"สูงกว่าข้อกำหนด เอกสาร {cat} หน้า 6 ข้อ {section} ข้อย่อย 6.",
        # sub7 - Web Browser
        "ยินดีปฏิบัติตามข้อกำหนด",
        # sub8 - Syslog
        "ยินดีปฏิบัติตามข้อกำหนด",
        # sub9 - IPv6
        "ยินดีปฏิบัติตามข้อกำหนด",
    ]
    return vals


def update_xlsx(wb_path):
    """Update xlsx for all 4 sections (5.2.1.6, 5.2.2.2, 5.2.3.2, 5.2.4.2)."""
    wb = openpyxl.load_workbook(wb_path)
    ws = wb['Trio_SR_Solution']

    all_sections = [SECTION_516] + SISTERS
    total_rows = 0

    for sec_info in all_sections:
        section = sec_info['section']
        parent_row = sec_info['parent_row']
        sub_rows = sec_info['sub_rows']

        d_vals = col_d_values(section)

        # Parent row
        ws.cell(row=parent_row, column=4).value = d_vals[0]
        ws.cell(row=parent_row, column=5).value = "TRIO"
        ws.cell(row=parent_row, column=6).value = "รอ user ตรวจสอบ"
        total_rows += 1

        # Sub rows 1-9
        for i, row_num in enumerate(sub_rows):
            ws.cell(row=row_num, column=4).value = d_vals[i + 1]
            ws.cell(row=row_num, column=5).value = "TRIO"
            ws.cell(row=row_num, column=6).value = "รอ user ตรวจสอบ"
            total_rows += 1

    # Save with O_TRUNC workaround
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    wb.save(tmp.name)
    with open(tmp.name, 'rb') as src:
        data = src.read()
    fd = os.open(wb_path, os.O_WRONLY | os.O_TRUNC)
    os.write(fd, data)
    os.close(fd)
    os.unlink(tmp.name)

    return total_rows


def main():
    print("Collecting annotations from reference PDF...")
    ref_annots = collect_annotations(REF_PDF)
    print(f"  Found annotations on {len([p for p in ref_annots if p])} pages")

    files_created = 0
    for sister in SISTERS:
        section = sister['section']
        print(f"\nCloning for {section}...")
        out_path = clone_pdf(sister, ref_annots)
        print(f"  Created: {os.path.basename(out_path)}")
        files_created += 1

    print("\nUpdating xlsx...")
    total_rows = update_xlsx(XLSX_PATH)
    print(f"  Updated {total_rows} rows across 4 sections")

    print(f"\nDone. Files created: {files_created}, xlsx rows updated: {total_rows}")


if __name__ == '__main__':
    main()
