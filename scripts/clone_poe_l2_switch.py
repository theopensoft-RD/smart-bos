#!/usr/bin/env python3
"""
Clone PoE L2 Switch reference PDF to 5 sister sections + update xlsx.
"""
import os
import shutil

import fitz
import openpyxl

fitz.TOOLS.mupdf_display_errors(False)

BASE = "/Users/ppxndpxdd/Library/CloudStorage/GoogleDrive-smartsolution.in.th@gmail.com/My Drive/Smart Solution/Project/Pattaya Project/Smart Plant 1/co-work/claude-code"

SOURCE_CATALOG = os.path.join(BASE, "catalog/left/อุปกรณ์กระจายสัญญาณแบบ PoE (PoE L2 Switch) ขนาด 16 ช่อง -CS4220-16GT-240_V2_datasheet_20250429 ICT.pdf")

REF_PDF = os.path.join(BASE, "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์/5.2.1. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารสำนักงาน/5.2.1.-8/5.2.1.8. อุปกรณ์กระจายสัญญาณแบบ PoE (PoE L2 Switch) ขนาด 16 ช่อง Ruijie CS4220-16GT-240.pdf")

REF_SECTION = "5.2.1.8"

XLSX_PATH = os.path.join(BASE, "output/Comply spec Smart Plant 1.xlsx")

FILENAME_BASE = "อุปกรณ์กระจายสัญญาณแบบ PoE (PoE L2 Switch) ขนาด 16 ช่อง Ruijie CS4220-16GT-240.pdf"

SISTERS = [
    {
        "section": "5.2.2.3",
        "parent_dir": os.path.join(BASE, "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์/5.2.2. งานระบบเครือข่ายคอมพิวเตอร์ ภายในบ้านพักข้าราชการ/5.2.2.-3"),
    },
    {
        "section": "5.2.3.3",
        "parent_dir": os.path.join(BASE, "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์/5.2.3. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารบำบัดน้ำเสีย/5.2.3.-3"),
    },
    {
        "section": "5.2.4.3",
        "parent_dir": os.path.join(BASE, "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์/5.2.4. งานระบบเครือข่ายคอมพิวเตอร์ ภายในอาคารซ่อมบำรุง/5.2.4.-3"),
    },
    {
        "section": "5.2.5.5",
        "parent_dir": os.path.join(BASE, "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์/5.2.5. งานระบบเครือข่ายคอมพิวเตอร์ ภายในบ่อสัมผัสคลอรีน/5.2.5.-5"),
    },
    {
        "section": "5.2.6.5",
        "parent_dir": os.path.join(BASE, "output/5.2 งานระบบเครือข่ายคอมพิวเตอร์/5.2.6. งานระบบเครือข่ายคอมพิวเตอร์ ป้อม รปภ. และภายนอกอาคาร/5.2.6.-5"),
    },
]

# xlsx row map: section -> (parent_row, [sub_rows])
XLSX_ROWS = {
    "5.2.1.8": (341, list(range(342, 349))),
    "5.2.2.3": (400, list(range(401, 408))),
    "5.2.3.3": (463, list(range(464, 471))),
    "5.2.4.3": (526, list(range(527, 534))),
    "5.2.5.5": (586, list(range(587, 594))),
    "5.2.6.5": (632, list(range(633, 640))),
}

# ── PDF CLONING ──────────────────────────────────────────────────────────────

def read_ref_annotations(ref_path):
    """Read all annotations from reference PDF, grouped by page."""
    doc = fitz.open(ref_path)
    pages_annots = []
    for pno in range(len(doc)):
        page = doc[pno]
        annots_data = []
        for a in page.annots():
            annots_data.append({
                "type": a.type,
                "rect": tuple(a.rect),
                "content": a.info.get("content", ""),
                "colors": a.colors,
                "border": a.border,
                "flags": a.flags,
                "opacity": a.opacity,
            })
        pages_annots.append(annots_data)
    doc.close()
    return pages_annots


def make_sister_pdf(sister_section, out_dir, ref_annots):
    """Create sister PDF: catalog copy + header + cloned annotations with section substitution."""
    os.makedirs(out_dir, exist_ok=True)

    out_filename = f"{sister_section}. {FILENAME_BASE}"
    out_path = os.path.join(out_dir, out_filename)

    # Open source catalog
    doc = fitz.open(SOURCE_CATALOG)

    header_text_base = f"{sister_section}. {FILENAME_BASE[:-4]}"  # strip .pdf

    for pno in range(len(doc)):
        page = doc[pno]
        page_label = f"{header_text_base} หน้า {pno + 1}"

        # Add header FreeText annotation (14pt red centered)
        header_rect = fitz.Rect(15, 10, 597, 50)
        header_annot = page.add_freetext_annot(
            header_rect,
            page_label,
            fontsize=14,
            fontname="Helvetica",
            text_color=(1, 0, 0),
            fill_color=None,
            align=fitz.TEXT_ALIGN_CENTER,
        )
        header_annot.set_border(width=0)
        header_annot.update()

        # Clone annotations from reference for this page (skip the header FreeText which we just added)
        if pno < len(ref_annots):
            for ad in ref_annots[pno]:
                # Skip header FreeText (rect matches 15,10,597,50 and content starts with section)
                r = ad["rect"]
                if (ad["type"][0] == 2 and
                        abs(r[0]-15) < 1 and abs(r[1]-10) < 1 and
                        abs(r[2]-597) < 1 and abs(r[3]-50) < 1):
                    continue  # already added header above

                content = ad["content"]
                # Replace section references
                content = content.replace(REF_SECTION, sister_section)

                annot_type = ad["type"][0]
                rect = fitz.Rect(ad["rect"])
                colors = ad["colors"]
                stroke_color = colors.get("stroke") or None
                fill_color = colors.get("fill") or None

                if annot_type == 2:  # FreeText
                    a = page.add_freetext_annot(
                        rect,
                        content,
                        fontsize=9,
                        fontname="Helvetica",
                        text_color=(0, 0, 0),
                        fill_color=None,
                        align=fitz.TEXT_ALIGN_LEFT,
                    )
                    border = ad.get("border", {})
                    bw = border.get("width", 0) if border else 0
                    if bw is None or bw < 0:
                        bw = 0
                    a.set_border(width=bw)
                    a.update()

                elif annot_type == 4:  # Square
                    a = page.add_rect_annot(rect)
                    if stroke_color and len(stroke_color) == 3:
                        a.set_colors(stroke=stroke_color)
                    border = ad.get("border", {})
                    bw = border.get("width", 0.8) if border else 0.8
                    if bw is None or bw < 0:
                        bw = 0.8
                    a.set_border(width=bw)
                    if content:
                        a.set_info(content=content)
                    a.update()

                elif annot_type == 8:  # Highlight
                    quads = fitz.Rect(ad["rect"]).quad
                    a = page.add_highlight_annot(quads)
                    if stroke_color and len(stroke_color) == 3:
                        a.set_colors(stroke=stroke_color)
                    a.update()

    # Save with os.O_TRUNC workaround
    tmp_path = out_path + ".tmp"
    doc.save(tmp_path, garbage=4, clean=True, deflate=True)
    doc.close()

    # Move tmp to final (handles Google Drive / extended attribute issues)
    if os.path.exists(out_path):
        os.remove(out_path)
    shutil.move(tmp_path, out_path)
    print(f"  Created: {out_path}")
    return out_path


# ── XLSX UPDATE ───────────────────────────────────────────────────────────────

def update_xlsx(xlsx_path, all_sections):
    """Update Col D, E, F for all 6 sections."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Trio_SR_Solution"]

    for section, (parent_row, sub_rows) in XLSX_ROWS.items():
        cat_name = f"{section} อุปกรณ์กระจายสัญญาณแบบ PoE (PoE L2 Switch) ขนาด 16 ช่อง Ruijie CS4220-16GT-240"

        # Parent row
        ws.cell(parent_row, 4).value = "ยี่ห้อ Ruijie รุ่น CS4220-16GT-240"
        ws.cell(parent_row, 5).value = "TRIO"
        ws.cell(parent_row, 6).value = "รอ user ตรวจสอบ"

        # Sub-rows
        sub_templates = [
            f"เทียบเท่าข้อกำหนด เอกสาร {cat_name} หน้า 2 ข้อ {section} ข้อย่อย 1.",
            f"เทียบเท่าข้อกำหนด เอกสาร {cat_name} หน้า 2 ข้อ {section} ข้อย่อย 2.",
            f"เทียบเท่าข้อกำหนด เอกสาร {cat_name} หน้า 2 ข้อ {section} ข้อย่อย 3.",
            f"เทียบเท่าข้อกำหนด เอกสาร {cat_name} หน้า 2 ข้อ {section} ข้อย่อย 4.",
            f"เทียบเท่าข้อกำหนด เอกสาร {cat_name} หน้า 2 ข้อ {section} ข้อย่อย 5.",
            f"เทียบเท่าข้อกำหนด เอกสาร {cat_name} หน้า 2 ข้อ {section} ข้อย่อย 6.",
            "ยินดีปฏิบัติตามข้อกำหนด",
        ]

        for i, row in enumerate(sub_rows):
            ws.cell(row, 4).value = sub_templates[i]
            ws.cell(row, 5).value = "TRIO"
            ws.cell(row, 6).value = "รอ user ตรวจสอบ"

        print(f"  Updated xlsx rows for {section}: parent R{parent_row}, subs R{sub_rows[0]}-R{sub_rows[-1]}")

    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=== Reading reference annotations ===")
    ref_annots = read_ref_annotations(REF_PDF)
    print(f"  Pages: {len(ref_annots)}, annots per page: {[len(p) for p in ref_annots]}")

    print("\n=== Creating sister PDFs ===")
    created_pdfs = []
    for s in SISTERS:
        print(f"  Processing {s['section']}...")
        path = make_sister_pdf(s["section"], s["parent_dir"], ref_annots)
        created_pdfs.append(path)

    print("\n=== Updating xlsx ===")
    update_xlsx(XLSX_PATH, XLSX_ROWS)

    print("\n=== Done ===")
    print(f"Created {len(created_pdfs)} PDFs, updated xlsx with 6 sections (48 rows total).")
